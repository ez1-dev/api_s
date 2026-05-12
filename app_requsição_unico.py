from fastapi import FastAPI, HTTPException, Depends, Body, Query
from fastapi.responses import HTMLResponse, StreamingResponse, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
import pyodbc
from jose import jwt
from datetime import datetime, timedelta
from typing import Optional, Any, List
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter
import io
import csv
import re

# =========================================
# CONFIGURAÇÕES
# =========================================

SECRET_KEY = "ERP_SECRET"
ALGORITHM = "HS256"

SQL_SERVER   = "172.16.137.100"
SQL_DATABASE = "sapiens"
SQL_USER     = "sapiens"
SQL_PASSWORD = "0n%lV'g0F94"
EMPRESA_PADRAO = 1

USERS = {
    "ADMIN":  "123",
    "RENATO": "123"
}

AUDITORIA_EXPORT_BATCH_SIZE = 5000
AUDITORIA_EXPORT_MAX_PAGES  = 5000
AUDITORIA_EXCEL_MAX_ROWS    = 100000

app = FastAPI(title="ERP Estoque, Onde Usa, BOM, Compras, Produção e Requisição")
security          = HTTPBearer()
optional_security = HTTPBearer(auto_error=False)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================================
# CONEXÃO BANCO
# =========================================

def get_connection():
    try:
        conn = pyodbc.connect(
            f"DRIVER={{SQL Server}};"
            f"SERVER={SQL_SERVER};"
            f"DATABASE={SQL_DATABASE};"
            f"UID={SQL_USER};"
            f"PWD={SQL_PASSWORD};",
            timeout=10
        )
        return conn
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def normalizar_situacao_cadastro(valor: Optional[str]) -> str:
    valor_norm = (valor or 'TODOS').strip().upper()
    if valor_norm in ('A', 'ATIVO', 'ATIVOS'):   return 'ATIVO'
    if valor_norm in ('I', 'INATIVO', 'INATIVOS'): return 'INATIVO'
    return 'TODOS'


def clausula_sql_situacao_cadastro(alias: str, situacao_cadastro: Optional[str],
                                   nulo_como_ativo: bool = True) -> str:
    situacao_norm = normalizar_situacao_cadastro(situacao_cadastro)
    default_val = 'A' if nulo_como_ativo else ''
    if situacao_norm == 'ATIVO':
        return f" AND COALESCE({alias}.SITPRO, '{default_val}') = 'A'"
    if situacao_norm == 'INATIVO':
        return f" AND COALESCE({alias}.SITPRO, '{default_val}') = 'I'"
    return ''

# =========================================
# TOKEN / AUTH
# =========================================

def gerar_token(usuario):
    payload = {
        "sub": usuario,
        "exp": datetime.now().astimezone(__import__("datetime").timezone.utc) + timedelta(hours=8)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def validar_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload["sub"]
    except Exception:
        raise HTTPException(status_code=401, detail="Token inválido")


def validar_token_download(
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(optional_security),
    access_token: Optional[str] = None
):
    token = access_token or (credentials.credentials if credentials else None)
    if not token:
        raise HTTPException(status_code=401, detail="Token não informado")
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload["sub"]
    except Exception:
        raise HTTPException(status_code=401, detail="Token inválido")


class LoginRequest(BaseModel):
    usuario: str
    senha: str

# =========================================
# LOGIN / HEALTH
# =========================================

@app.post("/login")
def login(
    payload: Optional[LoginRequest] = Body(default=None),
    usuario: Optional[str] = Query(default=None),
    senha:   Optional[str] = Query(default=None),
):
    usuario_final = ""
    senha_final   = ""
    if payload:
        usuario_final = (payload.usuario or "").upper().strip()
        senha_final   = (payload.senha   or "").strip()
    else:
        usuario_final = (usuario or "").upper().strip()
        senha_final   = (senha   or "").strip()

    if not usuario_final or not senha_final:
        raise HTTPException(status_code=400, detail="Usuário e senha são obrigatórios")

    if usuario_final in USERS and USERS[usuario_final] == senha_final:
        return {
            "access_token": gerar_token(usuario_final),
            "token_type":   "bearer",
            "usuario":      usuario_final
        }
    raise HTTPException(status_code=401, detail="Login inválido")


@app.get("/health")
def health():
    return {"status": "ok", "app": "ERP Web"}


# =========================================
# FAMÍLIAS / ORIGENS
# =========================================

@app.get("/api/familias")
def listar_familias(q: Optional[str] = None, limite: int = 200,
                    usuario=Depends(validar_token)):
    limite = max(1, min(limite, 500))
    conn = get_connection(); cursor = conn.cursor()
    where_sql = """
        FROM E012FAM F WHERE F.CODEMP = ?
        AND EXISTS (SELECT 1 FROM E075PRO P WHERE P.CODEMP=F.CODEMP AND P.CODFAM=F.CODFAM)
    """
    params = [EMPRESA_PADRAO]
    if q:
        where_sql += " AND (F.CODFAM LIKE ? OR F.DESFAM LIKE ?)"
        params.extend([f"%{q.strip()}%", f"%{q.strip()}%"])
    cursor.execute(f"SELECT TOP {limite} F.CODFAM AS codigo, F.DESFAM AS descricao "
                   f"{where_sql} ORDER BY F.CODFAM", params)
    rows = cursor.fetchall(); conn.close()
    resultado = []
    for row in rows:
        codigo    = (row[0] or '').strip()
        descricao = (row[1] or '').strip()
        label     = f"{codigo} - {descricao}" if descricao else codigo
        resultado.append({"codigo": codigo, "descricao": descricao,
                           "value": codigo, "label": label})
    return resultado


@app.get("/api/origens")
def listar_origens(q: Optional[str] = None, limite: int = 200,
                   usuario=Depends(validar_token)):
    limite = max(1, min(limite, 500))
    conn = get_connection(); cursor = conn.cursor()
    where_sql = """
        FROM E083ORI O WHERE O.CODEMP = ?
        AND EXISTS (SELECT 1 FROM E075PRO P WHERE P.CODEMP=O.CODEMP AND P.CODORI=O.CODORI)
    """
    params = [EMPRESA_PADRAO]
    if q:
        where_sql += " AND (O.CODORI LIKE ? OR O.DESORI LIKE ?)"
        params.extend([f"%{q.strip()}%", f"%{q.strip()}%"])
    cursor.execute(f"SELECT TOP {limite} O.CODORI AS codigo, O.DESORI AS descricao "
                   f"{where_sql} ORDER BY O.CODORI", params)
    rows = cursor.fetchall(); conn.close()
    resultado = []
    for row in rows:
        codigo    = (row[0] or '').strip()
        descricao = (row[1] or '').strip()
        label     = f"{codigo} - {descricao}" if descricao else codigo
        resultado.append({"codigo": codigo, "descricao": descricao,
                           "value": codigo, "label": label})
    return resultado


# =========================================
# ESTOQUE
# =========================================

@app.get("/api/estoque")
def consultar_estoque(
    codpro:            Optional[str] = None,
    despro:            Optional[str] = None,
    codfam:            Optional[str] = None,
    codori:            Optional[str] = None,
    coddep:            Optional[str] = None,
    situacao_cadastro: str  = 'TODOS',
    somente_com_estoque: bool = True,
    pagina:            int  = 1,
    tamanho_pagina:    int  = 100,
    usuario=Depends(validar_token)
):
    pagina         = max(pagina, 1)
    tamanho_pagina = min(max(tamanho_pagina, 1), 100)
    offset         = (pagina - 1) * tamanho_pagina
    conn = get_connection(); cursor = conn.cursor()
    where_sql = """
        FROM E210EST E
        INNER JOIN E075PRO P ON P.CODEMP=E.CODEMP AND P.CODPRO=E.CODPRO
        WHERE E.CODEMP = ?
    """
    params = [EMPRESA_PADRAO]
    if codpro: where_sql += " AND E.CODPRO LIKE ?";  params.append(f"%{codpro}%")
    if despro: where_sql += " AND P.DESPRO LIKE ?";  params.append(f"%{despro}%")
    if codfam: where_sql += " AND P.CODFAM = ?";     params.append(codfam)
    if codori: where_sql += " AND P.CODORI = ?";     params.append(codori)
    if coddep: where_sql += " AND E.CODDEP = ?";     params.append(coddep)
    where_sql += clausula_sql_situacao_cadastro('P', situacao_cadastro)
    if somente_com_estoque: where_sql += " AND E.QTDEST > 0"

    cursor.execute(f"SELECT COUNT(*) {where_sql}", params)
    total_registros = cursor.fetchone()[0]

    cursor.execute(f"""
        SELECT E.CODPRO AS codigo, P.DESPRO AS descricao,
               P.CODFAM AS familia, P.CODORI AS origem,
               P.TIPPRO AS tipo,   P.UNIMED AS unidade_medida,
               E.CODDER AS derivacao, E.CODDEP AS deposito,
               CAST(E.QTDEST AS FLOAT) AS saldo
        {where_sql}
        ORDER BY E.CODPRO, E.CODDER, E.CODDEP
        OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
    """, params + [offset, tamanho_pagina])
    rows    = cursor.fetchall()
    columns = [col[0] for col in cursor.description]
    conn.close()

    resultado = []
    for row in rows:
        item = {col: row[i] for i, col in enumerate(columns)}
        tipo = (item.get("tipo") or "").strip().upper()
        item["tipo_descricao"] = ("Produzido" if tipo == "P" else
                                  "Comprado"  if tipo == "C" else
                                  item.get("tipo") or "Não informado")
        item["origem"]         = (item.get("origem") or "").strip()
        item["unidade_medida"] = (item.get("unidade_medida") or "").strip()
        resultado.append(item)

    total_paginas = (total_registros + tamanho_pagina - 1) // tamanho_pagina if total_registros > 0 else 1
    return {"pagina": pagina, "tamanho_pagina": tamanho_pagina,
            "total_registros": total_registros, "total_paginas": total_paginas,
            "dados": resultado}


# =========================================
# ONDE USA
# =========================================

@app.get("/api/onde-usa")
def consultar_onde_usa(
    codcmp:            Optional[str] = None,
    dercmp:            Optional[str] = None,
    codmod:            Optional[str] = None,
    situacao_cadastro: str = 'TODOS',
    pagina:            int = 1,
    tamanho_pagina:    int = 100,
    usuario=Depends(validar_token)
):
    pagina         = max(pagina, 1)
    tamanho_pagina = min(max(tamanho_pagina, 1), 100)
    offset         = (pagina - 1) * tamanho_pagina
    conn = get_connection(); cursor = conn.cursor()

    where_sql = """
        FROM E700CTM CT
        INNER JOIN E700MOD M ON M.CODEMP=CT.CODEMP AND M.CODMOD=CT.CODMOD
        LEFT  JOIN E075PRO PC ON PC.CODEMP=CT.CODEMP AND PC.CODPRO=CT.CODCMP
        WHERE CT.CODEMP = ?
    """
    params = [EMPRESA_PADRAO]
    if codcmp: where_sql += " AND CT.CODCMP LIKE ?"; params.append(f"%{codcmp}%")
    if dercmp: where_sql += " AND COALESCE(CT.DERCMP,'') LIKE ?"; params.append(f"%{dercmp}%")
    if codmod: where_sql += " AND CT.CODMOD LIKE ?"; params.append(f"%{codmod}%")
    where_sql += clausula_sql_situacao_cadastro('PC', situacao_cadastro)

    cursor.execute(f"SELECT COUNT(*) {where_sql}", params)
    total_registros = cursor.fetchone()[0]

    cursor.execute(f"""
        SELECT CT.CODCMP AS codigo_componente,
               COALESCE(PC.DESPRO,'') AS descricao_componente,
               COALESCE(PC.UNIMED,CT.UNIME2,'') AS unidade_componente,
               COALESCE(CT.DERCMP,'') AS derivacao_componente,
               CT.CODMOD AS codigo_modelo, M.DESMOD AS descricao_modelo,
               CT.CODDER AS derivacao_modelo, CT.CODETG AS estagio,
               CT.SEQMOD AS sequencia,
               CAST(CT.QTDUTI AS FLOAT) AS quantidade_utilizada,
               COALESCE(CT.UNIME2,'') AS unidade_utilizada,
               CAST(COALESCE(CT.PERPRD,0) AS FLOAT) AS perda_percentual,
               CAST(COALESCE(CT.PRDQTD,0) AS FLOAT) AS quantidade_perda
        {where_sql}
        ORDER BY CT.CODCMP, CT.CODMOD, CT.CODDER, CT.CODETG, CT.SEQMOD
        OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
    """, params + [offset, tamanho_pagina])
    rows    = cursor.fetchall()
    columns = [col[0] for col in cursor.description]
    conn.close()

    resultado = []
    for row in rows:
        item = {col: (row[i].strip() if isinstance(row[i], str) else row[i])
                for i, col in enumerate(columns)}
        resultado.append(item)

    total_paginas = (total_registros + tamanho_pagina - 1) // tamanho_pagina if total_registros > 0 else 1
    return {"pagina": pagina, "tamanho_pagina": tamanho_pagina,
            "total_registros": total_registros, "total_paginas": total_paginas,
            "dados": resultado}


# =========================================
# BOM — helpers
# =========================================

def _fetch_children_bom(cursor, codemp, codmod, codder, situacao_cadastro='TODOS'):
    sql = """
        SELECT CT.CODMOD AS modelo_pai,
               COALESCE(M.DESMOD,'') AS descricao_modelo_pai,
               COALESCE(CT.CODDER,'') AS derivacao_modelo,
               CT.CODETG AS estagio, CT.SEQMOD AS sequencia,
               CT.CODCMP AS codigo_componente,
               COALESCE(CT.DERCMP,'') AS derivacao_componente,
               COALESCE(P.DESPRO,'') AS descricao_componente,
               COALESCE(P.UNIMED,'') AS unidade_medida,
               COALESCE(P.TIPPRO,'') AS tipo_produto,
               CAST(CT.QTDUTI AS FLOAT) AS quantidade_utilizada,
               CAST(COALESCE(CT.PERPRD,0) AS FLOAT) AS perda_percentual,
               CAST(COALESCE(CT.PRDQTD,0) AS FLOAT) AS quantidade_perda
        FROM E700CTM CT
        LEFT JOIN E700MOD M ON M.CODEMP=CT.CODEMP AND M.CODMOD=CT.CODMOD
        LEFT JOIN E075PRO P  ON P.CODEMP=CT.CODEMP AND P.CODPRO=CT.CODCMP
        WHERE CT.CODEMP=? AND CT.CODMOD=?
    """
    params = [codemp, codmod]
    if codder:
        sql += " AND COALESCE(CT.CODDER,'')=?"
        params.append(codder)
    sql += clausula_sql_situacao_cadastro('P', situacao_cadastro)
    sql += " ORDER BY CT.CODETG, CT.SEQMOD, CT.CODCMP, CT.DERCMP"
    cursor.execute(sql, params)
    rows = cursor.fetchall()
    cols = [col[0] for col in cursor.description]
    result = []
    for row in rows:
        item = {col: (row[i].strip() if isinstance(row[i], str) else row[i])
                for i, col in enumerate(cols)}
        result.append(item)
    return result


def _model_exists(cursor, codemp, codmod, cache):
    key = (codemp, (codmod or '').strip())
    if key in cache: return cache[key]
    cursor.execute("SELECT TOP 1 1 FROM E700MOD WHERE CODEMP=? AND CODMOD=?",
                   [codemp, key[1]])
    exists      = cursor.fetchone() is not None
    cache[key]  = exists
    return exists


def _model_description(cursor, codemp, codmod, cache):
    key = (codemp, (codmod or '').strip())
    if key in cache: return cache[key]
    cursor.execute("SELECT TOP 1 COALESCE(DESMOD,'') FROM E700MOD WHERE CODEMP=? AND CODMOD=?",
                   [codemp, key[1]])
    row         = cursor.fetchone()
    desc        = row[0].strip() if row and isinstance(row[0], str) else (row[0] or '') if row else ''
    cache[key]  = desc
    return desc


def _build_bom_rows(cursor, codemp, codmod, codder, max_nivel, situacao_cadastro='TODOS'):
    rows, exists_cache, desc_cache, max_rows = [], {}, {}, 5000

    def walk(modelo_atual, derivacao_atual, nivel, caminho):
        if nivel > max_nivel or len(rows) >= max_rows: return
        filhos = _fetch_children_bom(cursor, codemp, modelo_atual, derivacao_atual, situacao_cadastro)
        for item in filhos:
            comp     = (item.get('codigo_componente') or '').strip()
            der_comp = (item.get('derivacao_componente') or '').strip()
            path_key = f"{comp}|{der_comp}"
            ciclo    = path_key in caminho
            possui_filhos = _model_exists(cursor, codemp, comp, exists_cache) if comp and not ciclo else False
            item['nivel']              = nivel
            item['caminho']            = ' > '.join([x.split('|')[0] for x in caminho] + [comp])
            item['possui_filhos']      = possui_filhos
            item['ciclo_detectado']    = ciclo
            item['descricao_modelo_pai'] = item.get('descricao_modelo_pai') or \
                                           _model_description(cursor, codemp, modelo_atual, desc_cache)
            rows.append(item)
            if possui_filhos and not ciclo and nivel < max_nivel and len(rows) < max_rows:
                walk(comp, der_comp or None, nivel + 1, list(caminho) + [path_key])

    walk(codmod, codder or None, 1, [f"{codmod}|{codder or ''}"])
    return rows


@app.get('/api/bom')
def consultar_bom(
    codmod:            str,
    codder:            Optional[str] = None,
    situacao_cadastro: str  = 'TODOS',
    max_nivel:         int  = 10,
    usuario=Depends(validar_token)
):
    codmod = (codmod or '').strip()
    codder = (codder or '').strip() or None
    if not codmod:
        raise HTTPException(status_code=400, detail='Informe o código do modelo.')
    max_nivel = min(max(max_nivel, 1), 15)
    situacao_cadastro = normalizar_situacao_cadastro(situacao_cadastro)

    conn = get_connection(); cursor = conn.cursor()
    if not _model_exists(cursor, EMPRESA_PADRAO, codmod, {}):
        conn.close()
        raise HTTPException(status_code=404, detail='Modelo não encontrado.')

    descricao_raiz = _model_description(cursor, EMPRESA_PADRAO, codmod, {})
    cursor.execute("SELECT TOP 1 COALESCE(UNIMED,'') FROM E075PRO WHERE CODEMP=? AND CODPRO=?",
                   [EMPRESA_PADRAO, codmod])
    row_um     = cursor.fetchone()
    unidade_r  = row_um[0].strip() if row_um and isinstance(row_um[0], str) else ''
    dados      = _build_bom_rows(cursor, EMPRESA_PADRAO, codmod, codder, max_nivel, situacao_cadastro)
    conn.close()

    return {
        'cabecalho': {
            'codigo_modelo':    codmod, 'descricao_modelo': descricao_raiz,
            'derivacao_modelo': codder or '', 'unidade_modelo': unidade_r,
            'max_nivel': max_nivel,
        },
        'total_itens':          len(dados),
        'total_niveis':         max([item.get('nivel', 0) for item in dados], default=0),
        'total_modelos_filhos': len({item.get('codigo_componente') for item in dados if item.get('possui_filhos')}),
        'dados': dados,
    }


# =========================================
# COMPRAS / CUSTOS
# =========================================

@app.get('/api/compras-produto')
def consultar_compras_produto(
    codpro:            Optional[str]   = None,
    despro:            Optional[str]   = None,
    codfam:            Optional[str]   = None,
    codori:            Optional[str]   = None,
    codder:            Optional[str]   = None,
    numero_projeto:    Optional[str]   = None,
    centro_custo:      Optional[str]   = None,
    situacao_cadastro: str             = 'TODOS',
    somente_com_oc_aberta: bool        = False,
    pagina:            int             = 1,
    tamanho_pagina:    int             = 100,
    usuario=Depends(validar_token)
):
    pagina         = max(pagina, 1)
    tamanho_pagina = min(max(tamanho_pagina, 1), 100)
    offset         = (pagina - 1) * tamanho_pagina
    conn           = get_connection(); cursor = conn.cursor()

    from_count = """
        FROM E075PRO P
        LEFT JOIN E075DER D ON D.CODEMP=P.CODEMP AND D.CODPRO=P.CODPRO
        LEFT JOIN (
            SELECT E.CODEMP, E.CODPRO, COALESCE(E.CODDER,'') AS CODDER,
                   SUM(CAST(E.QTDEST AS FLOAT)) AS SALDO_TOTAL
            FROM E210EST E GROUP BY E.CODEMP, E.CODPRO, COALESCE(E.CODDER,'')
        ) EST ON EST.CODEMP=P.CODEMP AND EST.CODPRO=P.CODPRO AND EST.CODDER=COALESCE(D.CODDER,'')
        WHERE P.CODEMP = ?
    """
    params = [EMPRESA_PADRAO]
    if codpro: from_count += " AND P.CODPRO LIKE ?"; params.append(f"%{codpro}%")
    if despro: from_count += " AND P.DESPRO LIKE ?"; params.append(f"%{despro}%")
    if codfam: from_count += " AND P.CODFAM = ?";   params.append(codfam)
    if codori: from_count += " AND P.CODORI = ?";   params.append(codori)
    if codder: from_count += " AND COALESCE(D.CODDER,'') LIKE ?"; params.append(f"%{codder}%")
    from_count += clausula_sql_situacao_cadastro('P', situacao_cadastro)
    if somente_com_oc_aberta:
        from_count += """
            AND EXISTS (SELECT 1 FROM E420IPO I WHERE I.CODEMP=P.CODEMP AND I.CODPRO=P.CODPRO
                        AND COALESCE(I.CODDER,'')=COALESCE(D.CODDER,'') AND COALESCE(I.QTDABE,0)>0)
        """

    cursor.execute(f"SELECT COUNT(*) {from_count}", params)
    total_registros = cursor.fetchone()[0]

    cursor.execute(f"""
        SELECT P.CODPRO AS codigo, P.DESPRO AS descricao,
               P.CODFAM AS familia, P.CODORI AS origem,
               P.TIPPRO AS tipo,   P.UNIMED AS unidade_medida,
               COALESCE(D.CODDER,'') AS derivacao,
               CAST(COALESCE(EST.SALDO_TOTAL,0) AS FLOAT) AS saldo_total,
               CAST(COALESCE(D.PREMED,0) AS FLOAT) AS preco_medio,
               CAST(COALESCE(D.PRECUS,0) AS FLOAT) AS custo_calculado,
               D.DATCUS AS data_custo_calculado,
               COALESCE(D.HORCUS,0) AS hora_custo_calculado,
               COALESCE(D.ORICUS,'') AS origem_custo,
               COALESCE(D.INFCUS,'') AS informacao_custo,
               CAST(COALESCE(D.QTDCUS,0) AS FLOAT) AS quantidade_base_custo,
               CAST(COALESCE(D.CUSSAL,0) AS FLOAT) AS custo_salarial,
               CAST(COALESCE(D.CUSENC,0) AS FLOAT) AS custo_encargos,
               CAST(COALESCE(D.PREUEN,0) AS FLOAT) AS preco_ultima_entrada_cadastro,
               D.DATUEN AS data_ultima_entrada_cadastro,
               COALESCE(ULTNF.NUMERO_NF_ULTIMA_COMPRA,0) AS numero_nf_ultima_compra,
               COALESCE(ULTNF.SERIE_NF_ULTIMA_COMPRA,'') AS serie_nf_ultima_compra,
               ULTNF.DATA_ENTRADA_NF_ULTIMA_COMPRA AS data_entrada_nf_ultima_compra,
               COALESCE(ULTNF.FORNECEDOR_ULTIMA_COMPRA,'') AS fornecedor_ultima_compra,
               CAST(COALESCE(ULTNF.PRECO_NF_ULTIMA_COMPRA,0) AS FLOAT) AS preco_nf_ultima_compra,
               CASE WHEN COALESCE(OCA.QTD_OC_ABERTA_TOTAL,0)>0 THEN 'Sim' ELSE 'Não' END AS possui_oc_aberta,
               CAST(COALESCE(OCA.QTD_OC_ABERTA_TOTAL,0) AS FLOAT) AS quantidade_oc_aberta,
               COALESCE(OCA.QTDE_OCS_ABERTAS,0) AS quantidade_ocs_abertas,
               COALESCE(ULTPO.ULTIMA_OC_ABERTA,0) AS ultima_oc_aberta,
               CAST(COALESCE(ULTPO.ULTIMA_OC_ABERTA, ULTNF.NUMERO_OC_ORIGEM_ULTIMA_COMPRA,0) AS INT) AS numero_oc_ultima
        {from_count}
        OUTER APPLY (
            SELECT TOP 1 H.NUMNFC AS NUMERO_NF_ULTIMA_COMPRA, H.CODSNF AS SERIE_NF_ULTIMA_COMPRA,
                         H.DATENT AS DATA_ENTRADA_NF_ULTIMA_COMPRA,
                         COALESCE(F.APEFOR,F.NOMFOR,'') AS FORNECEDOR_ULTIMA_COMPRA,
                         CAST(I.PREUNI AS FLOAT) AS PRECO_NF_ULTIMA_COMPRA,
                         I.NUMOCP AS NUMERO_OC_ORIGEM_ULTIMA_COMPRA
            FROM E440IPC I INNER JOIN E440NFC H ON H.CODEMP=I.CODEMP AND H.CODFIL=I.CODFIL
                 AND H.CODFOR=I.CODFOR AND H.NUMNFC=I.NUMNFC AND H.CODSNF=I.CODSNF
            LEFT JOIN E095FOR F ON F.CODFOR=H.CODFOR
            WHERE I.CODEMP=P.CODEMP AND I.CODPRO=P.CODPRO
                  AND COALESCE(I.CODDER,'')=COALESCE(D.CODDER,'')
            ORDER BY COALESCE(H.DATENT,H.DATEMI) DESC, H.NUMNFC DESC
        ) ULTNF
        OUTER APPLY (
            SELECT COUNT(DISTINCT O.NUMOCP) AS QTDE_OCS_ABERTAS,
                   SUM(CAST(I.QTDABE AS FLOAT)) AS QTD_OC_ABERTA_TOTAL
            FROM E420IPO I INNER JOIN E420OCP O ON O.CODEMP=I.CODEMP AND O.CODFIL=I.CODFIL AND O.NUMOCP=I.NUMOCP
            WHERE I.CODEMP=P.CODEMP AND I.CODPRO=P.CODPRO
                  AND COALESCE(I.CODDER,'')=COALESCE(D.CODDER,'') AND COALESCE(I.QTDABE,0)>0
        ) OCA
        OUTER APPLY (
            SELECT TOP 1 O.NUMOCP AS ULTIMA_OC_ABERTA
            FROM E420IPO I INNER JOIN E420OCP O ON O.CODEMP=I.CODEMP AND O.CODFIL=I.CODFIL AND O.NUMOCP=I.NUMOCP
            WHERE I.CODEMP=P.CODEMP AND I.CODPRO=P.CODPRO
                  AND COALESCE(I.CODDER,'')=COALESCE(D.CODDER,'') AND COALESCE(I.QTDABE,0)>0
            ORDER BY O.DATEMI DESC, O.NUMOCP DESC
        ) ULTPO
        ORDER BY P.CODPRO, COALESCE(D.CODDER,'')
        OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
    """, params + [offset, tamanho_pagina])

    rows    = cursor.fetchall()
    columns = [col[0] for col in cursor.description]
    conn.close()

    resultado = []
    for row in rows:
        item = {col: (row[i].strip() if isinstance(row[i], str) else row[i])
                for i, col in enumerate(columns)}
        tipo = (item.get('tipo') or '').strip().upper()
        item['tipo_descricao'] = ('Produzido' if tipo == 'P' else
                                  'Comprado'  if tipo == 'C' else 'Não informado')
        resultado.append(item)

    total_paginas = (total_registros + tamanho_pagina - 1) // tamanho_pagina if total_registros > 0 else 1
    return {'pagina': pagina, 'tamanho_pagina': tamanho_pagina,
            'total_registros': total_registros, 'total_paginas': total_paginas,
            'dados': resultado}


# =========================================
# HELPERS EXPORT / EXCEL
# =========================================

def _excel_label(coluna: str) -> str:
    mapa = {
        'codigo': 'Código', 'descricao': 'Descrição', 'familia': 'Família',
        'origem': 'Origem',  'tipo': 'Tipo',           'tipo_descricao': 'Tipo',
        'unidade_medida': 'UM', 'deposito': 'Depósito', 'derivacao': 'Derivação',
        'saldo': 'Saldo',    'saldo_total': 'Saldo Total',
        'preco_medio': 'Preço Médio', 'custo_calculado': 'Custo Calculado',
        'numero_oc_ultima': 'Última OC',
    }
    return mapa.get(coluna, coluna.replace('_', ' ').strip().title())


def _excel_value(valor):
    if valor is None:    return ''
    if isinstance(valor, bool): return 'Sim' if valor else 'Não'
    if isinstance(valor, datetime): return valor.strftime('%d/%m/%Y %H:%M:%S')
    texto = str(valor)
    texto = ILLEGAL_CHARACTERS_RE.sub('', texto).replace(chr(0), '')
    return texto


def _excel_sheet_title(titulo: str) -> str:
    for char in ['\\', '/', '*', '?', ':', '[', ']']:
        titulo = titulo.replace(char, ' ')
    return (' '.join(titulo.split()).strip() or 'Planilha')[:31]


def _append_excel_sheet(wb, titulo, linhas, cabecalhos=None):
    ws = wb.create_sheet(title=_excel_sheet_title(titulo))
    cabecalhos = cabecalhos or {}
    if not linhas:
        ws.append(['Sem dados para exportar'])
        ws['A1'].font = Font(bold=True)
        return ws
    colunas = []
    for linha in linhas:
        for chave in linha.keys():
            if chave not in colunas: colunas.append(chave)
    ws.append([cabecalhos.get(col, _excel_label(col)) for col in colunas])
    fill = PatternFill(fill_type='solid', fgColor='1D4ED8')
    for cell in ws[1]:
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for linha in linhas:
        ws.append([_excel_value(linha.get(col)) for col in colunas])
    ws.freeze_panes  = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for idx, col in enumerate(colunas, start=1):
        largura = len(cabecalhos.get(col, _excel_label(col)))
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            largura = max(largura, len(str(row[0].value)) if row[0].value is not None else 0)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(largura + 2, 12), 42)
    return ws


def _xlsx_response(nome_arquivo, sheets):
    wb = Workbook(); wb.remove(wb.active)
    for titulo, linhas, cabecalhos in sheets:
        _append_excel_sheet(wb, titulo, linhas, cabecalhos)
    buffer = io.BytesIO(); wb.save(buffer); payload = buffer.getvalue()
    headers = {
        'Content-Disposition': f'attachment; filename="{nome_arquivo}"',
        'Access-Control-Expose-Headers': 'Content-Disposition',
        'Content-Length': str(len(payload)),
        'Cache-Control': 'no-store, max-age=0',
    }
    return Response(content=payload,
                    media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers=headers)


def _iter_paginated_pages(func, usuario, batch_size=100, max_pages=5000, **kwargs):
    pagina = 1; total_paginas = 1
    while pagina <= total_paginas:
        resposta = func(**kwargs, pagina=pagina, tamanho_pagina=batch_size, usuario=usuario)
        yield resposta
        total_paginas = int(resposta.get('total_paginas', 1) or 1)
        pagina += 1
        if pagina > max_pages:
            raise HTTPException(status_code=400, detail='Exportação excedeu o limite de páginas.')


def _collect_paginated_data(func, usuario, batch_size=100, max_pages=5000, **kwargs):
    dados = []
    for resposta in _iter_paginated_pages(func, usuario, batch_size=batch_size,
                                          max_pages=max_pages, **kwargs):
        dados.extend(resposta.get('dados', []))
    return dados


# =========================================
# EXPORTS BÁSICOS
# =========================================

@app.get('/api/export/estoque')
def exportar_estoque_excel(
    codpro: Optional[str] = None, despro: Optional[str] = None,
    codfam: Optional[str] = None, codori: Optional[str] = None,
    coddep: Optional[str] = None, situacao_cadastro: str = 'TODOS',
    somente_com_estoque: bool = True,
    usuario=Depends(validar_token)
):
    dados = _collect_paginated_data(consultar_estoque, usuario, codpro=codpro,
                                    despro=despro, codfam=codfam, codori=codori,
                                    coddep=coddep, situacao_cadastro=situacao_cadastro,
                                    somente_com_estoque=somente_com_estoque)
    return _xlsx_response('estoque.xlsx', [('Estoque', dados, None)])


@app.get('/api/export/onde-usa')
def exportar_onde_usa_excel(
    codcmp: Optional[str] = None, dercmp: Optional[str] = None,
    codmod: Optional[str] = None, situacao_cadastro: str = 'TODOS',
    usuario=Depends(validar_token)
):
    dados = _collect_paginated_data(consultar_onde_usa, usuario, codcmp=codcmp,
                                    dercmp=dercmp, codmod=codmod,
                                    situacao_cadastro=situacao_cadastro)
    return _xlsx_response('onde_usa.xlsx', [('Onde Usa', dados, None)])


@app.get('/api/export/bom')
def exportar_bom_excel(
    codmod: str, codder: Optional[str] = None,
    situacao_cadastro: str = 'TODOS', max_nivel: int = 10,
    usuario=Depends(validar_token)
):
    resposta = consultar_bom(codmod=codmod, codder=codder,
                             situacao_cadastro=situacao_cadastro,
                             max_nivel=max_nivel, usuario=usuario)
    resumo = [{'campo': _excel_label(k), 'valor': v}
              for k, v in resposta.get('cabecalho', {}).items()]
    resumo += [{'campo': 'Total Itens',   'valor': resposta.get('total_itens', 0)},
               {'campo': 'Total Níveis',  'valor': resposta.get('total_niveis', 0)},
               {'campo': 'Submodelos',    'valor': resposta.get('total_modelos_filhos', 0)}]
    return _xlsx_response('estrutura_bom.xlsx',
                          [('Resumo BOM', resumo, {'campo': 'Campo', 'valor': 'Valor'}),
                           ('Estrutura BOM', resposta.get('dados', []), None)])


@app.get('/api/export/compras-produto')
def exportar_compras_produto_excel(
    codpro: Optional[str] = None, despro: Optional[str] = None,
    codfam: Optional[str] = None, codori: Optional[str] = None,
    codder: Optional[str] = None, numero_projeto: Optional[str] = None,
    centro_custo: Optional[str] = None, situacao_cadastro: str = 'TODOS',
    somente_com_oc_aberta: bool = False, usuario=Depends(validar_token)
):
    dados = _collect_paginated_data(consultar_compras_produto, usuario,
                                    codpro=codpro, despro=despro, codfam=codfam,
                                    codori=codori, codder=codder,
                                    numero_projeto=numero_projeto,
                                    centro_custo=centro_custo,
                                    situacao_cadastro=situacao_cadastro,
                                    somente_com_oc_aberta=somente_com_oc_aberta)
    return _xlsx_response('compras_custos.xlsx', [('Compras Custos', dados, None)])



def normalizar_situacao_nf_entrada(valor):
    v = (valor or '').strip().upper()
    mapa = {
        '1': 'DIGITADA', '2': 'FECHADA', '3': 'CANCELADA',
        '4': 'DOCUMENTO FISCAL EMITIDO (SAÍDA)',
        '5': 'AGUARDANDO FECHAMENTO (PÓS-SAÍDA)',
        '6': 'AGUARDANDO INTEGRAÇÃO WMS',
        '7': 'DIGITADA INTEGRAÇÃO', '8': 'AGRUPADA',
    }
    if v in mapa: return mapa[v]
    if 'CANCEL' in v: return 'CANCELADA'
    if 'FECHAD' in v: return 'FECHADA'
    if 'DIGITAD' in v and 'INTEGRA' in v: return 'DIGITADA INTEGRAÇÃO'
    if 'DIGITAD' in v: return 'DIGITADA'
    if 'AGRUP' in v: return 'AGRUPADA'
    if 'WMS' in v: return 'AGUARDANDO INTEGRAÇÃO WMS'
    if 'PÓS' in v or 'POS' in v: return 'AGUARDANDO FECHAMENTO (PÓS-SAÍDA)'
    if 'EMIT' in v and 'SAÍDA' in v: return 'DOCUMENTO FISCAL EMITIDO (SAÍDA)'
    return v if v else 'SEM_STATUS'


# =========================================
# NOTAS DE RECEBIMENTO
# =========================================

@app.get('/api/notas-recebimento')
def consultar_notas_recebimento(
    fornecedor: Optional[str] = None, situacao_nf: Optional[str] = None,
    numero_nf: Optional[str] = None, numero_nf_exato: bool = True,
    serie_nf: Optional[str] = None, codigo_filial: Optional[str] = None,
    codigo_item: Optional[str] = None, descricao_item: Optional[str] = None,
    centro_custo: Optional[str] = None, numero_projeto: Optional[str] = None,
    transacao: Optional[str] = None, origem_material: Optional[str] = None,
    familia: Optional[str] = None, deposito: Optional[str] = None,
    numero_oc_origem: Optional[str] = None, situacao_cadastro: str = 'TODOS',
    data_emissao_ini: Optional[str] = None, data_emissao_fim: Optional[str] = None,
    data_recebimento_ini: Optional[str] = None, data_recebimento_fim: Optional[str] = None,
    tipo_item: str = 'TODOS', valor_min: Optional[float] = None,
    valor_max: Optional[float] = None, pagina: int = 1, tamanho_pagina: int = 100,
    usuario=Depends(validar_token)
):
    pagina = max(pagina, 1); tamanho_pagina = min(max(tamanho_pagina, 1), 100)
    offset = (pagina - 1) * tamanho_pagina
    conn = get_connection(); cursor = conn.cursor()

    base_union = """
        SELECT H.CodEmp AS codigo_empresa, H.CodFil AS codigo_filial,
            H.NumNfc AS numero_nf, H.CodSnf AS serie_nf, H.CodFor AS codigo_fornecedor,
            COALESCE(F.NomFor,'') AS nome_fornecedor, COALESCE(F.ApeFor,'') AS fantasia_fornecedor,
            H.DatEmi AS data_emissao, H.DatEnt AS data_recebimento,
            CASE WHEN LTRIM(RTRIM(CAST(COALESCE(H.SitNfc,'') AS VARCHAR(40))))='1' THEN 'DIGITADA'
                 WHEN LTRIM(RTRIM(CAST(COALESCE(H.SitNfc,'') AS VARCHAR(40))))='2' THEN 'FECHADA'
                 WHEN LTRIM(RTRIM(CAST(COALESCE(H.SitNfc,'') AS VARCHAR(40))))='3' THEN 'CANCELADA'
                 WHEN LTRIM(RTRIM(CAST(COALESCE(H.SitNfc,'') AS VARCHAR(40))))='4' THEN 'DOCUMENTO FISCAL EMITIDO (SAÍDA)'
                 WHEN LTRIM(RTRIM(CAST(COALESCE(H.SitNfc,'') AS VARCHAR(40))))='5' THEN 'AGUARDANDO FECHAMENTO (PÓS-SAÍDA)'
                 WHEN LTRIM(RTRIM(CAST(COALESCE(H.SitNfc,'') AS VARCHAR(40))))='6' THEN 'AGUARDANDO INTEGRAÇÃO WMS'
                 WHEN LTRIM(RTRIM(CAST(COALESCE(H.SitNfc,'') AS VARCHAR(40))))='7' THEN 'DIGITADA INTEGRAÇÃO'
                 WHEN LTRIM(RTRIM(CAST(COALESCE(H.SitNfc,'') AS VARCHAR(40))))='8' THEN 'AGRUPADA'
                 WHEN COALESCE(CAST(H.SitNfc AS VARCHAR(40)),'')='' THEN 'SEM_STATUS'
                 ELSE UPPER(CAST(COALESCE(H.SitNfc,'') AS VARCHAR(40))) END AS situacao_nf,
            'PRODUTO' AS tipo_item, CAST(I.SeqIpc AS INT) AS sequencia_item,
            COALESCE(I.CodPro,'') AS codigo_item,
            COALESCE(P.DesPro,I.CplIpc,'') AS descricao_item,
            COALESCE(I.CodDer,'') AS derivacao, COALESCE(I.UniMed,P.UniMed,'') AS unidade_medida,
            COALESCE(I.TnsPro,'') AS transacao, COALESCE(I.CodDep,'') AS deposito,
            COALESCE(I.CodCcu,'') AS codigo_centro_custo, COALESCE(CC.DesCcu,'') AS descricao_centro_custo,
            CAST(COALESCE(I.NumPrj,0) AS INT) AS numero_projeto,
            COALESCE(PRJ.NomPrj,'') AS nome_projeto,
            COALESCE(I.CodFam,P.CodFam,'') AS codigo_familia,
            COALESCE(P.CodOri,'') AS origem_material, COALESCE(P.SitPro,'') AS situacao_cadastro_produto,
            CAST(COALESCE(I.QtdRec,0) AS FLOAT) AS quantidade_recebida,
            CAST(COALESCE(I.PreUni,0) AS FLOAT) AS preco_unitario,
            CAST(COALESCE(I.VlrBru,0) AS FLOAT) AS valor_bruto,
            CAST(COALESCE(I.VlrLiq,0) AS FLOAT) AS valor_liquido,
            CAST(COALESCE(I.VlrIpi,0) AS FLOAT) AS valor_ipi,
            CAST(COALESCE(I.VlrIcm,0) AS FLOAT) AS valor_icms,
            CAST(0 AS FLOAT) AS valor_iss,
            CAST(COALESCE(I.NumOcp,0) AS INT) AS numero_oc_origem
        FROM E440NFC H
        INNER JOIN E440IPC I ON I.CodEmp=H.CodEmp AND I.CodFil=H.CodFil
            AND I.CodFor=H.CodFor AND I.NumNfc=H.NumNfc AND I.CodSnf=H.CodSnf
        LEFT JOIN E095FOR F ON F.CodFor=H.CodFor
        LEFT JOIN E075PRO P ON P.CodEmp=I.CodEmp AND P.CodPro=I.CodPro
        LEFT JOIN E044CCU CC ON CC.CodEmp=I.CodEmp AND CC.CodCcu=I.CodCcu
        LEFT JOIN E615PRJ PRJ ON PRJ.CodEmp=I.CodEmp AND PRJ.NumPrj=I.NumPrj

        UNION ALL

        SELECT H.CodEmp AS codigo_empresa, H.CodFil AS codigo_filial,
            H.NumNfc AS numero_nf, H.CodSnf AS serie_nf, H.CodFor AS codigo_fornecedor,
            COALESCE(F.NomFor,'') AS nome_fornecedor, COALESCE(F.ApeFor,'') AS fantasia_fornecedor,
            H.DatEmi AS data_emissao, H.DatEnt AS data_recebimento,
            CASE WHEN LTRIM(RTRIM(CAST(COALESCE(H.SitNfc,'') AS VARCHAR(40))))='1' THEN 'DIGITADA'
                 WHEN LTRIM(RTRIM(CAST(COALESCE(H.SitNfc,'') AS VARCHAR(40))))='2' THEN 'FECHADA'
                 WHEN LTRIM(RTRIM(CAST(COALESCE(H.SitNfc,'') AS VARCHAR(40))))='3' THEN 'CANCELADA'
                 WHEN COALESCE(CAST(H.SitNfc AS VARCHAR(40)),'')='' THEN 'SEM_STATUS'
                 ELSE UPPER(CAST(COALESCE(H.SitNfc,'') AS VARCHAR(40))) END AS situacao_nf,
            'SERVIÇO' AS tipo_item, CAST(SI.SeqIsc AS INT) AS sequencia_item,
            COALESCE(SI.CodSer,'') AS codigo_item,
            COALESCE(S.DesSer,SI.CplIsc,'') AS descricao_item,
            '' AS derivacao, COALESCE(SI.UniMed,S.UniMed,'') AS unidade_medida,
            COALESCE(SI.TnsSer,'') AS transacao, '' AS deposito,
            COALESCE(SI.CodCcu,'') AS codigo_centro_custo, COALESCE(CC.DesCcu,'') AS descricao_centro_custo,
            CAST(COALESCE(SI.NumPrj,0) AS INT) AS numero_projeto,
            COALESCE(PRJ.NomPrj,'') AS nome_projeto,
            COALESCE(SI.CodFam,S.CodFam,'') AS codigo_familia,
            COALESCE(S.OriMer,'') AS origem_material, '' AS situacao_cadastro_produto,
            CAST(COALESCE(SI.QtdRec,0) AS FLOAT) AS quantidade_recebida,
            CAST(COALESCE(SI.PreUni,0) AS FLOAT) AS preco_unitario,
            CAST(COALESCE(SI.VlrBru,0) AS FLOAT) AS valor_bruto,
            CAST(COALESCE(SI.VlrLiq,0) AS FLOAT) AS valor_liquido,
            CAST(COALESCE(SI.VlrIpi,0) AS FLOAT) AS valor_ipi,
            CAST(COALESCE(SI.VlrIcm,0) AS FLOAT) AS valor_icms,
            CAST(COALESCE(SI.VlrIss,0) AS FLOAT) AS valor_iss,
            CAST(COALESCE(SI.NumOcp,0) AS INT) AS numero_oc_origem
        FROM E440NFC H
        INNER JOIN E440ISC SI ON SI.CodEmp=H.CodEmp AND SI.CodFil=H.CodFil
            AND SI.CodFor=H.CodFor AND SI.NumNfc=H.NumNfc AND SI.CodSnf=H.CodSnf
        LEFT JOIN E095FOR F ON F.CodFor=H.CodFor
        LEFT JOIN E080SER S ON S.CodEmp=SI.CodEmp AND S.CodSer=SI.CodSer
        LEFT JOIN E044CCU CC ON CC.CodEmp=SI.CodEmp AND CC.CodCcu=SI.CodCcu
        LEFT JOIN E615PRJ PRJ ON PRJ.CodEmp=SI.CodEmp AND PRJ.NumPrj=SI.NumPrj
    """

    cte_sql = f"WITH BASE AS ({base_union})"
    where_parts = ["BASE.codigo_empresa = ?"]
    params_where = [EMPRESA_PADRAO]
    tipo_item = (tipo_item or 'TODOS').strip().upper()
    situacao_nf = (situacao_nf or '').strip()
    numero_nf = (numero_nf or '').strip()
    serie_nf = (serie_nf or '').strip()
    codigo_filial = (codigo_filial or '').strip()
    codigo_item = (codigo_item or '').strip()
    descricao_item = (descricao_item or '').strip()
    centro_custo = (centro_custo or '').strip()
    numero_projeto = (numero_projeto or '').strip()
    transacao = (transacao or '').strip()
    origem_material = (origem_material or '').strip()
    familia = (familia or '').strip()
    deposito = (deposito or '').strip()
    numero_oc_origem = (numero_oc_origem or '').strip()
    fornecedor = (fornecedor or '').strip()

    if fornecedor:
        like = f"%{fornecedor}%"
        where_parts.append("(CAST(BASE.codigo_fornecedor AS VARCHAR(20)) LIKE ? OR BASE.nome_fornecedor LIKE ? OR BASE.fantasia_fornecedor LIKE ?)")
        params_where.extend([like, like, like])
    if codigo_filial:
        if codigo_filial.isdigit():
            where_parts.append("BASE.codigo_filial = ?"); params_where.append(int(codigo_filial))
        else:
            where_parts.append("CAST(BASE.codigo_filial AS VARCHAR(20)) LIKE ?"); params_where.append(f"%{codigo_filial}%")
    if numero_nf:
        if numero_nf_exato and numero_nf.isdigit():
            where_parts.append("BASE.numero_nf = ?"); params_where.append(int(numero_nf))
        else:
            where_parts.append("CAST(BASE.numero_nf AS VARCHAR(20)) LIKE ?"); params_where.append(f"%{numero_nf}%")
    if situacao_nf:
        where_parts.append("UPPER(CAST(COALESCE(BASE.situacao_nf,'') AS VARCHAR(20))) LIKE ?")
        params_where.append(f"%{situacao_nf.upper()}%")
    if serie_nf: where_parts.append("BASE.serie_nf LIKE ?"); params_where.append(f"%{serie_nf}%")
    if codigo_item: where_parts.append("BASE.codigo_item LIKE ?"); params_where.append(f"%{codigo_item}%")
    if descricao_item: where_parts.append("BASE.descricao_item LIKE ?"); params_where.append(f"%{descricao_item}%")
    if centro_custo:
        like = f"%{centro_custo}%"
        where_parts.append("(BASE.codigo_centro_custo LIKE ? OR BASE.descricao_centro_custo LIKE ?)")
        params_where.extend([like, like])
    if numero_projeto: where_parts.append("CAST(BASE.numero_projeto AS VARCHAR(20)) LIKE ?"); params_where.append(f"%{numero_projeto}%")
    if transacao: where_parts.append("BASE.transacao LIKE ?"); params_where.append(f"%{transacao}%")
    if origem_material: where_parts.append("BASE.origem_material LIKE ?"); params_where.append(f"%{origem_material}%")
    if familia: where_parts.append("BASE.codigo_familia LIKE ?"); params_where.append(f"%{familia}%")
    if deposito: where_parts.append("BASE.deposito LIKE ?"); params_where.append(f"%{deposito}%")
    if numero_oc_origem:
        if numero_oc_origem.isdigit():
            where_parts.append("BASE.numero_oc_origem = ?"); params_where.append(int(numero_oc_origem))
        else:
            where_parts.append("CAST(BASE.numero_oc_origem AS VARCHAR(20)) LIKE ?"); params_where.append(f"%{numero_oc_origem}%")
    sc = normalizar_situacao_cadastro(situacao_cadastro)
    if sc == 'ATIVO': where_parts.append("(BASE.tipo_item<>'PRODUTO' OR COALESCE(BASE.situacao_cadastro_produto,'A')='A')")
    elif sc == 'INATIVO': where_parts.append("(BASE.tipo_item<>'PRODUTO' OR COALESCE(BASE.situacao_cadastro_produto,'A')='I')")
    if data_emissao_ini: where_parts.append("CAST(BASE.data_emissao AS DATE) >= ?"); params_where.append(data_emissao_ini)
    if data_emissao_fim: where_parts.append("CAST(BASE.data_emissao AS DATE) <= ?"); params_where.append(data_emissao_fim)
    if data_recebimento_ini: where_parts.append("CAST(BASE.data_recebimento AS DATE) >= ?"); params_where.append(data_recebimento_ini)
    if data_recebimento_fim: where_parts.append("CAST(BASE.data_recebimento AS DATE) <= ?"); params_where.append(data_recebimento_fim)
    if tipo_item in ('PRODUTO', 'SERVIÇO'): where_parts.append("BASE.tipo_item = ?"); params_where.append(tipo_item)
    if valor_min is not None: where_parts.append("COALESCE(BASE.valor_liquido,0) >= ?"); params_where.append(valor_min)
    if valor_max is not None: where_parts.append("COALESCE(BASE.valor_liquido,0) <= ?"); params_where.append(valor_max)

    where_sql = " AND ".join(where_parts)
    sql_total = cte_sql + f" SELECT COUNT(*) FROM BASE WHERE {where_sql}"
    sql_dados = cte_sql + f" SELECT * FROM BASE WHERE {where_sql} ORDER BY BASE.data_recebimento DESC, BASE.data_emissao DESC, BASE.numero_nf DESC, BASE.sequencia_item ASC OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"

    try:
        cursor.execute(sql_total, params_where); total_registros = cursor.fetchone()[0]
        cursor.execute(sql_dados, params_where + [offset, tamanho_pagina])
        rows = cursor.fetchall(); columns = [col[0] for col in cursor.description]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro SQL Notas de Recebimento: {str(e)}")
    finally:
        conn.close()

    dados = [dict(zip(columns, [v.strip() if isinstance(v,str) else v for v in row])) for row in rows]
    total_paginas = (total_registros + tamanho_pagina - 1) // tamanho_pagina if total_registros > 0 else 1
    return {'pagina': pagina, 'tamanho_pagina': tamanho_pagina,
            'total_registros': total_registros, 'total_paginas': total_paginas, 'dados': dados}


# =========================================
# CONCILIAÇÃO ERP x EDOCS
# =========================================

@app.get('/api/notas-edocs-conciliacao')
def consultar_notas_edocs_conciliacao(
    tipo_nota: str = 'TODOS', numero_nf: Optional[str] = None,
    serie_nf: Optional[str] = None, codigo_filial: Optional[str] = None,
    codigo_pessoa: Optional[str] = None, nome_pessoa: Optional[str] = None,
    numero_lote: Optional[str] = None, situacao_erp: Optional[str] = None,
    situacao_edocs: Optional[str] = None, status_conciliacao: str = 'TODOS',
    data_ini: Optional[str] = None, data_fim: Optional[str] = None,
    somente_divergencia: bool = False, somente_sem_edocs: bool = False,
    somente_com_erro: bool = False, pagina: int = 1, tamanho_pagina: int = 100,
    usuario=Depends(validar_token)
):
    pagina = max(pagina, 1); tamanho_pagina = min(max(tamanho_pagina, 1), 100)
    offset = (pagina - 1) * tamanho_pagina
    tipo_nota = (tipo_nota or 'TODOS').strip().upper()
    if tipo_nota not in ('TODOS', 'ENTRADA', 'SAIDA'): tipo_nota = 'TODOS'
    status_conciliacao = (status_conciliacao or 'TODOS').strip().upper()

    conn = get_connection(); cursor = conn.cursor()

    # Simplified CTE for ERP x EDocs conciliation
    cte_entrada = """
        SELECT nfc.CODEMP AS codigo_empresa, nfc.CODFIL AS codigo_filial,
            'ENTRADA' AS tipo_nota, nfc.CODSNF AS serie_nf, nfc.NUMNFC AS numero_nf,
            CAST(COALESCE(nfc.SITNFC,'') AS VARCHAR(40)) AS situacao_erp,
            CASE WHEN LTRIM(RTRIM(CAST(COALESCE(nfc.SITNFC,'') AS VARCHAR(40))))='1' THEN 'DIGITADA'
                 WHEN LTRIM(RTRIM(CAST(COALESCE(nfc.SITNFC,'') AS VARCHAR(40))))='2' THEN 'FECHADA'
                 WHEN LTRIM(RTRIM(CAST(COALESCE(nfc.SITNFC,'') AS VARCHAR(40))))='3' THEN 'CANCELADA'
                 WHEN COALESCE(CAST(nfc.SITNFC AS VARCHAR(40)),'')='' THEN 'SEM_STATUS'
                 ELSE UPPER(CAST(COALESCE(nfc.SITNFC,'') AS VARCHAR(40)))
            END AS situacao_erp_normalizada,
            CAST(COALESCE(nfc.NUMLOT,0) AS INT) AS numero_lote,
            CAST(nfc.CODFOR AS INT) AS codigo_pessoa,
            COALESCE(forx.APEFOR,forx.NOMFOR,'') AS nome_pessoa,
            CAST(COALESCE(nfc.VLRLIQ,0) AS FLOAT) AS valor_liquido,
            COALESCE(nfc.CHVNEL,'') AS chave_nota,
            COALESCE(nex.IDEUNI,'') AS edocs_ideuni_ponte,
            CAST(COALESCE(edc.NUMNFC,0) AS INT) AS edocs_numero_nf,
            COALESCE(edc.CODSNF,'') AS edocs_serie_nf,
            COALESCE(edc.CHVNEL,'') AS edocs_chave_acesso,
            COALESCE(cix.SITIEX,'') AS situacao_edocs,
            CASE WHEN UPPER(COALESCE(cix.SITIEX,'')) LIKE '%CANCEL%' THEN 'CANCELADA'
                 WHEN UPPER(COALESCE(cix.SITIEX,'')) LIKE '%AUTORIZ%' THEN 'AUTORIZADA'
                 WHEN COALESCE(cix.SITIEX,'')='' THEN 'SEM_STATUS'
                 ELSE UPPER(COALESCE(cix.SITIEX,''))
            END AS situacao_edocs_normalizada,
            COALESCE(cix.MSGERR,'') AS mensagem_edocs,
            nfc.DATENT AS data_documento,
            CASE WHEN COALESCE(cix.MSGERR,'') <> '' OR COALESCE(eis.DESMOT,'') <> '' THEN 'ERRO_EDOCS'
                 WHEN COALESCE(nex.IDEUNI,'')='' AND COALESCE(edc.CHVNEL,'')='' THEN 'SEM_EDOCS'
                 WHEN ISNULL(NULLIF(LTRIM(RTRIM(nfc.CHVNEL)),''),'SEM') <> ISNULL(NULLIF(LTRIM(RTRIM(edc.CHVNEL)),''),'SEM') AND COALESCE(edc.CHVNEL,'') <> '' THEN 'CHAVE_DIVERGENTE'
                 ELSE 'OK'
            END AS status_conciliacao,
            COALESCE(eis.DESMOT,'') AS descricao_motivo_edocs,
            CAST(COALESCE(cix.CODINT,'') AS VARCHAR(20)) AS codint_edocs,
            CAST(COALESCE(cix.IDEINT,'') AS VARCHAR(50)) AS ideint_edocs
        FROM E440NFC nfc
        LEFT JOIN E095FOR forx ON forx.CODFOR=nfc.CODFOR
        LEFT JOIN E000NEX nex ON nex.CODEMP=nfc.CODEMP AND nex.CODFIL=nfc.CODFIL
            AND nex.CODFOR=nfc.CODFOR AND nex.NUMNFC=nfc.NUMNFC
            AND ISNULL(LTRIM(RTRIM(nex.CODSNF)),'')=ISNULL(LTRIM(RTRIM(nfc.CODSNF)),'')
        LEFT JOIN E000NFC edc ON (NULLIF(LTRIM(RTRIM(edc.CHVNEL)),'') IS NOT NULL
            AND NULLIF(LTRIM(RTRIM(nfc.CHVNEL)),'') IS NOT NULL
            AND LTRIM(RTRIM(edc.CHVNEL))=LTRIM(RTRIM(nfc.CHVNEL)))
        LEFT JOIN E000CIX cix ON cix.IDEINT=nex.IDEUNI
        LEFT JOIN E000CIM cim ON cim.CODINT=cix.CODINT AND cim.IDEINT=cix.IDEINT
        LEFT JOIN E000EIS eis ON eis.CODINT=cim.CODINT AND eis.IDTREQ=cim.IDTREQ
        WHERE nfc.TIPNFE=1
        GROUP BY nfc.CODEMP, nfc.CODFIL, nfc.CODFOR, forx.APEFOR, forx.NOMFOR,
            nfc.NUMNFC, nfc.CODSNF, nfc.SITNFC, nfc.DATEMI, nfc.DATENT, nfc.DATGER,
            nfc.NUMLOT, nfc.VLRLIQ, nfc.VLRFIN, nfc.CHVNEL, nex.IDEUNI,
            edc.NUMNFC, edc.CODSNF, edc.CHVNEL, cix.CODINT, cix.IDEINT,
            cix.SITIEX, cix.MSGERR, cix.INDEXP, cix.DATEXP, cix.HOREXP,
            cix.REGINT, cix.VERREG, cim.IDTREQ, cim.CHVREQ, cim.DATREQ, cim.HORREQ,
            eis.TIPEIS, eis.CODMOT, eis.DESMOT, eis.DATGER, eis.HORGER
    """

    cte_sql = f"WITH BASE AS ({cte_entrada})"
    where_parts = ["BASE.codigo_empresa = ?"]
    params_where = [EMPRESA_PADRAO]

    if tipo_nota in ('ENTRADA', 'SAIDA'): where_parts.append("BASE.tipo_nota = ?"); params_where.append(tipo_nota)
    if numero_nf: where_parts.append("CAST(BASE.numero_nf AS VARCHAR(20)) LIKE ?"); params_where.append(f"%{numero_nf}%")
    if serie_nf: where_parts.append("BASE.serie_nf LIKE ?"); params_where.append(f"%{serie_nf}%")
    if codigo_filial: where_parts.append("CAST(BASE.codigo_filial AS VARCHAR(20)) LIKE ?"); params_where.append(f"%{codigo_filial}%")
    if codigo_pessoa: where_parts.append("CAST(BASE.codigo_pessoa AS VARCHAR(20)) LIKE ?"); params_where.append(f"%{codigo_pessoa}%")
    if nome_pessoa: where_parts.append("BASE.nome_pessoa LIKE ?"); params_where.append(f"%{nome_pessoa}%")
    if numero_lote: where_parts.append("CAST(BASE.numero_lote AS VARCHAR(20)) LIKE ?"); params_where.append(f"%{numero_lote}%")
    if situacao_erp: where_parts.append("COALESCE(BASE.situacao_erp_normalizada,'') = ?"); params_where.append(situacao_erp.upper())
    if situacao_edocs: where_parts.append("COALESCE(BASE.situacao_edocs_normalizada,'') = ?"); params_where.append(situacao_edocs.upper())
    if status_conciliacao != 'TODOS': where_parts.append("BASE.status_conciliacao = ?"); params_where.append(status_conciliacao)
    if data_ini: where_parts.append("CAST(BASE.data_documento AS DATE) >= ?"); params_where.append(data_ini)
    if data_fim: where_parts.append("CAST(BASE.data_documento AS DATE) <= ?"); params_where.append(data_fim)
    if somente_divergencia: where_parts.append("BASE.status_conciliacao IN ('SEM_EDOCS','ERRO_EDOCS','DIVERGENCIA_SITUACAO','CHAVE_DIVERGENTE','NUMERO_DIVERGENTE','SERIE_DIVERGENTE')")
    if somente_sem_edocs: where_parts.append("BASE.status_conciliacao = 'SEM_EDOCS'")
    if somente_com_erro: where_parts.append("BASE.status_conciliacao = 'ERRO_EDOCS'")

    where_sql = " AND ".join(where_parts)
    sql_resumo = cte_sql + f"""
        SELECT COUNT(*) AS total_registros,
            SUM(CASE WHEN BASE.status_conciliacao='OK' THEN 1 ELSE 0 END) AS total_ok,
            SUM(CASE WHEN BASE.status_conciliacao='SEM_EDOCS' THEN 1 ELSE 0 END) AS total_sem_edocs,
            SUM(CASE WHEN BASE.status_conciliacao='ERRO_EDOCS' THEN 1 ELSE 0 END) AS total_com_erro,
            SUM(CASE WHEN BASE.status_conciliacao='DIVERGENCIA_SITUACAO' THEN 1 ELSE 0 END) AS total_divergencia_situacao,
            SUM(CASE WHEN BASE.status_conciliacao='CHAVE_DIVERGENTE' THEN 1 ELSE 0 END) AS total_chave_divergente
        FROM BASE WHERE {where_sql}"""
    sql_dados = cte_sql + f"""
        SELECT * FROM BASE WHERE {where_sql}
        ORDER BY CASE WHEN BASE.status_conciliacao='SEM_EDOCS' THEN 0 WHEN BASE.status_conciliacao='ERRO_EDOCS' THEN 1 ELSE 6 END,
            BASE.data_documento DESC, BASE.numero_nf DESC
        OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"""

    try:
        cursor.execute(sql_resumo, params_where); row = cursor.fetchone()
        resumo = {
            'total_registros': int(row[0] or 0), 'total_ok': int(row[1] or 0),
            'total_sem_edocs': int(row[2] or 0), 'total_com_erro': int(row[3] or 0),
            'total_divergencia_situacao': int(row[4] or 0), 'total_chave_divergente': int(row[5] or 0),
        }
        cursor.execute(sql_dados, params_where + [offset, tamanho_pagina])
        rows = cursor.fetchall(); columns = [col[0] for col in cursor.description]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro SQL Conciliação ERP x EDocs: {str(e)}")
    finally:
        conn.close()

    dados = [dict(zip(columns, [v.strip() if isinstance(v,str) else v for v in row])) for row in rows]
    total_registros = resumo['total_registros']
    total_paginas = (total_registros + tamanho_pagina - 1) // tamanho_pagina if total_registros > 0 else 1
    return {'pagina': pagina, 'tamanho_pagina': tamanho_pagina,
            'total_registros': total_registros, 'total_paginas': total_paginas,
            'resumo': resumo, 'dados': dados}


# =========================================
# PAINEL DE COMPRAS
# =========================================

@app.get('/api/painel-compras')
def consultar_painel_compras(
    codigo_produto: Optional[str] = None, codigo_item: Optional[str] = None,
    descricao_item: Optional[str] = None, fornecedor: Optional[str] = None,
    numero_oc: Optional[str] = None, numero_projeto: Optional[str] = None,
    centro_custo: Optional[str] = None, transacao: Optional[str] = None,
    valor_min: Optional[float] = None, valor_max: Optional[float] = None,
    tipo_item: str = 'TODOS', tipo_oc: str = 'TODOS',
    situacao_oc: Optional[int] = None, codigo_motivo_oc: Optional[str] = None,
    observacao_oc: Optional[str] = None,
    data_emissao_ini: Optional[str] = None, data_emissao_fim: Optional[str] = None,
    data_entrega_ini: Optional[str] = None, data_entrega_fim: Optional[str] = None,
    origem_material: Optional[str] = None, familia: Optional[str] = None,
    situacao_cadastro: str = 'TODOS', somente_pendentes: bool = True,
    agrupar_por_fornecedor: bool = False,
    pagina: int = 1, tamanho_pagina: int = 100,
    usuario=Depends(validar_token)
):
    pagina = max(pagina, 1); tamanho_pagina = min(max(tamanho_pagina, 1), 100)
    offset = (pagina - 1) * tamanho_pagina
    conn = get_connection(); cursor = conn.cursor()

    base_union = """
        SELECT O.CodEmp AS codigo_empresa, O.CodFil AS codigo_filial,
            O.NumOcp AS numero_oc, CAST(I.SeqIpo AS INT) AS sequencia_item,
            'PRODUTO' AS tipo_item, CAST(O.CodFor AS INT) AS codigo_fornecedor,
            COALESCE(F.NomFor,'') AS nome_fornecedor, COALESCE(F.ApeFor,'') AS fantasia_fornecedor,
            COALESCE(I.CodPro,'') AS codigo_item,
            COALESCE(P.DesPro,I.DesFor,I.CplIpo,'') AS descricao_item,
            COALESCE(I.CodDer,'') AS derivacao, COALESCE(I.UniMed,P.UniMed,'') AS unidade_medida,
            COALESCE(I.CodFam,P.CodFam,'') AS codigo_familia,
            COALESCE(P.CodOri,'') AS origem_material, COALESCE(P.SitPro,'') AS situacao_cadastro_produto,
            COALESCE(I.TnsPro,O.TnsPro,'') AS transacao,
            O.DatEmi AS data_emissao, I.DatEnt AS data_entrega,
            CAST(COALESCE(I.QtdPed,0) AS FLOAT) AS quantidade_pedida,
            CAST(COALESCE(I.QtdRec,0) AS FLOAT) AS quantidade_recebida,
            CAST(COALESCE(I.QtdAbe,0) AS FLOAT) AS saldo_pendente,
            CAST(COALESCE(I.PreUni,0) AS FLOAT) AS preco_unitario,
            CAST(COALESCE(I.VlrBru,0) AS FLOAT) AS valor_bruto,
            CAST(COALESCE(I.VlrLiq,0) AS FLOAT) AS valor_liquido,
            CAST(COALESCE(NULLIF(O.VlrFin,0),NULLIF(O.VlrLiq,0),NULLIF(O.VlrOri,0),0) AS FLOAT) AS valor_total_oc_cabecalho,
            CAST(COALESCE(I.VlrIpi,0) AS FLOAT) AS valor_ipi,
            CAST(COALESCE(I.VlrIcm,0) AS FLOAT) AS valor_icms,
            CAST(0 AS FLOAT) AS valor_iss,
            COALESCE(I.CodCcu,'') AS codigo_centro_custo,
            COALESCE(CC.DesCcu,'') AS descricao_centro_custo,
            CAST(COALESCE(I.NumPrj,0) AS INT) AS numero_projeto,
            COALESCE(PRJ.NomPrj,'') AS nome_projeto,
            CAST(COALESCE(O.SitOcp,0) AS INT) AS situacao_oc,
            COALESCE(O.SitApr,'') AS situacao_aprovacao
        FROM E420OCP O
        INNER JOIN E420IPO I ON I.CodEmp=O.CodEmp AND I.CodFil=O.CodFil AND I.NumOcp=O.NumOcp
        LEFT JOIN E095FOR F ON F.CodFor=O.CodFor
        LEFT JOIN E075PRO P ON P.CodEmp=I.CodEmp AND P.CodPro=I.CodPro
        LEFT JOIN E044CCU CC ON CC.CodEmp=I.CodEmp AND CC.CodCcu=I.CodCcu
        LEFT JOIN E615PRJ PRJ ON PRJ.CodEmp=I.CodEmp AND PRJ.NumPrj=I.NumPrj

        UNION ALL

        SELECT O.CodEmp AS codigo_empresa, O.CodFil AS codigo_filial,
            O.NumOcp AS numero_oc, CAST(SI.SeqIso AS INT) AS sequencia_item,
            'SERVIÇO' AS tipo_item, CAST(O.CodFor AS INT) AS codigo_fornecedor,
            COALESCE(F.NomFor,'') AS nome_fornecedor, COALESCE(F.ApeFor,'') AS fantasia_fornecedor,
            COALESCE(SI.CodSer,'') AS codigo_item,
            COALESCE(S.DesSer,SI.CplIso,'') AS descricao_item,
            '' AS derivacao, COALESCE(SI.UniMed,S.UniMed,'') AS unidade_medida,
            COALESCE(SI.CodFam,S.CodFam,'') AS codigo_familia,
            COALESCE(S.OriMer,'') AS origem_material, '' AS situacao_cadastro_produto,
            COALESCE(SI.TnsSer,O.TnsSer,'') AS transacao,
            O.DatEmi AS data_emissao, SI.DatEnt AS data_entrega,
            CAST(COALESCE(SI.QtdPed,0) AS FLOAT) AS quantidade_pedida,
            CAST(COALESCE(SI.QtdRec,0) AS FLOAT) AS quantidade_recebida,
            CAST(COALESCE(SI.QtdAbe,0) AS FLOAT) AS saldo_pendente,
            CAST(COALESCE(SI.PreUni,0) AS FLOAT) AS preco_unitario,
            CAST(COALESCE(SI.VlrBru,0) AS FLOAT) AS valor_bruto,
            CAST(COALESCE(SI.VlrLiq,0) AS FLOAT) AS valor_liquido,
            CAST(COALESCE(NULLIF(O.VlrFin,0),NULLIF(O.VlrLiq,0),NULLIF(O.VlrOri,0),0) AS FLOAT) AS valor_total_oc_cabecalho,
            CAST(COALESCE(SI.VlrIpi,0) AS FLOAT) AS valor_ipi,
            CAST(COALESCE(SI.VlrIcm,0) AS FLOAT) AS valor_icms,
            CAST(COALESCE(SI.VlrIss,0) AS FLOAT) AS valor_iss,
            COALESCE(SI.CodCcu,'') AS codigo_centro_custo,
            COALESCE(CC.DesCcu,'') AS descricao_centro_custo,
            CAST(COALESCE(SI.NumPrj,0) AS INT) AS numero_projeto,
            COALESCE(PRJ.NomPrj,'') AS nome_projeto,
            CAST(COALESCE(O.SitOcp,0) AS INT) AS situacao_oc,
            COALESCE(O.SitApr,'') AS situacao_aprovacao
        FROM E420OCP O
        INNER JOIN E420ISO SI ON SI.CodEmp=O.CodEmp AND SI.CodFil=O.CodFil AND SI.NumOcp=O.NumOcp
        LEFT JOIN E095FOR F ON F.CodFor=O.CodFor
        LEFT JOIN E080SER S ON S.CodEmp=SI.CodEmp AND S.CodSer=SI.CodSer
        LEFT JOIN E044CCU CC ON CC.CodEmp=SI.CodEmp AND CC.CodCcu=SI.CodCcu
        LEFT JOIN E615PRJ PRJ ON PRJ.CodEmp=SI.CodEmp AND PRJ.NumPrj=SI.NumPrj
    """

    cte_sql = f"""
        WITH BASE AS ({base_union}),
        TIPO_OC AS (
            SELECT codigo_empresa, codigo_filial, numero_oc,
                CASE WHEN SUM(CASE WHEN tipo_item='PRODUTO' THEN 1 ELSE 0 END)>0
                      AND SUM(CASE WHEN tipo_item='SERVIÇO' THEN 1 ELSE 0 END)>0 THEN 'MISTA'
                     WHEN SUM(CASE WHEN tipo_item='SERVIÇO' THEN 1 ELSE 0 END)>0 THEN 'SERVIÇO'
                     ELSE 'PRODUTO' END AS tipo_oc
            FROM BASE GROUP BY codigo_empresa, codigo_filial, numero_oc
        ),
        Q AS (
            SELECT BASE.*, TIPO_OC.tipo_oc,
                CAST(COALESCE(BASE.saldo_pendente,0)*COALESCE(BASE.preco_unitario,0) AS FLOAT) AS valor_pendente_estimado,
                CASE WHEN COALESCE(BASE.saldo_pendente,0)>0 AND BASE.data_entrega IS NOT NULL
                      AND CAST(BASE.data_entrega AS DATE)<CAST(GETDATE() AS DATE)
                    THEN DATEDIFF(DAY, CAST(BASE.data_entrega AS DATE), CAST(GETDATE() AS DATE))
                    ELSE 0 END AS dias_atraso
            FROM BASE
            LEFT JOIN TIPO_OC ON TIPO_OC.codigo_empresa=BASE.codigo_empresa
                AND TIPO_OC.codigo_filial=BASE.codigo_filial AND TIPO_OC.numero_oc=BASE.numero_oc
        )
    """

    where_parts = ["Q.codigo_empresa = ?"]
    params_where = [EMPRESA_PADRAO]
    tipo_item = (tipo_item or 'TODOS').strip().upper()
    tipo_oc = (tipo_oc or 'TODOS').strip().upper()

    if codigo_item: where_parts.append("Q.codigo_item LIKE ?"); params_where.append(f"%{codigo_item}%")
    if codigo_produto: where_parts.append("(Q.tipo_item='PRODUTO' AND Q.codigo_item LIKE ?)"); params_where.append(f"%{codigo_produto}%")
    if descricao_item: where_parts.append("Q.descricao_item LIKE ?"); params_where.append(f"%{descricao_item}%")
    if fornecedor:
        like = f"%{fornecedor}%"
        where_parts.append("(CAST(Q.codigo_fornecedor AS VARCHAR(20)) LIKE ? OR Q.nome_fornecedor LIKE ? OR Q.fantasia_fornecedor LIKE ?)")
        params_where.extend([like, like, like])
    if numero_oc: where_parts.append("CAST(Q.numero_oc AS VARCHAR(20)) LIKE ?"); params_where.append(f"%{numero_oc}%")
    if numero_projeto: where_parts.append("CAST(Q.numero_projeto AS VARCHAR(20)) LIKE ?"); params_where.append(f"%{numero_projeto}%")
    if centro_custo:
        like = f"%{centro_custo}%"
        where_parts.append("(Q.codigo_centro_custo LIKE ? OR Q.descricao_centro_custo LIKE ?)")
        params_where.extend([like, like])
    if transacao: where_parts.append("Q.transacao LIKE ?"); params_where.append(f"%{transacao}%")
    if valor_min is not None: where_parts.append("COALESCE(Q.valor_liquido,0) >= ?"); params_where.append(valor_min)
    if valor_max is not None: where_parts.append("COALESCE(Q.valor_liquido,0) <= ?"); params_where.append(valor_max)
    if tipo_item in ('PRODUTO', 'SERVIÇO'): where_parts.append("Q.tipo_item = ?"); params_where.append(tipo_item)
    if tipo_oc in ('PRODUTO', 'SERVIÇO', 'MISTA'): where_parts.append("Q.tipo_oc = ?"); params_where.append(tipo_oc)
    if situacao_oc is not None: where_parts.append("COALESCE(Q.situacao_oc,0) = ?"); params_where.append(situacao_oc)
    if data_emissao_ini: where_parts.append("CAST(Q.data_emissao AS DATE) >= ?"); params_where.append(data_emissao_ini)
    if data_emissao_fim: where_parts.append("CAST(Q.data_emissao AS DATE) <= ?"); params_where.append(data_emissao_fim)
    if data_entrega_ini: where_parts.append("CAST(Q.data_entrega AS DATE) >= ?"); params_where.append(data_entrega_ini)
    if data_entrega_fim: where_parts.append("CAST(Q.data_entrega AS DATE) <= ?"); params_where.append(data_entrega_fim)
    if origem_material: where_parts.append("Q.origem_material LIKE ?"); params_where.append(f"%{origem_material}%")
    if familia: where_parts.append("Q.codigo_familia LIKE ?"); params_where.append(f"%{familia}%")
    sc = normalizar_situacao_cadastro(situacao_cadastro)
    if sc == 'ATIVO': where_parts.append("(Q.tipo_item<>'PRODUTO' OR COALESCE(Q.situacao_cadastro_produto,'A')='A')")
    elif sc == 'INATIVO': where_parts.append("(Q.tipo_item<>'PRODUTO' OR COALESCE(Q.situacao_cadastro_produto,'A')='I')")
    if somente_pendentes: where_parts.append("COALESCE(Q.saldo_pendente,0) > 0")
    where_sql = " AND ".join(where_parts)

    try:
        if agrupar_por_fornecedor:
            sql_total = cte_sql + f"SELECT COUNT(*) FROM (SELECT Q.codigo_fornecedor FROM Q WHERE {where_sql} GROUP BY Q.codigo_fornecedor) G"
            sql_dados = cte_sql + f"""
                SELECT CAST(Q.codigo_fornecedor AS INT) AS codigo_fornecedor,
                    MAX(Q.nome_fornecedor) AS nome_fornecedor, MAX(Q.fantasia_fornecedor) AS fantasia_fornecedor,
                    COUNT(*) AS quantidade_itens, COUNT(DISTINCT Q.numero_oc) AS quantidade_ocs,
                    CAST(SUM(COALESCE(Q.saldo_pendente,0)) AS FLOAT) AS saldo_pendente_total,
                    CAST(SUM(COALESCE(Q.valor_pendente_estimado,0)) AS FLOAT) AS valor_pendente_total,
                    CAST(SUM(COALESCE(Q.valor_liquido,0)) AS FLOAT) AS valor_liquido_total,
                    CAST(SUM(COALESCE(Q.valor_bruto,0)) AS FLOAT) AS valor_bruto_total,
                    SUM(CASE WHEN COALESCE(Q.dias_atraso,0)>0 THEN 1 ELSE 0 END) AS itens_atrasados
                FROM Q WHERE {where_sql}
                GROUP BY Q.codigo_fornecedor ORDER BY MAX(Q.nome_fornecedor)
                OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"""
        else:
            sql_total = cte_sql + f"SELECT COUNT(*) FROM Q WHERE {where_sql}"
            sql_dados = cte_sql + f"""
                SELECT Q.* FROM Q WHERE {where_sql}
                ORDER BY Q.data_emissao DESC, Q.numero_oc DESC, Q.sequencia_item ASC
                OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"""

        cursor.execute(sql_total, params_where); total_registros = cursor.fetchone()[0]
        cursor.execute(sql_dados, params_where + [offset, tamanho_pagina])
        rows = cursor.fetchall(); columns = [col[0] for col in cursor.description]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro SQL Painel de Compras: {str(e)}")
    finally:
        conn.close()

    dados = [dict(zip(columns, [v.strip() if isinstance(v,str) else v for v in row])) for row in rows]
    total_paginas = (total_registros + tamanho_pagina - 1) // tamanho_pagina if total_registros > 0 else 1
    return {'pagina': pagina, 'tamanho_pagina': tamanho_pagina,
            'total_registros': total_registros, 'total_paginas': total_paginas,
            'agrupar_por_fornecedor': agrupar_por_fornecedor, 'dados': dados}


# =========================================
# CONTAS A RECEBER / PAGAR — HELPERS
# =========================================

def _safe_strip(value): return (value or '').strip() if isinstance(value, str) else value

def normalizar_status_conta_receber(valor):
    v = (valor or 'TODOS').strip().upper()
    mapa = {'ABERTO':'EM_ABERTO','EM_ABERTO':'EM_ABERTO','PARCIAL':'PARCIAL',
            'LIQUIDADO':'LIQUIDADO','RECEBIDO':'LIQUIDADO','VENCIDO':'VENCIDO',
            'A_VENCER':'A_VENCER','AVENCER':'A_VENCER','TODOS':'TODOS'}
    return mapa.get(v, 'TODOS')

def normalizar_status_conta_pagar(valor):
    v = (valor or 'TODOS').strip().upper()
    mapa = {'ABERTO':'EM_ABERTO','EM_ABERTO':'EM_ABERTO','PARCIAL':'PARCIAL',
            'LIQUIDADO':'LIQUIDADO','PAGO':'LIQUIDADO','VENCIDO':'VENCIDO',
            'A_VENCER':'A_VENCER','AVENCER':'A_VENCER','TODOS':'TODOS'}
    return mapa.get(v, 'TODOS')

def _montar_cte_contas_receber():
    return """
        WITH MOV AS (
            SELECT M.CodEmp, M.CodFil, M.CodTpt, M.NumTit,
                MAX(M.DatMov) AS ultima_data_movimento,
                SUM(CAST(COALESCE(M.VlrMov,0) AS FLOAT)) AS valor_movimentado,
                COUNT(*) AS quantidade_movimentos
            FROM E301MCR M GROUP BY M.CodEmp, M.CodFil, M.CodTpt, M.NumTit
        ),
        BASE AS (
            SELECT T.CodEmp AS codigo_empresa, T.CodFil AS codigo_filial,
                COALESCE(T.CodTpt,'') AS tipo_titulo, COALESCE(T.NumTit,'') AS numero_titulo,
                CAST(COALESCE(T.CodCli,0) AS INT) AS codigo_cliente,
                COALESCE(C.NomCli,'') AS nome_cliente, COALESCE(C.ApeCli,'') AS fantasia_cliente,
                COALESCE(T.CodSnf,'') AS serie_nf, CAST(COALESCE(T.NumNfv,0) AS INT) AS numero_nf,
                T.DatEmi AS data_emissao, T.VctPro AS data_vencimento,
                MOV.ultima_data_movimento AS data_ultimo_movimento,
                CAST(COALESCE(T.VlrOri,0) AS FLOAT) AS valor_original,
                CAST(COALESCE(T.VlrAbe,0) AS FLOAT) AS valor_aberto,
                CAST(COALESCE(MOV.valor_movimentado,0) AS FLOAT) AS valor_movimentado,
                CAST(COALESCE(T.VlrOri,0)-COALESCE(T.VlrAbe,0) AS FLOAT) AS valor_recebido,
                CAST(COALESCE(T.VlrCom,0) AS FLOAT) AS valor_comissao,
                CAST(COALESCE(T.VlrBco,0) AS FLOAT) AS base_comissao,
                COALESCE(MOV.quantidade_movimentos,0) AS quantidade_movimentos,
                CASE WHEN COALESCE(T.VlrAbe,0)<=0 THEN 'LIQUIDADO'
                     WHEN COALESCE(T.VlrAbe,0)<COALESCE(T.VlrOri,0) THEN
                        CASE WHEN T.VctPro IS NOT NULL AND CAST(T.VctPro AS DATE)<CAST(GETDATE() AS DATE) THEN 'VENCIDO' ELSE 'PARCIAL' END
                     WHEN T.VctPro IS NOT NULL AND CAST(T.VctPro AS DATE)<CAST(GETDATE() AS DATE) THEN 'VENCIDO'
                     ELSE 'A_VENCER' END AS status_titulo,
                CASE WHEN T.VctPro IS NOT NULL AND COALESCE(T.VlrAbe,0)>0
                      AND CAST(T.VctPro AS DATE)<CAST(GETDATE() AS DATE)
                    THEN DATEDIFF(DAY, CAST(T.VctPro AS DATE), CAST(GETDATE() AS DATE))
                    ELSE 0 END AS dias_atraso
            FROM E301TCR T
            LEFT JOIN MOV ON MOV.CodEmp=T.CodEmp AND MOV.CodFil=T.CodFil
                AND MOV.CodTpt=T.CodTpt AND MOV.NumTit=T.NumTit
            LEFT JOIN E085CLI C ON C.CodCli=T.CodCli
        )
    """

def _montar_cte_contas_pagar():
    return """
        WITH MOV AS (
            SELECT M.CodEmp, M.CodFil, M.CodTpt, M.NumTit,
                MAX(M.DatMov) AS ultima_data_movimento,
                SUM(CAST(COALESCE(M.VlrMov,0) AS FLOAT)) AS valor_movimentado,
                COUNT(*) AS quantidade_movimentos
            FROM E501MCP M GROUP BY M.CodEmp, M.CodFil, M.CodTpt, M.NumTit
        ),
        BASE AS (
            SELECT T.CodEmp AS codigo_empresa, T.CodFil AS codigo_filial,
                COALESCE(T.CodTpt,'') AS tipo_titulo, COALESCE(T.NumTit,'') AS numero_titulo,
                CAST(COALESCE(T.CodFor,0) AS INT) AS codigo_fornecedor,
                COALESCE(F.NomFor,'') AS nome_fornecedor, COALESCE(F.ApeFor,'') AS fantasia_fornecedor,
                T.DatEmi AS data_emissao, T.VctPro AS data_vencimento,
                MOV.ultima_data_movimento AS data_ultimo_movimento,
                CAST(COALESCE(T.VlrOri,0) AS FLOAT) AS valor_original,
                CAST(COALESCE(T.VlrAbe,0) AS FLOAT) AS valor_aberto,
                CAST(COALESCE(MOV.valor_movimentado,0) AS FLOAT) AS valor_movimentado,
                CAST(COALESCE(T.VlrOri,0)-COALESCE(T.VlrAbe,0) AS FLOAT) AS valor_pago,
                COALESCE(MOV.quantidade_movimentos,0) AS quantidade_movimentos,
                CASE WHEN COALESCE(T.VlrAbe,0)<=0 THEN 'PAGO'
                     WHEN COALESCE(T.VlrAbe,0)<COALESCE(T.VlrOri,0) THEN
                        CASE WHEN T.VctPro IS NOT NULL AND CAST(T.VctPro AS DATE)<CAST(GETDATE() AS DATE) THEN 'VENCIDO' ELSE 'PARCIAL' END
                     WHEN T.VctPro IS NOT NULL AND CAST(T.VctPro AS DATE)<CAST(GETDATE() AS DATE) THEN 'VENCIDO'
                     ELSE 'A_VENCER' END AS status_titulo,
                CASE WHEN T.VctPro IS NOT NULL AND COALESCE(T.VlrAbe,0)>0
                      AND CAST(T.VctPro AS DATE)<CAST(GETDATE() AS DATE)
                    THEN DATEDIFF(DAY, CAST(T.VctPro AS DATE), CAST(GETDATE() AS DATE))
                    ELSE 0 END AS dias_atraso
            FROM E501TCP T
            LEFT JOIN MOV ON MOV.CodEmp=T.CodEmp AND MOV.CodFil=T.CodFil
                AND MOV.CodTpt=T.CodTpt AND MOV.NumTit=T.NumTit
            LEFT JOIN E095FOR F ON F.CodFor=T.CodFor
        )
    """

def _where_contas(where_parts, params_where, cliente=None, numero_titulo=None,
                  tipo_titulo=None, numero_nf=None, serie_nf=None, codigo_filial=None,
                  status_titulo='TODOS', somente_vencidos=False, somente_saldo_aberto=False,
                  somente_cheques=False, data_emissao_ini=None, data_emissao_fim=None,
                  data_vencimento_ini=None, data_vencimento_fim=None,
                  data_movimento_ini=None, data_movimento_fim=None,
                  valor_min=None, valor_max=None, entidade='cliente'):
    cod_f = 'codigo_cliente' if entidade == 'cliente' else 'codigo_fornecedor'
    nom_f = 'nome_cliente' if entidade == 'cliente' else 'nome_fornecedor'
    fan_f = 'fantasia_cliente' if entidade == 'cliente' else 'fantasia_fornecedor'
    pessoa = _safe_strip(cliente)
    if pessoa:
        like = f"%{pessoa}%"
        where_parts.append(f"(CAST(BASE.{cod_f} AS VARCHAR(20)) LIKE ? OR BASE.{nom_f} LIKE ? OR BASE.{fan_f} LIKE ?)")
        params_where.extend([like, like, like])
    if _safe_strip(numero_titulo): where_parts.append("BASE.numero_titulo LIKE ?"); params_where.append(f"%{_safe_strip(numero_titulo)}%")
    if _safe_strip(tipo_titulo): where_parts.append("BASE.tipo_titulo LIKE ?"); params_where.append(f"%{_safe_strip(tipo_titulo)}%")
    if _safe_strip(numero_nf): where_parts.append("CAST(BASE.numero_nf AS VARCHAR(20)) LIKE ?"); params_where.append(f"%{_safe_strip(numero_nf)}%")
    if _safe_strip(serie_nf): where_parts.append("BASE.serie_nf LIKE ?"); params_where.append(f"%{_safe_strip(serie_nf)}%")
    if _safe_strip(codigo_filial):
        cf = _safe_strip(codigo_filial)
        if cf.isdigit(): where_parts.append("BASE.codigo_filial = ?"); params_where.append(int(cf))
        else: where_parts.append("CAST(BASE.codigo_filial AS VARCHAR(20)) LIKE ?"); params_where.append(f"%{cf}%")
    if status_titulo != 'TODOS':
        if status_titulo == 'EM_ABERTO': where_parts.append("COALESCE(BASE.valor_aberto,0) > 0")
        else: where_parts.append("BASE.status_titulo = ?"); params_where.append(status_titulo)
    if somente_vencidos: where_parts.append("BASE.status_titulo = 'VENCIDO'")
    if somente_saldo_aberto: where_parts.append("COALESCE(BASE.valor_aberto,0) > 0")
    if somente_cheques: where_parts.append("BASE.tipo_titulo IN ('CHQ','CH1','CH2')")
    if data_emissao_ini: where_parts.append("CAST(BASE.data_emissao AS DATE) >= ?"); params_where.append(data_emissao_ini)
    if data_emissao_fim: where_parts.append("CAST(BASE.data_emissao AS DATE) <= ?"); params_where.append(data_emissao_fim)
    if data_vencimento_ini: where_parts.append("CAST(BASE.data_vencimento AS DATE) >= ?"); params_where.append(data_vencimento_ini)
    if data_vencimento_fim: where_parts.append("CAST(BASE.data_vencimento AS DATE) <= ?"); params_where.append(data_vencimento_fim)
    if data_movimento_ini: where_parts.append("CAST(BASE.data_ultimo_movimento AS DATE) >= ?"); params_where.append(data_movimento_ini)
    if data_movimento_fim: where_parts.append("CAST(BASE.data_ultimo_movimento AS DATE) <= ?"); params_where.append(data_movimento_fim)
    if valor_min is not None: where_parts.append("COALESCE(BASE.valor_original,0) >= ?"); params_where.append(valor_min)
    if valor_max is not None: where_parts.append("COALESCE(BASE.valor_original,0) <= ?"); params_where.append(valor_max)


@app.get('/api/contas-receber')
def consultar_contas_receber(
    cliente: Optional[str] = None, numero_titulo: Optional[str] = None,
    tipo_titulo: Optional[str] = None, numero_nf: Optional[str] = None,
    serie_nf: Optional[str] = None, codigo_filial: Optional[str] = None,
    status_titulo: str = 'TODOS', somente_vencidos: bool = False,
    somente_saldo_aberto: bool = False, somente_cheques: bool = False,
    data_emissao_ini: Optional[str] = None, data_emissao_fim: Optional[str] = None,
    data_vencimento_ini: Optional[str] = None, data_vencimento_fim: Optional[str] = None,
    data_movimento_ini: Optional[str] = None, data_movimento_fim: Optional[str] = None,
    valor_min: Optional[float] = None, valor_max: Optional[float] = None,
    agrupar_por_cliente: bool = False, pagina: int = 1, tamanho_pagina: int = 100,
    usuario=Depends(validar_token)
):
    if pagina < 1: pagina = 1
    if tamanho_pagina < 1 or tamanho_pagina > 100: tamanho_pagina = 100
    offset = (pagina - 1) * tamanho_pagina
    status_titulo = normalizar_status_conta_receber(status_titulo)
    conn = get_connection(); cursor = conn.cursor()
    cte_sql = _montar_cte_contas_receber()
    where_parts = ["BASE.codigo_empresa = ?"]; params_where = [EMPRESA_PADRAO]
    _where_contas(where_parts, params_where, cliente=cliente, numero_titulo=numero_titulo,
                  tipo_titulo=tipo_titulo, numero_nf=numero_nf, serie_nf=serie_nf,
                  codigo_filial=codigo_filial, status_titulo=status_titulo,
                  somente_vencidos=somente_vencidos, somente_saldo_aberto=somente_saldo_aberto,
                  somente_cheques=somente_cheques, data_emissao_ini=data_emissao_ini,
                  data_emissao_fim=data_emissao_fim, data_vencimento_ini=data_vencimento_ini,
                  data_vencimento_fim=data_vencimento_fim, data_movimento_ini=data_movimento_ini,
                  data_movimento_fim=data_movimento_fim, valor_min=valor_min, valor_max=valor_max,
                  entidade='cliente')
    where_sql = " AND ".join(where_parts)
    try:
        if agrupar_por_cliente:
            sql_total = cte_sql + f"SELECT COUNT(*) FROM (SELECT BASE.codigo_cliente FROM BASE WHERE {where_sql} GROUP BY BASE.codigo_cliente) X"
            sql_dados = cte_sql + f"""
                SELECT CAST(BASE.codigo_cliente AS INT) AS codigo_cliente,
                    MAX(BASE.nome_cliente) AS nome_cliente, MAX(BASE.fantasia_cliente) AS fantasia_cliente,
                    COUNT(*) AS quantidade_titulos,
                    SUM(CASE WHEN COALESCE(BASE.valor_aberto,0)>0 THEN 1 ELSE 0 END) AS titulos_em_aberto,
                    SUM(CASE WHEN BASE.status_titulo='VENCIDO' THEN 1 ELSE 0 END) AS titulos_vencidos,
                    CAST(SUM(COALESCE(BASE.valor_original,0)) AS FLOAT) AS valor_original_total,
                    CAST(SUM(COALESCE(BASE.valor_aberto,0)) AS FLOAT) AS valor_aberto_total,
                    CAST(SUM(COALESCE(BASE.valor_recebido,0)) AS FLOAT) AS valor_recebido_total,
                    MAX(COALESCE(BASE.dias_atraso,0)) AS maior_atraso_dias
                FROM BASE WHERE {where_sql}
                GROUP BY BASE.codigo_cliente ORDER BY MAX(BASE.nome_cliente)
                OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"""
        else:
            sql_total = cte_sql + f"SELECT COUNT(*) FROM BASE WHERE {where_sql}"
            sql_dados = cte_sql + f"""
                SELECT BASE.* FROM BASE WHERE {where_sql}
                ORDER BY CASE WHEN BASE.status_titulo='VENCIDO' THEN 0 ELSE 1 END,
                    BASE.data_vencimento ASC, BASE.numero_titulo ASC
                OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"""
        cursor.execute(sql_total, params_where); total_registros = cursor.fetchone()[0]
        cursor.execute(sql_dados, params_where + [offset, tamanho_pagina])
        rows = cursor.fetchall(); columns = [col[0] for col in cursor.description]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro SQL Contas a Receber: {str(e)}")
    finally:
        conn.close()
    dados = [dict(zip(columns, [v.strip() if isinstance(v,str) else v for v in row])) for row in rows]
    total_paginas = max(1, (total_registros + tamanho_pagina - 1) // tamanho_pagina)
    return {'pagina': pagina, 'tamanho_pagina': tamanho_pagina, 'total_registros': total_registros,
            'total_paginas': total_paginas, 'agrupar_por_cliente': agrupar_por_cliente, 'dados': dados}


@app.get('/api/export/contas-receber')
def exportar_contas_receber_excel(
    cliente: Optional[str] = None, numero_titulo: Optional[str] = None,
    tipo_titulo: Optional[str] = None, numero_nf: Optional[str] = None,
    serie_nf: Optional[str] = None, codigo_filial: Optional[str] = None,
    status_titulo: str = 'TODOS', somente_vencidos: bool = False,
    somente_saldo_aberto: bool = False, somente_cheques: bool = False,
    data_emissao_ini: Optional[str] = None, data_emissao_fim: Optional[str] = None,
    data_vencimento_ini: Optional[str] = None, data_vencimento_fim: Optional[str] = None,
    data_movimento_ini: Optional[str] = None, data_movimento_fim: Optional[str] = None,
    valor_min: Optional[float] = None, valor_max: Optional[float] = None,
    agrupar_por_cliente: bool = False, usuario=Depends(validar_token)
):
    dados = _collect_paginated_data(consultar_contas_receber, usuario,
        cliente=cliente, numero_titulo=numero_titulo, tipo_titulo=tipo_titulo,
        numero_nf=numero_nf, serie_nf=serie_nf, codigo_filial=codigo_filial,
        status_titulo=status_titulo, somente_vencidos=somente_vencidos,
        somente_saldo_aberto=somente_saldo_aberto, somente_cheques=somente_cheques,
        data_emissao_ini=data_emissao_ini, data_emissao_fim=data_emissao_fim,
        data_vencimento_ini=data_vencimento_ini, data_vencimento_fim=data_vencimento_fim,
        data_movimento_ini=data_movimento_ini, data_movimento_fim=data_movimento_fim,
        valor_min=valor_min, valor_max=valor_max, agrupar_por_cliente=agrupar_por_cliente)
    nome = 'contas_receber_agrupado.xlsx' if agrupar_por_cliente else 'contas_receber.xlsx'
    return _xlsx_response(nome, [('Contas Receber', dados, None)])


@app.get('/api/contas-pagar')
def consultar_contas_pagar(
    fornecedor: Optional[str] = None, numero_titulo: Optional[str] = None,
    tipo_titulo: Optional[str] = None, codigo_filial: Optional[str] = None,
    status_titulo: str = 'TODOS', somente_vencidos: bool = False,
    somente_saldo_aberto: bool = False, somente_cheques: bool = False,
    data_emissao_ini: Optional[str] = None, data_emissao_fim: Optional[str] = None,
    data_vencimento_ini: Optional[str] = None, data_vencimento_fim: Optional[str] = None,
    data_movimento_ini: Optional[str] = None, data_movimento_fim: Optional[str] = None,
    valor_min: Optional[float] = None, valor_max: Optional[float] = None,
    agrupar_por_fornecedor: bool = False, pagina: int = 1, tamanho_pagina: int = 100,
    usuario=Depends(validar_token)
):
    if pagina < 1: pagina = 1
    if tamanho_pagina < 1 or tamanho_pagina > 100: tamanho_pagina = 100
    offset = (pagina - 1) * tamanho_pagina
    status_titulo = normalizar_status_conta_pagar(status_titulo)
    conn = get_connection(); cursor = conn.cursor()
    cte_sql = _montar_cte_contas_pagar()
    where_parts = ["BASE.codigo_empresa = ?"]; params_where = [EMPRESA_PADRAO]
    _where_contas(where_parts, params_where, cliente=fornecedor, numero_titulo=numero_titulo,
                  tipo_titulo=tipo_titulo, codigo_filial=codigo_filial, status_titulo=status_titulo,
                  somente_vencidos=somente_vencidos, somente_saldo_aberto=somente_saldo_aberto,
                  somente_cheques=somente_cheques, data_emissao_ini=data_emissao_ini,
                  data_emissao_fim=data_emissao_fim, data_vencimento_ini=data_vencimento_ini,
                  data_vencimento_fim=data_vencimento_fim, data_movimento_ini=data_movimento_ini,
                  data_movimento_fim=data_movimento_fim, valor_min=valor_min, valor_max=valor_max,
                  entidade='fornecedor')
    where_sql = " AND ".join(where_parts)
    try:
        if agrupar_por_fornecedor:
            sql_total = cte_sql + f"SELECT COUNT(*) FROM (SELECT BASE.codigo_fornecedor FROM BASE WHERE {where_sql} GROUP BY BASE.codigo_fornecedor) X"
            sql_dados = cte_sql + f"""
                SELECT CAST(BASE.codigo_fornecedor AS INT) AS codigo_fornecedor,
                    MAX(BASE.nome_fornecedor) AS nome_fornecedor, MAX(BASE.fantasia_fornecedor) AS fantasia_fornecedor,
                    COUNT(*) AS quantidade_titulos,
                    SUM(CASE WHEN COALESCE(BASE.valor_aberto,0)>0 THEN 1 ELSE 0 END) AS titulos_em_aberto,
                    SUM(CASE WHEN BASE.status_titulo='VENCIDO' THEN 1 ELSE 0 END) AS titulos_vencidos,
                    CAST(SUM(COALESCE(BASE.valor_original,0)) AS FLOAT) AS valor_original_total,
                    CAST(SUM(COALESCE(BASE.valor_aberto,0)) AS FLOAT) AS valor_aberto_total,
                    CAST(SUM(COALESCE(BASE.valor_pago,0)) AS FLOAT) AS valor_pago_total,
                    MAX(COALESCE(BASE.dias_atraso,0)) AS maior_atraso_dias
                FROM BASE WHERE {where_sql}
                GROUP BY BASE.codigo_fornecedor ORDER BY MAX(BASE.nome_fornecedor)
                OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"""
        else:
            sql_total = cte_sql + f"SELECT COUNT(*) FROM BASE WHERE {where_sql}"
            sql_dados = cte_sql + f"""
                SELECT BASE.* FROM BASE WHERE {where_sql}
                ORDER BY CASE WHEN BASE.status_titulo='VENCIDO' THEN 0 ELSE 1 END,
                    BASE.data_vencimento ASC, BASE.numero_titulo ASC
                OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"""
        cursor.execute(sql_total, params_where); total_registros = cursor.fetchone()[0]
        cursor.execute(sql_dados, params_where + [offset, tamanho_pagina])
        rows = cursor.fetchall(); columns = [col[0] for col in cursor.description]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro SQL Contas a Pagar: {str(e)}")
    finally:
        conn.close()
    dados = [dict(zip(columns, [v.strip() if isinstance(v,str) else v for v in row])) for row in rows]
    total_paginas = max(1, (total_registros + tamanho_pagina - 1) // tamanho_pagina)
    return {'pagina': pagina, 'tamanho_pagina': tamanho_pagina, 'total_registros': total_registros,
            'total_paginas': total_paginas, 'agrupar_por_fornecedor': agrupar_por_fornecedor, 'dados': dados}


@app.get('/api/export/contas-pagar')
def exportar_contas_pagar_excel(
    fornecedor: Optional[str] = None, numero_titulo: Optional[str] = None,
    tipo_titulo: Optional[str] = None, codigo_filial: Optional[str] = None,
    status_titulo: str = 'TODOS', somente_vencidos: bool = False,
    somente_saldo_aberto: bool = False, somente_cheques: bool = False,
    data_emissao_ini: Optional[str] = None, data_emissao_fim: Optional[str] = None,
    data_vencimento_ini: Optional[str] = None, data_vencimento_fim: Optional[str] = None,
    data_movimento_ini: Optional[str] = None, data_movimento_fim: Optional[str] = None,
    valor_min: Optional[float] = None, valor_max: Optional[float] = None,
    agrupar_por_fornecedor: bool = False, usuario=Depends(validar_token)
):
    dados = _collect_paginated_data(consultar_contas_pagar, usuario,
        fornecedor=fornecedor, numero_titulo=numero_titulo, tipo_titulo=tipo_titulo,
        codigo_filial=codigo_filial, status_titulo=status_titulo, somente_vencidos=somente_vencidos,
        somente_saldo_aberto=somente_saldo_aberto, somente_cheques=somente_cheques,
        data_emissao_ini=data_emissao_ini, data_emissao_fim=data_emissao_fim,
        data_vencimento_ini=data_vencimento_ini, data_vencimento_fim=data_vencimento_fim,
        data_movimento_ini=data_movimento_ini, data_movimento_fim=data_movimento_fim,
        valor_min=valor_min, valor_max=valor_max, agrupar_por_fornecedor=agrupar_por_fornecedor)
    nome = 'contas_pagar_agrupado.xlsx' if agrupar_por_fornecedor else 'contas_pagar.xlsx'
    return _xlsx_response(nome, [('Contas Pagar', dados, None)])


# =========================================
# EXPORTS EXTRAS
# =========================================

@app.get('/api/export/notas-recebimento')
def exportar_notas_recebimento_excel(
    fornecedor: Optional[str] = None, situacao_nf: Optional[str] = None,
    numero_nf: Optional[str] = None, numero_nf_exato: bool = True,
    serie_nf: Optional[str] = None, codigo_filial: Optional[str] = None,
    codigo_item: Optional[str] = None, descricao_item: Optional[str] = None,
    centro_custo: Optional[str] = None, numero_projeto: Optional[str] = None,
    transacao: Optional[str] = None, origem_material: Optional[str] = None,
    familia: Optional[str] = None, deposito: Optional[str] = None,
    numero_oc_origem: Optional[str] = None, situacao_cadastro: str = 'TODOS',
    data_emissao_ini: Optional[str] = None, data_emissao_fim: Optional[str] = None,
    data_recebimento_ini: Optional[str] = None, data_recebimento_fim: Optional[str] = None,
    tipo_item: str = 'TODOS', valor_min: Optional[float] = None,
    valor_max: Optional[float] = None, usuario=Depends(validar_token)
):
    dados = _collect_paginated_data(consultar_notas_recebimento, usuario,
        fornecedor=fornecedor, situacao_nf=situacao_nf, numero_nf=numero_nf,
        numero_nf_exato=numero_nf_exato, serie_nf=serie_nf, codigo_filial=codigo_filial,
        codigo_item=codigo_item, descricao_item=descricao_item, centro_custo=centro_custo,
        numero_projeto=numero_projeto, transacao=transacao, origem_material=origem_material,
        familia=familia, deposito=deposito, numero_oc_origem=numero_oc_origem,
        situacao_cadastro=situacao_cadastro, data_emissao_ini=data_emissao_ini,
        data_emissao_fim=data_emissao_fim, data_recebimento_ini=data_recebimento_ini,
        data_recebimento_fim=data_recebimento_fim, tipo_item=tipo_item,
        valor_min=valor_min, valor_max=valor_max)
    return _xlsx_response('notas_recebimento.xlsx', [('Notas Recebimento', dados, None)])


@app.get('/api/export/painel-compras')
def exportar_painel_compras_excel(
    codigo_produto: Optional[str] = None, codigo_item: Optional[str] = None,
    descricao_item: Optional[str] = None, fornecedor: Optional[str] = None,
    numero_oc: Optional[str] = None, numero_projeto: Optional[str] = None,
    centro_custo: Optional[str] = None, transacao: Optional[str] = None,
    valor_min: Optional[float] = None, valor_max: Optional[float] = None,
    tipo_item: str = 'TODOS', tipo_oc: str = 'TODOS',
    situacao_oc: Optional[int] = None, data_emissao_ini: Optional[str] = None,
    data_emissao_fim: Optional[str] = None, data_entrega_ini: Optional[str] = None,
    data_entrega_fim: Optional[str] = None, origem_material: Optional[str] = None,
    familia: Optional[str] = None, situacao_cadastro: str = 'TODOS',
    somente_pendentes: bool = True, agrupar_por_fornecedor: bool = False,
    usuario=Depends(validar_token)
):
    dados = _collect_paginated_data(consultar_painel_compras, usuario,
        codigo_produto=codigo_produto, codigo_item=codigo_item, descricao_item=descricao_item,
        fornecedor=fornecedor, numero_oc=numero_oc, numero_projeto=numero_projeto,
        centro_custo=centro_custo, transacao=transacao, valor_min=valor_min,
        valor_max=valor_max, tipo_item=tipo_item, tipo_oc=tipo_oc, situacao_oc=situacao_oc,
        data_emissao_ini=data_emissao_ini, data_emissao_fim=data_emissao_fim,
        data_entrega_ini=data_entrega_ini, data_entrega_fim=data_entrega_fim,
        origem_material=origem_material, familia=familia, situacao_cadastro=situacao_cadastro,
        somente_pendentes=somente_pendentes, agrupar_por_fornecedor=agrupar_por_fornecedor)
    nome = 'painel_compras_agrupado.xlsx' if agrupar_por_fornecedor else 'painel_compras.xlsx'
    return _xlsx_response(nome, [('Painel Compras', dados, None)])


@app.get('/api/export/notas-edocs-conciliacao')
def exportar_notas_edocs_conciliacao_excel(
    tipo_nota: str = 'TODOS', numero_nf: Optional[str] = None,
    serie_nf: Optional[str] = None, codigo_filial: Optional[str] = None,
    codigo_pessoa: Optional[str] = None, nome_pessoa: Optional[str] = None,
    numero_lote: Optional[str] = None, situacao_erp: Optional[str] = None,
    situacao_edocs: Optional[str] = None, status_conciliacao: str = 'TODOS',
    data_ini: Optional[str] = None, data_fim: Optional[str] = None,
    somente_divergencia: bool = False, somente_sem_edocs: bool = False,
    somente_com_erro: bool = False, usuario=Depends(validar_token)
):
    dados = _collect_paginated_data(consultar_notas_edocs_conciliacao, usuario,
        tipo_nota=tipo_nota, numero_nf=numero_nf, serie_nf=serie_nf,
        codigo_filial=codigo_filial, codigo_pessoa=codigo_pessoa, nome_pessoa=nome_pessoa,
        numero_lote=numero_lote, situacao_erp=situacao_erp, situacao_edocs=situacao_edocs,
        status_conciliacao=status_conciliacao, data_ini=data_ini, data_fim=data_fim,
        somente_divergencia=somente_divergencia, somente_sem_edocs=somente_sem_edocs,
        somente_com_erro=somente_com_erro)
    return _xlsx_response('notas_edocs_conciliacao.xlsx', [('Notas x EDocs', dados, None)])


# =========================================
# ENGENHARIA x PRODUÇÃO
# =========================================

@app.get('/api/engenharia-producao')
def consultar_engenharia_producao(
    unidade_negocio: str = 'TODAS', numero_projeto: Optional[str] = None,
    numero_desenho: Optional[str] = None, revisao: Optional[str] = None,
    numero_op: Optional[str] = None, origem: Optional[str] = None,
    familia: Optional[str] = None, data_entrega_ini: Optional[str] = None,
    data_entrega_fim: Optional[str] = None, status_atendimento: str = 'TODOS',
    status_producao: str = 'TODOS', status_estoque: str = 'TODOS',
    pagina: int = 1, tamanho_pagina: int = 100, usuario=Depends(validar_token)
):
    pagina = max(pagina, 1); tamanho_pagina = min(max(tamanho_pagina, 1), 100)
    offset = (pagina - 1) * tamanho_pagina
    unidade_negocio = (unidade_negocio or 'TODAS').strip().upper()
    if unidade_negocio not in ('TODAS', 'ESTRUTURAL', 'GENIUS'): unidade_negocio = 'TODAS'

    genius_sql = "'210', '220', '230', '235', '240', '250'"

    cte_sql = f"""
        WITH chaves AS (
            SELECT DISTINCT usu_codemp, usu_numprj, usu_numdes, usu_revdes FROM usu_t900prj
            UNION SELECT DISTINCT usu_codemp, usu_numprj, usu_numdes, usu_revdes FROM usu_t900cop
            UNION SELECT DISTINCT usu_codemp, usu_numprj, usu_numdes, usu_revdes FROM usu_t900qdo
        ),
        cop_aggr AS (
            SELECT c.usu_codemp, c.usu_numprj, c.usu_numdes, c.usu_revdes,
                COUNT(DISTINCT CASE WHEN COALESCE(c.usu_numorp,0)>0 THEN c.usu_numorp END) AS qtd_ops,
                STRING_AGG(CAST(COALESCE(c.usu_numorp,0) AS VARCHAR(20)),', ') AS ops,
                STRING_AGG(COALESCE(NULLIF(c.usu_codori,''),'SEM'),', ') AS origens,
                STRING_AGG(COALESCE(NULLIF(c.usu_codfam,''),'SEM'),', ') AS familias,
                MAX(CASE WHEN COALESCE(c.usu_codori,'') IN ({genius_sql}) THEN 1 ELSE 0 END) AS flag_genius,
                MAX(CASE WHEN COALESCE(c.usu_codori,'') NOT IN ({genius_sql}) OR COALESCE(c.usu_codori,'')='' THEN 1 ELSE 0 END) AS flag_estrutural
            FROM usu_t900cop c GROUP BY c.usu_codemp, c.usu_numprj, c.usu_numdes, c.usu_revdes
        ),
        qdo_aggr AS (
            SELECT q.usu_codemp, q.usu_numprj, q.usu_numdes, q.usu_revdes,
                CAST(SUM(COALESCE(q.usu_totkgs,0)) AS FLOAT) AS kg_estrutura
            FROM usu_t900qdo q GROUP BY q.usu_codemp, q.usu_numprj, q.usu_numdes, q.usu_revdes
        ),
        eep_aggr AS (
            SELECT e.usu_codemp, e.usu_numprj, e.usu_numdes,
                CAST(SUM(COALESCE(e.usu_qtdent,0)*COALESCE(e.usu_pesliq,0)) AS FLOAT) AS kg_entrada_estoque_calc
            FROM usu_t900eep e GROUP BY e.usu_codemp, e.usu_numprj, e.usu_numdes
        ),
        mvf_aggr AS (
            SELECT m.usu_codemp, m.usu_numprj, m.usu_numdes,
                CAST(SUM(COALESCE(m.usu_pesteo,0)) AS FLOAT) AS kg_produzido_calc
            FROM usu_t900mvf m GROUP BY m.usu_codemp, m.usu_numprj, m.usu_numdes
        )
    """

    from_sql = f"""
        FROM (
            SELECT k.usu_codemp AS codigo_empresa, k.usu_numprj AS numero_projeto,
                k.usu_numdes AS numero_desenho, COALESCE(k.usu_revdes,'') AS revisao,
                COALESCE(p.usu_desprj,'') AS descricao_projeto,
                COALESCE(p.usu_desdes,'') AS descricao_desenho,
                CAST(COALESCE(NULLIF(p.usu_pescpr,0), NULLIF(qd.kg_estrutura,0), 0) AS FLOAT) AS kg_engenharia,
                CAST(COALESCE(qd.kg_estrutura,0) AS FLOAT) AS kg_estrutura,
                CAST(COALESCE(NULLIF(p.usu_pesfab,0),mv.kg_produzido_calc,0) AS FLOAT) AS kg_produzido,
                CAST(COALESCE(NULLIF(p.usu_pesrcp,0),ee.kg_entrada_estoque_calc,0) AS FLOAT) AS kg_entrada_estoque,
                CAST(COALESCE(NULLIF(p.usu_pescpr,0),NULLIF(qd.kg_estrutura,0),0)-COALESCE(NULLIF(p.usu_pesfab,0),mv.kg_produzido_calc,0) AS FLOAT) AS kg_gap_producao,
                CAST(COALESCE(NULLIF(p.usu_pescpr,0),NULLIF(qd.kg_estrutura,0),0)-COALESCE(NULLIF(p.usu_pesrcp,0),ee.kg_entrada_estoque_calc,0) AS FLOAT) AS kg_gap_estoque,
                CAST(CASE WHEN COALESCE(NULLIF(p.usu_pescpr,0),NULLIF(qd.kg_estrutura,0),0)>0
                    THEN (COALESCE(NULLIF(p.usu_pesfab,0),mv.kg_produzido_calc,0)/COALESCE(NULLIF(p.usu_pescpr,0),NULLIF(qd.kg_estrutura,0),0))*100
                    ELSE 0 END AS FLOAT) AS perc_atendimento_producao,
                CAST(CASE WHEN COALESCE(NULLIF(p.usu_pescpr,0),NULLIF(qd.kg_estrutura,0),0)>0
                    THEN (COALESCE(NULLIF(p.usu_pesrcp,0),ee.kg_entrada_estoque_calc,0)/COALESCE(NULLIF(p.usu_pescpr,0),NULLIF(qd.kg_estrutura,0),0))*100
                    ELSE 0 END AS FLOAT) AS perc_atendimento_estoque,
                CAST(COALESCE(ca.qtd_ops,0) AS INT) AS qtd_ops,
                COALESCE(ca.ops,'') AS ops, COALESCE(ca.origens,'') AS origens,
                COALESCE(ca.familias,'') AS familias,
                CASE WHEN COALESCE(NULLIF(p.usu_pescpr,0),NULLIF(qd.kg_estrutura,0),0)<=0 THEN 'SEM BASE'
                     WHEN COALESCE(NULLIF(p.usu_pesfab,0),mv.kg_produzido_calc,0)>=COALESCE(NULLIF(p.usu_pescpr,0),NULLIF(qd.kg_estrutura,0),0) THEN 'ATENDEU'
                     WHEN COALESCE(NULLIF(p.usu_pesfab,0),mv.kg_produzido_calc,0)>0 THEN 'PARCIAL'
                     ELSE 'SEM PRODUÇÃO' END AS status_producao,
                CASE WHEN COALESCE(NULLIF(p.usu_pescpr,0),NULLIF(qd.kg_estrutura,0),0)<=0 THEN 'SEM BASE'
                     WHEN COALESCE(NULLIF(p.usu_pesrcp,0),ee.kg_entrada_estoque_calc,0)>=COALESCE(NULLIF(p.usu_pescpr,0),NULLIF(qd.kg_estrutura,0),0) THEN 'ATENDEU'
                     WHEN COALESCE(NULLIF(p.usu_pesrcp,0),ee.kg_entrada_estoque_calc,0)>0 THEN 'PARCIAL'
                     ELSE 'SEM ENTRADA' END AS status_estoque,
                COALESCE(ca.flag_genius,0) AS flag_genius,
                COALESCE(ca.flag_estrutural,1) AS flag_estrutural,
                p.usu_datetr AS data_entrega_engenharia_raw
            FROM chaves k
            LEFT JOIN usu_t900prj p ON p.usu_codemp=k.usu_codemp AND p.usu_numprj=k.usu_numprj
                AND p.usu_numdes=k.usu_numdes AND p.usu_revdes=k.usu_revdes
            LEFT JOIN cop_aggr ca ON ca.usu_codemp=k.usu_codemp AND ca.usu_numprj=k.usu_numprj
                AND ca.usu_numdes=k.usu_numdes AND ca.usu_revdes=k.usu_revdes
            LEFT JOIN qdo_aggr qd ON qd.usu_codemp=k.usu_codemp AND qd.usu_numprj=k.usu_numprj
                AND qd.usu_numdes=k.usu_numdes AND qd.usu_revdes=k.usu_revdes
            LEFT JOIN eep_aggr ee ON ee.usu_codemp=k.usu_codemp AND ee.usu_numprj=k.usu_numprj
                AND ee.usu_numdes=k.usu_numdes
            LEFT JOIN mvf_aggr mv ON mv.usu_codemp=k.usu_codemp AND mv.usu_numprj=k.usu_numprj
                AND mv.usu_numdes=k.usu_numdes
        ) base
        WHERE base.codigo_empresa = ?
          AND (COALESCE(base.kg_engenharia,0)>0 OR COALESCE(base.kg_estrutura,0)>0
               OR COALESCE(base.kg_produzido,0)>0 OR COALESCE(base.kg_entrada_estoque,0)>0)
    """
    params_where = [EMPRESA_PADRAO]

    if numero_projeto: from_sql += " AND CAST(base.numero_projeto AS VARCHAR(20)) LIKE ?"; params_where.append(f"%{numero_projeto}%")
    if numero_desenho: from_sql += " AND CAST(base.numero_desenho AS VARCHAR(20)) LIKE ?"; params_where.append(f"%{numero_desenho}%")
    if revisao: from_sql += " AND base.revisao LIKE ?"; params_where.append(f"%{revisao}%")
    if numero_op: from_sql += " AND COALESCE(base.ops,'') LIKE ?"; params_where.append(f"%{numero_op}%")
    if origem:
        origens_filtradas = [x.strip() for x in re.split(r'[;,\s]+', origem) if x.strip()]
        if origens_filtradas:
            clausulas = []; [clausulas.append("COALESCE(base.origens,'') LIKE ?") or params_where.append(f"%{o}%") for o in origens_filtradas]
            from_sql += " AND (" + " OR ".join(clausulas) + ")"
    if familia: from_sql += " AND COALESCE(base.familias,'') LIKE ?"; params_where.append(f"%{familia}%")
    if data_entrega_ini: from_sql += " AND base.data_entrega_engenharia_raw >= ?"; params_where.append(data_entrega_ini)
    if data_entrega_fim: from_sql += " AND base.data_entrega_engenharia_raw < DATEADD(day,1,?)"; params_where.append(data_entrega_fim)
    if unidade_negocio == 'ESTRUTURAL': from_sql += " AND COALESCE(base.flag_estrutural,1)=1"
    elif unidade_negocio == 'GENIUS': from_sql += " AND COALESCE(base.flag_genius,0)=1"
    if status_producao != 'TODOS': from_sql += " AND COALESCE(base.status_producao,'')=?"; params_where.append(status_producao)
    if status_estoque != 'TODOS': from_sql += " AND COALESCE(base.status_estoque,'')=?"; params_where.append(status_estoque)

    conn = get_connection(); cursor = conn.cursor()
    try:
        sql_resumo = cte_sql + f"""
            SELECT COUNT(*) AS total_registros,
                CAST(COALESCE(SUM(base.kg_engenharia),0) AS FLOAT) AS kg_engenharia_total,
                CAST(COALESCE(SUM(base.kg_produzido),0) AS FLOAT) AS kg_produzido_total,
                CAST(COALESCE(SUM(base.kg_entrada_estoque),0) AS FLOAT) AS kg_entrada_estoque_total,
                SUM(CASE WHEN base.status_producao='ATENDEU' THEN 1 ELSE 0 END) AS projetos_atendidos_producao,
                SUM(CASE WHEN base.status_estoque='ATENDEU' THEN 1 ELSE 0 END) AS projetos_atendidos_estoque
            {from_sql}"""
        cursor.execute(sql_resumo, params_where); row_r = cursor.fetchone()
        total_registros = int(row_r[0] or 0)
        resumo = {
            'total_registros': total_registros, 'total_projetos': total_registros,
            'kg_engenharia_total': float(row_r[1] or 0), 'kg_produzido_total': float(row_r[2] or 0),
            'kg_entrada_estoque_total': float(row_r[3] or 0),
            'projetos_atendidos_producao': int(row_r[4] or 0), 'projetos_atendidos_estoque': int(row_r[5] or 0),
        }
        kg_e = resumo['kg_engenharia_total']
        resumo['perc_atendimento_producao_total'] = resumo['kg_produzido_total']/kg_e*100 if kg_e else 0
        resumo['perc_atendimento_estoque_total'] = resumo['kg_entrada_estoque_total']/kg_e*100 if kg_e else 0

        sql_dados = cte_sql + f"""
            SELECT base.codigo_empresa, base.numero_projeto, base.numero_desenho, base.revisao,
                base.descricao_projeto, base.descricao_desenho, base.kg_engenharia,
                base.kg_estrutura, base.kg_produzido, base.kg_entrada_estoque,
                base.perc_atendimento_producao, base.perc_atendimento_estoque,
                base.kg_gap_producao, base.kg_gap_estoque, base.qtd_ops, base.ops,
                base.origens, base.familias, base.status_producao, base.status_estoque,
                CAST(NULL AS DATETIME) AS ultima_data_estoque
            {from_sql}
            ORDER BY CASE WHEN base.data_entrega_engenharia_raw IS NULL THEN 1 ELSE 0 END,
                base.data_entrega_engenharia_raw DESC, base.numero_projeto DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"""
        cursor.execute(sql_dados, params_where + [offset, tamanho_pagina])
        rows = cursor.fetchall(); columns = [col[0] for col in cursor.description]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro SQL Engenharia x Produção: {str(e)}")
    finally:
        conn.close()

    dados = [{col: row[i] for i, col in enumerate(columns)} for row in rows]
    total_paginas = (total_registros + tamanho_pagina - 1) // tamanho_pagina if total_registros > 0 else 1
    return {'pagina': pagina, 'tamanho_pagina': tamanho_pagina,
            'total_registros': total_registros, 'total_paginas': total_paginas,
            'dados': dados, 'resumo': resumo}


@app.get('/api/export/engenharia-producao')
def exportar_engenharia_producao_excel(
    unidade_negocio: str = 'ESTRUTURAL', numero_projeto: Optional[str] = None,
    numero_desenho: Optional[str] = None, revisao: Optional[str] = None,
    numero_op: Optional[str] = None, origem: Optional[str] = None,
    familia: Optional[str] = None, data_entrega_ini: Optional[str] = None,
    data_entrega_fim: Optional[str] = None, status_atendimento: str = 'TODOS',
    status_producao: str = 'TODOS', status_estoque: str = 'TODOS',
    usuario=Depends(validar_token)
):
    dados = _collect_paginated_data(consultar_engenharia_producao, usuario,
        unidade_negocio=unidade_negocio, numero_projeto=numero_projeto,
        numero_desenho=numero_desenho, revisao=revisao, numero_op=numero_op,
        origem=origem, familia=familia, data_entrega_ini=data_entrega_ini,
        data_entrega_fim=data_entrega_fim, status_atendimento=status_atendimento,
        status_producao=status_producao, status_estoque=status_estoque)
    return _xlsx_response('engenharia_producao.xlsx', [('Engenharia Producao', dados, None)])


# =============================================================
# MÓDULO DE REQUISIÇÃO DE MATERIAIS
# =============================================================
# Tabelas necessárias — execute UMA VEZ no SQL Server:
#
# IF OBJECT_ID('dbo.USU_REQ_SEQ','U') IS NULL
# BEGIN
#   CREATE TABLE dbo.USU_REQ_SEQ (
#       USU_CODEMP INT NOT NULL, USU_CODFIL INT NOT NULL,
#       USU_NUMREQ INT NOT NULL DEFAULT 0,
#       CONSTRAINT PK_USU_REQ_SEQ PRIMARY KEY (USU_CODEMP, USU_CODFIL)
#   );
#   INSERT INTO dbo.USU_REQ_SEQ VALUES (1, 1, 0);
# END;
#
# IF OBJECT_ID('dbo.USU_REQ_CAB','U') IS NULL
# CREATE TABLE dbo.USU_REQ_CAB (
#     USU_IDREQ  INT IDENTITY(1,1) PRIMARY KEY,
#     USU_CODEMP INT NOT NULL DEFAULT 1, USU_CODFIL INT NOT NULL DEFAULT 1,
#     USU_NUMREQ INT NOT NULL, USU_CODUSU VARCHAR(30) NOT NULL,
#     USU_NOMUSU VARCHAR(100) NULL, USU_DATEMI DATETIME NOT NULL DEFAULT GETDATE(),
#     USU_CODCCU VARCHAR(30) NULL, USU_DESCCU VARCHAR(100) NULL,
#     USU_NUMPRJ INT NULL,    USU_NOMPRJ VARCHAR(100) NULL,
#     USU_CODDEP VARCHAR(10) NULL,  USU_DATNEC DATETIME NULL,
#     USU_MOTIVO VARCHAR(500) NULL, USU_STATUS VARCHAR(30) NOT NULL DEFAULT 'DIGITACAO',
#     USU_ORIGEM VARCHAR(30) NOT NULL DEFAULT 'APP_REQ',
#     USU_OBS    VARCHAR(1000) NULL
# );
#
# IF OBJECT_ID('dbo.USU_REQ_ITE','U') IS NULL
# CREATE TABLE dbo.USU_REQ_ITE (
#     USU_IDREQITE INT IDENTITY(1,1) PRIMARY KEY,
#     USU_IDREQ  INT NOT NULL, USU_SEQITE INT NOT NULL,
#     USU_CODPRO VARCHAR(30) NOT NULL,  USU_CODDER VARCHAR(30) NULL,
#     USU_DESPRO VARCHAR(200) NOT NULL, USU_CODFAM VARCHAR(10) NULL,
#     USU_CODORI VARCHAR(10) NULL,      USU_UNIMED VARCHAR(10) NULL,
#     USU_QTDSOL DECIMAL(18,6) NOT NULL, USU_QTDAPR DECIMAL(18,6) NULL,
#     USU_QTDATE DECIMAL(18,6) NULL,     USU_QTDSAL DECIMAL(18,6) NULL,
#     USU_INDCMP CHAR(1) NOT NULL DEFAULT 'N', USU_OBS VARCHAR(500) NULL
# );
#
# IF OBJECT_ID('dbo.USU_REQ_STATUS','U') IS NULL
# CREATE TABLE dbo.USU_REQ_STATUS (
#     USU_IDREQSTS INT IDENTITY(1,1) PRIMARY KEY,
#     USU_IDREQ  INT NOT NULL, USU_STSANT VARCHAR(30) NULL,
#     USU_STSNOV VARCHAR(30) NOT NULL, USU_DATSTS DATETIME NOT NULL DEFAULT GETDATE(),
#     USU_CODUSU VARCHAR(30) NOT NULL, USU_OBS    VARCHAR(500) NULL
# );
# =============================================================

# --- Modelos Pydantic ---

class _ItemReqBody(BaseModel):
    codpro: str
    codder: Optional[str] = None
    despro: Optional[str] = None
    codfam: Optional[str] = None
    codori: Optional[str] = None
    unimed: Optional[str] = None
    qtdsol: float
    obs:    Optional[str] = None

class _CriarReqBody(BaseModel):
    codccu: Optional[str] = None
    desccu: Optional[str] = None
    numprj: Optional[int] = None
    nomprj: Optional[str] = None
    coddep: Optional[str] = None
    datnec: Optional[str] = None
    motivo: Optional[str] = None
    obs:    Optional[str] = None
    itens:  list = []

class _AtualizarReqBody(BaseModel):
    codccu: Optional[str] = None
    desccu: Optional[str] = None
    numprj: Optional[int] = None
    nomprj: Optional[str] = None
    coddep: Optional[str] = None
    datnec: Optional[str] = None
    motivo: Optional[str] = None
    obs:    Optional[str] = None

class _AtualizarItemReqBody(BaseModel):
    qtdsol: Optional[float] = None
    obs:    Optional[str]   = None

class _StatusReqBody(BaseModel):
    obs: Optional[str] = None

# --- Helpers internos ---

def _req_proximo_numero(cursor, codemp: int = 1, codfil: int = 1) -> int:
    cursor.execute(
        "UPDATE dbo.USU_REQ_SEQ SET USU_NUMREQ = USU_NUMREQ + 1 "
        "OUTPUT INSERTED.USU_NUMREQ WHERE USU_CODEMP=? AND USU_CODFIL=?",
        [codemp, codfil]
    )
    row = cursor.fetchone()
    if not row:
        cursor.execute(
            "INSERT INTO dbo.USU_REQ_SEQ (USU_CODEMP,USU_CODFIL,USU_NUMREQ) VALUES (?,?,1)",
            [codemp, codfil]
        )
        return 1
    return int(row[0])


def _req_log_status(cursor, idreq: int, ant: Optional[str],
                    nov: str, usuario: str, obs: Optional[str] = None):
    cursor.execute(
        "INSERT INTO dbo.USU_REQ_STATUS "
        "(USU_IDREQ,USU_STSANT,USU_STSNOV,USU_CODUSU,USU_OBS) VALUES (?,?,?,?,?)",
        [idreq, ant, nov, usuario, obs]
    )


def _req_saldo_produto(cursor, codemp: int, codpro: str,
                       codder: Optional[str] = None) -> float:
    try:
        if codder:
            cursor.execute(
                "SELECT COALESCE(SUM(CAST(QTDEST AS FLOAT)),0) FROM E210EST "
                "WHERE CODEMP=? AND CODPRO=? AND COALESCE(CODDER,'')=?",
                [codemp, codpro, codder or '']
            )
        else:
            cursor.execute(
                "SELECT COALESCE(SUM(CAST(QTDEST AS FLOAT)),0) FROM E210EST "
                "WHERE CODEMP=? AND CODPRO=?", [codemp, codpro]
            )
        row = cursor.fetchone()
        return float(row[0] or 0) if row else 0.0
    except Exception:
        return 0.0


def _req_dados_produto(cursor, codemp: int, codpro: str) -> dict:
    try:
        cursor.execute(
            "SELECT TOP 1 COALESCE(DESPRO,''),COALESCE(UNIMED,''),"
            "COALESCE(CODFAM,''),COALESCE(CODORI,'') "
            "FROM E075PRO WHERE CODEMP=? AND CODPRO=?", [codemp, codpro]
        )
        row = cursor.fetchone()
        if row:
            return {'despro': (row[0] or '').strip(), 'unimed': (row[1] or '').strip(),
                    'codfam': (row[2] or '').strip(), 'codori': (row[3] or '').strip()}
    except Exception:
        pass
    return {'despro': codpro, 'unimed': '', 'codfam': '', 'codori': ''}


def _req_row(cursor, row) -> dict:
    cols = [col[0] for col in cursor.description]
    return {col: (row[i].strip() if isinstance(row[i], str) else row[i])
            for i, col in enumerate(cols)}


def _req_verificar_status(cursor, idreq: int, codemp: int,
                           permitidos: Optional[list] = None) -> str:
    cursor.execute(
        "SELECT USU_STATUS FROM dbo.USU_REQ_CAB WHERE USU_IDREQ=? AND USU_CODEMP=?",
        [idreq, codemp]
    )
    row = cursor.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail='Requisição não encontrada.')
    status = row[0]
    if permitidos and status not in permitidos:
        raise HTTPException(
            status_code=400,
            detail=f'Operação não permitida. Status atual: {status}. '
                   f'Permitido em: {", ".join(permitidos)}.'
        )
    return status


# --- Rotas — Cabeçalho ---

@app.get('/api/requisicoes')
def req_listar(
    status:         Optional[str] = None,
    codusu:         Optional[str] = None,
    codccu:         Optional[str] = None,
    numprj:         Optional[int] = None,
    datnec_ini:     Optional[str] = None,
    datnec_fim:     Optional[str] = None,
    datemi_ini:     Optional[str] = None,
    datemi_fim:     Optional[str] = None,
    pagina:         int           = 1,
    tamanho_pagina: int           = 100,
    usuario=Depends(validar_token),
):
    pagina         = max(pagina, 1)
    tamanho_pagina = min(max(tamanho_pagina, 1), 200)
    offset         = (pagina - 1) * tamanho_pagina

    where  = "WHERE C.USU_CODEMP = ?"
    params = [EMPRESA_PADRAO]

    if status:     where += " AND C.USU_STATUS = ?";         params.append(status.upper().strip())
    if codusu:     where += " AND UPPER(C.USU_CODUSU) LIKE ?"; params.append(f"%{codusu.upper().strip()}%")
    if codccu:     where += " AND C.USU_CODCCU LIKE ?";      params.append(f"%{codccu.strip()}%")
    if numprj:     where += " AND C.USU_NUMPRJ = ?";         params.append(numprj)
    if datnec_ini: where += " AND CAST(C.USU_DATNEC AS DATE) >= ?"; params.append(datnec_ini)
    if datnec_fim: where += " AND CAST(C.USU_DATNEC AS DATE) <= ?"; params.append(datnec_fim)
    if datemi_ini: where += " AND CAST(C.USU_DATEMI AS DATE) >= ?"; params.append(datemi_ini)
    if datemi_fim: where += " AND CAST(C.USU_DATEMI AS DATE) <= ?"; params.append(datemi_fim)

    conn = get_connection(); cursor = conn.cursor()
    try:
        cursor.execute(f"SELECT COUNT(*) FROM dbo.USU_REQ_CAB C {where}", params)
        total = int(cursor.fetchone()[0] or 0)

        cursor.execute(f"""
            SELECT C.USU_IDREQ, C.USU_NUMREQ, C.USU_CODUSU, C.USU_NOMUSU,
                   C.USU_DATEMI, C.USU_DATNEC, C.USU_STATUS, C.USU_MOTIVO,
                   C.USU_CODCCU, C.USU_DESCCU, C.USU_NUMPRJ, C.USU_NOMPRJ,
                   C.USU_CODDEP, C.USU_OBS, C.USU_ORIGEM,
                   (SELECT COUNT(*) FROM dbo.USU_REQ_ITE I
                    WHERE I.USU_IDREQ=C.USU_IDREQ) AS QTD_ITENS
            FROM dbo.USU_REQ_CAB C {where}
            ORDER BY C.USU_DATEMI DESC, C.USU_IDREQ DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """, params + [offset, tamanho_pagina])
        dados = [_req_row(cursor, r) for r in cursor.fetchall()]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro SQL Requisições: {e}")
    finally:
        conn.close()

    total_paginas = (total + tamanho_pagina - 1) // tamanho_pagina if total > 0 else 1
    return {'pagina': pagina, 'tamanho_pagina': tamanho_pagina,
            'total_registros': total, 'total_paginas': total_paginas, 'dados': dados}


@app.get('/api/requisicoes/dashboard')
def req_dashboard(usuario=Depends(validar_token)):
    conn = get_connection(); cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT USU_STATUS, COUNT(*) AS QUANTIDADE,
                   COUNT(DISTINCT USU_CODUSU) AS USUARIOS
            FROM dbo.USU_REQ_CAB WHERE USU_CODEMP=? GROUP BY USU_STATUS
        """, [EMPRESA_PADRAO])
        por_status = [_req_row(cursor, r) for r in cursor.fetchall()]

        cursor.execute("""
            SELECT TOP 5 USU_CODUSU AS usuario, COUNT(*) AS total_requisicoes
            FROM dbo.USU_REQ_CAB
            WHERE USU_CODEMP=? AND USU_DATEMI >= DATEADD(DAY,-30,GETDATE())
            GROUP BY USU_CODUSU ORDER BY COUNT(*) DESC
        """, [EMPRESA_PADRAO])
        top_usuarios = [_req_row(cursor, r) for r in cursor.fetchall()]

        cursor.execute("""
            SELECT
                COUNT(*) AS total_req,
                SUM(CASE WHEN USU_STATUS='DIGITACAO'  THEN 1 ELSE 0 END) AS em_digitacao,
                SUM(CASE WHEN USU_STATUS='ENVIADA'    THEN 1 ELSE 0 END) AS enviadas,
                SUM(CASE WHEN USU_STATUS IN ('EM_APROVACAO','APROVADA') THEN 1 ELSE 0 END) AS em_aprovacao,
                SUM(CASE WHEN USU_STATUS='CANCELADA'  THEN 1 ELSE 0 END) AS canceladas,
                SUM(CASE WHEN USU_STATUS='ENTREGUE'   THEN 1 ELSE 0 END) AS entregues,
                SUM(CASE WHEN USU_STATUS='EM_COMPRA'  THEN 1 ELSE 0 END) AS em_compra
            FROM dbo.USU_REQ_CAB
            WHERE USU_CODEMP=? AND USU_DATEMI >= DATEADD(DAY,-30,GETDATE())
        """, [EMPRESA_PADRAO])
        kpis = _req_row(cursor, cursor.fetchone())
    finally:
        conn.close()
    return {'periodo': 'Últimos 30 dias', 'kpis': kpis,
            'por_status': por_status, 'top_usuarios': top_usuarios}


@app.get('/api/requisicoes/{idreq}')
def req_detalhe(idreq: int, usuario=Depends(validar_token)):
    conn = get_connection(); cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT * FROM dbo.USU_REQ_CAB WHERE USU_IDREQ=? AND USU_CODEMP=?",
            [idreq, EMPRESA_PADRAO]
        )
        row = cursor.fetchone()
        if not row: raise HTTPException(status_code=404, detail='Requisição não encontrada.')
        cab = _req_row(cursor, row)

        cursor.execute("SELECT * FROM dbo.USU_REQ_ITE WHERE USU_IDREQ=? ORDER BY USU_SEQITE", [idreq])
        itens = [_req_row(cursor, r) for r in cursor.fetchall()]

        cursor.execute("SELECT * FROM dbo.USU_REQ_STATUS WHERE USU_IDREQ=? ORDER BY USU_DATSTS", [idreq])
        historico = [_req_row(cursor, r) for r in cursor.fetchall()]
    finally:
        conn.close()
    return {'cabecalho': cab, 'itens': itens, 'historico': historico}


@app.post('/api/requisicoes')
def req_criar(body: _CriarReqBody, usuario=Depends(validar_token)):
    conn = get_connection(); cursor = conn.cursor()
    try:
        numreq = _req_proximo_numero(cursor)
        cursor.execute(
            """
            INSERT INTO dbo.USU_REQ_CAB
            (USU_CODEMP,USU_CODFIL,USU_NUMREQ,USU_CODUSU,USU_NOMUSU,
             USU_CODCCU,USU_DESCCU,USU_NUMPRJ,USU_NOMPRJ,USU_CODDEP,
             USU_DATNEC,USU_MOTIVO,USU_OBS,USU_STATUS,USU_ORIGEM)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,'DIGITACAO','APP_REQ')
            """,
            [EMPRESA_PADRAO, 1, numreq, usuario, usuario,
             body.codccu, body.desccu, body.numprj, body.nomprj,
             body.coddep, body.datnec, body.motivo, body.obs]
        )
        cursor.execute("SELECT SCOPE_IDENTITY()")
        idreq = int(cursor.fetchone()[0])
        _req_log_status(cursor, idreq, None, 'DIGITACAO', usuario, 'Requisição criada')

        for seq, it in enumerate(body.itens or [], start=1):
            it_dict = it if isinstance(it, dict) else (it.dict() if hasattr(it, 'dict') else vars(it))
            codpro  = it_dict.get('codpro', '')
            codder  = it_dict.get('codder')
            qtdsol  = float(it_dict.get('qtdsol', 0))
            if qtdsol <= 0: continue
            dp  = _req_dados_produto(cursor, EMPRESA_PADRAO, codpro)
            sal = _req_saldo_produto(cursor, EMPRESA_PADRAO, codpro, codder)
            cursor.execute(
                """
                INSERT INTO dbo.USU_REQ_ITE
                (USU_IDREQ,USU_SEQITE,USU_CODPRO,USU_CODDER,USU_DESPRO,
                 USU_CODFAM,USU_CODORI,USU_UNIMED,USU_QTDSOL,USU_QTDSAL,
                 USU_INDCMP,USU_OBS)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                [idreq, seq, codpro, codder or None,
                 it_dict.get('despro') or dp['despro'],
                 it_dict.get('codfam') or dp['codfam'],
                 it_dict.get('codori') or dp['codori'],
                 it_dict.get('unimed') or dp['unimed'],
                 qtdsol, sal, 'N' if sal >= qtdsol else 'S', it_dict.get('obs')]
            )
        conn.commit()
        return {'mensagem': 'Requisição criada com sucesso.', 'idreq': idreq, 'numreq': numreq}
    except HTTPException:
        conn.rollback(); raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f'Erro ao criar requisição: {e}')
    finally:
        conn.close()


@app.put('/api/requisicoes/{idreq}')
def req_atualizar(idreq: int, body: _AtualizarReqBody, usuario=Depends(validar_token)):
    conn = get_connection(); cursor = conn.cursor()
    try:
        _req_verificar_status(cursor, idreq, EMPRESA_PADRAO, ['DIGITACAO'])
        cursor.execute(
            """
            UPDATE dbo.USU_REQ_CAB SET
                USU_CODCCU=?,USU_DESCCU=?,USU_NUMPRJ=?,USU_NOMPRJ=?,
                USU_CODDEP=?,USU_DATNEC=?,USU_MOTIVO=?,USU_OBS=?
            WHERE USU_IDREQ=?
            """,
            [body.codccu, body.desccu, body.numprj, body.nomprj,
             body.coddep, body.datnec, body.motivo, body.obs, idreq]
        )
        conn.commit()
        return {'mensagem': 'Requisição atualizada com sucesso.'}
    except HTTPException:
        conn.rollback(); raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f'Erro ao atualizar: {e}')
    finally:
        conn.close()


@app.post('/api/requisicoes/{idreq}/enviar')
def req_enviar(idreq: int, body: _StatusReqBody = Body(default=None),
               usuario=Depends(validar_token)):
    conn = get_connection(); cursor = conn.cursor()
    try:
        _req_verificar_status(cursor, idreq, EMPRESA_PADRAO, ['DIGITACAO'])
        cursor.execute("SELECT COUNT(*) FROM dbo.USU_REQ_ITE WHERE USU_IDREQ=?", [idreq])
        if int(cursor.fetchone()[0] or 0) == 0:
            raise HTTPException(status_code=400, detail='Adicione ao menos um item antes de enviar.')
        cursor.execute("UPDATE dbo.USU_REQ_CAB SET USU_STATUS='ENVIADA' WHERE USU_IDREQ=?", [idreq])
        _req_log_status(cursor, idreq, 'DIGITACAO', 'ENVIADA', usuario,
                        body.obs if body else None)
        conn.commit()
        return {'mensagem': 'Requisição enviada com sucesso.', 'status': 'ENVIADA'}
    except HTTPException:
        conn.rollback(); raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f'Erro ao enviar: {e}')
    finally:
        conn.close()


@app.post('/api/requisicoes/{idreq}/cancelar')
def req_cancelar(idreq: int, body: _StatusReqBody = Body(default=None),
                 usuario=Depends(validar_token)):
    CANCELAVEIS = {'DIGITACAO', 'ENVIADA', 'EM_APROVACAO'}
    conn = get_connection(); cursor = conn.cursor()
    try:
        status_ant = _req_verificar_status(cursor, idreq, EMPRESA_PADRAO, list(CANCELAVEIS))
        cursor.execute("UPDATE dbo.USU_REQ_CAB SET USU_STATUS='CANCELADA' WHERE USU_IDREQ=?", [idreq])
        _req_log_status(cursor, idreq, status_ant, 'CANCELADA', usuario,
                        body.obs if body else None)
        conn.commit()
        return {'mensagem': 'Requisição cancelada.', 'status': 'CANCELADA'}
    except HTTPException:
        conn.rollback(); raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f'Erro ao cancelar: {e}')
    finally:
        conn.close()


@app.post('/api/requisicoes/{idreq}/duplicar')
def req_duplicar(idreq: int, usuario=Depends(validar_token)):
    conn = get_connection(); cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT * FROM dbo.USU_REQ_CAB WHERE USU_IDREQ=? AND USU_CODEMP=?",
            [idreq, EMPRESA_PADRAO]
        )
        row = cursor.fetchone()
        if not row: raise HTTPException(status_code=404, detail='Requisição original não encontrada.')
        origem = _req_row(cursor, row)

        cursor.execute("SELECT * FROM dbo.USU_REQ_ITE WHERE USU_IDREQ=? ORDER BY USU_SEQITE", [idreq])
        itens = [_req_row(cursor, r) for r in cursor.fetchall()]

        numreq_novo = _req_proximo_numero(cursor)
        cursor.execute(
            """
            INSERT INTO dbo.USU_REQ_CAB
            (USU_CODEMP,USU_CODFIL,USU_NUMREQ,USU_CODUSU,USU_NOMUSU,
             USU_CODCCU,USU_DESCCU,USU_NUMPRJ,USU_NOMPRJ,USU_CODDEP,
             USU_DATNEC,USU_MOTIVO,USU_OBS,USU_STATUS,USU_ORIGEM)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,'DIGITACAO','APP_REQ')
            """,
            [EMPRESA_PADRAO, 1, numreq_novo, usuario, usuario,
             origem.get('USU_CODCCU'), origem.get('USU_DESCCU'),
             origem.get('USU_NUMPRJ'), origem.get('USU_NOMPRJ'),
             origem.get('USU_CODDEP'), origem.get('USU_DATNEC'),
             f"Duplicada da requisição {origem.get('USU_NUMREQ')}",
             origem.get('USU_OBS')]
        )
        cursor.execute("SELECT SCOPE_IDENTITY()")
        idreq_novo = int(cursor.fetchone()[0])
        _req_log_status(cursor, idreq_novo, None, 'DIGITACAO', usuario,
                        f"Duplicada da requisição ID {idreq}")

        for it in itens:
            codpro = it.get('USU_CODPRO', '')
            codder = it.get('USU_CODDER')
            qtdsol = float(it.get('USU_QTDSOL') or 0)
            sal    = _req_saldo_produto(cursor, EMPRESA_PADRAO, codpro, codder)
            cursor.execute(
                """
                INSERT INTO dbo.USU_REQ_ITE
                (USU_IDREQ,USU_SEQITE,USU_CODPRO,USU_CODDER,USU_DESPRO,
                 USU_CODFAM,USU_CODORI,USU_UNIMED,USU_QTDSOL,USU_QTDSAL,
                 USU_INDCMP,USU_OBS)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                [idreq_novo, it.get('USU_SEQITE'), codpro, codder or None,
                 it.get('USU_DESPRO'), it.get('USU_CODFAM'), it.get('USU_CODORI'),
                 it.get('USU_UNIMED'), qtdsol, sal,
                 'N' if sal >= qtdsol else 'S', it.get('USU_OBS')]
            )
        conn.commit()
        return {'mensagem': 'Requisição duplicada.', 'idreq': idreq_novo, 'numreq': numreq_novo}
    except HTTPException:
        conn.rollback(); raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f'Erro ao duplicar: {e}')
    finally:
        conn.close()


# --- Rotas — Itens ---

@app.get('/api/requisicoes/{idreq}/itens')
def req_itens_listar(idreq: int, usuario=Depends(validar_token)):
    conn = get_connection(); cursor = conn.cursor()
    try:
        cursor.execute("SELECT 1 FROM dbo.USU_REQ_CAB WHERE USU_IDREQ=? AND USU_CODEMP=?",
                       [idreq, EMPRESA_PADRAO])
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail='Requisição não encontrada.')
        cursor.execute("SELECT * FROM dbo.USU_REQ_ITE WHERE USU_IDREQ=? ORDER BY USU_SEQITE", [idreq])
        dados = [_req_row(cursor, r) for r in cursor.fetchall()]
    finally:
        conn.close()
    return {'idreq': idreq, 'total': len(dados), 'dados': dados}


@app.post('/api/requisicoes/{idreq}/itens')
def req_itens_adicionar(idreq: int, body: _ItemReqBody, usuario=Depends(validar_token)):
    conn = get_connection(); cursor = conn.cursor()
    try:
        _req_verificar_status(cursor, idreq, EMPRESA_PADRAO, ['DIGITACAO'])
        if body.qtdsol <= 0:
            raise HTTPException(status_code=400, detail='Quantidade deve ser maior que zero.')
        cursor.execute(
            "SELECT COALESCE(MAX(USU_SEQITE),0)+1 FROM dbo.USU_REQ_ITE WHERE USU_IDREQ=?", [idreq]
        )
        seq = int(cursor.fetchone()[0] or 1)
        dp  = _req_dados_produto(cursor, EMPRESA_PADRAO, body.codpro)
        sal = _req_saldo_produto(cursor, EMPRESA_PADRAO, body.codpro, body.codder)
        cursor.execute(
            """
            INSERT INTO dbo.USU_REQ_ITE
            (USU_IDREQ,USU_SEQITE,USU_CODPRO,USU_CODDER,USU_DESPRO,
             USU_CODFAM,USU_CODORI,USU_UNIMED,USU_QTDSOL,USU_QTDSAL,
             USU_INDCMP,USU_OBS)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            [idreq, seq, body.codpro, body.codder or None,
             body.despro or dp['despro'], body.codfam or dp['codfam'],
             body.codori or dp['codori'], body.unimed or dp['unimed'],
             body.qtdsol, sal, 'N' if sal >= body.qtdsol else 'S', body.obs]
        )
        conn.commit()
        return {'mensagem': 'Item adicionado.', 'seqite': seq, 'saldo_atual': sal,
                'precisa_compra': 'S' if sal < body.qtdsol else 'N'}
    except HTTPException:
        conn.rollback(); raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f'Erro ao adicionar item: {e}')
    finally:
        conn.close()


@app.put('/api/requisicoes/{idreq}/itens/{seq}')
def req_itens_atualizar(idreq: int, seq: int, body: _AtualizarItemReqBody,
                        usuario=Depends(validar_token)):
    conn = get_connection(); cursor = conn.cursor()
    try:
        _req_verificar_status(cursor, idreq, EMPRESA_PADRAO, ['DIGITACAO'])
        sets, params = [], []
        if body.qtdsol is not None:
            if body.qtdsol <= 0:
                raise HTTPException(status_code=400, detail='Quantidade deve ser maior que zero.')
            sets.append('USU_QTDSOL=?'); params.append(body.qtdsol)
        if body.obs is not None:
            sets.append('USU_OBS=?'); params.append(body.obs)
        if not sets:
            raise HTTPException(status_code=400, detail='Nenhum campo para atualizar.')
        params += [idreq, seq]
        cursor.execute(
            f"UPDATE dbo.USU_REQ_ITE SET {', '.join(sets)} WHERE USU_IDREQ=? AND USU_SEQITE=?",
            params
        )
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail='Item não encontrado.')
        conn.commit()
        return {'mensagem': 'Item atualizado.'}
    except HTTPException:
        conn.rollback(); raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f'Erro ao atualizar item: {e}')
    finally:
        conn.close()


@app.delete('/api/requisicoes/{idreq}/itens/{seq}')
def req_itens_remover(idreq: int, seq: int, usuario=Depends(validar_token)):
    conn = get_connection(); cursor = conn.cursor()
    try:
        _req_verificar_status(cursor, idreq, EMPRESA_PADRAO, ['DIGITACAO'])
        cursor.execute(
            "DELETE FROM dbo.USU_REQ_ITE WHERE USU_IDREQ=? AND USU_SEQITE=?", [idreq, seq]
        )
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail='Item não encontrado.')
        conn.commit()
        return {'mensagem': 'Item removido.'}
    except HTTPException:
        conn.rollback(); raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f'Erro ao remover item: {e}')
    finally:
        conn.close()


# --- Catálogo de itens (busca para nova requisição) ---

@app.get('/api/catalogo-itens')
def req_catalogo_itens(
    q:             Optional[str] = None,
    codpro:        Optional[str] = None,
    codfam:        Optional[str] = None,
    codori:        Optional[str] = None,
    somente_ativos: bool         = True,
    com_saldo:      bool         = False,
    limite:         int          = 50,
    usuario=Depends(validar_token),
):
    limite = max(1, min(limite, 200))
    where  = "WHERE P.CODEMP = ?"
    params = [EMPRESA_PADRAO]

    if somente_ativos: where += " AND COALESCE(P.SITPRO,'A') = 'A'"
    if q:
        where += " AND (P.CODPRO LIKE ? OR P.DESPRO LIKE ?)"
        params.extend([f"%{q.strip()}%", f"%{q.strip()}%"])
    if codpro: where += " AND P.CODPRO LIKE ?"; params.append(f"%{codpro.strip()}%")
    if codfam: where += " AND P.CODFAM = ?";    params.append(codfam.strip())
    if codori: where += " AND P.CODORI = ?";    params.append(codori.strip())
    if com_saldo:
        where += (" AND EXISTS (SELECT 1 FROM E210EST E "
                  "WHERE E.CODEMP=P.CODEMP AND E.CODPRO=P.CODPRO AND E.QTDEST>0)")

    conn = get_connection(); cursor = conn.cursor()
    try:
        cursor.execute(f"""
            SELECT TOP {limite}
                P.CODPRO AS codigo, P.DESPRO AS descricao,
                COALESCE(P.CODFAM,'') AS familia, COALESCE(P.CODORI,'') AS origem,
                COALESCE(P.UNIMED,'') AS unidade_medida,
                COALESCE(P.TIPPRO,'') AS tipo, COALESCE(P.SITPRO,'A') AS situacao,
                COALESCE((SELECT SUM(CAST(QTDEST AS FLOAT)) FROM E210EST E
                          WHERE E.CODEMP=P.CODEMP AND E.CODPRO=P.CODPRO), 0.0) AS saldo_total
            {where} ORDER BY P.CODPRO
        """, params)
        dados = [_req_row(cursor, r) for r in cursor.fetchall()]
    finally:
        conn.close()
    return {'total': len(dados), 'dados': dados}


# --- Export Excel de requisições ---

@app.get('/api/export/requisicoes')
def req_exportar_excel(
    status:         Optional[str] = None,
    codusu:         Optional[str] = None,
    codccu:         Optional[str] = None,
    numprj:         Optional[int] = None,
    datnec_ini:     Optional[str] = None,
    datnec_fim:     Optional[str] = None,
    datemi_ini:     Optional[str] = None,
    datemi_fim:     Optional[str] = None,
    usuario=Depends(validar_token),
):
    dados = _collect_paginated_data(
        req_listar, usuario,
        status=status, codusu=codusu, codccu=codccu, numprj=numprj,
        datnec_ini=datnec_ini, datnec_fim=datnec_fim,
        datemi_ini=datemi_ini, datemi_fim=datemi_fim,
    )
    return _xlsx_response('requisicoes.xlsx', [('Requisições', dados, None)])


# =============================================================
# INICIALIZAÇÃO
# =============================================================


if __name__ == '__main__':
    import uvicorn
    uvicorn.run(app, host='0.0.0.0', port=8010)
