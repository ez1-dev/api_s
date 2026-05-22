import os
from dotenv import load_dotenv

from fastapi import FastAPI, HTTPException, Depends, Body, Query, BackgroundTasks
from fastapi.responses import HTMLResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from jose import jwt
from datetime import datetime, timedelta, timezone
from typing import Optional, List, Dict, Any
from pydantic import BaseModel
from decimal import Decimal
import pyodbc
import json
import re
import logging
import httpx
import google.generativeai as genai

# =========================================================
# CONFIG
# =========================================================

SECRET_KEY = "ERP_SECRET"
ALGORITHM = "HS256"

SQL_SERVER = "172.16.137.100"
SQL_DATABASE = "sapiens"
SQL_USER = "sapiens"
SQL_PASSWORD = "0n%lV'g0F94"
EMPRESA_PADRAO = 1

# Quando True, o endpoint /api/controle-fiscal-produtos/salvar é bloqueado.
# Toda alteração passa pelo workflow:
#   POST /api/auditoria-tributaria/solicitacoes
#   POST /api/auditoria-tributaria/solicitacoes/{id}/aprovar (x2)
#   POST /api/auditoria-tributaria/solicitacoes/{id}/aplicar
BLOQUEAR_SALVAR_DIRETO = True

API_HOST = "0.0.0.0"
API_PORT = 8055

# carrega variáveis do arquivo .env que estiver ao lado do app
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = os.path.join(BASE_DIR, ".env")

print(f"[ENV] BASE_DIR = {BASE_DIR}")
print(f"[ENV] procurando .env em = {ENV_PATH}")
print(f"[ENV] arquivo existe? {os.path.exists(ENV_PATH)}")

load_dotenv(ENV_PATH, override=True)

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip()
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-1.5-flash").strip()

print(f"[ENV] GEMINI_API_KEY carregada? {bool(GEMINI_API_KEY)}")
print(f"[ENV] GEMINI_MODEL = {GEMINI_MODEL}")

GEMINI_DISPONIVEL = bool(GEMINI_API_KEY)

if GEMINI_DISPONIVEL:
    genai.configure(api_key=GEMINI_API_KEY)
    print(f"[GEMINI] configurado com modelo: {GEMINI_MODEL}")
else:
    print("[GEMINI] indisponível: GEMINI_API_KEY não encontrada no .env")

# Webhook para novo item fiscal
FASTAPI_WEBHOOK_SECRET = os.getenv("FASTAPI_WEBHOOK_SECRET", "").strip()
WEBHOOK_NOVO_ITEM_URL = os.getenv(
    "WEBHOOK_NOVO_ITEM_URL",
    "https://mlhpwdrsxkxelltwzisi.supabase.co/functions/v1/notificar-novo-item-fiscal"
).strip()

logger = logging.getLogger("controle_fiscal_webhook")

if FASTAPI_WEBHOOK_SECRET and WEBHOOK_NOVO_ITEM_URL:
    print(f"[WEBHOOK] configurado para: {WEBHOOK_NOVO_ITEM_URL[:60]}...")
else:
    print("[WEBHOOK] não configurado (FASTAPI_WEBHOOK_SECRET ou WEBHOOK_NOVO_ITEM_URL ausentes)")

USERS = {
    "ADMIN": "123",
    "RENATO": "123",
    "TRIBUTOS": "123456"
}

# =========================================================
# APP
# =========================================================

app = FastAPI(title="Auditoria Tributária ERP Senior")
security = HTTPBearer()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================================================
# DB
# =========================================================

def get_connection():
    try:
        conn = pyodbc.connect(
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={SQL_SERVER},1433;"
            f"DATABASE={SQL_DATABASE};"
            f"UID={SQL_USER};"
            f"PWD={SQL_PASSWORD};"
            "Encrypt=no;"
            "TrustServerCertificate=yes;",
            timeout=20
        )
        return conn
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro conexão SQL: {e}")


def row_to_dict(cursor, row):
    cols = [c[0] for c in cursor.description]
    out = {}
    for i, col in enumerate(cols):
        val = row[i]
        if isinstance(val, Decimal):
            val = float(val)
        elif isinstance(val, datetime):
            val = val.isoformat()
        elif isinstance(val, str):
            val = val.strip()
        out[col] = val
    return out


def is_str_diff(a, b):
    aa = "" if a is None else str(a).strip()
    bb = "" if b is None else str(b).strip()
    return aa != bb and (aa != "" or bb != "")


def is_num_diff(a, b, tol=0.000001):
    if a is None and b is None:
        return False
    av = 0 if a is None else float(a)
    bv = 0 if b is None else float(b)
    return abs(av - bv) > tol

def _clean_str(v) -> str:
    """Converte qualquer valor para str limpa. Seguro para int, float e None."""
    if v is None:
        return ""
    return str(v).strip()



# =========================================================
# IA HELPERS
# =========================================================

def chamar_gemini(prompt: str, instrucao_sistema: str = "") -> Optional[Dict[str, Any]]:
    """Chama Gemini e extrai JSON da resposta. Retorna None se falhar, mas loga o motivo real."""
    if not GEMINI_DISPONIVEL:
        print("[GEMINI] indisponível: chave ausente")
        return None

    try:
        print(f"[GEMINI] chamando modelo: {GEMINI_MODEL}")

        model = genai.GenerativeModel(
            GEMINI_MODEL,
            system_instruction=instrucao_sistema or (
                "Você é um auditor tributário sênior especialista em legislação fiscal brasileira "
                "(PIS/COFINS, ICMS, IPI, INSS, ISS). Responda sempre em JSON válido, sem markdown extra, "
                "sem texto antes ou depois do JSON. Seja preciso e objetivo."
            )
        )

        resposta = model.generate_content(prompt)
        texto = (getattr(resposta, "text", "") or "").strip()

        if not texto:
            print("[GEMINI] resposta vazia")
            return None

        match = re.search(r"```json\s*([\s\S]*?)```", texto)
        if match:
            texto = match.group(1).strip()
        elif texto.startswith("```"):
            texto = re.sub(r"^```[a-z]*\n?", "", texto).rstrip("`").strip()

        try:
            return json.loads(texto)
        except Exception as json_err:
            print(f"[GEMINI] erro ao converter resposta em JSON: {json_err}")
            print(f"[GEMINI] resposta recebida: {texto[:2000]}")
            return None

    except Exception as e:
        print(f"[GEMINI] erro real: {type(e).__name__}: {e}")
        return None


def calcular_score_risco(item: Dict[str, Any]) -> Dict[str, Any]:
    """Calcula score de risco fiscal 0-100 localmente, sem API.

    Hierarquia de peso:
      - divergências reais (erro fiscal da NF): peso alto.
      - avisos cadastrais (saneamento de cadastro): peso reduzido.
      - pendências de mapeamento (campos não coletados): ignoradas.
    """
    score = 0
    fatores: List[str] = []
    divergencias = item.get("divergencias_reais", []) or []
    avisos = item.get("avisos_cadastrais", []) or []
    impostos = item.get("impostos", {}) or {}
    qtd_div = len(divergencias)
    qtd_avi = len(avisos)

    icms = impostos.get("icms", {}) or {}
    cadastro = impostos.get("cadastro_produto", {}) or {}

    # ICMS - apenas se houver divergência real
    if any("ICMS" in m.upper() for m in divergencias):
        score += 30; fatores.append("Divergência de ICMS")
    elif icms.get("item_aliq_icms") is None and item.get("tipo_item") == "PRODUTO" and item.get("movimento") in ("ENTRADA", "SAIDA"):
        score += 15; fatores.append("Alíquota ICMS ausente no item")

    # PIS/COFINS - apenas divergência real
    if any(("PIS" in m.upper() or "COFINS" in m.upper()) for m in divergencias):
        score += 25; fatores.append("Divergência PIS/COFINS")

    # NCM ausente (real)
    if any("sem NCM" in m for m in divergencias):
        score += 20; fatores.append("Produto sem NCM")

    # Documento sem transação fiscal (real)
    if any("sem transação" in m.lower() for m in divergencias):
        score += 15; fatores.append("Documento sem transação fiscal")

    # Conflito família x item (após hierarquia) - real
    if any("família" in m.lower() and "difere" in m.lower() for m in divergencias):
        score += 20; fatores.append("Conflito família x item (não justificado pela transação)")

    # Avisos cadastrais (saneamento) - impacto pequeno e limitado
    if qtd_avi > 0:
        peso_avisos = min(10, qtd_avi)  # no máximo +10 pontos
        score += peso_avisos
        fatores.append(f"{qtd_avi} aviso(s) de saneamento cadastral")

    # CodTrd ausente é cadastral (saneamento) - só pesa se tiver divergência junto
    if not _clean_str(cadastro.get("cad_codtrd")) and item.get("tipo_item") == "PRODUTO" and qtd_div > 0:
        score += 5; fatores.append("Cadastro sem CodTrd")

    # Acúmulo de divergências reais
    if qtd_div > 5:
        score += 10; fatores.append(f"{qtd_div} divergências reais acumuladas")
    elif qtd_div > 2:
        score += 5

    score = min(100, score)
    nivel = "CRITICO" if score >= 75 else "ALTO" if score >= 50 else "MEDIO" if score >= 25 else "BAIXO"
    return {
        "score_risco": score,
        "nivel_risco": nivel,
        "fatores_risco": fatores,
        "qtd_divergencias_reais": qtd_div,
        "qtd_avisos_cadastrais": qtd_avi,
    }


def pick_prefixed_fields(data: Dict[str, Any], prefix: str, extra_keys=None) -> Dict[str, Any]:
    """Coleta todos os campos com prefixo dado + chaves extras de um dict de linha."""
    extra = set(extra_keys or [])
    return {k: v for k, v in data.items() if k.startswith(prefix) or k in extra}

# Campos fiscais da familia (E012FAM) â€” usados nas queries de produto com JOIN FAM
FAMILIA_SQL_PROD = """
                FAM.CSTPIS AS fam_cst_pis,
                FAM.CSTCOF AS fam_cst_cofins,
                FAM.RECPIS AS fam_recpis,
                FAM.PERPIS AS fam_perpis,
                FAM.PERCOF AS fam_percof,
                FAM.CSTIPI AS fam_cst_ipi,
                FAM.PERIPI AS fam_peripi,
                FAM.PROIMP AS fam_proimp,
                FAM.TIPPRO AS fam_tippro,
                FAM.CODORI AS fam_codori,
                FAM.PERICM AS fam_pericm,
                FAM.CODSTR AS fam_codstr,
                FAM.TMIICM AS fam_tmiicm,
                FAM.CODTRD AS fam_codtrd,
                FAM.CODTST AS fam_codtst,
                FAM.CODSTP AS fam_codstp,
                FAM.TEMICM AS fam_temicm,
                FAM.RECICM AS fam_recicm,
                FAM.RECIPI AS fam_recipi,
                FAM.RECCOF AS fam_reccof"""

# Versao NULL da familia para queries sem JOIN E012FAM (servicos)
FAMILIA_SQL_NULL = """
                CAST(NULL AS VARCHAR(10)) AS fam_cst_pis,
                CAST(NULL AS VARCHAR(10)) AS fam_cst_cofins,
                CAST(NULL AS VARCHAR(1)) AS fam_recpis,
                CAST(NULL AS NUMERIC(15,4)) AS fam_perpis,
                CAST(NULL AS NUMERIC(15,4)) AS fam_percof,
                CAST(NULL AS VARCHAR(10)) AS fam_cst_ipi,
                CAST(NULL AS NUMERIC(15,4)) AS fam_peripi,
                CAST(NULL AS VARCHAR(1)) AS fam_proimp,
                CAST(NULL AS VARCHAR(1)) AS fam_tippro,
                CAST(NULL AS VARCHAR(20)) AS fam_codori,
                CAST(NULL AS NUMERIC(15,4)) AS fam_pericm,
                CAST(NULL AS VARCHAR(20)) AS fam_codstr,
                CAST(NULL AS NUMERIC(15,4)) AS fam_tmiicm,
                CAST(NULL AS VARCHAR(20)) AS fam_codtrd,
                CAST(NULL AS VARCHAR(20)) AS fam_codtst,
                CAST(NULL AS VARCHAR(20)) AS fam_codstp,
                CAST(NULL AS VARCHAR(1)) AS fam_temicm,
                CAST(NULL AS VARCHAR(1)) AS fam_recicm,
                CAST(NULL AS VARCHAR(1)) AS fam_recipi,
                CAST(NULL AS VARCHAR(1)) AS fam_reccof"""

# Campos fiscais completos do produto (E075PRO) â€” usados em sql_ent_prod, sql_sai_prod, sql_cad_prod
PRODUTO_FISCAL_SQL = """
                P.CODSTR AS cad_codstr,
                P.CODTIC AS cad_codtic,
                P.CODSTC AS cad_codstc,
                P.BASREC AS cad_basrec,
                P.BASCRE AS cad_bascre_produto,
                P.TRIPIS AS cad_tripis,
                P.TRICOF AS cad_tricof,
                P.CSTIPI AS cad_cstipi_produto,
                P.CSTPIS AS cad_cstpis_produto,
                P.CSTCOF AS cad_cstcof_produto,
                P.TPRPIS AS cad_tprpis,
                P.TPRCOF AS cad_tprcof,
                P.TPRIPI AS cad_tpripi,
                P.REGTRI AS cad_regtri,
                P.CSTIPC AS cad_cstipc,
                P.CSTPIC AS cad_cstpic,
                P.CSTCOC AS cad_cstcoc,
                P.ORIMER AS cad_orimer,
                P.NATPIS AS cad_natpis,
                P.NATCOF AS cad_natcof,
                P.TPRPII AS cad_tprpii,
                P.TPRCOI AS cad_tprcoi,
                P.PERIFP AS cad_perifp,
                P.PDIFCP AS cad_pdifcp,
                P.CODENQ AS cad_codenq,
                P.CODCES AS cad_codces,
                P.CODDFS AS cad_coddfs,
                P.ORIGTI AS cad_origti,
                P.CATPRO AS cad_catpro,
                P.ITEFIS AS cad_itefis,
                P.DESFIS AS cad_desfis,
                P.IMPSCF AS cad_impscf,
                P.PERDIF AS cad_perdif,
                P.EMIREC AS cad_emirec,
                P.IDEPAR AS cad_idepar,
                P.TIPCIC AS cad_tipcic,
                P.FICCAT AS cad_ficcat,
                P.USU_MCGRCP AS cad_usu_mcgrcp"""

# Versao NULL para queries sem acesso a E075PRO (servicos, etc.)
# Tipos espelham exatamente o schema da E075PRO para evitar erro de conversao no UNION ALL
PRODUTO_FISCAL_NULL = """
                CAST(NULL AS VARCHAR(20)) AS cad_codstr,
                CAST(NULL AS VARCHAR(20)) AS cad_codtic,
                CAST(NULL AS VARCHAR(20)) AS cad_codstc,
                CAST(NULL AS VARCHAR(20)) AS cad_basrec,
                CAST(NULL AS NUMERIC(15,4)) AS cad_bascre_produto,
                CAST(NULL AS VARCHAR(3))  AS cad_tripis,
                CAST(NULL AS VARCHAR(3))  AS cad_tricof,
                CAST(NULL AS VARCHAR(10)) AS cad_cstipi_produto,
                CAST(NULL AS VARCHAR(10)) AS cad_cstpis_produto,
                CAST(NULL AS VARCHAR(10)) AS cad_cstcof_produto,
                CAST(NULL AS VARCHAR(4))  AS cad_tprpis,
                CAST(NULL AS VARCHAR(4))  AS cad_tprcof,
                CAST(NULL AS VARCHAR(4))  AS cad_tpripi,
                CAST(NULL AS VARCHAR(2))  AS cad_regtri,
                CAST(NULL AS VARCHAR(10)) AS cad_cstipc,
                CAST(NULL AS VARCHAR(10)) AS cad_cstpic,
                CAST(NULL AS VARCHAR(10)) AS cad_cstcoc,
                CAST(NULL AS VARCHAR(2))  AS cad_orimer,
                CAST(NULL AS VARCHAR(10)) AS cad_natpis,
                CAST(NULL AS VARCHAR(10)) AS cad_natcof,
                CAST(NULL AS VARCHAR(4))  AS cad_tprpii,
                CAST(NULL AS VARCHAR(4))  AS cad_tprcoi,
                CAST(NULL AS NUMERIC(15,4)) AS cad_perifp,
                CAST(NULL AS NUMERIC(15,4)) AS cad_pdifcp,
                CAST(NULL AS VARCHAR(20)) AS cad_codenq,
                CAST(NULL AS VARCHAR(20)) AS cad_codces,
                CAST(NULL AS VARCHAR(20)) AS cad_coddfs,
                CAST(NULL AS VARCHAR(2))  AS cad_origti,
                CAST(NULL AS VARCHAR(2))  AS cad_catpro,
                CAST(NULL AS VARCHAR(2))  AS cad_itefis,
                CAST(NULL AS VARCHAR(250)) AS cad_desfis,
                CAST(NULL AS VARCHAR(2))  AS cad_impscf,
                CAST(NULL AS NUMERIC(15,4)) AS cad_perdif,
                CAST(NULL AS VARCHAR(2))  AS cad_emirec,
                CAST(NULL AS VARCHAR(2))  AS cad_idepar,
                CAST(NULL AS VARCHAR(2))  AS cad_tipcic,
                CAST(NULL AS VARCHAR(30)) AS cad_ficcat,
                CAST(NULL AS VARCHAR(30)) AS cad_usu_mcgrcp"""

# =========================================================
# METADADOS DE CAMPOS â€” label, tabela, campo ERP, grupo
# =========================================================

FIELD_META = {
    # === E075PRO â€” Produto (cad_) ===
    "cod_classificacao":      {"label": "Classificação fiscal",               "table": "E075PRO", "field": "CODCLF",  "group": "Produto"},
    "ncm":                    {"label": "NCM",                                  "table": "E022CLF", "field": "CLAFIS",  "group": "Produto"},
    "cad_recpis":             {"label": "Recupera PIS",                         "table": "E075PRO", "field": "RECPIS",  "group": "PIS/COFINS"},
    "cad_reccof":             {"label": "Recupera COFINS",                      "table": "E075PRO", "field": "RECCOF",  "group": "PIS/COFINS"},
    "cad_peripi":             {"label": "Alíquota IPI",                         "table": "E075PRO", "field": "PERIPI",  "group": "IPI"},
    "cad_recipi":             {"label": "Recupera IPI",                         "table": "E075PRO", "field": "RECIPI",  "group": "IPI"},
    "cad_temicm":             {"label": "Tem ICMS",                             "table": "E075PRO", "field": "TEMICM",  "group": "ICMS"},
    "cad_codtrd":             {"label": "Código de tributação/diferencial ICMS","table": "E075PRO", "field": "CODTRD",  "group": "ICMS"},
    "cad_codtst":             {"label": "Código TST (ICMS ST)",               "table": "E075PRO", "field": "CODTST",  "group": "ICMS ST"},
    "cad_codstp":             {"label": "Código STP (ICMS ST pauta)",          "table": "E075PRO", "field": "CODSTP",  "group": "ICMS ST"},
    "cad_recicm":             {"label": "Recupera ICMS",                        "table": "E075PRO", "field": "RECICM",  "group": "ICMS"},
    "cad_codstr":             {"label": "Situação/Estratégia ICMS (CodStr)",  "table": "E075PRO", "field": "CODSTR",  "group": "ICMS"},
    "cad_codtic":             {"label": "Código TIC",                          "table": "E075PRO", "field": "CODTIC",  "group": "ICMS"},
    "cad_codstc":             {"label": "Código STC",                          "table": "E075PRO", "field": "CODSTC",  "group": "ICMS"},
    "cad_basrec":             {"label": "Base de recuperação (BasRec)",        "table": "E075PRO", "field": "BASREC",  "group": "PIS/COFINS"},
    "cad_bascre_produto":     {"label": "Base de crédito do produto (BasCre)",  "table": "E075PRO", "field": "BASCRE",  "group": "PIS/COFINS"},
    "cad_tripis":             {"label": "Tipo tributação PIS",                 "table": "E075PRO", "field": "TRIPIS",  "group": "PIS/COFINS"},
    "cad_tricof":             {"label": "Tipo tributação COFINS",              "table": "E075PRO", "field": "TRICOF",  "group": "PIS/COFINS"},
    "cad_cstipi_produto":     {"label": "CST IPI do cadastro",                 "table": "E075PRO", "field": "CSTIPI",  "group": "IPI"},
    "cad_cstpis_produto":     {"label": "CST PIS do cadastro",                 "table": "E075PRO", "field": "CSTPIS",  "group": "PIS/COFINS"},
    "cad_cstcof_produto":     {"label": "CST COFINS do cadastro",              "table": "E075PRO", "field": "CSTCOF",  "group": "PIS/COFINS"},
    "cad_tprpis":             {"label": "TPR PIS",                             "table": "E075PRO", "field": "TPRPIS",  "group": "PIS/COFINS"},
    "cad_tprcof":             {"label": "TPR COFINS",                          "table": "E075PRO", "field": "TPRCOF",  "group": "PIS/COFINS"},
    "cad_tpripi":             {"label": "TPR IPI",                             "table": "E075PRO", "field": "TPRIPI",  "group": "IPI"},
    "cad_regtri":             {"label": "Regime tributário",                   "table": "E075PRO", "field": "REGTRI",  "group": "Geral"},
    "cad_cstipc":             {"label": "CST IPC",                             "table": "E075PRO", "field": "CSTIPC",  "group": "ICMS"},
    "cad_cstpic":             {"label": "CST PIC",                             "table": "E075PRO", "field": "CSTPIC",  "group": "PIS/COFINS"},
    "cad_cstcoc":             {"label": "CST COC",                             "table": "E075PRO", "field": "CSTCOC",  "group": "PIS/COFINS"},
    "cad_orimer":             {"label": "Origem mercadoria",                   "table": "E075PRO", "field": "ORIMER",  "group": "Produto"},
    "cad_natpis":             {"label": "Natureza receita PIS",                "table": "E075PRO", "field": "NATPIS",  "group": "PIS/COFINS"},
    "cad_natcof":             {"label": "Natureza receita COFINS",             "table": "E075PRO", "field": "NATCOF",  "group": "PIS/COFINS"},
    "cad_tprpii":             {"label": "TPRPII",                              "table": "E075PRO", "field": "TPRPII",  "group": "IPI"},
    "cad_tprcoi":             {"label": "TPRCOI",                              "table": "E075PRO", "field": "TPRCOI",  "group": "PIS/COFINS"},
    "cad_perifp":             {"label": "% IPI fixo",                          "table": "E075PRO", "field": "PERIFP",  "group": "IPI"},
    "cad_pdifcp":             {"label": "% Diferimento ICMS",                  "table": "E075PRO", "field": "PDIFCP",  "group": "ICMS"},
    "cad_codenq":             {"label": "Código de enquadramento",             "table": "E075PRO", "field": "CODENQ",  "group": "IPI"},
    "cad_codces":             {"label": "CEST",                                "table": "E075PRO", "field": "CODCES",  "group": "ICMS ST"},
    "cad_coddfs":             {"label": "Código DFS",                          "table": "E075PRO", "field": "CODDFS",  "group": "Geral"},
    "cad_origti":             {"label": "Origem TI",                           "table": "E075PRO", "field": "ORIGTI",  "group": "Produto"},
    "cad_catpro":             {"label": "Categoria do produto",               "table": "E075PRO", "field": "CATPRO",  "group": "Produto"},
    "cad_itefis":             {"label": "Item fiscal",                         "table": "E075PRO", "field": "ITEFIS",  "group": "Geral"},
    "cad_desfis":             {"label": "Descrição fiscal",                    "table": "E075PRO", "field": "DESFIS",  "group": "Geral"},
    "cad_impscf":             {"label": "Imposto SCF",                         "table": "E075PRO", "field": "IMPSCF",  "group": "Geral"},
    "cad_perdif":             {"label": "% Diferimento",                       "table": "E075PRO", "field": "PERDIF",  "group": "ICMS"},
    "cad_emirec":             {"label": "Emissao recebimento",                  "table": "E075PRO", "field": "EMIREC",  "group": "Geral"},
    "cad_idepar":             {"label": "Identificador parc.",                 "table": "E075PRO", "field": "IDEPAR",  "group": "Geral"},
    "cad_tipcic":             {"label": "Tipo ciclo",                          "table": "E075PRO", "field": "TIPCIC",  "group": "Geral"},
    "cad_ficcat":             {"label": "Ficha catálogo",                      "table": "E075PRO", "field": "FICCAT",  "group": "Produto"},
    "cad_usu_mcgrcp":         {"label": "USU MCGRCP",                          "table": "E075PRO", "field": "USU_MCGRCP","group": "Customizado"},
    # === E012FAM â€” Família (fam_) ===
    "familia_codigo":         {"label": "Código família",                      "table": "E012FAM", "field": "CODFAM",  "group": "Família"},
    "familia_descricao":      {"label": "Descrição família",                   "table": "E012FAM", "field": "DESFAM",  "group": "Família"},
    "fam_cst_pis":            {"label": "CST PIS família",                     "table": "E012FAM", "field": "CSTPIS",  "group": "PIS/COFINS"},
    "fam_cst_cofins":         {"label": "CST COFINS família",                  "table": "E012FAM", "field": "CSTCOF",  "group": "PIS/COFINS"},
    "fam_cst_ipi":            {"label": "CST IPI família",                     "table": "E012FAM", "field": "CSTIPI",  "group": "IPI"},
    "fam_recpis":             {"label": "Recupera PIS família",                "table": "E012FAM", "field": "RECPIS",  "group": "PIS/COFINS"},
    "fam_reccof":             {"label": "Recupera COFINS família",             "table": "E012FAM", "field": "RECCOF",  "group": "PIS/COFINS"},
    "fam_perpis":             {"label": "% PIS família",                       "table": "E012FAM", "field": "PERPIS",  "group": "PIS/COFINS"},
    "fam_percof":             {"label": "% COFINS família",                    "table": "E012FAM", "field": "PERCOF",  "group": "PIS/COFINS"},
    "fam_peripi":             {"label": "% IPI família",                       "table": "E012FAM", "field": "PERIPI",  "group": "IPI"},
    "fam_pericm":             {"label": "% ICMS família",                      "table": "E012FAM", "field": "PERICM",  "group": "ICMS"},
    "fam_codstr":             {"label": "Estratégia ICMS família",             "table": "E012FAM", "field": "CODSTR",  "group": "ICMS"},
    "fam_codtrd":             {"label": "CodTrd família",                      "table": "E012FAM", "field": "CODTRD",  "group": "ICMS"},
    "fam_codtst":             {"label": "CodTST família",                      "table": "E012FAM", "field": "CODTST",  "group": "ICMS ST"},
    "fam_codstp":             {"label": "CodSTP família",                      "table": "E012FAM", "field": "CODSTP",  "group": "ICMS ST"},
    "fam_temicm":             {"label": "Tem ICMS família",                    "table": "E012FAM", "field": "TEMICM",  "group": "ICMS"},
    "fam_recicm":             {"label": "Recupera ICMS família",               "table": "E012FAM", "field": "RECICM",  "group": "ICMS"},
    "fam_recipi":             {"label": "Recupera IPI família",                "table": "E012FAM", "field": "RECIPI",  "group": "IPI"},
    "fam_proimp":             {"label": "Proímport família",                   "table": "E012FAM", "field": "PROIMP",  "group": "Produto"},
    "fam_tippro":             {"label": "Tipo produto família",                "table": "E012FAM", "field": "TIPPRO",  "group": "Produto"},
    "fam_codori":             {"label": "Origem família",                      "table": "E012FAM", "field": "CODORI",  "group": "Produto"},
    "fam_tmiicm":             {"label": "TMI ICMS família",                    "table": "E012FAM", "field": "TMIICM",  "group": "ICMS"},
    # === E083ORI â€” Origem (ori_) ===
    "origem_codigo":          {"label": "Código origem",                        "table": "E083ORI", "field": "CODORI",  "group": "Origem"},
    "origem_descricao":       {"label": "Descrição origem",                     "table": "E083ORI", "field": "DESORI",  "group": "Origem"},
    "ori_codreg":             {"label": "Regra fiscal origem",                  "table": "E083ORI", "field": "CODREG",  "group": "Origem"},
    "ori_codms1":             {"label": "Mercado 1 origem",                     "table": "E083ORI", "field": "CODMS1",  "group": "Origem"},
    "ori_codms2":             {"label": "Mercado 2 origem",                     "table": "E083ORI", "field": "CODMS2",  "group": "Origem"},
    "ori_codms3":             {"label": "Mercado 3 origem",                     "table": "E083ORI", "field": "CODMS3",  "group": "Origem"},
    "ori_codms4":             {"label": "Mercado 4 origem",                     "table": "E083ORI", "field": "CODMS4",  "group": "Origem"},
    "ori_proimp":             {"label": "Proímport origem",                     "table": "E083ORI", "field": "PROIMP",  "group": "Origem"},
    # === E085CLI â€” Cliente ===
    "cliente_codigo":         {"label": "Código cliente",                       "table": "E085CLI", "field": "CODCLI",  "group": "Cliente"},
    "cliente_nome":           {"label": "Nome cliente",                          "table": "E085CLI", "field": "NOMCLI",  "group": "Cliente"},
    "cliente_uf":             {"label": "UF cliente",                            "table": "E085CLI", "field": "SIGUFS",  "group": "Cliente"},
    "cliente_situacao":       {"label": "Situação cliente",                     "table": "E085CLI", "field": "SITCLI",  "group": "Cliente"},
    "cliente_endereco":       {"label": "Endereço cliente",                     "table": "E085CLI", "field": "ENDCLI",  "group": "Cliente"},
    "cliente_cidade":         {"label": "Cidade cliente",                        "table": "E085CLI", "field": "CIDCLI",  "group": "Cliente"},
    "cliente_cep":            {"label": "CEP cliente",                           "table": "E085CLI", "field": "CEPCLI",  "group": "Cliente"},
    "cliente_bairro":         {"label": "Bairro cliente",                        "table": "E085CLI", "field": "BAICLI",  "group": "Cliente"},
    "cliente_redsai_pis":     {"label": "RedSai PIS (E019RED)",                  "table": "E019RED", "field": "REDSAI",  "group": "PIS/COFINS"},
    "cliente_redsai_cofins":  {"label": "RedSai COFINS (E019RED)",               "table": "E019RED", "field": "REDSAI",  "group": "PIS/COFINS"},
    # === E095FOR â€” Fornecedor ===
    "fornecedor_codigo":      {"label": "Código fornecedor",                    "table": "E095FOR", "field": "CODFOR",  "group": "Fornecedor"},
    "fornecedor_nome":        {"label": "Nome fornecedor",                       "table": "E095FOR", "field": "NOMFOR",  "group": "Fornecedor"},
    "fornecedor_uf":          {"label": "UF fornecedor",                         "table": "E095FOR", "field": "SIGUFS",  "group": "Fornecedor"},
    "fornecedor_codtri":      {"label": "Cód. tributação serviço (CodTri)",    "table": "E095FOR", "field": "CODTRI",  "group": "Fornecedor"},
    "fornecedor_tipfor":      {"label": "Tipo fornecedor (PF/PJ)",               "table": "E095FOR", "field": "TIPFOR",  "group": "Fornecedor"},
    "fornecedor_situacao":    {"label": "Situação fornecedor",                  "table": "E095FOR", "field": "SITFOR",  "group": "Fornecedor"},
}


def enrich_field(key: str, value):
    """Retorna campo enriquecido com metadados: label, tabela, campo ERP, grupo e valor."""
    meta = FIELD_META.get(key, {})
    return {
        "key": key,
        "label": meta.get("label", key),
        "table": meta.get("table"),
        "field": meta.get("field"),
        "group": meta.get("group"),
        "value": value
    }


def build_enriched_block(data: Dict[str, Any], keys: list) -> Dict[str, Any]:
    """Monta dict de campos enriquecidos para um conjunto de chaves."""
    return {key: enrich_field(key, data.get(key)) for key in keys}


# =========================================================
# AUTH
# =========================================================

class LoginRequest(BaseModel):
    usuario: str
    senha: str


class ControleFiscalSalvarRequest(BaseModel):
    codemp: int = EMPRESA_PADRAO
    codpro: str
    codder: Optional[str] = None
    campos_pro: Dict[str, Any] = {}
    campos_der: Dict[str, Any] = {}
    ativar_apos_salvar: bool = False


class ControleFiscalAtivarRequest(BaseModel):
    codemp: int = EMPRESA_PADRAO
    codpro: str
    codder: Optional[str] = None


CAMPOS_EDITAVEIS_E075PRO = {
    "CODCLF", "CODSTR", "CODTIC", "CODTRD", "CODTST", "CODSTP", "CODSTC",
    "PERIPI", "RECIPI", "TEMICM", "RECICM", "RECPIS", "TRIPIS", "TRICOF", "RECCOF",
    "PERIRF", "PERPIS", "PERCOF", "PERCSL", "PEROUR", "BASCRE", "BASREC",
    "CSTIPI", "CSTPIS", "CSTCOF", "TPRPIS", "TPRCOF", "TPRIPI", "REGTRI",
    "CSTIPC", "CSTPIC", "CSTCOC", "ORIMER", "NATPIS", "NATCOF", "CODANP", "PROIMP"
}

CAMPOS_EDITAVEIS_E075DER = {
    "ITEFIS", "DESFIS", "CODFIF", "CODFIE", "CODFIM", "BSTUFC", "ASTFCP", "VSTUFC", "CODCES"
}


def _normalizar_chaves(d: Dict[str, Any]) -> Dict[str, Any]:
    out = {}
    for k, v in (d or {}).items():
        if not k:
            continue
        out[str(k).strip().upper()] = v
    return out


def _montar_update_sql(tabela: str, campos: Dict[str, Any], where_sql: str, where_params: List[Any]):
    sets = []
    params = []
    for campo, valor in campos.items():
        sets.append(f"{campo} = ?")
        params.append(valor)
    if not sets:
        return None, None
    sql = f"UPDATE {tabela} SET " + ", ".join(sets) + " " + where_sql
    params.extend(where_params)
    return sql, params


def _buscar_cadastro_atual(cursor, codemp: int, codpro: str, codder=None):
    cursor.execute(
        """
        SELECT
            P.CODEMP, P.CODPRO, P.SITPRO, P.CODCLF, P.CODTRD, P.RECPIS, P.RECCOF,
            P.CSTPIS, P.CSTCOF, P.BASCRE,
            D.CODDER, D.SITDER
        FROM E075PRO P
        LEFT JOIN E075DER D
               ON D.CODEMP = P.CODEMP
              AND D.CODPRO = P.CODPRO
              AND (? IS NULL OR D.CODDER = ?)
        WHERE P.CODEMP = ?
          AND P.CODPRO = ?
        """,
        [codder, codder, codemp, codpro]
    )
    return cursor.fetchone()


def _validar_minimos_ativacao(row, campos_pro: Dict[str, Any], campos_der: Dict[str, Any]):
    def valor(nome, idx):
        if nome in campos_pro:
            return campos_pro.get(nome)
        if nome in campos_der:
            return campos_der.get(nome)
        return row[idx] if row and len(row) > idx else None

    pendencias = []
    codclf = valor("CODCLF", 3)
    codtrd = valor("CODTRD", 4)
    recpis = valor("RECPIS", 5)
    reccof = valor("RECCOF", 6)
    cstpis = valor("CSTPIS", 7)
    cstcof = valor("CSTCOF", 8)
    # bascre = valor("BASCRE", 9)  # não bloqueia mais

    if codclf in (None, ""):
        pendencias.append("CODCLF")
    if codtrd in (None, ""):
        pendencias.append("CODTRD")
    if recpis in (None, ""):
        pendencias.append("RECPIS")
    if reccof in (None, ""):
        pendencias.append("RECCOF")
    if cstpis in (None, ""):
        pendencias.append("CSTPIS")
    if cstcof in (None, ""):
        pendencias.append("CSTCOF")
    # BASCRE removido da trava

    if str(cstpis).strip() != str(cstcof).strip():
        pendencias.append("CSTPIS/CSTCOF divergentes")
    if str(recpis).strip() != str(reccof).strip():
        pendencias.append("RECPIS/RECCOF divergentes")
    return pendencias


NUMERIC_FIELDS_E075PRO = {
    "PERIPI", "PERIRF", "PERPIS", "PERCOF", "PERCSL", "PEROUR",
    "BASCRE", "BASREC"
}

NUMERIC_FIELDS_E075DER = {
    "BSTUFC", "ASTFCP", "VSTUFC"
}


def _valor_vazio_para_none(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    return v


def _normalizar_numero_sql(v):
    v = _valor_vazio_para_none(v)
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    return float(s)


def _tratar_campos_para_update(campos: Dict[str, Any], campos_numericos: set) -> Dict[str, Any]:
    tratados = {}
    for campo, valor in (campos or {}).items():
        campo = str(campo).strip().upper()
        if campo in campos_numericos:
            tratados[campo] = _normalizar_numero_sql(valor)
        else:
            tratados[campo] = _valor_vazio_para_none(valor)
    return tratados


def gerar_token(usuario: str):
    payload = {
        "sub": usuario,
        "exp": datetime.now(timezone.utc) + timedelta(hours=8)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def validar_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload["sub"]
    except Exception:
        raise HTTPException(status_code=401, detail="Token inválido")


@app.post("/login")
def login(
    payload: Optional[LoginRequest] = Body(default=None),
    usuario: Optional[str] = Query(default=None),
    senha: Optional[str] = Query(default=None),
):
    usuario_final = ""
    senha_final = ""

    if payload:
        usuario_final = (payload.usuario or "").upper().strip()
        senha_final = (payload.senha or "").strip()
    else:
        usuario_final = (usuario or "").upper().strip()
        senha_final = (senha or "").strip()

    if not usuario_final or not senha_final:
        raise HTTPException(status_code=400, detail="Usuário e senha são obrigatórios")

    if usuario_final in USERS and USERS[usuario_final] == senha_final:
        return {
            "access_token": gerar_token(usuario_final),
            "token_type": "bearer",
            "usuario": usuario_final
        }

    raise HTTPException(status_code=401, detail="Login inválido")


@app.get("/health")
def health():
    return {
        "status": "ok",
        "app": "Auditoria Tributária ERP Senior",
        "port": API_PORT
    }


# =========================================================
# FILTROS AUXILIARES
# =========================================================

@app.get("/api/familias")
def listar_familias(
    q: Optional[str] = None,
    limite: int = 200,
    usuario=Depends(validar_token)
):
    limite = max(1, min(limite, 500))
    conn = get_connection()
    cursor = conn.cursor()

    sql = f"""
        SELECT TOP {limite}
            F.CODFAM AS codigo,
            F.DESFAM AS descricao
        FROM E012FAM F
        WHERE F.CODEMP = ?
          AND EXISTS (
              SELECT 1
              FROM E075PRO P
              WHERE P.CODEMP = F.CODEMP
                AND P.CODFAM = F.CODFAM
          )
    """
    params = [EMPRESA_PADRAO]

    if q:
        sql += " AND (F.CODFAM LIKE ? OR F.DESFAM LIKE ?)"
        termo = f"%{q.strip()}%"
        params.extend([termo, termo])

    sql += " ORDER BY F.CODFAM"

    cursor.execute(sql, params)
    rows = cursor.fetchall()
    conn.close()

    resultado = []
    for row in rows:
        codigo = (row[0] or "").strip()
        descricao = (row[1] or "").strip()
        resultado.append({
            "codigo": codigo,
            "descricao": descricao,
            "label": f"{codigo} - {descricao}" if descricao else codigo
        })
    return resultado


@app.get("/api/origens")
def listar_origens(
    q: Optional[str] = None,
    limite: int = 200,
    usuario=Depends(validar_token)
):
    limite = max(1, min(limite, 500))
    conn = get_connection()
    cursor = conn.cursor()

    sql = f"""
        SELECT TOP {limite}
            O.CODORI AS codigo,
            O.DESORI AS descricao
        FROM E083ORI O
        WHERE O.CODEMP = ?
          AND EXISTS (
              SELECT 1
              FROM E075PRO P
              WHERE P.CODEMP = O.CODEMP
                AND P.CODORI = O.CODORI
          )
    """
    params = [EMPRESA_PADRAO]

    if q:
        sql += " AND (O.CODORI LIKE ? OR O.DESORI LIKE ?)"
        termo = f"%{q.strip()}%"
        params.extend([termo, termo])

    sql += " ORDER BY O.CODORI"

    cursor.execute(sql, params)
    rows = cursor.fetchall()
    conn.close()

    resultado = []
    for row in rows:
        codigo = (row[0] or "").strip()
        descricao = (row[1] or "").strip()
        resultado.append({
            "codigo": codigo,
            "descricao": descricao,
            "label": f"{codigo} - {descricao}" if descricao else codigo
        })
    return resultado



# =========================================================
# CONTROLE FISCAL DE PRODUTOS
# =========================================================

@app.get("/api/controle-fiscal-produtos")
def api_controle_fiscal_produtos(
    codigo_produto: Optional[str] = Query(None),
    descricao: Optional[str] = Query(None),
    derivacao: Optional[str] = Query(None),
    tipo_produto: Optional[str] = Query(None),
    origem: Optional[str] = Query(None),
    familia: Optional[str] = Query(None),
    usuario_geracao: Optional[int] = Query(None),
    data_ini: Optional[str] = Query(None),
    data_fim: Optional[str] = Query(None),
    visao: str = Query("AMBOS"),   # PRODUTO | DERIVACAO | AMBOS
    pagina: int = Query(1),
    tamanho_pagina: int = Query(100),
    usuario=Depends(validar_token)
):
    try:
        pagina = max(1, pagina)
        tamanho_pagina = min(max(1, tamanho_pagina), 200)
        offset = (pagina - 1) * tamanho_pagina
        visao = (visao or "AMBOS").upper().strip()

        filtros = ["P.CODEMP = ?"]
        params = [EMPRESA_PADRAO]

        if codigo_produto:
            filtros.append("P.CODPRO LIKE ?")
            params.append(f"%{codigo_produto.strip()}%")

        if descricao:
            filtros.append("P.DESPRO LIKE ?")
            params.append(f"%{descricao.strip()}%")

        if derivacao:
            filtros.append("D.CODDER LIKE ?")
            params.append(f"%{derivacao.strip()}%")

        if tipo_produto:
            filtros.append("P.TIPPRO = ?")
            params.append(tipo_produto.strip())

        if origem:
            filtros.append("P.CODORI = ?")
            params.append(origem.strip())

        if familia:
            filtros.append("P.CODFAM = ?")
            params.append(familia.strip())

        if usuario_geracao is not None:
            filtros.append("(P.USUGER = ? OR D.USUGER = ?)")
            params.append(usuario_geracao)
            params.append(usuario_geracao)

        if data_ini:
            filtros.append("""
                (
                    (P.DATGER IS NOT NULL AND CAST(P.DATGER AS DATE) >= ?)
                    OR
                    (D.DATGER IS NOT NULL AND CAST(D.DATGER AS DATE) >= ?)
                )
            """)
            params.append(data_ini)
            params.append(data_ini)

        if data_fim:
            filtros.append("""
                (
                    (P.DATGER IS NOT NULL AND CAST(P.DATGER AS DATE) <= ?)
                    OR
                    (D.DATGER IS NOT NULL AND CAST(D.DATGER AS DATE) <= ?)
                )
            """)
            params.append(data_fim)
            params.append(data_fim)

        if visao == "PRODUTO":
            filtros.append("P.SITPRO = 'I'")
        elif visao == "DERIVACAO":
            filtros.append("D.SITDER = 'I'")
        else:
            filtros.append("(P.SITPRO = 'I' OR D.SITDER = 'I')")

        where_sql = " AND ".join(filtros)

        cte_sql = f"""
            ;WITH BASE AS (
                SELECT
                    P.CODEMP,
                    P.CODPRO,
                    P.DESPRO,
                    P.TIPPRO,
                    P.CODORI,
                    P.CODFAM,
                    P.SITPRO,
                    P.USUGER AS USUGER_PRO,
                    P.DATGER AS DATGER_PRO,

                    D.CODDER,
                    D.DESDER,
                    D.SITDER,
                    D.USUGER AS USUGER_DER,
                    D.DATGER AS DATGER_DER,

                    UPRO.NOMUSU AS NOME_USUGER_PRO,
                    UDER.NOMUSU AS NOME_USUGER_DER,

                    CLF.CLAFIS AS NCM,

                    P.CODCLF,
                    P.CODSTR,
                    P.CODTIC,
                    P.CODTRD,
                    P.CODTST,
                    P.CODSTP,
                    P.CODSTC,
                    P.PERIPI,
                    P.RECIPI,
                    P.TEMICM,
                    P.RECICM,
                    P.RECPIS,
                    P.TRIPIS,
                    P.TRICOF,
                    P.RECCOF,
                    P.PERIRF,
                    P.PERPIS,
                    P.PERCOF,
                    P.PERCSL,
                    P.PEROUR,
                    P.BASCRE,
                    P.BASREC,
                    P.CSTIPI,
                    P.CSTPIS,
                    P.CSTCOF,
                    P.TPRPIS,
                    P.TPRCOF,
                    P.TPRIPI,
                    P.REGTRI,
                    P.CSTIPC,
                    P.CSTPIC,
                    P.CSTCOC,
                    P.ORIMER,
                    P.NATPIS,
                    P.NATCOF,
                    P.CODANP,
                    P.PROIMP,

                    D.ITEFIS,
                    D.DESFIS,
                    D.CODFIF,
                    D.CODFIE,
                    D.CODFIM,
                    D.BSTUFC,
                    D.ASTFCP,
                    D.VSTUFC,
                    D.CODCES,

                    CASE
                        WHEN P.SITPRO = 'I' AND ISNULL(D.SITDER, 'I') = 'I' THEN 'PRODUTO_E_DERIVACAO_INATIVOS'
                        WHEN P.SITPRO = 'I' THEN 'PRODUTO_INATIVO'
                        WHEN D.SITDER = 'I' THEN 'DERIVACAO_INATIVA'
                        ELSE 'ATIVO'
                    END AS STATUS_CONTROLE
                FROM E075PRO P
                LEFT JOIN E075DER D
                    ON D.CODEMP = P.CODEMP
                   AND D.CODPRO = P.CODPRO
                LEFT JOIN E099USU UPRO
                    ON UPRO.CODEMP = P.CODEMP
                   AND UPRO.CODUSU = P.USUGER
                LEFT JOIN E099USU UDER
                    ON UDER.CODEMP = D.CODEMP
                   AND UDER.CODUSU = D.USUGER
                LEFT JOIN E022CLF CLF
                    ON CLF.CODCLF = P.CODCLF
                WHERE {where_sql}
            )
        """

        sql_total = cte_sql + """
            SELECT COUNT(1) AS TOTAL
            FROM BASE
        """

        sql_resumo = cte_sql + """
            SELECT
                COUNT(1) AS TOTAL_REGISTROS,
                SUM(CASE WHEN SITPRO = 'I' THEN 1 ELSE 0 END) AS PRODUTO_INATIVO,
                SUM(CASE WHEN SITDER = 'I' THEN 1 ELSE 0 END) AS DERIVACAO_INATIVA,
                SUM(CASE WHEN STATUS_CONTROLE = 'PRODUTO_E_DERIVACAO_INATIVOS' THEN 1 ELSE 0 END) AS PRODUTO_E_DERIVACAO_INATIVOS
            FROM BASE
        """

        sql_paginado = cte_sql + """
            SELECT *
            FROM (
                SELECT
                    BASE.*,
                    ROW_NUMBER() OVER (ORDER BY CODPRO, CODDER) AS RN
                FROM BASE
            ) Z
            WHERE Z.RN > ? AND Z.RN <= ?
            ORDER BY Z.RN
        """

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute(sql_total, params)
        total_registros = int(cursor.fetchone()[0] or 0)

        cursor.execute(sql_resumo, params)
        row_resumo = cursor.fetchone()

        resumo = {
            "total_registros": int(row_resumo[0] or 0),
            "produto_inativo": int(row_resumo[1] or 0),
            "derivacao_inativa": int(row_resumo[2] or 0),
            "produto_e_derivacao_inativos": int(row_resumo[3] or 0),
        }

        cursor.execute(sql_paginado, params + [offset, offset + tamanho_pagina])
        rows = cursor.fetchall()
        itens = [row_to_dict(cursor, row) for row in rows]

        conn.close()

        total_paginas = max(1, (total_registros + tamanho_pagina - 1) // tamanho_pagina) if total_registros else 1

        return {
            "resumo": resumo,
            "itens": itens,
            "pagina": pagina,
            "tamanho_pagina": tamanho_pagina,
            "total_registros": total_registros,
            "total_paginas": total_paginas
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro no controle fiscal de produtos: {str(e)}")


@app.get("/api/controle-fiscal-produtos/diagnostico")
def diagnostico_controle_fiscal(usuario=Depends(validar_token)):
    """Retorna contagens brutas para diagnóstico â€” sem filtros."""
    try:
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                (SELECT COUNT(1) FROM E075PRO WHERE CODEMP = ?) AS total_produtos,
                (SELECT COUNT(1) FROM E075PRO WHERE CODEMP = ? AND SITPRO = 'I') AS produtos_inativos,
                (SELECT COUNT(1) FROM E075DER WHERE CODEMP = ? AND SITDER = 'I') AS derivacoes_inativas,
                (SELECT COUNT(DISTINCT P.CODPRO) FROM E075PRO P
                    INNER JOIN E075DER D ON D.CODEMP = P.CODEMP AND D.CODPRO = P.CODPRO
                    WHERE P.CODEMP = ? AND (P.SITPRO = 'I' OR D.SITDER = 'I')
                ) AS produtos_com_algum_inativo
        """, [EMPRESA_PADRAO, EMPRESA_PADRAO, EMPRESA_PADRAO, EMPRESA_PADRAO])
        row = cursor.fetchone()

        cursor.execute("""
            SELECT DISTINCT SITPRO FROM E075PRO WHERE CODEMP = ?
        """, [EMPRESA_PADRAO])
        valores_sitpro = [r[0] for r in cursor.fetchall()]

        cursor.execute("""
            SELECT DISTINCT SITDER FROM E075DER WHERE CODEMP = ?
        """, [EMPRESA_PADRAO])
        valores_sitder = [r[0] for r in cursor.fetchall()]

        cursor.execute("""
            SELECT TOP 5 P.CODPRO, P.DESPRO, P.SITPRO, D.CODDER, D.SITDER
            FROM E075PRO P
            LEFT JOIN E075DER D ON D.CODEMP = P.CODEMP AND D.CODPRO = P.CODPRO
            WHERE P.CODEMP = ? AND (P.SITPRO = 'I' OR D.SITDER = 'I')
            ORDER BY P.CODPRO
        """, [EMPRESA_PADRAO])
        amostra = [row_to_dict(cursor, r) for r in cursor.fetchall()]

        conn.close()

        return {
            "empresa": EMPRESA_PADRAO,
            "total_produtos": int(row[0] or 0),
            "produtos_inativos_sitpro_I": int(row[1] or 0),
            "derivacoes_inativas_sitder_I": int(row[2] or 0),
            "produtos_com_algum_inativo": int(row[3] or 0),
            "valores_sitpro_existentes": valores_sitpro,
            "valores_sitder_existentes": valores_sitder,
            "amostra_inativos": amostra,
            "dica": (
                "Se produtos_inativos e derivacoes_inativas forem 0, "
                "não existem registros com SITPRO='I' ou SITDER='I' na empresa. "
                "Verifique valores_sitpro_existentes e valores_sitder_existentes "
                "para ver quais status realmente existem no banco."
            )
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro no diagnóstico: {str(e)}")


# =========================================================
# AUDITORIA DO CONTROLE FISCAL
# =========================================================

_AUDIT_TABLE_CHECKED = False

def _ensure_audit_table():
    """Cria a tabela USU_AUDIT_FISCAL se não existir."""
    global _AUDIT_TABLE_CHECKED
    if _AUDIT_TABLE_CHECKED:
        return
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            IF NOT EXISTS (
                SELECT * FROM sys.objects
                WHERE object_id = OBJECT_ID(N'USU_AUDIT_FISCAL') AND type = 'U'
            )
            CREATE TABLE USU_AUDIT_FISCAL (
                ID          INT IDENTITY(1,1) PRIMARY KEY,
                CREATED_AT  DATETIME DEFAULT GETDATE(),
                ACAO        VARCHAR(50) NOT NULL,
                CODEMP      INT,
                CODPRO      VARCHAR(50) NOT NULL,
                CODDER      VARCHAR(50),
                USUARIO     VARCHAR(100),
                PAYLOAD_ANTES NVARCHAR(MAX),
                PAYLOAD_DEPOIS NVARCHAR(MAX),
                SUCESSO     BIT DEFAULT 1,
                ERRO        NVARCHAR(MAX)
            )
        """)
        conn.commit()
        conn.close()
        _AUDIT_TABLE_CHECKED = True
        print("[AUDIT] Tabela USU_AUDIT_FISCAL verificada/criada.")
    except Exception as e:
        print(f"[AUDIT] Erro ao criar tabela: {e}")


def _snapshot_produto(cursor, codemp: int, codpro: str, codder: str = None) -> dict:
    """Captura snapshot dos campos atuais de E075PRO + E075DER."""
    snap = {}
    try:
        cursor.execute(
            "SELECT CODCLF,CODTRD,CODTST,CODSTP,CODSTC,CODSTR,CODTIC,"
            "CSTPIS,CSTCOF,RECPIS,RECCOF,BASCRE,PERIPI,RECIPI,SITPRO "
            "FROM E075PRO WHERE CODEMP=? AND CODPRO=?",
            [codemp, codpro]
        )
        row = cursor.fetchone()
        if row:
            cols = [d[0] for d in cursor.description]
            snap["e075pro"] = {c: (str(v) if v is not None else None) for c, v in zip(cols, row)}

        if codder:
            cursor.execute(
                "SELECT CODCES,BSTUFC,ASTFCP,VSTUFC,SITDER "
                "FROM E075DER WHERE CODEMP=? AND CODPRO=? AND CODDER=?",
                [codemp, codpro, codder]
            )
            row2 = cursor.fetchone()
            if row2:
                cols2 = [d[0] for d in cursor.description]
                snap["e075der"] = {c: (str(v) if v is not None else None) for c, v in zip(cols2, row2)}
    except Exception as e:
        print(f"[AUDIT] Erro no snapshot: {e}")
    return snap


def registrar_auditoria_fiscal(
    acao: str,
    codemp: int,
    codpro: str,
    codder: str = None,
    usuario: str = None,
    payload_antes: dict = None,
    payload_depois: dict = None,
    sucesso: bool = True,
    erro: str = None
):
    """Registra evento na trilha de auditoria fiscal."""
    _ensure_audit_table()
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO USU_AUDIT_FISCAL "
            "(ACAO, CODEMP, CODPRO, CODDER, USUARIO, PAYLOAD_ANTES, PAYLOAD_DEPOIS, SUCESSO, ERRO) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [
                acao,
                codemp,
                codpro,
                codder,
                usuario,
                json.dumps(payload_antes, default=str, ensure_ascii=False) if payload_antes else None,
                json.dumps(payload_depois, default=str, ensure_ascii=False) if payload_depois else None,
                1 if sucesso else 0,
                erro
            ]
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[AUDIT] Erro ao registrar auditoria: {e}")

def disparar_webhook_novo_item_fiscal(
    codemp: int,
    codpro: str,
    codder: str = None,
    descricao: str = None,
    ncm: str = None,
    uf: str = None,
    origem: str = None,
    familia: str = None,
    usuario_nome: str = None,
    dados_extras: dict = None,
):
    """Dispara webhook para Supabase edge function quando um novo item fiscal entra."""
    if not WEBHOOK_NOVO_ITEM_URL:
        logger.error("webhook FAIL codpro=%s motivo=WEBHOOK_NOVO_ITEM_URL ausente", codpro)
        return

    if not FASTAPI_WEBHOOK_SECRET:
        logger.error("webhook FAIL codpro=%s motivo=FASTAPI_WEBHOOK_SECRET ausente", codpro)
        return

    payload = {
        "Codemp": codemp,
        "CodPro": codpro,
        "Codder": codder,
        "Descricao": descricao,
        "NCM": ncm,
        "UF": uf,
        "Origem": origem,
        "Familia": familia,
        "UsuarioNome": usuario_nome,
        "dados_extras": dados_extras or {},
    }

    headers = {
        "Content-Type": "application/json",
        "x-webhook-secret": FASTAPI_WEBHOOK_SECRET,
    }

    try:
        with httpx.Client(timeout=15) as client:
            response = client.post(
                WEBHOOK_NOVO_ITEM_URL,
                headers=headers,
                json=payload,
            )
            response.raise_for_status()

        logger.info(
            "webhook OK codpro=%s codder=%s status=%s",
            codpro,
            codder or "-",
            response.status_code
        )

    except Exception as e:
        logger.exception(
            "webhook FAIL codpro=%s codder=%s erro=%s",
            codpro,
            codder or "-",
            e
        )


def buscar_dados_item_controle_fiscal(codemp: int, codpro: str, codder: str = None) -> dict:
    """Busca dados reais do produto no banco para enriquecer o webhook."""
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT P.CODEMP, P.CODPRO, P.DESPRO, P.CODCLF, P.CODORI, P.CODFAM, "
            "D.CODDER, D.DESDER "
            "FROM E075PRO P "
            "LEFT JOIN E075DER D ON D.CODEMP = P.CODEMP AND D.CODPRO = P.CODPRO "
            "AND (? IS NULL OR D.CODDER = ?) "
            "WHERE P.CODEMP = ? AND P.CODPRO = ?",
            [codder, codder, codemp, codpro]
        )
        row = cursor.fetchone()
        conn.close()
        if not row:
            return {}
        cols = [d[0] for d in cursor.description]
        return {c: (str(v) if v is not None else None) for c, v in zip(cols, row)}
    except Exception as e:
        logger.exception("Erro ao buscar dados do produto codpro=%s: %s", codpro, e)
        return {}


@app.get("/api/controle-fiscal-produtos/auditoria")
def auditoria_controle_fiscal(
    codpro: str = Query(...),
    codder: Optional[str] = Query(None),
    limite: int = Query(50),
    usuario=Depends(validar_token)
):
    """Retorna histórico de auditoria de um produto."""
    _ensure_audit_table()
    try:
        conn = get_connection()
        cursor = conn.cursor()

        sql = """
            SELECT TOP (?) ID, CREATED_AT, ACAO, CODEMP, CODPRO, CODDER,
                   USUARIO, PAYLOAD_ANTES, PAYLOAD_DEPOIS, SUCESSO, ERRO
            FROM USU_AUDIT_FISCAL
            WHERE CODPRO = ?
        """
        params = [limite, codpro.strip().upper()]

        if codder:
            sql += " AND CODDER = ?"
            params.append(codder.strip().upper())

        sql += " ORDER BY CREATED_AT DESC"

        cursor.execute(sql, params)
        cols = [d[0] for d in cursor.description]
        rows = []
        for row in cursor.fetchall():
            item = {}
            for c, v in zip(cols, row):
                if c in ("PAYLOAD_ANTES", "PAYLOAD_DEPOIS") and v:
                    try:
                        item[c] = json.loads(v)
                    except Exception:
                        item[c] = v
                elif isinstance(v, datetime):
                    item[c] = v.strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(v, Decimal):
                    item[c] = float(v)
                else:
                    item[c] = v
            rows.append(item)

        conn.close()
        return {"itens": rows, "total": len(rows)}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao buscar auditoria: {str(e)}")


@app.post("/api/controle-fiscal-produtos/salvar")
def salvar_controle_fiscal_produtos(
    payload: ControleFiscalSalvarRequest,
    background_tasks: BackgroundTasks,
    usuario=Depends(validar_token)
):
    # Quando o workflow de dupla aprovação está habilitado, este endpoint vira
    # apenas um atalho que CRIA uma solicitação (status PENDENTE). Nada é
    # gravado no ERP até que duas aprovações distintas e o /aplicar sejam feitos.
    if BLOQUEAR_SALVAR_DIRETO:
        return _salvar_como_solicitacao_fiscal(payload, usuario)
    try:
        codemp = payload.codemp
        codpro = payload.codpro.strip().upper()
        codder = payload.codder.strip().upper() if payload.codder else None

        campos_pro = _normalizar_chaves(payload.campos_pro)
        campos_der = _normalizar_chaves(payload.campos_der)

        campos_pro = {k: v for k, v in campos_pro.items() if k in CAMPOS_EDITAVEIS_E075PRO}
        campos_der = {k: v for k, v in campos_der.items() if k in CAMPOS_EDITAVEIS_E075DER}

        campos_pro = _tratar_campos_para_update(campos_pro, NUMERIC_FIELDS_E075PRO)
        campos_der = _tratar_campos_para_update(campos_der, NUMERIC_FIELDS_E075DER)

        if not codpro:
            raise HTTPException(status_code=400, detail="codpro é obrigatório")

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute(
            "SELECT COUNT(1) FROM E075PRO WHERE CODEMP = ? AND CODPRO = ?",
            [codemp, codpro]
        )
        existe_pro = int(cursor.fetchone()[0] or 0)
        if existe_pro == 0:
            conn.close()
            raise HTTPException(status_code=404, detail="Produto não encontrado em E075PRO")

        if codder:
            cursor.execute(
                "SELECT COUNT(1) FROM E075DER WHERE CODEMP = ? AND CODPRO = ? AND CODDER = ?",
                [codemp, codpro, codder]
            )
            existe_der = int(cursor.fetchone()[0] or 0)
            if existe_der == 0:
                conn.close()
                raise HTTPException(status_code=404, detail="Derivação não encontrada em E075DER")
        # Snapshot antes das alterações
        snap_antes = _snapshot_produto(cursor, codemp, codpro, codder)

        updates_executados = []

        if campos_pro:
            sql_pro, params_pro = _montar_update_sql(
                "E075PRO", campos_pro,
                "WHERE CODEMP = ? AND CODPRO = ?", [codemp, codpro]
            )
            cursor.execute(sql_pro, params_pro)
            updates_executados.append("E075PRO")

        if codder and campos_der:
            sql_der, params_der = _montar_update_sql(
                "E075DER", campos_der,
                "WHERE CODEMP = ? AND CODPRO = ? AND CODDER = ?", [codemp, codpro, codder]
            )
            cursor.execute(sql_der, params_der)
            updates_executados.append("E075DER")

        ativado = False
        pendencias_ativacao = []

        if payload.ativar_apos_salvar:
            row = _buscar_cadastro_atual(cursor, codemp, codpro, codder)
            pendencias_ativacao = _validar_minimos_ativacao(row, campos_pro, campos_der)

            if pendencias_ativacao:
                conn.rollback()
                conn.close()
                raise HTTPException(
                    status_code=400,
                    detail=f"Não foi possível ativar. Pendências: {', '.join(pendencias_ativacao)}"
                )

            cursor.execute(
                "UPDATE E075PRO SET SITPRO = 'A' WHERE CODEMP = ? AND CODPRO = ?",
                [codemp, codpro]
            )
            if codder:
                cursor.execute(
                    "UPDATE E075DER SET SITDER = 'A' WHERE CODEMP = ? AND CODPRO = ? AND CODDER = ?",
                    [codemp, codpro, codder]
                )
            ativado = True

        conn.commit()
        conn.close()

        # Registrar auditoria
        acao = "SALVAR_E_ATIVAR" if ativado else "SALVAR"
        registrar_auditoria_fiscal(
            acao=acao,
            codemp=codemp,
            codpro=codpro,
            codder=codder,
            usuario=usuario,
            payload_antes=snap_antes,
            payload_depois={"campos_pro": campos_pro, "campos_der": campos_der, "ativado": ativado},
            sucesso=True
        )

        # Webhook em background: disparar se entrou em pendência (não ativou)
        entrou_em_pendencia = not ativado

        if entrou_em_pendencia:
            try:
                dados_item = buscar_dados_item_controle_fiscal(codemp, codpro, codder) or {}

                background_tasks.add_task(
                    disparar_webhook_novo_item_fiscal,
                    codemp,
                    codpro,
                    codder,
                    dados_item.get("DESPRO"),
                    dados_item.get("CODCLF"),
                    None,  # UF
                    dados_item.get("CODORI"),
                    dados_item.get("CODFAM"),
                    str(usuario),
                    {
                        "acao": "NOVO_ITEM_CONTROLE_FISCAL",
                        "ativado": ativado,
                        "updates": updates_executados,
                        "descricao_derivacao": dados_item.get("DESDER"),
                    }
                )
            except Exception as e:
                logger.exception(
                    "falha ao agendar webhook novo item codpro=%s codder=%s erro=%s",
                    codpro, codder or "-", e
                )

        return {
            "ok": True,
            "mensagem": "Cadastro fiscal salvo com sucesso" + (" e ativado" if ativado else ""),
            "codemp": codemp,
            "codpro": codpro,
            "codder": codder,
            "updates": updates_executados,
            "ativado": ativado,
            "usuario_logado": usuario
        }

    except HTTPException:
        raise
    except Exception as e:
        # Registrar falha na auditoria
        try:
            registrar_auditoria_fiscal(
                acao="SALVAR", codemp=payload.codemp,
                codpro=payload.codpro, codder=payload.codder,
                usuario=usuario, sucesso=False, erro=str(e)
            )
        except Exception:
            pass
        raise HTTPException(status_code=500, detail=f"Erro ao salvar controle fiscal: {str(e)}")


@app.post("/api/controle-fiscal-produtos/ativar")
def ativar_controle_fiscal_produtos(
    payload: ControleFiscalAtivarRequest,
    usuario=Depends(validar_token)
):
    try:
        codemp = payload.codemp
        codpro = payload.codpro.strip().upper()
        codder = payload.codder.strip().upper() if payload.codder else None

        if not codpro:
            raise HTTPException(status_code=400, detail="codpro é obrigatório")

        conn = get_connection()
        cursor = conn.cursor()

        # Snapshot antes
        snap_antes = _snapshot_produto(cursor, codemp, codpro, codder)

        row = _buscar_cadastro_atual(cursor, codemp, codpro, codder)
        if not row:
            conn.close()
            raise HTTPException(status_code=404, detail="Produto/derivação não encontrados")

        pendencias = _validar_minimos_ativacao(row, {}, {})
        if pendencias:
            conn.close()
            raise HTTPException(
                status_code=400,
                detail=f"Não foi possível ativar. Pendências: {', '.join(pendencias)}"
            )

        cursor.execute(
            "UPDATE E075PRO SET SITPRO = 'A' WHERE CODEMP = ? AND CODPRO = ?",
            [codemp, codpro]
        )
        if codder:
            cursor.execute(
                "UPDATE E075DER SET SITDER = 'A' WHERE CODEMP = ? AND CODPRO = ? AND CODDER = ?",
                [codemp, codpro, codder]
            )

        conn.commit()
        conn.close()

        # Registrar auditoria
        registrar_auditoria_fiscal(
            acao="ATIVAR",
            codemp=codemp,
            codpro=codpro,
            codder=codder,
            usuario=usuario,
            payload_antes=snap_antes,
            payload_depois={"SITPRO": "A", "SITDER": "A" if codder else None},
            sucesso=True
        )

        return {
            "ok": True,
            "mensagem": "Produto ativado com sucesso" + (f" e derivação {codder} ativada" if codder else ""),
            "codemp": codemp,
            "codpro": codpro,
            "codder": codder,
            "usuario_logado": usuario
        }

    except HTTPException:
        raise
    except Exception as e:
        try:
            registrar_auditoria_fiscal(
                acao="ATIVAR", codemp=payload.codemp,
                codpro=payload.codpro, codder=payload.codder,
                usuario=usuario, sucesso=False, erro=str(e)
            )
        except Exception:
            pass
        raise HTTPException(status_code=500, detail=f"Erro ao ativar produto: {str(e)}")


# =========================================================
# WORKFLOW DE SOLICITAÇÃO FISCAL — MOTOR GENÉRICO MULTI-ENTIDADE
# =========================================================
#
# Motor único que substitui qualquer alteração direta em cadastros fiscais.
# Suporta: CAD_PRODUTO, CAD_FAMILIA, ORIGEM, CLIENTE, FORNECEDOR,
#          OP_FISCAL, CLASS_FISCAL.
# Fluxo: PENDENTE → APROVADO_1 → APROVADO_2 → APLICADO_ERP.
# Tabelas próprias: USU_TAUDIMP_SOL, USU_TAUDIMP_ALT, USU_TAUDIMP_LOG.

class AlteracaoFiscalRequest(BaseModel):
    camada: str
    grupo_imposto: Optional[str] = None
    campo_tela: str
    campo_trilha: Optional[str] = None
    valor_atual: Optional[Any] = None
    valor_novo: Optional[Any] = None


class SolicitacaoFiscalRequest(BaseModel):
    codemp: int = EMPRESA_PADRAO
    codfil: Optional[int] = None

    origem_solicitacao: str
    entidade: str

    # chaves possíveis — usadas conforme a entidade
    codpro: Optional[str] = None
    codder: Optional[str] = None
    codfam: Optional[str] = None
    codori: Optional[str] = None
    codcli: Optional[int] = None
    codfor: Optional[int] = None
    codtns: Optional[str] = None
    codclf: Optional[str] = None

    grupo_imposto: Optional[str] = None
    motivo: str

    trilha: Optional[Dict[str, Any]] = None
    alteracoes: List[AlteracaoFiscalRequest]


class AprovacaoFiscalRequest(BaseModel):
    observacao: Optional[str] = None


# ---------- Configuração de entidades fiscais ----------
# Cada entidade conhece sua tabela primária, suas chaves e como montar
# o WHERE para UPDATE. Tabelas adicionais (ex.: E075DER) são listadas em
# `tabelas_extras` e tratadas por TABLE_KEY_CONFIG abaixo.
ENTITY_CONFIG: Dict[str, Dict[str, Any]] = {
    "CAD_PRODUTO": {
        "tabela_primaria": "E075PRO",
        "tabelas_extras": ["E075DER"],
        "chaves": ["CODEMP", "CODPRO"],
        "campos_contexto": ["codemp", "codpro", "codder"],
    },
    "CAD_FAMILIA": {
        "tabela_primaria": "E012FAM",
        "tabelas_extras": [],
        "chaves": ["CODEMP", "CODFAM"],
        "campos_contexto": ["codemp", "codfam"],
    },
    "ORIGEM": {
        "tabela_primaria": "E083ORI",
        "tabelas_extras": [],
        "chaves": ["CODEMP", "CODORI"],
        "campos_contexto": ["codemp", "codori"],
    },
    "CLIENTE": {
        "tabela_primaria": "E085CLI",
        "tabelas_extras": [],
        "chaves": ["CODCLI"],
        "campos_contexto": ["codcli"],
    },
    "FORNECEDOR": {
        "tabela_primaria": "E095FOR",
        "tabelas_extras": [],
        "chaves": ["CODFOR"],
        "campos_contexto": ["codfor"],
    },
    "OP_FISCAL": {
        "tabela_primaria": "E001TNS",
        "tabelas_extras": ["E001TCP", "E001TNC"],
        "chaves": ["CODEMP", "CODTNS"],
        "campos_contexto": ["codemp", "codtns"],
    },
    "CLASS_FISCAL": {
        "tabela_primaria": "E022CLF",
        "tabelas_extras": [],
        "chaves": ["CODCLF"],
        "campos_contexto": ["codclf"],
    },
}

# Por tabela ERP: como montar WHERE e quais atributos do solicitação são
# os parâmetros (na ordem). Usado pelo aplicador genérico.
TABLE_KEY_CONFIG: Dict[str, Dict[str, Any]] = {
    "E075PRO": {"entidade": "CAD_PRODUTO", "where": "CODEMP = ? AND CODPRO = ?",                   "params": ("CODEMP", "CODPRO")},
    "E075DER": {"entidade": "CAD_PRODUTO", "where": "CODEMP = ? AND CODPRO = ? AND CODDER = ?",    "params": ("CODEMP", "CODPRO", "CODDER")},
    "E012FAM": {"entidade": "CAD_FAMILIA", "where": "CODEMP = ? AND CODFAM = ?",                   "params": ("CODEMP", "CODFAM")},
    "E083ORI": {"entidade": "ORIGEM",      "where": "CODEMP = ? AND CODORI = ?",                   "params": ("CODEMP", "CODORI")},
    "E085CLI": {"entidade": "CLIENTE",     "where": "CODCLI = ?",                                  "params": ("CODCLI",)},
    "E095FOR": {"entidade": "FORNECEDOR",  "where": "CODFOR = ?",                                  "params": ("CODFOR",)},
    "E001TNS": {"entidade": "OP_FISCAL",   "where": "CODEMP = ? AND CODTNS = ?",                   "params": ("CODEMP", "CODTNS")},
    "E001TCP": {"entidade": "OP_FISCAL",   "where": "CODEMP = ? AND CODTNS = ?",                   "params": ("CODEMP", "CODTNS")},
    "E001TNC": {"entidade": "OP_FISCAL",   "where": "CODEMP = ? AND CODTNS = ?",                   "params": ("CODEMP", "CODTNS")},
    "E022CLF": {"entidade": "CLASS_FISCAL", "where": "CODCLF = ?",                                 "params": ("CODCLF",)},
}


# ---------- Mapa oficial de campos fiscais ----------
# campo_tela  -> entidade + tabela ERP + campo ERP + grupo de imposto.
# É a ÚNICA fonte de verdade para alterações: a API só atualiza campos que
# existem aqui, e o nome físico do campo nunca vem do payload do usuário.
FISCAL_FIELD_MAP: Dict[str, Dict[str, str]] = {
    # --- Produto / PIS-COFINS ---
    "cad_recpis":         {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "RECPIS", "grupo": "PIS_COFINS"},
    "cad_reccof":         {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "RECCOF", "grupo": "PIS_COFINS"},
    "cad_cstpis_produto": {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "CSTPIS", "grupo": "PIS_COFINS"},
    "cad_cstcof_produto": {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "CSTCOF", "grupo": "PIS_COFINS"},
    "cad_perpis":         {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "PERPIS", "grupo": "PIS_COFINS"},
    "cad_percof":         {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "PERCOF", "grupo": "PIS_COFINS"},
    "cad_basrec":         {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "BASREC", "grupo": "PIS_COFINS"},
    "cad_bascre_produto": {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "BASCRE", "grupo": "PIS_COFINS"},
    # --- Produto / IPI ---
    "cad_cstipi_produto": {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "CSTIPI", "grupo": "IPI"},
    "cad_peripi":         {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "PERIPI", "grupo": "IPI"},
    "cad_recipi":         {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "RECIPI", "grupo": "IPI"},
    "cad_codenq":         {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "CODENQ", "grupo": "IPI"},
    # --- Produto / ICMS ---
    "cad_temicm":         {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "TEMICM", "grupo": "ICMS"},
    "cad_recicm":         {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "RECICM", "grupo": "ICMS"},
    "cad_codtrd":         {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "CODTRD", "grupo": "ICMS"},
    "cad_codstr":         {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "CODSTR", "grupo": "ICMS"},
    # --- Produto / ICMS-ST ---
    "cad_codtst":         {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "CODTST", "grupo": "ICMS_ST"},
    "cad_codstp":         {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "CODSTP", "grupo": "ICMS_ST"},
    # --- Produto / Geral ---
    "cad_regtri":         {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "REGTRI", "grupo": "GERAL"},
    "cad_codclf":         {"entidade": "CAD_PRODUTO", "tabela": "E075PRO", "campo": "CODCLF", "grupo": "CLASS_FISCAL"},
    # --- Produto / Derivação (E075DER) ---
    "cad_codces":         {"entidade": "CAD_PRODUTO", "tabela": "E075DER", "campo": "CODCES", "grupo": "ICMS_ST"},
    "cad_itefis":         {"entidade": "CAD_PRODUTO", "tabela": "E075DER", "campo": "ITEFIS", "grupo": "GERAL"},
    "cad_codfif":         {"entidade": "CAD_PRODUTO", "tabela": "E075DER", "campo": "CODFIF", "grupo": "GERAL"},
    "cad_codfie":         {"entidade": "CAD_PRODUTO", "tabela": "E075DER", "campo": "CODFIE", "grupo": "GERAL"},
    "cad_codfim":         {"entidade": "CAD_PRODUTO", "tabela": "E075DER", "campo": "CODFIM", "grupo": "GERAL"},
    # --- Família ---
    "fam_recpis":         {"entidade": "CAD_FAMILIA", "tabela": "E012FAM", "campo": "RECPIS", "grupo": "PIS_COFINS"},
    "fam_reccof":         {"entidade": "CAD_FAMILIA", "tabela": "E012FAM", "campo": "RECCOF", "grupo": "PIS_COFINS"},
    "fam_cst_pis":        {"entidade": "CAD_FAMILIA", "tabela": "E012FAM", "campo": "CSTPIS", "grupo": "PIS_COFINS"},
    "fam_cst_cofins":     {"entidade": "CAD_FAMILIA", "tabela": "E012FAM", "campo": "CSTCOF", "grupo": "PIS_COFINS"},
    "fam_codtrd":         {"entidade": "CAD_FAMILIA", "tabela": "E012FAM", "campo": "CODTRD", "grupo": "ICMS"},
    "fam_codtst":         {"entidade": "CAD_FAMILIA", "tabela": "E012FAM", "campo": "CODTST", "grupo": "ICMS_ST"},
    # --- Origem ---
    "ori_codreg":         {"entidade": "ORIGEM",     "tabela": "E083ORI", "campo": "CODREG", "grupo": "ORIGEM"},
    "ori_proimp":         {"entidade": "ORIGEM",     "tabela": "E083ORI", "campo": "PROIMP", "grupo": "ORIGEM"},
    # --- Fornecedor ---
    "fornecedor_codtri":  {"entidade": "FORNECEDOR", "tabela": "E095FOR", "campo": "CODTRI", "grupo": "FORNECEDOR"},
    "fornecedor_tipfor":  {"entidade": "FORNECEDOR", "tabela": "E095FOR", "campo": "TIPFOR", "grupo": "FORNECEDOR"},
}

# (tabela_erp, campo_erp) -> {campo_tela, entidade, grupo, tabela, campo}
# Permite mapeamento reverso (ex.: o legado /salvar manda campos por nome ERP).
FISCAL_FIELD_REVERSE_MAP: Dict[tuple, Dict[str, str]] = {
    (m["tabela"].upper(), m["campo"].upper()): {"campo_tela": k, **m}
    for k, m in FISCAL_FIELD_MAP.items()
}

# Conjunto de tabelas habilitadas para aplicação automática.
TABELAS_APLICACAO_PERMITIDAS = {m["tabela"].upper() for m in FISCAL_FIELD_MAP.values()}

# Status finais — não aceitam mais aprovação/reprovação/aplicação.
STATUS_FINAIS_SOLICITACAO = ("APLICADO_ERP", "REPROVADO", "CANCELADO")

# Numéricos por tabela (estende caso necessário no futuro).
NUMERIC_FIELDS_POR_TABELA: Dict[str, set] = {
    "E075PRO": NUMERIC_FIELDS_E075PRO,
    "E075DER": NUMERIC_FIELDS_E075DER,
}

_WORKFLOW_TABLES_READY = False


def _ensure_workflow_tables():
    """Cria/atualiza tabelas do workflow fiscal (idempotente)."""
    global _WORKFLOW_TABLES_READY
    if _WORKFLOW_TABLES_READY:
        return

    conn = get_connection()
    cursor = conn.cursor()
    try:
        # Tabela principal de solicitações
        cursor.execute("""
            IF OBJECT_ID('dbo.USU_TAUDIMP_SOL', 'U') IS NULL
            BEGIN
                CREATE TABLE dbo.USU_TAUDIMP_SOL (
                    ID INT IDENTITY(1,1) PRIMARY KEY,
                    CODEMP INT NOT NULL,
                    CODFIL INT NULL,
                    ORIGEM_SOLICITACAO VARCHAR(50) NOT NULL,
                    ENTIDADE VARCHAR(50) NOT NULL,
                    STATUS VARCHAR(30) NOT NULL DEFAULT 'PENDENTE',
                    CHAVE_REGISTRO NVARCHAR(1000) NOT NULL,
                    CODPRO VARCHAR(50) NULL,
                    CODDER VARCHAR(50) NULL,
                    CODFAM VARCHAR(50) NULL,
                    CODORI VARCHAR(50) NULL,
                    CODCLI INT NULL,
                    CODFOR INT NULL,
                    CODTNS VARCHAR(20) NULL,
                    CODCLF VARCHAR(50) NULL,
                    GRUPO_IMPOSTO VARCHAR(50) NULL,
                    MOTIVO NVARCHAR(1000) NULL,
                    USUARIO_SOLICITANTE VARCHAR(100) NOT NULL,
                    DATA_SOLICITACAO DATETIME2 NOT NULL DEFAULT SYSDATETIME(),
                    USUARIO_APROVADOR_1 VARCHAR(100) NULL,
                    DATA_APROVACAO_1 DATETIME2 NULL,
                    USUARIO_APROVADOR_2 VARCHAR(100) NULL,
                    DATA_APROVACAO_2 DATETIME2 NULL,
                    DATA_APLICACAO_ERP DATETIME2 NULL,
                    MSG_ERRO NVARCHAR(MAX) NULL,
                    SNAPSHOT_ANTES_JSON NVARCHAR(MAX) NULL,
                    TRILHA_JSON NVARCHAR(MAX) NULL
                );
            END
        """)
        # ALTERs idempotentes para upgrade in-place de instalações antigas
        for col_name, col_ddl in [
            ("ENTIDADE",            "VARCHAR(50) NOT NULL DEFAULT 'CAD_PRODUTO'"),
            ("CHAVE_REGISTRO",      "NVARCHAR(1000) NOT NULL DEFAULT ''"),
            ("CODFAM",              "VARCHAR(50) NULL"),
            ("CODORI",              "VARCHAR(50) NULL"),
            ("CODTNS",              "VARCHAR(20) NULL"),
            ("CODCLF",              "VARCHAR(50) NULL"),
            ("GRUPO_IMPOSTO",       "VARCHAR(50) NULL"),
            ("SNAPSHOT_ANTES_JSON", "NVARCHAR(MAX) NULL"),
        ]:
            cursor.execute(f"""
                IF COL_LENGTH('dbo.USU_TAUDIMP_SOL', '{col_name}') IS NULL
                    ALTER TABLE dbo.USU_TAUDIMP_SOL ADD {col_name} {col_ddl};
            """)

        # Tabela de alterações por solicitação
        cursor.execute("""
            IF OBJECT_ID('dbo.USU_TAUDIMP_ALT', 'U') IS NULL
            BEGIN
                CREATE TABLE dbo.USU_TAUDIMP_ALT (
                    ID INT IDENTITY(1,1) PRIMARY KEY,
                    ID_SOLICITACAO INT NOT NULL,
                    CAMADA VARCHAR(50) NOT NULL,
                    GRUPO_IMPOSTO VARCHAR(50) NULL,
                    CAMPO_TELA VARCHAR(100) NOT NULL,
                    CAMPO_TRILHA VARCHAR(100) NULL,
                    TABELA_ERP VARCHAR(50) NOT NULL,
                    CAMPO_ERP VARCHAR(100) NOT NULL,
                    VALOR_ATUAL NVARCHAR(500) NULL,
                    VALOR_NOVO NVARCHAR(500) NULL,
                    APLICADO BIT NOT NULL DEFAULT 0,
                    CONSTRAINT FK_USU_TAUDIMP_ALT_SOL
                        FOREIGN KEY (ID_SOLICITACAO)
                        REFERENCES dbo.USU_TAUDIMP_SOL(ID)
                );
            END
        """)
        for col_name, col_ddl in [
            ("CAMPO_TELA",    "VARCHAR(100) NOT NULL DEFAULT ''"),
            ("GRUPO_IMPOSTO", "VARCHAR(50) NULL"),
        ]:
            cursor.execute(f"""
                IF COL_LENGTH('dbo.USU_TAUDIMP_ALT', '{col_name}') IS NULL
                    ALTER TABLE dbo.USU_TAUDIMP_ALT ADD {col_name} {col_ddl};
            """)

        # Tabela de log de eventos
        cursor.execute("""
            IF OBJECT_ID('dbo.USU_TAUDIMP_LOG', 'U') IS NULL
            BEGIN
                CREATE TABLE dbo.USU_TAUDIMP_LOG (
                    ID INT IDENTITY(1,1) PRIMARY KEY,
                    ID_SOLICITACAO INT NOT NULL,
                    ACAO VARCHAR(50) NOT NULL,
                    USUARIO VARCHAR(100) NOT NULL,
                    DATA_HORA DATETIME2 NOT NULL DEFAULT SYSDATETIME(),
                    OBSERVACAO NVARCHAR(1000) NULL,
                    PAYLOAD_JSON NVARCHAR(MAX) NULL
                );
            END
        """)

        conn.commit()
        _WORKFLOW_TABLES_READY = True
    except Exception as e:
        conn.rollback()
        logger.exception("falha ao criar/atualizar tabelas de workflow fiscal: %s", e)
        raise HTTPException(status_code=500, detail=f"Erro ao preparar tabelas de workflow: {e}")
    finally:
        conn.close()


# ---------- Helpers do motor ----------

def _entidade_chaves_dict(payload_or_dict: Any) -> Dict[str, Any]:
    """Retorna as chaves de contexto disponíveis em um payload ou dict, em uppercase."""
    keys = ("codemp", "codfil", "codpro", "codder", "codfam", "codori",
            "codcli", "codfor", "codtns", "codclf")
    out = {}
    for k in keys:
        if isinstance(payload_or_dict, BaseModel):
            v = getattr(payload_or_dict, k, None)
        else:
            v = payload_or_dict.get(k) if payload_or_dict else None
        if v is not None and v != "":
            out[k.upper()] = v
    return out


def _montar_chave_registro(payload: SolicitacaoFiscalRequest) -> str:
    """Texto curto identificando o registro alvo, usado para listagens/UI."""
    cfg = ENTITY_CONFIG.get(payload.entidade.upper().strip())
    if not cfg:
        return f"{payload.entidade}"
    partes = [payload.entidade]
    chaves = _entidade_chaves_dict(payload)
    for k in cfg["campos_contexto"]:
        v = chaves.get(k.upper())
        if v is not None:
            partes.append(f"{k.upper()}={v}")
    return " | ".join(partes)


def _buscar_snapshot_entidade(cursor, payload: SolicitacaoFiscalRequest) -> Dict[str, Any]:
    """Lê o estado atual dos campos fiscais da entidade alvo (apenas tabela primária)."""
    entidade = payload.entidade.upper().strip()
    cfg = ENTITY_CONFIG.get(entidade)
    if not cfg:
        return {}

    tabela = cfg["tabela_primaria"].upper()
    table_cfg = TABLE_KEY_CONFIG.get(tabela)
    if not table_cfg:
        return {}

    chaves = _entidade_chaves_dict(payload)
    params = []
    for col in table_cfg["params"]:
        v = chaves.get(col)
        if v is None:
            return {}  # contexto incompleto, sem snapshot
        params.append(v)

    campos = sorted({m["campo"] for m in FISCAL_FIELD_MAP.values()
                     if m["entidade"] == entidade and m["tabela"].upper() == tabela})
    if not campos:
        return {}
    campos_select = list(table_cfg["params"]) + campos

    sql = f"SELECT {', '.join(campos_select)} FROM {tabela} WHERE {table_cfg['where']}"
    try:
        cursor.execute(sql, params)
        row = cursor.fetchone()
    except Exception as e:
        logger.warning("snapshot %s falhou: %s", tabela, e)
        return {}
    if not row:
        return {}
    return row_to_dict(cursor, row)


def _resolver_campo_alteracao(alt: AlteracaoFiscalRequest, entidade: str) -> Dict[str, str]:
    """Valida e resolve o campo da alteração contra FISCAL_FIELD_MAP."""
    chave = alt.campo_tela or alt.campo_trilha
    if not chave:
        raise HTTPException(status_code=400, detail="Alteração sem campo_tela/campo_trilha")
    meta = FISCAL_FIELD_MAP.get(chave)
    if not meta:
        raise HTTPException(status_code=400, detail=f"Campo fiscal não permitido: {chave}")
    if meta["entidade"] != entidade:
        raise HTTPException(
            status_code=400,
            detail=f"Campo {chave} pertence a {meta['entidade']}, não a {entidade}"
        )
    return meta


def _coercer_valor_para_update(tabela: str, campo: str, valor: Any) -> Any:
    numericos = NUMERIC_FIELDS_POR_TABELA.get(tabela.upper(), set())
    if campo.upper() in numericos:
        return _normalizar_numero_sql(valor)
    return _valor_vazio_para_none(valor)


def _aplicar_alteracao_fiscal(cursor, solicitacao: Dict[str, Any], alteracao: Dict[str, Any]) -> None:
    """Aplica uma única alteração no ERP — segura contra injeção: tabela/campo
    são validados contra TABLE_KEY_CONFIG / FISCAL_FIELD_REVERSE_MAP."""
    tabela = (alteracao.get("TABELA_ERP") or "").upper()
    campo = (alteracao.get("CAMPO_ERP") or "").upper()
    valor = alteracao.get("VALOR_NOVO")

    if (tabela, campo) not in FISCAL_FIELD_REVERSE_MAP:
        raise HTTPException(status_code=400, detail=f"Campo não permitido para update: {tabela}.{campo}")

    table_cfg = TABLE_KEY_CONFIG.get(tabela)
    if not table_cfg:
        raise HTTPException(status_code=400, detail=f"Tabela não configurada: {tabela}")

    valor_param = _coercer_valor_para_update(tabela, campo, valor)

    where_params: List[Any] = []
    for col in table_cfg["params"]:
        v = solicitacao.get(col)
        if v is None:
            raise HTTPException(
                status_code=400,
                detail=f"Solicitação sem chave {col} para alteração em {tabela}"
            )
        where_params.append(v)

    sql = f"UPDATE {tabela} SET {campo} = ? WHERE {table_cfg['where']}"
    cursor.execute(sql, [valor_param] + where_params)


# ---------- Helpers de leitura ----------

def _carregar_solicitacao(cursor, id_solicitacao: int) -> Optional[Dict[str, Any]]:
    cursor.execute("""
        SELECT ID, CODEMP, CODFIL, ORIGEM_SOLICITACAO, ENTIDADE, STATUS, CHAVE_REGISTRO,
               CODPRO, CODDER, CODFAM, CODORI, CODCLI, CODFOR, CODTNS, CODCLF,
               GRUPO_IMPOSTO, MOTIVO,
               USUARIO_SOLICITANTE, DATA_SOLICITACAO,
               USUARIO_APROVADOR_1, DATA_APROVACAO_1,
               USUARIO_APROVADOR_2, DATA_APROVACAO_2,
               DATA_APLICACAO_ERP, MSG_ERRO,
               SNAPSHOT_ANTES_JSON, TRILHA_JSON
        FROM USU_TAUDIMP_SOL
        WHERE ID = ?
    """, [id_solicitacao])
    row = cursor.fetchone()
    if not row:
        return None
    return row_to_dict(cursor, row)


def _carregar_alteracoes(cursor, id_solicitacao: int) -> List[Dict[str, Any]]:
    cursor.execute("""
        SELECT ID, ID_SOLICITACAO, CAMADA, GRUPO_IMPOSTO,
               CAMPO_TELA, CAMPO_TRILHA, TABELA_ERP, CAMPO_ERP,
               VALOR_ATUAL, VALOR_NOVO, APLICADO
        FROM USU_TAUDIMP_ALT
        WHERE ID_SOLICITACAO = ?
        ORDER BY ID
    """, [id_solicitacao])
    cols = [c[0] for c in cursor.description]
    return [dict(zip(cols, r)) for r in cursor.fetchall()]


# ---------- Endpoints públicos ----------

@app.post("/api/auditoria-tributaria/solicitacoes")
def criar_solicitacao_fiscal(
    payload: SolicitacaoFiscalRequest,
    usuario=Depends(validar_token),
):
    _ensure_workflow_tables()

    if not payload.alteracoes:
        raise HTTPException(status_code=400, detail="Informe ao menos uma alteração")
    if not (payload.motivo or "").strip():
        raise HTTPException(status_code=400, detail="Motivo é obrigatório")

    entidade = payload.entidade.upper().strip()
    if entidade not in ENTITY_CONFIG:
        raise HTTPException(status_code=400, detail=f"Entidade não permitida: {entidade}")
    payload.entidade = entidade

    # validação de campos antes de transação
    metas = []
    for alt in payload.alteracoes:
        metas.append(_resolver_campo_alteracao(alt, entidade))

    # exige chaves mínimas da entidade
    chaves = _entidade_chaves_dict(payload)
    for col in ENTITY_CONFIG[entidade]["chaves"]:
        if col not in chaves:
            raise HTTPException(
                status_code=400,
                detail=f"Chave obrigatória ausente para {entidade}: {col}"
            )

    conn = get_connection()
    cursor = conn.cursor()
    try:
        snapshot_antes = _buscar_snapshot_entidade(cursor, payload)
        chave_registro = _montar_chave_registro(payload)

        cursor.execute("""
            INSERT INTO USU_TAUDIMP_SOL (
                CODEMP, CODFIL, ORIGEM_SOLICITACAO, ENTIDADE, STATUS, CHAVE_REGISTRO,
                CODPRO, CODDER, CODFAM, CODORI, CODCLI, CODFOR, CODTNS, CODCLF,
                GRUPO_IMPOSTO, MOTIVO, USUARIO_SOLICITANTE,
                SNAPSHOT_ANTES_JSON, TRILHA_JSON
            )
            OUTPUT INSERTED.ID
            VALUES (?, ?, ?, ?, 'PENDENTE', ?,
                    ?, ?, ?, ?, ?, ?, ?, ?,
                    ?, ?, ?,
                    ?, ?)
        """, [
            payload.codemp,
            payload.codfil,
            payload.origem_solicitacao,
            entidade,
            chave_registro,
            payload.codpro,
            payload.codder,
            payload.codfam,
            payload.codori,
            payload.codcli,
            payload.codfor,
            payload.codtns,
            payload.codclf,
            payload.grupo_imposto,
            payload.motivo,
            usuario,
            json.dumps(snapshot_antes, ensure_ascii=False, default=str),
            json.dumps(payload.trilha or {}, ensure_ascii=False, default=str),
        ])
        id_solicitacao = int(cursor.fetchone()[0])

        for alt, meta in zip(payload.alteracoes, metas):
            cursor.execute("""
                INSERT INTO USU_TAUDIMP_ALT (
                    ID_SOLICITACAO, CAMADA, GRUPO_IMPOSTO,
                    CAMPO_TELA, CAMPO_TRILHA, TABELA_ERP, CAMPO_ERP,
                    VALOR_ATUAL, VALOR_NOVO
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, [
                id_solicitacao,
                alt.camada,
                alt.grupo_imposto or meta.get("grupo"),
                alt.campo_tela or meta.get("campo_tela"),
                alt.campo_trilha,
                meta["tabela"],
                meta["campo"],
                None if alt.valor_atual is None else str(alt.valor_atual),
                None if alt.valor_novo is None else str(alt.valor_novo),
            ])

        cursor.execute("""
            INSERT INTO USU_TAUDIMP_LOG (
                ID_SOLICITACAO, ACAO, USUARIO, OBSERVACAO, PAYLOAD_JSON
            )
            VALUES (?, 'CRIADA', ?, ?, ?)
        """, [
            id_solicitacao,
            usuario,
            payload.motivo,
            json.dumps(payload.dict(), ensure_ascii=False, default=str),
        ])

        conn.commit()

        return {
            "ok": True,
            "id_solicitacao": id_solicitacao,
            "status": "PENDENTE",
            "entidade": entidade,
            "chave_registro": chave_registro,
            "mensagem": "Solicitação fiscal criada. Aguardando duas aprovações.",
        }
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao criar solicitação: {str(e)}")
    finally:
        conn.close()


@app.get("/api/auditoria-tributaria/solicitacoes")
def listar_solicitacoes_fiscais(
    status: Optional[str] = Query(default=None),
    entidade: Optional[str] = Query(default=None),
    codpro: Optional[str] = Query(default=None),
    codfam: Optional[str] = Query(default=None),
    codori: Optional[str] = Query(default=None),
    codcli: Optional[int] = Query(default=None),
    codfor: Optional[int] = Query(default=None),
    grupo_imposto: Optional[str] = Query(default=None),
    limite: int = Query(default=200, ge=1, le=1000),
    usuario=Depends(validar_token),
):
    _ensure_workflow_tables()

    conn = get_connection()
    cursor = conn.cursor()
    try:
        where = ["1 = 1"]
        params: List[Any] = []
        if status:
            where.append("STATUS = ?"); params.append(status.strip().upper())
        if entidade:
            where.append("ENTIDADE = ?"); params.append(entidade.strip().upper())
        if codpro:
            where.append("CODPRO = ?"); params.append(codpro.strip().upper())
        if codfam:
            where.append("CODFAM = ?"); params.append(codfam.strip().upper())
        if codori:
            where.append("CODORI = ?"); params.append(codori.strip().upper())
        if codcli is not None:
            where.append("CODCLI = ?"); params.append(codcli)
        if codfor is not None:
            where.append("CODFOR = ?"); params.append(codfor)
        if grupo_imposto:
            where.append("GRUPO_IMPOSTO = ?"); params.append(grupo_imposto.strip().upper())

        sql = f"""
            SELECT TOP {limite}
                ID, CODEMP, CODFIL, ORIGEM_SOLICITACAO, ENTIDADE, STATUS, CHAVE_REGISTRO,
                CODPRO, CODDER, CODFAM, CODORI, CODCLI, CODFOR, CODTNS, CODCLF,
                GRUPO_IMPOSTO, MOTIVO, USUARIO_SOLICITANTE, DATA_SOLICITACAO,
                USUARIO_APROVADOR_1, DATA_APROVACAO_1,
                USUARIO_APROVADOR_2, DATA_APROVACAO_2,
                DATA_APLICACAO_ERP
            FROM USU_TAUDIMP_SOL
            WHERE {' AND '.join(where)}
            ORDER BY ID DESC
        """
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        return {
            "ok": True,
            "total": len(rows),
            "solicitacoes": [row_to_dict(cursor, r) for r in rows],
        }
    finally:
        conn.close()


@app.get("/api/auditoria-tributaria/solicitacoes/{id_solicitacao}")
def obter_solicitacao_fiscal(
    id_solicitacao: int,
    usuario=Depends(validar_token),
):
    _ensure_workflow_tables()

    conn = get_connection()
    cursor = conn.cursor()
    try:
        sol = _carregar_solicitacao(cursor, id_solicitacao)
        if not sol:
            raise HTTPException(status_code=404, detail="Solicitação não encontrada")

        alteracoes = _carregar_alteracoes(cursor, id_solicitacao)

        cursor.execute("""
            SELECT ID, ACAO, USUARIO, DATA_HORA, OBSERVACAO
            FROM USU_TAUDIMP_LOG
            WHERE ID_SOLICITACAO = ?
            ORDER BY ID
        """, [id_solicitacao])
        log_cols = [c[0] for c in cursor.description]
        logs = [dict(zip(log_cols, r)) for r in cursor.fetchall()]

        for json_col, target in (("TRILHA_JSON", "TRILHA"), ("SNAPSHOT_ANTES_JSON", "SNAPSHOT_ANTES")):
            try:
                sol[target] = json.loads(sol.get(json_col) or "{}")
            except Exception:
                sol[target] = {}

        return {
            "ok": True,
            "solicitacao": sol,
            "alteracoes": alteracoes,
            "log": logs,
        }
    finally:
        conn.close()


@app.post("/api/auditoria-tributaria/solicitacoes/{id_solicitacao}/aprovar")
def aprovar_solicitacao_fiscal(
    id_solicitacao: int,
    payload: Optional[AprovacaoFiscalRequest] = Body(default=None),
    usuario=Depends(validar_token),
):
    _ensure_workflow_tables()
    observacao = payload.observacao if payload else None

    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT STATUS, USUARIO_SOLICITANTE, USUARIO_APROVADOR_1, USUARIO_APROVADOR_2
            FROM USU_TAUDIMP_SOL
            WHERE ID = ?
        """, [id_solicitacao])
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Solicitação não encontrada")

        status, solicitante, aprovador1, _aprovador2 = row

        if status in STATUS_FINAIS_SOLICITACAO:
            raise HTTPException(status_code=400, detail=f"Solicitação não pode ser aprovada no status {status}")
        if usuario == solicitante:
            raise HTTPException(status_code=400, detail="Solicitante não pode aprovar a própria alteração")
        if aprovador1 and usuario == aprovador1:
            raise HTTPException(status_code=400, detail="O mesmo usuário não pode fazer as duas aprovações")

        if not aprovador1:
            cursor.execute("""
                UPDATE USU_TAUDIMP_SOL
                SET STATUS = 'APROVADO_1',
                    USUARIO_APROVADOR_1 = ?,
                    DATA_APROVACAO_1 = SYSDATETIME()
                WHERE ID = ?
            """, [usuario, id_solicitacao])
            novo_status = "APROVADO_1"
        else:
            cursor.execute("""
                UPDATE USU_TAUDIMP_SOL
                SET STATUS = 'APROVADO_2',
                    USUARIO_APROVADOR_2 = ?,
                    DATA_APROVACAO_2 = SYSDATETIME()
                WHERE ID = ?
            """, [usuario, id_solicitacao])
            novo_status = "APROVADO_2"

        cursor.execute("""
            INSERT INTO USU_TAUDIMP_LOG (ID_SOLICITACAO, ACAO, USUARIO, OBSERVACAO)
            VALUES (?, 'APROVADA', ?, ?)
        """, [id_solicitacao, usuario, observacao])

        conn.commit()
        return {"ok": True, "id_solicitacao": id_solicitacao, "status": novo_status}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao aprovar solicitação: {str(e)}")
    finally:
        conn.close()


@app.post("/api/auditoria-tributaria/solicitacoes/{id_solicitacao}/reprovar")
def reprovar_solicitacao_fiscal(
    id_solicitacao: int,
    payload: Optional[AprovacaoFiscalRequest] = Body(default=None),
    usuario=Depends(validar_token),
):
    _ensure_workflow_tables()
    observacao = payload.observacao if payload else None

    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT STATUS FROM USU_TAUDIMP_SOL WHERE ID = ?", [id_solicitacao])
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Solicitação não encontrada")

        if row[0] in STATUS_FINAIS_SOLICITACAO:
            raise HTTPException(status_code=400, detail=f"Solicitação já está em status final: {row[0]}")

        cursor.execute("UPDATE USU_TAUDIMP_SOL SET STATUS = 'REPROVADO' WHERE ID = ?", [id_solicitacao])
        cursor.execute("""
            INSERT INTO USU_TAUDIMP_LOG (ID_SOLICITACAO, ACAO, USUARIO, OBSERVACAO)
            VALUES (?, 'REPROVADA', ?, ?)
        """, [id_solicitacao, usuario, observacao])

        conn.commit()
        return {"ok": True, "id_solicitacao": id_solicitacao, "status": "REPROVADO"}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao reprovar solicitação: {str(e)}")
    finally:
        conn.close()


@app.post("/api/auditoria-tributaria/solicitacoes/{id_solicitacao}/aplicar")
def aplicar_solicitacao_fiscal(
    id_solicitacao: int,
    usuario=Depends(validar_token),
):
    _ensure_workflow_tables()

    conn = get_connection()
    cursor = conn.cursor()
    try:
        sol = _carregar_solicitacao(cursor, id_solicitacao)
        if not sol:
            raise HTTPException(status_code=404, detail="Solicitação não encontrada")

        if sol.get("STATUS") != "APROVADO_2":
            raise HTTPException(
                status_code=400,
                detail="Solicitação só pode ser aplicada após duas aprovações (status APROVADO_2)"
            )

        cursor.execute("""
            SELECT ID, CAMADA, CAMPO_TELA, TABELA_ERP, CAMPO_ERP, VALOR_NOVO
            FROM USU_TAUDIMP_ALT
            WHERE ID_SOLICITACAO = ? AND APLICADO = 0
        """, [id_solicitacao])
        cols = [c[0] for c in cursor.description]
        alteracoes = [dict(zip(cols, r)) for r in cursor.fetchall()]

        if not alteracoes:
            raise HTTPException(status_code=400, detail="Nenhuma alteração pendente para aplicar")

        for alt in alteracoes:
            tabela_u = (alt["TABELA_ERP"] or "").upper()
            if tabela_u not in TABELAS_APLICACAO_PERMITIDAS:
                raise HTTPException(
                    status_code=400,
                    detail=f"Tabela ainda não liberada para aplicação automática: {tabela_u}"
                )
            _aplicar_alteracao_fiscal(cursor, sol, alt)
            cursor.execute("UPDATE USU_TAUDIMP_ALT SET APLICADO = 1 WHERE ID = ?", [alt["ID"]])

        cursor.execute("""
            UPDATE USU_TAUDIMP_SOL
            SET STATUS = 'APLICADO_ERP',
                DATA_APLICACAO_ERP = SYSDATETIME()
            WHERE ID = ?
        """, [id_solicitacao])
        cursor.execute("""
            INSERT INTO USU_TAUDIMP_LOG (ID_SOLICITACAO, ACAO, USUARIO, OBSERVACAO)
            VALUES (?, 'APLICADA_ERP', ?, 'Alteração aplicada no ERP após dupla aprovação')
        """, [id_solicitacao, usuario])

        conn.commit()

        try:
            registrar_auditoria_fiscal(
                acao="APLICAR_SOLICITACAO",
                codemp=sol.get("CODEMP"),
                codpro=sol.get("CODPRO"),
                codder=sol.get("CODDER"),
                usuario=usuario,
                payload_antes=None,
                payload_depois={
                    "id_solicitacao": id_solicitacao,
                    "entidade": sol.get("ENTIDADE"),
                    "alteracoes_aplicadas": len(alteracoes),
                },
                sucesso=True,
            )
        except Exception:
            pass

        return {
            "ok": True,
            "id_solicitacao": id_solicitacao,
            "status": "APLICADO_ERP",
            "alteracoes_aplicadas": len(alteracoes),
        }
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        try:
            cursor.execute("""
                UPDATE USU_TAUDIMP_SOL
                SET STATUS = 'ERRO_APLICACAO',
                    MSG_ERRO = ?
                WHERE ID = ?
            """, [str(e), id_solicitacao])
            cursor.execute("""
                INSERT INTO USU_TAUDIMP_LOG (ID_SOLICITACAO, ACAO, USUARIO, OBSERVACAO)
                VALUES (?, 'ERRO_APLICACAO', ?, ?)
            """, [id_solicitacao, usuario, str(e)[:1000]])
            conn.commit()
        except Exception:
            pass
        raise HTTPException(status_code=500, detail=f"Erro ao aplicar no ERP: {str(e)}")
    finally:
        conn.close()


# ---------- Atalho legado: /salvar -> cria solicitação ----------

def _salvar_como_solicitacao_fiscal(
    payload: ControleFiscalSalvarRequest,
    usuario: str,
) -> Dict[str, Any]:
    """Converte o payload legado de /api/controle-fiscal-produtos/salvar em uma
    solicitação no novo workflow. Não grava nada no ERP — apenas abre o ticket."""

    codemp = payload.codemp
    codpro = (payload.codpro or "").strip().upper() or None
    codder = (payload.codder or "").strip().upper() if payload.codder else None
    if not codpro:
        raise HTTPException(status_code=400, detail="codpro é obrigatório")

    campos_pro = _normalizar_chaves(payload.campos_pro)
    campos_der = _normalizar_chaves(payload.campos_der)

    alteracoes: List[AlteracaoFiscalRequest] = []

    def _add_alteracoes(tabela: str, campos: Dict[str, Any]):
        for campo, valor in campos.items():
            meta = FISCAL_FIELD_REVERSE_MAP.get((tabela.upper(), campo.upper()))
            if not meta:
                # campo não está no mapa — ignora silenciosamente para não bloquear
                # a UI antiga se enviar campos extras (ex.: SITPRO)
                continue
            alteracoes.append(AlteracaoFiscalRequest(
                camada="Cadastro",
                grupo_imposto=meta.get("grupo"),
                campo_tela=meta["campo_tela"],
                valor_atual=None,
                valor_novo=valor,
            ))

    _add_alteracoes("E075PRO", campos_pro)
    if codder:
        _add_alteracoes("E075DER", campos_der)

    if not alteracoes:
        raise HTTPException(status_code=400, detail="Nenhum campo válido para alteração")

    motivo = "Solicitação enviada pela tela de controle fiscal de produtos"
    if payload.ativar_apos_salvar:
        motivo += " (com pedido de ativação — ativação só após aplicação no ERP)"

    sol_payload = SolicitacaoFiscalRequest(
        codemp=codemp,
        origem_solicitacao="CONTROLE_FISCAL_PRODUTO",
        entidade="CAD_PRODUTO",
        codpro=codpro,
        codder=codder,
        motivo=motivo,
        alteracoes=alteracoes,
    )

    resp = criar_solicitacao_fiscal(sol_payload, usuario=usuario)  # type: ignore[arg-type]
    resp["mensagem"] = (
        "Alteração enviada para aprovação. Nenhum dado foi gravado no ERP ainda. "
        "Fluxo: 2 aprovações distintas + /aplicar."
    )
    return resp


# ---------- GET de cadastros fiscais individuais ----------

def _buscar_cadastro_fiscal_generico(
    cursor,
    entidade: str,
    chaves: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    cfg = ENTITY_CONFIG.get(entidade)
    if not cfg:
        raise HTTPException(status_code=400, detail=f"Entidade não suportada: {entidade}")

    tabela = cfg["tabela_primaria"].upper()
    table_cfg = TABLE_KEY_CONFIG.get(tabela)
    if not table_cfg:
        raise HTTPException(status_code=400, detail=f"Tabela {tabela} sem configuração de chave")

    params = []
    for col in table_cfg["params"]:
        v = chaves.get(col.upper())
        if v is None:
            raise HTTPException(status_code=400, detail=f"Chave obrigatória ausente: {col}")
        params.append(v)

    campos_fiscais = sorted({m["campo"] for m in FISCAL_FIELD_MAP.values()
                             if m["entidade"] == entidade and m["tabela"].upper() == tabela})
    campos_select = list(table_cfg["params"]) + campos_fiscais
    sql = f"SELECT {', '.join(campos_select)} FROM {tabela} WHERE {table_cfg['where']}"
    cursor.execute(sql, params)
    row = cursor.fetchone()
    return row_to_dict(cursor, row) if row else None


@app.get("/api/cadastros-fiscais/familia/{codfam}")
def consultar_cadastro_familia(
    codfam: str,
    codemp: int = Query(default=EMPRESA_PADRAO),
    usuario=Depends(validar_token),
):
    conn = get_connection(); cursor = conn.cursor()
    try:
        rec = _buscar_cadastro_fiscal_generico(
            cursor, "CAD_FAMILIA", {"CODEMP": codemp, "CODFAM": codfam.strip().upper()}
        )
        if not rec:
            raise HTTPException(status_code=404, detail="Família não encontrada")
        return {"ok": True, "entidade": "CAD_FAMILIA", "registro": rec}
    finally:
        conn.close()


@app.get("/api/cadastros-fiscais/origem/{codori}")
def consultar_cadastro_origem(
    codori: str,
    codemp: int = Query(default=EMPRESA_PADRAO),
    usuario=Depends(validar_token),
):
    conn = get_connection(); cursor = conn.cursor()
    try:
        rec = _buscar_cadastro_fiscal_generico(
            cursor, "ORIGEM", {"CODEMP": codemp, "CODORI": codori.strip().upper()}
        )
        if not rec:
            raise HTTPException(status_code=404, detail="Origem não encontrada")
        return {"ok": True, "entidade": "ORIGEM", "registro": rec}
    finally:
        conn.close()


@app.get("/api/cadastros-fiscais/cliente/{codcli}")
def consultar_cadastro_cliente(codcli: int, usuario=Depends(validar_token)):
    conn = get_connection(); cursor = conn.cursor()
    try:
        rec = _buscar_cadastro_fiscal_generico(cursor, "CLIENTE", {"CODCLI": codcli})
        if not rec:
            raise HTTPException(status_code=404, detail="Cliente não encontrado")
        return {"ok": True, "entidade": "CLIENTE", "registro": rec}
    finally:
        conn.close()


@app.get("/api/cadastros-fiscais/fornecedor/{codfor}")
def consultar_cadastro_fornecedor(codfor: int, usuario=Depends(validar_token)):
    conn = get_connection(); cursor = conn.cursor()
    try:
        rec = _buscar_cadastro_fiscal_generico(cursor, "FORNECEDOR", {"CODFOR": codfor})
        if not rec:
            raise HTTPException(status_code=404, detail="Fornecedor não encontrado")
        return {"ok": True, "entidade": "FORNECEDOR", "registro": rec}
    finally:
        conn.close()


@app.get("/api/cadastros-fiscais/operacao-fiscal/{codtns}")
def consultar_cadastro_operacao_fiscal(
    codtns: str,
    codemp: int = Query(default=EMPRESA_PADRAO),
    usuario=Depends(validar_token),
):
    conn = get_connection(); cursor = conn.cursor()
    try:
        rec = _buscar_cadastro_fiscal_generico(
            cursor, "OP_FISCAL", {"CODEMP": codemp, "CODTNS": codtns.strip().upper()}
        )
        if not rec:
            raise HTTPException(status_code=404, detail="Operação fiscal não encontrada")
        return {"ok": True, "entidade": "OP_FISCAL", "registro": rec}
    finally:
        conn.close()


@app.get("/api/cadastros-fiscais/classificacao-fiscal/{codclf}")
def consultar_cadastro_classificacao_fiscal(codclf: str, usuario=Depends(validar_token)):
    conn = get_connection(); cursor = conn.cursor()
    try:
        rec = _buscar_cadastro_fiscal_generico(
            cursor, "CLASS_FISCAL", {"CODCLF": codclf.strip().upper()}
        )
        if not rec:
            raise HTTPException(status_code=404, detail="Classificação fiscal não encontrada")
        return {"ok": True, "entidade": "CLASS_FISCAL", "registro": rec}
    finally:
        conn.close()


@app.get("/api/cadastros-fiscais/produto/{codpro}")
def consultar_cadastro_produto(
    codpro: str,
    codder: Optional[str] = Query(default=None),
    codemp: int = Query(default=EMPRESA_PADRAO),
    usuario=Depends(validar_token),
):
    conn = get_connection(); cursor = conn.cursor()
    try:
        rec = _buscar_cadastro_fiscal_generico(
            cursor, "CAD_PRODUTO", {"CODEMP": codemp, "CODPRO": codpro.strip().upper()}
        )
        if not rec:
            raise HTTPException(status_code=404, detail="Produto não encontrado")
        if codder:
            cder = codder.strip().upper()
            cursor.execute(
                "SELECT * FROM E075DER WHERE CODEMP = ? AND CODPRO = ? AND CODDER = ?",
                [codemp, codpro.strip().upper(), cder]
            )
            row = cursor.fetchone()
            rec["DERIVACAO"] = row_to_dict(cursor, row) if row else None
        return {"ok": True, "entidade": "CAD_PRODUTO", "registro": rec}
    finally:
        conn.close()


class ControleFiscalSugerirIARequest(BaseModel):
    codemp: int = EMPRESA_PADRAO
    codpro: str
    codder: Optional[str] = None

    # contexto opcional para melhorar a sugestão
    descricao_operacao: Optional[str] = None
    uf_origem: Optional[str] = None
    uf_destino: Optional[str] = None
    tipo_cliente: Optional[str] = None
    finalidade: Optional[str] = None
    transacao: Optional[str] = None

    # força foco em campos específicos
    campos_alvo: Optional[List[str]] = None


def _buscar_cadastro_controle_fiscal(codemp: int, codpro: str, codder: Optional[str] = None) -> Optional[Dict[str, Any]]:
    conn = get_connection()
    cursor = conn.cursor()

    sql = """
        SELECT
            P.CODEMP,
            P.CODPRO,
            P.DESPRO,
            P.TIPPRO,
            P.CODFAM,
            F.DESFAM,
            P.CODORI,
            O.DESORI,
            P.SITPRO,
            P.USUGER AS USUGER_PRO,
            P.DATGER AS DATGER_PRO,

            P.CODCLF,
            C.CLAFIS AS NCM,
            P.CODSTR,
            P.CODTIC,
            P.CODTRD,
            P.CODTST,
            P.CODSTP,
            P.CODSTC,
            P.PERIPI,
            P.RECIPI,
            P.TEMICM,
            P.RECICM,
            P.RECPIS,
            P.TRIPIS,
            P.TRICOF,
            P.RECCOF,
            P.PERIRF,
            P.PERPIS,
            P.PERCOF,
            P.PERCSL,
            P.PEROUR,
            P.BASCRE,
            P.BASREC,
            P.CSTIPI,
            P.CSTPIS,
            P.CSTCOF,
            P.TPRPIS,
            P.TPRCOF,
            P.TPRIPI,
            P.REGTRI,
            P.CSTIPC,
            P.CSTPIC,
            P.CSTCOC,
            P.ORIMER,
            P.NATPIS,
            P.NATCOF,
            P.CODANP,
            P.PROIMP,

            D.CODDER,
            D.DESDER,
            D.SITDER,
            D.USUGER AS USUGER_DER,
            D.DATGER AS DATGER_DER,
            D.ITEFIS,
            D.DESFIS,
            D.CODFIF,
            D.CODFIE,
            D.CODFIM,
            D.BSTUFC,
            D.ASTFCP,
            D.VSTUFC,
            D.CODCES
        FROM E075PRO P
        LEFT JOIN E075DER D
               ON D.CODEMP = P.CODEMP
              AND D.CODPRO = P.CODPRO
              AND (? IS NULL OR D.CODDER = ?)
        LEFT JOIN E012FAM F
               ON F.CODEMP = P.CODEMP
              AND F.CODFAM = P.CODFAM
        LEFT JOIN E083ORI O
               ON O.CODEMP = P.CODEMP
              AND O.CODORI = P.CODORI
        LEFT JOIN E022CLF C
               ON C.CODCLF = P.CODCLF
        WHERE P.CODEMP = ?
          AND P.CODPRO = ?
    """

    cursor.execute(sql, [codder, codder, codemp, codpro])
    row = cursor.fetchone()
    if not row:
        conn.close()
        return None

    data = row_to_dict(cursor, row)
    conn.close()

    if codder and not data.get("CODDER"):
        return None

    return data


def _pendencias_controle_fiscal_sugestao(cadastro: Dict[str, Any], campos_alvo: Optional[List[str]] = None) -> List[str]:
    pendencias = []

    def v(nome):
        return cadastro.get(nome)

    alvo = {str(x).strip().upper() for x in (campos_alvo or []) if str(x).strip()}

    def considerar(campo: str) -> bool:
        return not alvo or campo in alvo

    if considerar("CODCLF") and v("CODCLF") in (None, ""):
        pendencias.append("CODCLF")

    if considerar("CODTRD") and v("CODTRD") in (None, ""):
        pendencias.append("CODTRD")

    if considerar("RECPIS") and v("RECPIS") in (None, ""):
        pendencias.append("RECPIS")

    if considerar("RECCOF") and v("RECCOF") in (None, ""):
        pendencias.append("RECCOF")

    if considerar("CSTPIS") and v("CSTPIS") in (None, ""):
        pendencias.append("CSTPIS")

    if considerar("CSTCOF") and v("CSTCOF") in (None, ""):
        pendencias.append("CSTCOF")

    if considerar("BASCRE"):
        recpis = _clean_str(v("RECPIS"))
        cstpis = _clean_str(v("CSTPIS"))
        if v("BASCRE") in (None, "", 0, 0.0):
            if recpis == "S" or cstpis in {"50", "51", "52", "53", "54", "55", "56", "60", "61", "62", "63", "64", "65", "66"}:
                pendencias.append("BASCRE")

    if considerar("CSTPIS") and considerar("CSTCOF"):
        if _clean_str(v("CSTPIS")) and _clean_str(v("CSTCOF")) and _clean_str(v("CSTPIS")) != _clean_str(v("CSTCOF")):
            pendencias.append("CSTPIS/CSTCOF divergentes")

    if considerar("RECPIS") and considerar("RECCOF"):
        if _clean_str(v("RECPIS")) and _clean_str(v("RECCOF")) and _clean_str(v("RECPIS")) != _clean_str(v("RECCOF")):
            pendencias.append("RECPIS/RECCOF divergentes")

    return list(dict.fromkeys(pendencias))


def _buscar_regra_fiscal_controle(
    interpretacao: Optional[Dict[str, Any]],
    cadastro: Dict[str, Any],
    transacao_informada: Optional[str] = None
) -> Optional[Dict[str, Any]]:
    if not interpretacao:
        return None

    conn = get_connection()
    cursor = conn.cursor()

    ncm = _clean_str(cadastro.get("NCM"))
    familia = _clean_str(cadastro.get("CODFAM"))
    origem_prod = _clean_str(cadastro.get("CODORI"))
    transacao = _clean_str(transacao_informada)

    sql = """
        SELECT TOP 1
            ID,
            UF_ORIGEM,
            UF_DESTINO,
            TIPO_OPERACAO,
            NATUREZA,
            TIPO_CLIENTE,
            FINALIDADE,
            NCM,
            FAMILIA,
            ORIGEM_PRODUTO,
            TRANSACAO,
            CST_ICMS,
            CFOP,
            BENEFICIO_FISCAL,
            BASE_LEGAL,
            OBSERVACOES,
            PRIORIDADE
        FROM USU_TBTRIB_REGRA
        WHERE ATIVO = 1
          AND UF_ORIGEM = ?
          AND UF_DESTINO = ?
          AND TIPO_OPERACAO = ?
          AND NATUREZA = ?
          AND (TIPO_CLIENTE = ? OR TIPO_CLIENTE IS NULL OR TIPO_CLIENTE = '')
          AND (FINALIDADE = ? OR FINALIDADE IS NULL OR FINALIDADE = '')
          AND (NCM = ? OR NCM IS NULL OR NCM = '')
          AND (FAMILIA = ? OR FAMILIA IS NULL OR FAMILIA = '')
          AND (ORIGEM_PRODUTO = ? OR ORIGEM_PRODUTO IS NULL OR ORIGEM_PRODUTO = '')
          AND (TRANSACAO = ? OR TRANSACAO IS NULL OR TRANSACAO = '')
        ORDER BY PRIORIDADE ASC, ID ASC
    """

    params = [
        _upper(interpretacao.get("uf_origem")),
        _upper(interpretacao.get("uf_destino")),
        _upper(interpretacao.get("tipo_operacao")),
        _upper(interpretacao.get("natureza")),
        _upper(interpretacao.get("tipo_cliente")),
        _upper(interpretacao.get("finalidade")),
        ncm,
        familia,
        origem_prod,
        transacao,
    ]

    try:
        cursor.execute(sql, params)
        row = cursor.fetchone()
    except Exception:
        conn.close()
        return None

    conn.close()

    if not row:
        return None

    return {
        "id_regra": row[0],
        "uf_origem": row[1],
        "uf_destino": row[2],
        "tipo_operacao": row[3],
        "natureza": row[4],
        "tipo_cliente": row[5],
        "finalidade": row[6],
        "ncm": row[7],
        "familia": row[8],
        "origem_produto": row[9],
        "transacao": row[10],
        "cst_icms": row[11],
        "cfop": row[12],
        "beneficio_fiscal": row[13],
        "base_legal": row[14],
        "observacoes": row[15],
        "prioridade": row[16],
        "fonte_regra": "USU_TBTRIB_REGRA"
    }


@app.post("/api/controle-fiscal-produtos/sugerir-ia")
def sugerir_ia_controle_fiscal_produtos(
    payload: ControleFiscalSugerirIARequest,
    usuario=Depends(validar_token)
):
    try:
        codemp = payload.codemp
        codpro = _clean_str(payload.codpro).upper()
        codder = _clean_str(payload.codder).upper() or None

        if not codpro:
            raise HTTPException(status_code=400, detail="codpro é obrigatório")

        cadastro = _buscar_cadastro_controle_fiscal(codemp, codpro, codder)
        if not cadastro:
            raise HTTPException(status_code=404, detail="Produto/derivação não encontrados")

        pendencias = _pendencias_controle_fiscal_sugestao(cadastro, payload.campos_alvo)

        interpretacao = None
        regra_interna = None

        if _clean_str(payload.descricao_operacao) and _clean_str(payload.uf_origem) and _clean_str(payload.uf_destino):
            interpretacao = interpretar_operacao_fiscal(
                descricao_operacao=payload.descricao_operacao,
                uf_origem=payload.uf_origem,
                uf_destino=payload.uf_destino,
                tipo_cliente=payload.tipo_cliente,
                finalidade=payload.finalidade
            )
            regra_interna = _buscar_regra_fiscal_controle(
                interpretacao=interpretacao,
                cadastro=cadastro,
                transacao_informada=payload.transacao
            )

        campos_prioritarios = payload.campos_alvo or pendencias or ["CODTRD", "CSTPIS", "CSTCOF", "BASCRE"]

        prompt = f"""
Você é um especialista fiscal brasileiro com foco em parametrização ERP Senior.

Sua tarefa é sugerir preenchimento de campos fiscais do cadastro do produto/derivação.
A sugestão deve ser ASSISTIVA: nunca assumir certeza jurídica absoluta. Sempre exigir revisão humana final.

Retorne SOMENTE JSON válido no formato:
{{
  "resumo": "texto curto",
  "campos_sugeridos": [
    {{
      "campo_erp": "CODTRD",
      "valor_atual": "valor atual ou null",
      "valor_sugerido": "valor sugerido",
      "justificativa": "explicação objetiva",
      "base_legal_ou_regra": "base legal ou regra interna ou null",
      "confianca": "ALTA|MEDIA|BAIXA",
      "obrigatorio_para_ativacao": true
    }}
  ],
  "alertas": ["lista de alertas"],
  "pode_salvar_automaticamente": false
}}

Regras de comportamento:
- Priorize os campos pendentes.
- Se não houver base suficiente para sugerir valor exato, devolva valor_sugerido = null e explique o que falta.
- Se existir regra interna, ela tem prioridade sobre heurística geral.
- CSTPIS e CSTCOF precisam ficar coerentes.
- Se RECPIS = 'S' e houver crédito, BASCRE deve ser coerente com isso.
- Nunca responda com markdown.

Produto/cadastro atual:
{json.dumps(cadastro, ensure_ascii=False, default=str, indent=2)}

Pendências detectadas:
{json.dumps(pendencias, ensure_ascii=False)}

Campos prioritários:
{json.dumps(campos_prioritarios, ensure_ascii=False)}

Contexto operacional informado:
{json.dumps({
    "descricao_operacao": payload.descricao_operacao,
    "uf_origem": payload.uf_origem,
    "uf_destino": payload.uf_destino,
    "tipo_cliente": payload.tipo_cliente,
    "finalidade": payload.finalidade,
    "transacao": payload.transacao
}, ensure_ascii=False, indent=2)}

Interpretação operacional:
{json.dumps(interpretacao, ensure_ascii=False, indent=2) if interpretacao else "null"}

Regra interna encontrada:
{json.dumps(regra_interna, ensure_ascii=False, indent=2) if regra_interna else "null"}
        """.strip()

        resultado_ia = chamar_gemini(
            prompt,
            instrucao_sistema=(
                "Você é um especialista fiscal brasileiro em ERP Senior. "
                "Sugira preenchimento de campos fiscais cadastrais com prudência, "
                "priorizando regras internas da empresa quando existirem. "
                "Responda sempre em JSON válido."
            )
        )

        if not resultado_ia:
            return {
                "ok": True,
                "codemp": codemp,
                "codpro": codpro,
                "codder": codder,
                "cadastro_atual": cadastro,
                "pendencias": pendencias,
                "interpretacao": interpretacao,
                "regra_interna": regra_interna,
                "resultado_ia": None,
                "mensagem": "IA indisponível no momento. Verifique a configuração do Gemini."
            }

        return {
            "ok": True,
            "codemp": codemp,
            "codpro": codpro,
            "codder": codder,
            "cadastro_atual": cadastro,
            "pendencias": pendencias,
            "interpretacao": interpretacao,
            "regra_interna": regra_interna,
            "resultado_ia": resultado_ia,
            "usuario_logado": usuario
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao sugerir preenchimento com IA: {str(e)}")



# =========================================================
# AUDITORIA TRIBUTÁRIA
# =========================================================

# Classificação derivada de categoria do material (ajuste os códigos conforme seu ERP)
CATEGORIA_SQL = """
    CASE
        WHEN UPPER(ISNULL(P.CODFAM,'')) IN ('CONSUM','CONSUMO','CONS') THEN 'CONSUMO'
        WHEN UPPER(ISNULL(P.CODFAM,'')) IN ('EPI','EPIS','E.P.I') THEN 'EPI'
        WHEN UPPER(ISNULL(P.CODFAM,'')) IN ('IMOB','IMO','ATIVO','IMOBIL','AT.IMOB') THEN 'IMOBILIZADO'
        WHEN UPPER(ISNULL(P.CODORI,'')) IN ('MPM','MPG','MP','MATPRIMA','MATERIA_PRIMA') THEN 'MATERIA_PRIMA'
        WHEN UPPER(ISNULL(P.CODORI,'')) IN ('REV','REVENDA','REVENDAS') THEN 'REVENDA'
        WHEN UPPER(ISNULL(P.TIPPRO,'')) = 'P' THEN 'PRODUTO_PRODUZIDO'
        ELSE 'OUTROS'
    END
"""

# Bloco completo de campos da familia para queries de produto (E012FAM com alias FAM)
FAMILIA_SQL_PROD = """
                FAM.CODEMP AS fam_codemp,
                FAM.CODFAM AS fam_codfam,
                FAM.TIPPRO AS fam_tippro,
                FAM.CODORI AS fam_codori,
                FAM.DEPPAD AS fam_deppad,
                FAM.CTRVLD AS fam_ctrvld,
                FAM.CTRLOT AS fam_ctrlot,
                FAM.CTRSEP AS fam_ctrsep,
                FAM.POSPRO AS fam_pospro,
                FAM.CODMDP AS fam_codmdp,
                FAM.UNIMED AS fam_unimed,
                FAM.UNIME2 AS fam_unime2,
                FAM.UNIME3 AS fam_unime3,
                FAM.UTIDEC AS fam_utidec,
                FAM.QTDDEC AS fam_qtddec,
                FAM.CODET1 AS fam_codet1,
                FAM.CODET2 AS fam_codet2,
                FAM.CODET3 AS fam_codet3,
                FAM.CODET4 AS fam_codet4,
                FAM.CODET5 AS fam_codet5,
                FAM.CODET6 AS fam_codet6,
                FAM.CODET7 AS fam_codet7,
                FAM.CODET8 AS fam_codet8,
                FAM.CODET9 AS fam_codet9,
                FAM.TEMCTE AS fam_temcte,
                FAM.NUMORI AS fam_numori,
                FAM.QTDMLT AS fam_qtdmlt,
                FAM.QTDMIN AS fam_qtdmin,
                FAM.QTDMAX AS fam_qtdmax,
                FAM.QTDGOP AS fam_qtdgop,
                FAM.BXAORP AS fam_bxaorp,
                FAM.CODAGE AS fam_codage,
                FAM.CODAGP AS fam_codagp,
                FAM.CODAGU AS fam_codagu,
                FAM.CODAGC AS fam_codagc,
                FAM.CODAGT AS fam_codagt,
                FAM.CODAGF AS fam_codagf,
                FAM.CODCLF AS fam_codclf,
                FAM.CODSTR AS fam_codstr,
                FAM.RECIPI AS fam_recipi,
                FAM.RECCOF AS fam_reccof,
                FAM.TEMICM AS fam_temicm,
                FAM.CODTIC AS fam_codtic,
                FAM.CODTRD AS fam_codtrd,
                FAM.CODTST AS fam_codtst,
                FAM.CODSTP AS fam_codstp,
                FAM.CODSTC AS fam_codstc,
                FAM.RECICM AS fam_recicm,
                FAM.GEREAN AS fam_gerean,
                FAM.CODMP1 AS fam_codmp1,
                FAM.CODMP2 AS fam_codmp2,
                FAM.CODMP3 AS fam_codmp3,
                FAM.CODMP4 AS fam_codmp4,
                FAM.CODMP5 AS fam_codmp5,
                FAM.CODMP6 AS fam_codmp6,
                FAM.CODMP7 AS fam_codmp7,
                FAM.ROTPRO AS fam_rotpro,
                FAM.MATDIR AS fam_matdir,
                FAM.CODREG AS fam_codreg,
                FAM.CODNTG AS fam_codntg,
                FAM.CRIRAT AS fam_crirat,
                FAM.CTARED AS fam_ctared,
                FAM.CTARCR AS fam_ctarcr,
                FAM.CTAFDV AS fam_ctafdv,
                FAM.CTAFCR AS fam_ctafcr,
                FAM.CTADCD AS fam_ctadcd,
                FAM.CTADCI AS fam_ctadci,
                FAM.INDKIT AS fam_indkit,
                FAM.CODPIN AS fam_codpin,
                FAM.NOTFOR AS fam_notfor,
                FAM.INDMIS AS fam_indmis,
                FAM.EMIGTR AS fam_emigtr,
                FAM.SOMIIM AS fam_somiim,
                FAM.RECPIS AS fam_recpis,
                FAM.INDEXP AS fam_indexp,
                FAM.DATPAL AS fam_datpal,
                FAM.HORPAL AS fam_horpal,
                FAM.TIPINT AS fam_tipint,
                FAM.SOMIIL AS fam_somiil,
                FAM.CODMAR AS fam_codmar,
                FAM.CODCLC AS fam_codclc,
                FAM.NIVCBN AS fam_nivcbn,
                FAM.SITCAL AS fam_sitcal,
                FAM.GERORP AS fam_gerorp,
                FAM.PERIRF AS fam_perirf,
                FAM.PERPIS AS fam_perpis,
                FAM.PERCOF AS fam_percof,
                FAM.PERCSL AS fam_percsl,
                FAM.PEROUR AS fam_perour,
                FAM.SOMIPS AS fam_somips,
                FAM.SOMICO AS fam_somico,
                FAM.SOMIPL AS fam_somipl,
                FAM.SOMICL AS fam_somicl,
                FAM.INDOCT AS fam_indoct,
                FAM.INDSPR AS fam_indspr,
                FAM.PRECUS AS fam_precus,
                FAM.SITFAM AS fam_sitfam,
                FAM.PROIMP AS fam_proimp,
                FAM.USUGER AS fam_usuger,
                FAM.DATGER AS fam_datger,
                FAM.HORGER AS fam_horger,
                FAM.USUALT AS fam_usualt,
                FAM.DATALT AS fam_datalt,
                FAM.HORALT AS fam_horalt,
                FAM.INTAGR AS fam_intagr,
                FAM.CTRVIS AS fam_ctrvis,
                FAM.DATVIS AS fam_datvis,
                FAM.HORVIS AS fam_horvis,
                FAM.DIAREP AS fam_diarep,
                FAM.INDFRT AS fam_indfrt,
                FAM.FRTEQP AS fam_frteqp,
                FAM.GRPFRT AS fam_grpfrt,
                FAM.CSTIPI AS fam_cst_ipi,
                FAM.CSTPIS AS fam_cst_pis,
                FAM.CSTCOF AS fam_cst_cofins,
                FAM.CSTIPC AS fam_cstipc,
                FAM.CSTPIC AS fam_cstpic,
                FAM.CSTCOC AS fam_cstcoc,
                FAM.VARPRO AS fam_varpro,
                FAM.PROMON AS fam_promon,
                FAM.FINCRP AS fam_fincrp,
                FAM.FINCDP AS fam_fincdp,
                FAM.PERPIM AS fam_perpim,
                FAM.PERCIM AS fam_percim,
                FAM.APLATX AS fam_aplatx,
                FAM.CODATX AS fam_codatx,
                FAM.INDICP AS fam_indicp,
                FAM.MGCMIN AS fam_mgcmin,
                FAM.MGCLIM AS fam_mgclim,
                FAM.PERVEN AS fam_perven,
                FAM.PREREF AS fam_preref,
                FAM.CODBIC AS fam_codbic,
                FAM.INDVOL AS fam_indvol,
                FAM.CODMPH AS fam_codmph,
                FAM.MODFAB AS fam_modfab,
                FAM.CODPRI AS fam_codpri,
                FAM.CODPRC AS fam_codprc,
                FAM.TIPFTE AS fam_tipfte,
                FAM.INDENC AS fam_indenc,
                FAM.PERIFP AS fam_perifp,
                FAM.SEQHAS AS fam_seqhas,
                FAM.INDACO AS fam_indaco,
                FAM.INDM21 AS fam_indm21,
                FAM.REGTRI AS fam_regtri,
                FAM.ORIGTI AS fam_origti,
                FAM.INTWMW AS fam_intwmw,
                FAM.TEMRCI AS fam_temrci,
                FAM.SERREL AS fam_serrel,
                FAM.DPRFAM AS fam_dprfam,
                FAM.GEAD14 AS fam_gead14,
                FAM.CULIND AS fam_culind,
                FAM.DEVPRO AS fam_devpro,
                FAM.INTPOS AS fam_intpos,
                FAM.INSSAF AS fam_inssaf,
                FAM.SITWMW AS fam_sitwmw,
                FAM.USU_MCGRCP AS fam_usu_mcgrcp"""

# Bloco equivalente com NULLs para queries de servico (sem JOIN E012FAM)
FAMILIA_SQL_NULL = """
                CAST(NULL AS INT) AS fam_codemp,
                CAST(NULL AS VARCHAR(20)) AS fam_codfam,
                CAST(NULL AS VARCHAR(20)) AS fam_tippro,
                CAST(NULL AS VARCHAR(20)) AS fam_codori,
                CAST(NULL AS VARCHAR(1)) AS fam_deppad,
                CAST(NULL AS VARCHAR(1)) AS fam_ctrvld,
                CAST(NULL AS VARCHAR(1)) AS fam_ctrlot,
                CAST(NULL AS VARCHAR(1)) AS fam_ctrsep,
                CAST(NULL AS VARCHAR(20)) AS fam_pospro,
                CAST(NULL AS VARCHAR(20)) AS fam_codmdp,
                CAST(NULL AS VARCHAR(10)) AS fam_unimed,
                CAST(NULL AS VARCHAR(10)) AS fam_unime2,
                CAST(NULL AS VARCHAR(10)) AS fam_unime3,
                CAST(NULL AS VARCHAR(1)) AS fam_utidec,
                CAST(NULL AS SMALLINT) AS fam_qtddec,
                CAST(NULL AS VARCHAR(20)) AS fam_codet1,
                CAST(NULL AS VARCHAR(20)) AS fam_codet2,
                CAST(NULL AS VARCHAR(20)) AS fam_codet3,
                CAST(NULL AS VARCHAR(20)) AS fam_codet4,
                CAST(NULL AS VARCHAR(20)) AS fam_codet5,
                CAST(NULL AS VARCHAR(20)) AS fam_codet6,
                CAST(NULL AS VARCHAR(20)) AS fam_codet7,
                CAST(NULL AS VARCHAR(20)) AS fam_codet8,
                CAST(NULL AS VARCHAR(20)) AS fam_codet9,
                CAST(NULL AS VARCHAR(1)) AS fam_temcte,
                CAST(NULL AS SMALLINT) AS fam_numori,
                CAST(NULL AS NUMERIC(15,4)) AS fam_qtdmlt,
                CAST(NULL AS NUMERIC(15,4)) AS fam_qtdmin,
                CAST(NULL AS NUMERIC(15,4)) AS fam_qtdmax,
                CAST(NULL AS NUMERIC(15,4)) AS fam_qtdgop,
                CAST(NULL AS VARCHAR(1)) AS fam_bxaorp,
                CAST(NULL AS VARCHAR(20)) AS fam_codage,
                CAST(NULL AS VARCHAR(20)) AS fam_codagp,
                CAST(NULL AS VARCHAR(20)) AS fam_codagu,
                CAST(NULL AS VARCHAR(20)) AS fam_codagc,
                CAST(NULL AS VARCHAR(20)) AS fam_codagt,
                CAST(NULL AS VARCHAR(20)) AS fam_codagf,
                CAST(NULL AS VARCHAR(20)) AS fam_codclf,
                CAST(NULL AS VARCHAR(20)) AS fam_codstr,
                CAST(NULL AS VARCHAR(1)) AS fam_recipi,
                CAST(NULL AS VARCHAR(1)) AS fam_reccof,
                CAST(NULL AS VARCHAR(1)) AS fam_temicm,
                CAST(NULL AS VARCHAR(20)) AS fam_codtic,
                CAST(NULL AS VARCHAR(20)) AS fam_codtrd,
                CAST(NULL AS VARCHAR(20)) AS fam_codtst,
                CAST(NULL AS VARCHAR(20)) AS fam_codstp,
                CAST(NULL AS VARCHAR(20)) AS fam_codstc,
                CAST(NULL AS VARCHAR(1)) AS fam_recicm,
                CAST(NULL AS VARCHAR(1)) AS fam_gerean,
                CAST(NULL AS VARCHAR(20)) AS fam_codmp1,
                CAST(NULL AS VARCHAR(20)) AS fam_codmp2,
                CAST(NULL AS VARCHAR(20)) AS fam_codmp3,
                CAST(NULL AS VARCHAR(20)) AS fam_codmp4,
                CAST(NULL AS VARCHAR(20)) AS fam_codmp5,
                CAST(NULL AS VARCHAR(20)) AS fam_codmp6,
                CAST(NULL AS VARCHAR(20)) AS fam_codmp7,
                CAST(NULL AS VARCHAR(20)) AS fam_rotpro,
                CAST(NULL AS VARCHAR(1)) AS fam_matdir,
                CAST(NULL AS SMALLINT) AS fam_codreg,
                CAST(NULL AS VARCHAR(20)) AS fam_codntg,
                CAST(NULL AS VARCHAR(1)) AS fam_crirat,
                CAST(NULL AS VARCHAR(30)) AS fam_ctared,
                CAST(NULL AS VARCHAR(30)) AS fam_ctarcr,
                CAST(NULL AS VARCHAR(30)) AS fam_ctafdv,
                CAST(NULL AS VARCHAR(30)) AS fam_ctafcr,
                CAST(NULL AS VARCHAR(30)) AS fam_ctadcd,
                CAST(NULL AS VARCHAR(30)) AS fam_ctadci,
                CAST(NULL AS VARCHAR(1)) AS fam_indkit,
                CAST(NULL AS VARCHAR(20)) AS fam_codpin,
                CAST(NULL AS VARCHAR(1)) AS fam_notfor,
                CAST(NULL AS VARCHAR(1)) AS fam_indmis,
                CAST(NULL AS VARCHAR(1)) AS fam_emigtr,
                CAST(NULL AS VARCHAR(1)) AS fam_somiim,
                CAST(NULL AS VARCHAR(1)) AS fam_recpis,
                CAST(NULL AS VARCHAR(1)) AS fam_indexp,
                CAST(NULL AS DATETIME) AS fam_datpal,
                CAST(NULL AS VARCHAR(8)) AS fam_horpal,
                CAST(NULL AS VARCHAR(1)) AS fam_tipint,
                CAST(NULL AS VARCHAR(1)) AS fam_somiil,
                CAST(NULL AS VARCHAR(20)) AS fam_codmar,
                CAST(NULL AS VARCHAR(20)) AS fam_codclc,
                CAST(NULL AS SMALLINT) AS fam_nivcbn,
                CAST(NULL AS VARCHAR(1)) AS fam_sitcal,
                CAST(NULL AS VARCHAR(1)) AS fam_gerorp,
                CAST(NULL AS NUMERIC(15,4)) AS fam_perirf,
                CAST(NULL AS NUMERIC(15,4)) AS fam_perpis,
                CAST(NULL AS NUMERIC(15,4)) AS fam_percof,
                CAST(NULL AS NUMERIC(15,4)) AS fam_percsl,
                CAST(NULL AS NUMERIC(15,4)) AS fam_perour,
                CAST(NULL AS VARCHAR(1)) AS fam_somips,
                CAST(NULL AS VARCHAR(1)) AS fam_somico,
                CAST(NULL AS VARCHAR(1)) AS fam_somipl,
                CAST(NULL AS VARCHAR(1)) AS fam_somicl,
                CAST(NULL AS VARCHAR(1)) AS fam_indoct,
                CAST(NULL AS VARCHAR(1)) AS fam_indspr,
                CAST(NULL AS VARCHAR(1)) AS fam_precus,
                CAST(NULL AS VARCHAR(1)) AS fam_sitfam,
                CAST(NULL AS SMALLINT) AS fam_proimp,
                CAST(NULL AS VARCHAR(20)) AS fam_usuger,
                CAST(NULL AS DATETIME) AS fam_datger,
                CAST(NULL AS VARCHAR(8)) AS fam_horger,
                CAST(NULL AS VARCHAR(20)) AS fam_usualt,
                CAST(NULL AS DATETIME) AS fam_datalt,
                CAST(NULL AS VARCHAR(8)) AS fam_horalt,
                CAST(NULL AS VARCHAR(1)) AS fam_intagr,
                CAST(NULL AS VARCHAR(1)) AS fam_ctrvis,
                CAST(NULL AS DATETIME) AS fam_datvis,
                CAST(NULL AS VARCHAR(8)) AS fam_horvis,
                CAST(NULL AS SMALLINT) AS fam_diarep,
                CAST(NULL AS VARCHAR(1)) AS fam_indfrt,
                CAST(NULL AS VARCHAR(20)) AS fam_frteqp,
                CAST(NULL AS VARCHAR(20)) AS fam_grpfrt,
                CAST(NULL AS VARCHAR(10)) AS fam_cst_ipi,
                CAST(NULL AS VARCHAR(10)) AS fam_cst_pis,
                CAST(NULL AS VARCHAR(10)) AS fam_cst_cofins,
                CAST(NULL AS VARCHAR(10)) AS fam_cstipc,
                CAST(NULL AS VARCHAR(10)) AS fam_cstpic,
                CAST(NULL AS VARCHAR(10)) AS fam_cstcoc,
                CAST(NULL AS VARCHAR(20)) AS fam_varpro,
                CAST(NULL AS VARCHAR(1)) AS fam_promon,
                CAST(NULL AS VARCHAR(1)) AS fam_fincrp,
                CAST(NULL AS VARCHAR(1)) AS fam_fincdp,
                CAST(NULL AS NUMERIC(15,4)) AS fam_perpim,
                CAST(NULL AS NUMERIC(15,4)) AS fam_percim,
                CAST(NULL AS VARCHAR(1)) AS fam_aplatx,
                CAST(NULL AS VARCHAR(20)) AS fam_codatx,
                CAST(NULL AS VARCHAR(1)) AS fam_indicp,
                CAST(NULL AS NUMERIC(15,4)) AS fam_mgcmin,
                CAST(NULL AS NUMERIC(15,4)) AS fam_mgclim,
                CAST(NULL AS NUMERIC(15,4)) AS fam_perven,
                CAST(NULL AS VARCHAR(20)) AS fam_preref,
                CAST(NULL AS VARCHAR(20)) AS fam_codbic,
                CAST(NULL AS VARCHAR(1)) AS fam_indvol,
                CAST(NULL AS VARCHAR(20)) AS fam_codmph,
                CAST(NULL AS VARCHAR(1)) AS fam_modfab,
                CAST(NULL AS VARCHAR(20)) AS fam_codpri,
                CAST(NULL AS VARCHAR(20)) AS fam_codprc,
                CAST(NULL AS VARCHAR(1)) AS fam_tipfte,
                CAST(NULL AS VARCHAR(1)) AS fam_indenc,
                CAST(NULL AS NUMERIC(15,4)) AS fam_perifp,
                CAST(NULL AS INT) AS fam_seqhas,
                CAST(NULL AS VARCHAR(1)) AS fam_indaco,
                CAST(NULL AS VARCHAR(1)) AS fam_indm21,
                CAST(NULL AS VARCHAR(20)) AS fam_regtri,
                CAST(NULL AS VARCHAR(20)) AS fam_origti,
                CAST(NULL AS VARCHAR(1)) AS fam_intwmw,
                CAST(NULL AS VARCHAR(1)) AS fam_temrci,
                CAST(NULL AS VARCHAR(1)) AS fam_serrel,
                CAST(NULL AS VARCHAR(20)) AS fam_dprfam,
                CAST(NULL AS VARCHAR(1)) AS fam_gead14,
                CAST(NULL AS VARCHAR(1)) AS fam_culind,
                CAST(NULL AS VARCHAR(1)) AS fam_devpro,
                CAST(NULL AS VARCHAR(1)) AS fam_intpos,
                CAST(NULL AS VARCHAR(1)) AS fam_inssaf,
                CAST(NULL AS VARCHAR(1)) AS fam_sitwmw,
                CAST(NULL AS VARCHAR(20)) AS fam_usu_mcgrcp"""

@app.get("/api/auditoria-tributaria")
def auditoria_tributaria(
    tipo_item: str = Query("TODOS"),
    movimento: str = Query("TODOS"),
    numero_documento: Optional[int] = Query(None),
    serie: Optional[str] = Query(None),
    parceiro: Optional[int] = Query(None),
    codigo_item: Optional[str] = Query(None),
    codigo_produto: Optional[str] = Query(None),
    descricao: Optional[str] = Query(None),
    familia: Optional[str] = Query(None),
    origem: Optional[str] = Query(None),
    transacao: Optional[str] = Query(None),
    data_emissao_ini: Optional[str] = Query(None),
    data_emissao_fim: Optional[str] = Query(None),
    categoria_material: Optional[str] = Query(None),
    base_auditoria: str = Query("MOVIMENTOS"),
    apenas_divergencia: bool = Query(False),
    pagina: int = Query(1),
    tamanho_pagina: int = Query(100),
    usuario=Depends(validar_token)
):
    try:
        return _auditoria_tributaria_inner(
            tipo_item, movimento, numero_documento, serie, parceiro,
            codigo_item, codigo_produto, descricao, familia, origem,
            transacao, data_emissao_ini, data_emissao_fim,
            categoria_material, base_auditoria, apenas_divergencia, pagina, tamanho_pagina
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro na auditoria tributária: {str(e)}")


# =========================================================
# EXPORTACAO XLSX DA AUDITORIA TRIBUTARIA
# =========================================================
def _status_auditoria_export(item: Dict[str, Any]) -> str:
    """Status visual usado na exportacao - segue mesma regra acordada com o front."""
    div = item.get("divergencias_reais") or []
    avi = item.get("avisos_cadastrais") or []
    pen = item.get("pendencias_mapeamento") or []
    status_api = (item.get("status_auditoria") or "").upper()
    if status_api == "DIVERGENTE" or len(div) > 0:
        return "DIVERGENTE"
    if len(avi) > 0:
        return "OK_COM_AVISO"
    if len(pen) > 0:
        return "PENDENTE_MAPEAMENTO"
    return "OK"


def _coletar_itens_auditoria_para_export(filtros: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Itera o endpoint interno paginando ate trazer todos os itens compativeis com os filtros."""
    itens_total: List[Dict[str, Any]] = []
    pagina = 1
    tamanho_pagina = 200
    while True:
        resultado = _auditoria_tributaria_inner(
            filtros["tipo_item"], filtros["movimento"], filtros["numero_documento"],
            filtros["serie"], filtros["parceiro"], filtros["codigo_item"],
            filtros["codigo_produto"], filtros["descricao"], filtros["familia"],
            filtros["origem"], filtros["transacao"], filtros["data_emissao_ini"],
            filtros["data_emissao_fim"], filtros["categoria_material"],
            filtros["base_auditoria"], filtros["apenas_divergencia"],
            pagina, tamanho_pagina
        )
        itens_pagina = resultado.get("itens") or []
        itens_total.extend(itens_pagina)
        total_paginas = int(resultado.get("total_paginas") or 1)
        if pagina >= total_paginas or not itens_pagina:
            break
        pagina += 1
        # Salvaguarda - evita loop em payload corrompido
        if pagina > 500:
            break
    return itens_total


def _linha_export_auditoria(item: Dict[str, Any]) -> Dict[str, Any]:
    """Monta a linha de Excel com as colunas combinadas com o front."""
    div = item.get("divergencias_reais") or []
    avi = item.get("avisos_cadastrais") or []
    pen = item.get("pendencias_mapeamento") or []
    pis = (item.get("impostos") or {}).get("pis_cofins") or {}
    risco = item.get("risco") or {}
    fonte_efetiva = item.get("fonte_efetiva") or {}
    sep = " | "
    return {
        "Status Auditoria": _status_auditoria_export(item),
        "Movimento": item.get("movimento") or "",
        "Tipo Documento": item.get("documento_tipo") or "",
        "Nº Documento": item.get("numero_documento") or "",
        "Série": item.get("serie") or "",
        "Seq Item": item.get("seq_item") or "",
        "Data Emissão": item.get("data_emissao") or "",
        "Tipo Item": item.get("tipo_item") or "",
        "Código Item": item.get("codigo_item") or "",
        "Derivação": item.get("derivacao") or "",
        "Descrição Item": item.get("descricao_item") or "",
        "Fornecedor": item.get("fornecedor_nome") or "",
        "Cliente": item.get("cliente_nome") or "",
        "Transação": item.get("transacao") or "",
        "CFOP": item.get("cfop") or "",
        "Família": item.get("familia_codigo") or "",
        "Origem": item.get("origem_codigo") or "",
        "NCM": item.get("ncm") or "",
        "CEST": item.get("cest") or "",
        "Qtd Divergências Reais": len(div),
        "Divergências Reais": sep.join(div),
        "Qtd Avisos Cadastrais": len(avi),
        "Avisos Cadastrais": sep.join(avi),
        "Qtd Pendências Mapeamento": len(pen),
        "Pendências Mapeamento": sep.join(pen),
        "CST PIS Item NF": pis.get("item_cst_pis") or "",
        "CST PIS Transação": pis.get("tns_cst_pis") or "",
        "CST PIS Cadastro Produto": pis.get("cad_cstpis_produto") or "",
        "CST PIS Família": pis.get("fam_cst_pis") or "",
        "CST COFINS Item NF": pis.get("item_cst_cofins") or "",
        "CST COFINS Transação": pis.get("tns_cst_cofins") or "",
        "CST COFINS Cadastro Produto": pis.get("cad_cstcof_produto") or "",
        "CST COFINS Família": pis.get("fam_cst_cofins") or "",
        "Fonte Efetiva PIS": fonte_efetiva.get("pis") or "",
        "Fonte Efetiva COFINS": fonte_efetiva.get("cofins") or "",
        "Fonte Efetiva IPI": fonte_efetiva.get("ipi") or "",
        "Fonte Efetiva ICMS": fonte_efetiva.get("icms") or "",
        "Score Risco": risco.get("score_risco") if risco else "",
        "Nível Risco": risco.get("nivel_risco") or "" if risco else "",
        "Fatores Risco": sep.join(risco.get("fatores_risco") or []) if risco else "",
    }


@app.get("/api/auditoria-tributaria/export")
def auditoria_tributaria_export(
    tipo_item: str = Query("TODOS"),
    movimento: str = Query("TODOS"),
    numero_documento: Optional[int] = Query(None),
    serie: Optional[str] = Query(None),
    parceiro: Optional[int] = Query(None),
    codigo_item: Optional[str] = Query(None),
    codigo_produto: Optional[str] = Query(None),
    descricao: Optional[str] = Query(None),
    familia: Optional[str] = Query(None),
    origem: Optional[str] = Query(None),
    transacao: Optional[str] = Query(None),
    data_emissao_ini: Optional[str] = Query(None),
    data_emissao_fim: Optional[str] = Query(None),
    categoria_material: Optional[str] = Query(None),
    base_auditoria: str = Query("MOVIMENTOS"),
    apenas_divergencia: bool = Query(False),
    formato: str = Query("xlsx", description="xlsx | csv"),
    usuario=Depends(validar_token)
):
    """Exporta o resultado da auditoria tributaria em XLSX (default) ou CSV.

    Regras de classificacao seguem o backend:
      - status DIVERGENTE so quando ha divergencias_reais
      - avisos cadastrais nao derrubam a nota
      - filtro apenas_divergencia opera sobre divergencias_reais
    """
    try:
        from fastapi.responses import StreamingResponse
        from io import BytesIO, StringIO
        import csv as _csv
        from datetime import datetime as _dt

        filtros = {
            "tipo_item": tipo_item, "movimento": movimento,
            "numero_documento": numero_documento, "serie": serie, "parceiro": parceiro,
            "codigo_item": codigo_item, "codigo_produto": codigo_produto,
            "descricao": descricao, "familia": familia, "origem": origem,
            "transacao": transacao, "data_emissao_ini": data_emissao_ini,
            "data_emissao_fim": data_emissao_fim, "categoria_material": categoria_material,
            "base_auditoria": base_auditoria, "apenas_divergencia": apenas_divergencia,
        }

        itens = _coletar_itens_auditoria_para_export(filtros)

        # Garantia extra de filtro: mesmo que o backend ja filtre, aplicamos de novo aqui
        if apenas_divergencia:
            itens = [it for it in itens if len(it.get("divergencias_reais") or []) > 0]

        linhas = [_linha_export_auditoria(it) for it in itens]
        colunas: List[str] = list(linhas[0].keys()) if linhas else [
            "Status Auditoria", "Movimento", "Nº Documento", "Série", "Código Item",
            "Descrição Item", "Divergências Reais", "Avisos Cadastrais", "Pendências Mapeamento"
        ]

        timestamp = _dt.now().strftime("%Y%m%d_%H%M%S")
        formato_final = (formato or "xlsx").lower().strip()

        if formato_final == "csv":
            buf_txt = StringIO()
            writer = _csv.DictWriter(buf_txt, fieldnames=colunas, delimiter=";")
            writer.writeheader()
            for ln in linhas:
                writer.writerow(ln)
            data = buf_txt.getvalue().encode("utf-8-sig")
            filename = f"auditoria_tributaria_{timestamp}.csv"
            return StreamingResponse(
                BytesIO(data),
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'}
            )

        # XLSX (padrao) - exige openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="openpyxl nao instalado no servidor. Instale com 'pip install openpyxl' ou solicite o export em formato=csv"
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Auditoria"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F2937")
        center = Alignment(horizontal="left", vertical="center", wrap_text=False)

        ws.append(colunas)
        for col_idx, _ in enumerate(colunas, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # Cores por status
        fills_status = {
            "DIVERGENTE":         PatternFill("solid", fgColor="FECACA"),
            "OK_COM_AVISO":       PatternFill("solid", fgColor="FDE68A"),
            "PENDENTE_MAPEAMENTO":PatternFill("solid", fgColor="DDD6FE"),
            "OK":                 PatternFill("solid", fgColor="DCFCE7"),
        }

        for ln in linhas:
            valores = [ln.get(col, "") for col in colunas]
            ws.append(valores)
            row_idx = ws.max_row
            status = ln.get("Status Auditoria")
            fill = fills_status.get(status)
            if fill:
                ws.cell(row=row_idx, column=1).fill = fill

        # Larguras razoaveis
        larguras = {
            "Status Auditoria": 20, "Movimento": 12, "Tipo Documento": 14,
            "Nº Documento": 14, "Série": 8, "Seq Item": 8, "Data Emissão": 14,
            "Tipo Item": 12, "Código Item": 18, "Derivação": 12,
            "Descrição Item": 50, "Fornecedor": 35, "Cliente": 35,
            "Transação": 12, "CFOP": 10, "Família": 12, "Origem": 12,
            "NCM": 14, "CEST": 12,
            "Qtd Divergências Reais": 14, "Divergências Reais": 60,
            "Qtd Avisos Cadastrais": 14, "Avisos Cadastrais": 60,
            "Qtd Pendências Mapeamento": 14, "Pendências Mapeamento": 60,
            "CST PIS Item NF": 12, "CST PIS Transação": 14,
            "CST PIS Cadastro Produto": 16, "CST PIS Família": 12,
            "CST COFINS Item NF": 14, "CST COFINS Transação": 16,
            "CST COFINS Cadastro Produto": 18, "CST COFINS Família": 14,
            "Fonte Efetiva PIS": 16, "Fonte Efetiva COFINS": 18,
            "Fonte Efetiva IPI": 16, "Fonte Efetiva ICMS": 16,
            "Score Risco": 10, "Nível Risco": 12, "Fatores Risco": 50,
        }
        for col_idx, col_name in enumerate(colunas, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = larguras.get(col_name, 16)

        ws.freeze_panes = "A2"

        # Aba de resumo - facilita conferencia
        ws_resumo = wb.create_sheet("Resumo")
        total = len(linhas)
        total_div = sum(1 for ln in linhas if ln["Status Auditoria"] == "DIVERGENTE")
        total_aviso = sum(1 for ln in linhas if ln["Status Auditoria"] == "OK_COM_AVISO")
        total_pend = sum(1 for ln in linhas if ln["Status Auditoria"] == "PENDENTE_MAPEAMENTO")
        total_ok = sum(1 for ln in linhas if ln["Status Auditoria"] == "OK")
        resumo_rows = [
            ("Geração", _dt.now().strftime("%d/%m/%Y %H:%M:%S")),
            ("Filtro movimento", movimento or "TODOS"),
            ("Filtro tipo_item", tipo_item or "TODOS"),
            ("Período (ini)", data_emissao_ini or "-"),
            ("Período (fim)", data_emissao_fim or "-"),
            ("Apenas divergentes", "SIM" if apenas_divergencia else "NÃO"),
            ("Base auditoria", base_auditoria or "MOVIMENTOS"),
            ("", ""),
            ("Total itens exportados", total),
            ("Divergentes", total_div),
            ("OK com aviso", total_aviso),
            ("Pendentes de mapeamento", total_pend),
            ("OK", total_ok),
        ]
        for r_idx, (chave, valor) in enumerate(resumo_rows, start=1):
            ws_resumo.cell(row=r_idx, column=1, value=chave).font = Font(bold=True)
            ws_resumo.cell(row=r_idx, column=2, value=valor)
        ws_resumo.column_dimensions["A"].width = 28
        ws_resumo.column_dimensions["B"].width = 32

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"auditoria_tributaria_{timestamp}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao exportar auditoria tributária: {str(e)}")


def _auditoria_tributaria_inner(
    tipo_item, movimento, numero_documento, serie, parceiro,
    codigo_item, codigo_produto, descricao, familia, origem,
    transacao, data_emissao_ini, data_emissao_fim,
    categoria_material, base_auditoria, apenas_divergencia, pagina, tamanho_pagina
):
    base_auditoria = (base_auditoria or "MOVIMENTOS").upper().strip()
    pagina = max(1, pagina)
    tamanho_pagina = min(max(1, tamanho_pagina), 200)
    offset = (pagina - 1) * tamanho_pagina
    tipo_item = (tipo_item or "TODOS").upper().strip()
    movimento = (movimento or "TODOS").upper().strip()

    # Filtros que só fazem sentido em produtos: força tipo_item=PRODUTO se vier TODOS,
    # senão a query de serviços roda à toa e ainda retorna linhas que ignoram o filtro.
    filtro_produto_especifico = (
        bool(familia)
        or bool(origem)
        or bool(categoria_material and categoria_material.strip().upper() not in ("", "TODOS"))
    )
    if filtro_produto_especifico and tipo_item == "TODOS":
        tipo_item = "PRODUTO"

    filtros_ent_prod = ["I.CODEMP = ?"]
    params_ent_prod = [EMPRESA_PADRAO]
    filtros_ent_srv = ["S.CODEMP = ?"]
    params_ent_srv = [EMPRESA_PADRAO]
    filtros_sai_prod = ["H.CODEMP = ?"]
    params_sai_prod = [EMPRESA_PADRAO]
    filtros_sai_srv = ["H.CODEMP = ?"]
    params_sai_srv = [EMPRESA_PADRAO]

    if numero_documento:
        filtros_ent_prod.append("I.NUMNFC = ?"); filtros_ent_srv.append("S.NUMNFC = ?")
        filtros_sai_prod.append("H.NUMNFV = ?"); filtros_sai_srv.append("H.NUMNFV = ?")
        params_ent_prod.append(numero_documento); params_ent_srv.append(numero_documento)
        params_sai_prod.append(numero_documento); params_sai_srv.append(numero_documento)

    if serie:
        filtros_ent_prod.append("I.CODSNF = ?"); filtros_ent_srv.append("S.CODSNF = ?")
        filtros_sai_prod.append("H.CODSNF = ?"); filtros_sai_srv.append("H.CODSNF = ?")
        params_ent_prod.append(serie.strip()); params_ent_srv.append(serie.strip())
        params_sai_prod.append(serie.strip()); params_sai_srv.append(serie.strip())

    if parceiro:
        filtros_ent_prod.append("I.CODFOR = ?"); filtros_ent_srv.append("S.CODFOR = ?")
        filtros_sai_prod.append("H.CODCLI = ?"); filtros_sai_srv.append("H.CODCLI = ?")
        params_ent_prod.append(parceiro); params_ent_srv.append(parceiro)
        params_sai_prod.append(parceiro); params_sai_srv.append(parceiro)

    if codigo_item:
        termo_codigo = f"%{codigo_item.strip()}%"
        filtros_ent_prod.append("I.CODPRO LIKE ?"); filtros_ent_srv.append("S.CODSER LIKE ?")
        filtros_sai_prod.append("V.CODPRO LIKE ?"); filtros_sai_srv.append("SVI.CODSER LIKE ?")
        params_ent_prod.append(termo_codigo); params_ent_srv.append(termo_codigo)
        params_sai_prod.append(termo_codigo); params_sai_srv.append(termo_codigo)

    if codigo_produto:
        termo_produto = f"%{codigo_produto.strip()}%"
        filtros_ent_prod.append("I.CODPRO LIKE ?"); filtros_sai_prod.append("V.CODPRO LIKE ?")
        params_ent_prod.append(termo_produto); params_sai_prod.append(termo_produto)

    if descricao:
        termo_desc = f"%{descricao.strip()}%"
        filtros_ent_prod.append("P.DESPRO LIKE ?"); filtros_ent_srv.append("ES.DESSER LIKE ?")
        filtros_sai_prod.append("P.DESPRO LIKE ?"); filtros_sai_srv.append("ES.DESSER LIKE ?")
        params_ent_prod.append(termo_desc); params_ent_srv.append(termo_desc)
        params_sai_prod.append(termo_desc); params_sai_srv.append(termo_desc)

    if familia:
        filtros_ent_prod.append("P.CODFAM = ?"); filtros_sai_prod.append("P.CODFAM = ?")
        params_ent_prod.append(familia.strip()); params_sai_prod.append(familia.strip())

    if origem:
        filtros_ent_prod.append("P.CODORI = ?"); filtros_sai_prod.append("P.CODORI = ?")
        params_ent_prod.append(origem.strip()); params_sai_prod.append(origem.strip())

    if transacao:
        filtros_ent_prod.append("I.TNSPRO = ?"); filtros_ent_srv.append("S.TNSSER = ?")
        filtros_sai_prod.append("V.TNSPRO = ?"); filtros_sai_srv.append("SVI.TNSSER = ?")
        params_ent_prod.append(transacao.strip()); params_ent_srv.append(transacao.strip())
        params_sai_prod.append(transacao.strip()); params_sai_srv.append(transacao.strip())

    if data_emissao_ini:
        filtros_ent_prod.append("CAST(N.DATEMI AS DATE) >= ?")
        filtros_ent_srv.append("CAST(N.DATEMI AS DATE) >= ?")
        filtros_sai_prod.append("CAST(H.DATEMI AS DATE) >= ?")
        filtros_sai_srv.append("CAST(H.DATEMI AS DATE) >= ?")
        params_ent_prod.append(data_emissao_ini)
        params_ent_srv.append(data_emissao_ini)
        params_sai_prod.append(data_emissao_ini)
        params_sai_srv.append(data_emissao_ini)

    if data_emissao_fim:
        filtros_ent_prod.append("CAST(N.DATEMI AS DATE) <= ?")
        filtros_ent_srv.append("CAST(N.DATEMI AS DATE) <= ?")
        filtros_sai_prod.append("CAST(H.DATEMI AS DATE) <= ?")
        filtros_sai_srv.append("CAST(H.DATEMI AS DATE) <= ?")
        params_ent_prod.append(data_emissao_fim)
        params_ent_srv.append(data_emissao_fim)
        params_sai_prod.append(data_emissao_fim)
        params_sai_srv.append(data_emissao_fim)

    if categoria_material and categoria_material.strip().upper() not in ("", "TODOS"):
        cat = categoria_material.strip().upper()
        filtros_ent_prod.append(f"{CATEGORIA_SQL.strip()} = ?")
        filtros_sai_prod.append(f"{CATEGORIA_SQL.strip()} = ?")
        params_ent_prod.append(cat)
        params_sai_prod.append(cat)

    sql_parts = []

    if base_auditoria in ("MOVIMENTOS", "AMBOS") and movimento in ("TODOS", "ENTRADA") and tipo_item in ("TODOS", "PRODUTO"):
        sql_ent_prod = f"""
            SELECT
                'MOVIMENTO' AS origem_auditoria,
                'ENTRADA' AS movimento,
                'NF_ENTRADA' AS documento_tipo,
                'PRODUTO' AS tipo_item,
                I.CODEMP AS empresa,
                I.CODFIL AS filial,
                I.CODFOR AS fornecedor_codigo,
                FORN.NOMFOR AS fornecedor_nome,
                FORN.SIGUFS AS fornecedor_uf,
                FORN.CODTRI AS fornecedor_codtri,
                FORN.TIPFOR AS fornecedor_tipfor,
                FORN.SITFOR AS fornecedor_situacao,
                CAST(NULL AS INT) AS cliente_codigo,
                CAST(NULL AS VARCHAR(120)) AS cliente_nome,
                CAST(NULL AS VARCHAR(10)) AS cliente_uf,
                CAST(NULL AS VARCHAR(1)) AS cliente_situacao,
                CAST(NULL AS VARCHAR(120)) AS cliente_endereco,
                CAST(NULL AS VARCHAR(120)) AS cliente_complemento,
                CAST(NULL AS VARCHAR(20)) AS cliente_cep,
                CAST(NULL AS VARCHAR(120)) AS cliente_cidade,
                CAST(NULL AS VARCHAR(120)) AS cliente_bairro,
                CAST(NULL AS NUMERIC(15,4)) AS cliente_redsai_pis,
                CAST(NULL AS NUMERIC(15,4)) AS cliente_redsai_cofins,
                I.NUMNFC AS numero_documento,
                I.CODSNF AS serie,
                N.DATEMI AS data_emissao,
                I.SEQIPC AS seq_item,
                I.CODPRO AS codigo_item,
                I.CODDER AS derivacao,
                P.DESPRO AS descricao_item,
                P.CODFAM AS familia_codigo,
                FAM.DESFAM AS familia_descricao,
                P.CODORI AS origem_codigo,
                ORI.DESORI AS origem_descricao,
                ORI.CODREG AS ori_codreg,
                ORI.CODMS1 AS ori_codms1,
                ORI.CODMS2 AS ori_codms2,
                ORI.CODMS3 AS ori_codms3,
                ORI.CODMS4 AS ori_codms4,
                ORI.PROIMP AS ori_proimp,
                P.CODCLF AS cod_classificacao,
                CLF.CLAFIS AS ncm,
                I.TNSPRO AS transacao,
                I.CSTPIS AS item_cst_pis,
                I.CSTCOF AS item_cst_cofins,
                CAST(NULL AS NUMERIC(15,4)) AS item_bascre,
                I.VLRBPI AS item_base_pis,
                I.VLRBCR AS item_base_cofins,
                I.VLRPIS AS item_valor_pis,
                I.VLRCOR AS item_valor_cofins,
                TP.CSTPIS AS tns_cst_pis,
                TP.CSTCOF AS tns_cst_cofins,
                TNC.BASCRE AS tns_bascre,
                P.RECPIS AS cad_recpis,
                P.RECCOF AS cad_reccof,
                {FAMILIA_SQL_PROD},
                P.PERIPI AS cad_peripi,
                P.RECIPI AS cad_recipi,
                P.TEMICM AS cad_temicm,
                P.CODTRD AS cad_codtrd,
                P.CODTST AS cad_codtst,
                P.CODSTP AS cad_codstp,
                P.RECICM AS cad_recicm,
                {PRODUTO_FISCAL_SQL},
                I.VLRSUB AS item_valor_icms_st,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_inss,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_inss,
                CAST(NULL AS VARCHAR(20)) AS tns_inss_ref,
                I.CSTIPI AS item_cst_ipi,
                I.PERIPI AS item_aliq_ipi,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_ipi,
                I.VLRIPI AS item_valor_ipi,
                CAST(I.PERICM AS NUMERIC(15,4)) AS item_aliq_icms,
                CAST(I.VLRBIC AS NUMERIC(15,2)) AS item_base_icms,
                CAST(I.VLRICM AS NUMERIC(15,2)) AS item_valor_icms,
                CAST(I.CODSTR AS VARCHAR(20)) AS item_cst_icms,
                CAST(I.CODTST AS VARCHAR(20)) AS item_cod_tst_st,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_iss,
                CAST(NULL AS NUMERIC(15,4)) AS item_aliq_iss,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_iss,
                CAST(NULL AS VARCHAR(120)) AS municipio_iss,
                CAST(NULL AS VARCHAR(1)) AS iss_retido,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_irrf,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_csll,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_pis_ret,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_cofins_ret,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_difal,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_difal,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_fcp,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_fcp_st,
                COALESCE(TNP.COMNOP, TNS.COMNOP) AS cfop,
                TNS.DESTNS AS natureza_operacao,
                CAST(NULL AS VARCHAR(20)) AS cest,
                CASE
                    WHEN UPPER(LTRIM(RTRIM(P.CODFAM))) IN ('CONSUM','CONSUMO','CONS') THEN 'CONSUMO'
                    WHEN UPPER(LTRIM(RTRIM(P.CODFAM))) IN ('EPI','EPIS','E.P.I') THEN 'EPI'
                    WHEN UPPER(LTRIM(RTRIM(P.CODFAM))) IN ('IMOB','IMO','ATIVO','IMOBIL','AT.IMOB') THEN 'IMOBILIZADO'
                    WHEN UPPER(LTRIM(RTRIM(P.CODORI))) IN ('MPM','MPG','MP','MAT.PRIMA','MATERIA_PRIMA') THEN 'MATERIA_PRIMA'
                    WHEN UPPER(LTRIM(RTRIM(P.CODORI))) IN ('REV','REVENDA','REVENDAS') THEN 'REVENDA'
                    WHEN P.TIPPRO = 'P' THEN 'PRODUTO_PRODUZIDO'
                    ELSE 'OUTROS'
                END AS categoria_material
            FROM E440IPC I
            INNER JOIN E440NFC N
                ON N.CODEMP = I.CODEMP
               AND N.CODFIL = I.CODFIL
               AND N.CODFOR = I.CODFOR
               AND N.CODSNF = I.CODSNF
               AND N.NUMNFC = I.NUMNFC
            LEFT JOIN E075PRO P ON P.CODEMP = I.CODEMP AND P.CODPRO = I.CODPRO
            LEFT JOIN E001TCP TP ON TP.CODEMP = I.CODEMP AND TP.CODTNS = I.TNSPRO
            LEFT JOIN E001TNC TNC ON TNC.CODEMP = I.CODEMP AND TNC.CODTNS = I.TNSPRO
            LEFT JOIN E001TNS TNS ON TNS.CODEMP = I.CODEMP AND TNS.CODTNS = I.TNSPRO
            LEFT JOIN E012FAM FAM ON FAM.CODEMP = P.CODEMP AND FAM.CODFAM = P.CODFAM
            LEFT JOIN E022CLF CLF ON CLF.CODCLF = P.CODCLF
            LEFT JOIN E083ORI ORI ON ORI.CODEMP = P.CODEMP AND ORI.CODORI = P.CODORI
            LEFT JOIN E095FOR FORN ON FORN.CODFOR = I.CODFOR
            LEFT JOIN E001TNP TNP
                   ON TNP.CODEMP = I.CODEMP
                  AND TNP.CODFIL = I.CODFIL
                  AND TNP.CODTNS = I.TNSPRO
                  AND TNP.SIGUFS = FORN.SIGUFS
                  AND TNP.ENTSAI = 'E'
            WHERE {" AND ".join(filtros_ent_prod)}
        """
        sql_parts.append((sql_ent_prod, params_ent_prod))

    if base_auditoria in ("MOVIMENTOS", "AMBOS") and movimento in ("TODOS", "ENTRADA") and tipo_item in ("TODOS", "SERVICO", "SERVIÁ‡O"):
        sql_ent_srv = f"""
            SELECT
                'MOVIMENTO' AS origem_auditoria,
                'ENTRADA' AS movimento,
                'NF_ENTRADA' AS documento_tipo,
                'SERVICO' AS tipo_item,
                S.CODEMP AS empresa,
                S.CODFIL AS filial,
                S.CODFOR AS fornecedor_codigo,
                FORN.NOMFOR AS fornecedor_nome,
                FORN.SIGUFS AS fornecedor_uf,
                FORN.CODTRI AS fornecedor_codtri,
                FORN.TIPFOR AS fornecedor_tipfor,
                FORN.SITFOR AS fornecedor_situacao,
                CAST(NULL AS INT) AS cliente_codigo,
                CAST(NULL AS VARCHAR(120)) AS cliente_nome,
                CAST(NULL AS VARCHAR(10)) AS cliente_uf,
                CAST(NULL AS VARCHAR(1)) AS cliente_situacao,
                CAST(NULL AS VARCHAR(120)) AS cliente_endereco,
                CAST(NULL AS VARCHAR(120)) AS cliente_complemento,
                CAST(NULL AS VARCHAR(20)) AS cliente_cep,
                CAST(NULL AS VARCHAR(120)) AS cliente_cidade,
                CAST(NULL AS VARCHAR(120)) AS cliente_bairro,
                CAST(NULL AS NUMERIC(15,4)) AS cliente_redsai_pis,
                CAST(NULL AS NUMERIC(15,4)) AS cliente_redsai_cofins,
                S.NUMNFC AS numero_documento,
                S.CODSNF AS serie,
                N.DATEMI AS data_emissao,
                S.SEQISC AS seq_item,
                S.CODSER AS codigo_item,
                CAST(NULL AS VARCHAR(20)) AS derivacao,
                ES.DESSER AS descricao_item,
                CAST(NULL AS VARCHAR(20)) AS familia_codigo,
                CAST(NULL AS VARCHAR(120)) AS familia_descricao,
                CAST(NULL AS VARCHAR(20)) AS origem_codigo,
                CAST(NULL AS VARCHAR(120)) AS origem_descricao,
                CAST(NULL AS SMALLINT) AS ori_codreg,
                CAST(NULL AS VARCHAR(8)) AS ori_codms1,
                CAST(NULL AS VARCHAR(8)) AS ori_codms2,
                CAST(NULL AS VARCHAR(8)) AS ori_codms3,
                CAST(NULL AS VARCHAR(8)) AS ori_codms4,
                CAST(NULL AS SMALLINT) AS ori_proimp,
                CAST(NULL AS VARCHAR(20)) AS cod_classificacao,
                CAST(NULL AS VARCHAR(30)) AS ncm,
                S.TNSSER AS transacao,
                S.CSTPIS AS item_cst_pis,
                S.CSTCOF AS item_cst_cofins,
                CAST(NULL AS NUMERIC(15,4)) AS item_bascre,
                S.VLRBPI AS item_base_pis,
                S.VLRBCR AS item_base_cofins,
                S.VLRPIS AS item_valor_pis,
                S.VLRCOR AS item_valor_cofins,
                TP.CSTPIS AS tns_cst_pis,
                TP.CSTCOF AS tns_cst_cofins,
                TNC.BASCRE AS tns_bascre,
                ES.RECPIS AS cad_recpis,
                ES.RECCOF AS cad_reccof,
                {FAMILIA_SQL_NULL},
                ES.PERIPI AS cad_peripi,
                ES.RECIPI AS cad_recipi,
                ES.TEMICM AS cad_temicm,
                ES.CODTRD AS cad_codtrd,
                CAST(NULL AS VARCHAR(20)) AS cad_codtst,
                CAST(NULL AS VARCHAR(20)) AS cad_codstp,
                CAST(NULL AS VARCHAR(1)) AS cad_recicm,
                {PRODUTO_FISCAL_NULL},
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_icms_st,
                S.VLRBIN AS item_base_inss,
                S.VLRINS AS item_valor_inss,
                S.TNSSER AS tns_inss_ref,
                S.CSTIPI AS item_cst_ipi,
                S.PERPIF AS item_aliq_ipi,
                S.VLRBPF AS item_base_ipi,
                S.VLRPIF AS item_valor_ipi,
                CAST(NULL AS NUMERIC(15,4)) AS item_aliq_icms,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_icms,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_icms,
                CAST(NULL AS VARCHAR(20)) AS item_cst_icms,
                CAST(NULL AS VARCHAR(20)) AS item_cod_tst_st,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_iss,
                CAST(NULL AS NUMERIC(15,4)) AS item_aliq_iss,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_iss,
                CAST(NULL AS VARCHAR(120)) AS municipio_iss,
                CAST(NULL AS VARCHAR(1)) AS iss_retido,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_irrf,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_csll,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_pis_ret,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_cofins_ret,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_difal,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_difal,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_fcp,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_fcp_st,
                COALESCE(TNP.COMNOP, TNS.COMNOP) AS cfop,
                TNS.DESTNS AS natureza_operacao,
                CAST(NULL AS VARCHAR(20)) AS cest,
                'SERVICO' AS categoria_material
            FROM E440ISC S
            INNER JOIN E440NFC N
                ON N.CODEMP = S.CODEMP
               AND N.CODFIL = S.CODFIL
               AND N.CODFOR = S.CODFOR
               AND N.CODSNF = S.CODSNF
               AND N.NUMNFC = S.NUMNFC
            LEFT JOIN E080SER ES ON ES.CODEMP = S.CODEMP AND ES.CODSER = S.CODSER
            LEFT JOIN E001TCP TP ON TP.CODEMP = S.CODEMP AND TP.CODTNS = S.TNSSER
            LEFT JOIN E001TNC TNC ON TNC.CODEMP = S.CODEMP AND TNC.CODTNS = S.TNSSER
            LEFT JOIN E001TNS TNS ON TNS.CODEMP = S.CODEMP AND TNS.CODTNS = S.TNSSER
            LEFT JOIN E095FOR FORN ON FORN.CODFOR = S.CODFOR
            LEFT JOIN E001TNP TNP
                   ON TNP.CODEMP = S.CODEMP
                  AND TNP.CODFIL = S.CODFIL
                  AND TNP.CODTNS = S.TNSSER
                  AND TNP.SIGUFS = FORN.SIGUFS
                  AND TNP.ENTSAI = 'E'
            WHERE {" AND ".join(filtros_ent_srv)}
        """
        sql_parts.append((sql_ent_srv, params_ent_srv))

    if base_auditoria in ("MOVIMENTOS", "AMBOS") and movimento in ("TODOS", "SAIDA", "SAÁDA") and tipo_item in ("TODOS", "PRODUTO"):
        sql_sai_prod = f"""
            SELECT
                'MOVIMENTO' AS origem_auditoria,
                'SAIDA' AS movimento,
                'NF_SAIDA' AS documento_tipo,
                'PRODUTO' AS tipo_item,
                H.CODEMP AS empresa,
                H.CODFIL AS filial,
                CAST(NULL AS INT) AS fornecedor_codigo,
                CAST(NULL AS VARCHAR(120)) AS fornecedor_nome,
                CAST(NULL AS VARCHAR(10)) AS fornecedor_uf,
                CAST(NULL AS VARCHAR(20)) AS fornecedor_codtri,
                CAST(NULL AS VARCHAR(1)) AS fornecedor_tipfor,
                CAST(NULL AS VARCHAR(1)) AS fornecedor_situacao,
                H.CODCLI AS cliente_codigo,
                CLI.NOMCLI AS cliente_nome,
                CLI.SIGUFS AS cliente_uf,
                CLI.SITCLI AS cliente_situacao,
                CLI.ENDCLI AS cliente_endereco,
                CLI.CPLEND AS cliente_complemento,
                CLI.CEPCLI AS cliente_cep,
                CLI.CIDCLI AS cliente_cidade,
                CLI.BAICLI AS cliente_bairro,
                RED41.REDSAI AS cliente_redsai_pis,
                RED42.REDSAI AS cliente_redsai_cofins,
                H.NUMNFV AS numero_documento,
                H.CODSNF AS serie,
                H.DATEMI AS data_emissao,
                ROW_NUMBER() OVER (
                    PARTITION BY H.CODEMP, H.CODFIL, H.CODSNF, H.NUMNFV
                    ORDER BY V.CODPRO, V.CODDER
                ) AS seq_item,
                V.CODPRO AS codigo_item,
                V.CODDER AS derivacao,
                P.DESPRO AS descricao_item,
                P.CODFAM AS familia_codigo,
                FAM.DESFAM AS familia_descricao,
                P.CODORI AS origem_codigo,
                ORI.DESORI AS origem_descricao,
                ORI.CODREG AS ori_codreg,
                ORI.CODMS1 AS ori_codms1,
                ORI.CODMS2 AS ori_codms2,
                ORI.CODMS3 AS ori_codms3,
                ORI.CODMS4 AS ori_codms4,
                ORI.PROIMP AS ori_proimp,
                P.CODCLF AS cod_classificacao,
                CLF.CLAFIS AS ncm,
                V.TNSPRO AS transacao,
                V.CSTPIS AS item_cst_pis,
                V.CSTCOF AS item_cst_cofins,
                CAST(NULL AS NUMERIC(15,4)) AS item_bascre,
                V.VLRBPI AS item_base_pis,
                V.VLRBCR AS item_base_cofins,
                V.VLRPIS AS item_valor_pis,
                V.VLRCOR AS item_valor_cofins,
                TP.CSTPIS AS tns_cst_pis,
                TP.CSTCOF AS tns_cst_cofins,
                TNC.BASCRE AS tns_bascre,
                P.RECPIS AS cad_recpis,
                P.RECCOF AS cad_reccof,
                {FAMILIA_SQL_PROD},
                P.PERIPI AS cad_peripi,
                P.RECIPI AS cad_recipi,
                P.TEMICM AS cad_temicm,
                P.CODTRD AS cad_codtrd,
                P.CODTST AS cad_codtst,
                P.CODSTP AS cad_codstp,
                P.RECICM AS cad_recicm,
                {PRODUTO_FISCAL_SQL},
                V.VLRSUB AS item_valor_icms_st,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_inss,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_inss,
                CAST(NULL AS VARCHAR(20)) AS tns_inss_ref,
                V.CSTIPI AS item_cst_ipi,
                V.ALIIPI AS item_aliq_ipi,
                V.VLRBIP AS item_base_ipi,
                V.VLRIPI AS item_valor_ipi,
                CAST(V.PERICM AS NUMERIC(15,4)) AS item_aliq_icms,
                CAST(V.VLRBIC AS NUMERIC(15,2)) AS item_base_icms,
                CAST(V.VLRICM AS NUMERIC(15,2)) AS item_valor_icms,
                CAST(V.CODSTR AS VARCHAR(20)) AS item_cst_icms,
                CAST(V.CODTST AS VARCHAR(20)) AS item_cod_tst_st,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_iss,
                CAST(NULL AS NUMERIC(15,4)) AS item_aliq_iss,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_iss,
                CAST(NULL AS VARCHAR(120)) AS municipio_iss,
                CAST(NULL AS VARCHAR(1)) AS iss_retido,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_irrf,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_csll,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_pis_ret,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_cofins_ret,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_difal,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_difal,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_fcp,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_fcp_st,
                COALESCE(TNP.COMNOP, TNS.COMNOP) AS cfop,
                TNS.DESTNS AS natureza_operacao,
                CAST(NULL AS VARCHAR(20)) AS cest,
                CASE
                    WHEN UPPER(LTRIM(RTRIM(P.CODFAM))) IN ('CONSUM','CONSUMO','CONS') THEN 'CONSUMO'
                    WHEN UPPER(LTRIM(RTRIM(P.CODFAM))) IN ('EPI','EPIS','E.P.I') THEN 'EPI'
                    WHEN UPPER(LTRIM(RTRIM(P.CODFAM))) IN ('IMOB','IMO','ATIVO','IMOBIL','AT.IMOB') THEN 'IMOBILIZADO'
                    WHEN UPPER(LTRIM(RTRIM(P.CODORI))) IN ('MPM','MPG','MP','MAT.PRIMA','MATERIA_PRIMA') THEN 'MATERIA_PRIMA'
                    WHEN UPPER(LTRIM(RTRIM(P.CODORI))) IN ('REV','REVENDA','REVENDAS') THEN 'REVENDA'
                    WHEN P.TIPPRO = 'P' THEN 'PRODUTO_PRODUZIDO'
                    ELSE 'OUTROS'
                END AS categoria_material
            FROM E140NFV H
            INNER JOIN E140IPV V ON V.CODEMP = H.CODEMP AND V.CODFIL = H.CODFIL AND V.CODSNF = H.CODSNF AND V.NUMNFV = H.NUMNFV
            LEFT JOIN E075PRO P ON P.CODEMP = V.CODEMP AND P.CODPRO = V.CODPRO
            LEFT JOIN E001TCP TP ON TP.CODEMP = V.CODEMP AND TP.CODTNS = V.TNSPRO
            LEFT JOIN E001TNC TNC ON TNC.CODEMP = V.CODEMP AND TNC.CODTNS = V.TNSPRO
            LEFT JOIN E001TNS TNS ON TNS.CODEMP = V.CODEMP AND TNS.CODTNS = V.TNSPRO
            LEFT JOIN E012FAM FAM ON FAM.CODEMP = P.CODEMP AND FAM.CODFAM = P.CODFAM
            LEFT JOIN E022CLF CLF ON CLF.CODCLF = P.CODCLF
            LEFT JOIN E083ORI ORI ON ORI.CODEMP = P.CODEMP AND ORI.CODORI = P.CODORI
            LEFT JOIN E085CLI CLI ON CLI.CODCLI = H.CODCLI
            LEFT JOIN E001TNP TNP
                   ON TNP.CODEMP = V.CODEMP
                  AND TNP.CODFIL = V.CODFIL
                  AND TNP.CODTNS = V.TNSPRO
                  AND TNP.SIGUFS = CLI.SIGUFS
                  AND TNP.ENTSAI = 'S'
            LEFT JOIN E019RED RED41 ON RED41.CODEMP = H.CODEMP AND RED41.CODFIL = H.CODFIL AND RED41.CODTRD = P.CODTRD AND RED41.TIPIMP = 41 AND RED41.SIGUFS = CLI.SIGUFS
            LEFT JOIN E019RED RED42 ON RED42.CODEMP = H.CODEMP AND RED42.CODFIL = H.CODFIL AND RED42.CODTRD = P.CODTRD AND RED42.TIPIMP = 42 AND RED42.SIGUFS = CLI.SIGUFS
            WHERE {" AND ".join(filtros_sai_prod)}
        """
        sql_parts.append((sql_sai_prod, params_sai_prod))

    if base_auditoria in ("MOVIMENTOS", "AMBOS") and movimento in ("TODOS", "SAIDA", "SAÁDA") and tipo_item in ("TODOS", "SERVICO", "SERVIÁ‡O"):
        sql_sai_srv = f"""
            SELECT
                'MOVIMENTO' AS origem_auditoria,
                'SAIDA' AS movimento,
                'NF_SAIDA' AS documento_tipo,
                'SERVICO' AS tipo_item,
                H.CODEMP AS empresa,
                H.CODFIL AS filial,
                CAST(NULL AS INT) AS fornecedor_codigo,
                CAST(NULL AS VARCHAR(120)) AS fornecedor_nome,
                CAST(NULL AS VARCHAR(10)) AS fornecedor_uf,
                CAST(NULL AS VARCHAR(20)) AS fornecedor_codtri,
                CAST(NULL AS VARCHAR(1)) AS fornecedor_tipfor,
                CAST(NULL AS VARCHAR(1)) AS fornecedor_situacao,
                H.CODCLI AS cliente_codigo,
                CLI.NOMCLI AS cliente_nome,
                CLI.SIGUFS AS cliente_uf,
                CLI.SITCLI AS cliente_situacao,
                CLI.ENDCLI AS cliente_endereco,
                CLI.CPLEND AS cliente_complemento,
                CLI.CEPCLI AS cliente_cep,
                CLI.CIDCLI AS cliente_cidade,
                CLI.BAICLI AS cliente_bairro,
                RED41.REDSAI AS cliente_redsai_pis,
                RED42.REDSAI AS cliente_redsai_cofins,
                H.NUMNFV AS numero_documento,
                H.CODSNF AS serie,
                H.DATEMI AS data_emissao,
                ROW_NUMBER() OVER (
                    PARTITION BY H.CODEMP, H.CODFIL, H.CODSNF, H.NUMNFV
                    ORDER BY SVI.CODSER
                ) AS seq_item,
                SVI.CODSER AS codigo_item,
                CAST(NULL AS VARCHAR(20)) AS derivacao,
                ES.DESSER AS descricao_item,
                CAST(NULL AS VARCHAR(20)) AS familia_codigo,
                CAST(NULL AS VARCHAR(120)) AS familia_descricao,
                CAST(NULL AS VARCHAR(20)) AS origem_codigo,
                CAST(NULL AS VARCHAR(120)) AS origem_descricao,
                CAST(NULL AS SMALLINT) AS ori_codreg,
                CAST(NULL AS VARCHAR(8)) AS ori_codms1,
                CAST(NULL AS VARCHAR(8)) AS ori_codms2,
                CAST(NULL AS VARCHAR(8)) AS ori_codms3,
                CAST(NULL AS VARCHAR(8)) AS ori_codms4,
                CAST(NULL AS SMALLINT) AS ori_proimp,
                CAST(NULL AS VARCHAR(20)) AS cod_classificacao,
                CAST(NULL AS VARCHAR(30)) AS ncm,
                SVI.TNSSER AS transacao,
                SVI.CSTPIS AS item_cst_pis,
                SVI.CSTCOF AS item_cst_cofins,
                CAST(NULL AS NUMERIC(15,4)) AS item_bascre,
                SVI.VLRBPI AS item_base_pis,
                SVI.VLRBCR AS item_base_cofins,
                SVI.VLRPIS AS item_valor_pis,
                SVI.VLRCOR AS item_valor_cofins,
                TP.CSTPIS AS tns_cst_pis,
                TP.CSTCOF AS tns_cst_cofins,
                TNC.BASCRE AS tns_bascre,
                ES.RECPIS AS cad_recpis,
                ES.RECCOF AS cad_reccof,
                {FAMILIA_SQL_NULL},
                ES.PERIPI AS cad_peripi,
                ES.RECIPI AS cad_recipi,
                ES.TEMICM AS cad_temicm,
                ES.CODTRD AS cad_codtrd,
                CAST(NULL AS VARCHAR(20)) AS cad_codtst,
                CAST(NULL AS VARCHAR(20)) AS cad_codstp,
                CAST(NULL AS VARCHAR(1)) AS cad_recicm,
                {PRODUTO_FISCAL_NULL},
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_icms_st,
                SVI.VLRBIN AS item_base_inss,
                SVI.VLRINS AS item_valor_inss,
                SVI.TNSSER AS tns_inss_ref,
                SVI.CSTIPI AS item_cst_ipi,
                SVI.PERPIF AS item_aliq_ipi,
                SVI.VLRBPF AS item_base_ipi,
                SVI.VLRPIF AS item_valor_ipi,
                CAST(NULL AS NUMERIC(15,4)) AS item_aliq_icms,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_icms,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_icms,
                CAST(NULL AS VARCHAR(20)) AS item_cst_icms,
                CAST(NULL AS VARCHAR(20)) AS item_cod_tst_st,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_iss,
                CAST(NULL AS NUMERIC(15,4)) AS item_aliq_iss,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_iss,
                CAST(NULL AS VARCHAR(120)) AS municipio_iss,
                CAST(NULL AS VARCHAR(1)) AS iss_retido,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_irrf,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_csll,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_pis_ret,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_cofins_ret,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_difal,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_difal,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_fcp,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_fcp_st,
                COALESCE(TNP.COMNOP, TNS.COMNOP) AS cfop,
                TNS.DESTNS AS natureza_operacao,
                CAST(NULL AS VARCHAR(20)) AS cest,
                'SERVICO' AS categoria_material
            FROM E140NFV H
            INNER JOIN E140ISV SVI ON SVI.CODEMP = H.CODEMP AND SVI.CODFIL = H.CODFIL AND SVI.CODSNF = H.CODSNF AND SVI.NUMNFV = H.NUMNFV
            LEFT JOIN E080SER ES ON ES.CODEMP = SVI.CODEMP AND ES.CODSER = SVI.CODSER
            LEFT JOIN E001TCP TP ON TP.CODEMP = SVI.CODEMP AND TP.CODTNS = SVI.TNSSER
            LEFT JOIN E001TNC TNC ON TNC.CODEMP = SVI.CODEMP AND TNC.CODTNS = SVI.TNSSER
            LEFT JOIN E001TNS TNS ON TNS.CODEMP = SVI.CODEMP AND TNS.CODTNS = SVI.TNSSER
            LEFT JOIN E085CLI CLI ON CLI.CODCLI = H.CODCLI
            LEFT JOIN E001TNP TNP
                   ON TNP.CODEMP = SVI.CODEMP
                  AND TNP.CODFIL = SVI.CODFIL
                  AND TNP.CODTNS = SVI.TNSSER
                  AND TNP.SIGUFS = CLI.SIGUFS
                  AND TNP.ENTSAI = 'S'
            LEFT JOIN E019RED RED41 ON RED41.CODEMP = H.CODEMP AND RED41.CODFIL = H.CODFIL AND RED41.CODTRD = ES.CODTRD AND RED41.TIPIMP = 41 AND RED41.SIGUFS = CLI.SIGUFS
            LEFT JOIN E019RED RED42 ON RED42.CODEMP = H.CODEMP AND RED42.CODFIL = H.CODFIL AND RED42.CODTRD = ES.CODTRD AND RED42.TIPIMP = 42 AND RED42.SIGUFS = CLI.SIGUFS
            WHERE {" AND ".join(filtros_sai_srv)}
        """
        sql_parts.append((sql_sai_srv, params_sai_srv))

    # ---- Auditoria de cadastro fiscal (sem movimento) ----
    if base_auditoria in ("CADASTRO", "AMBOS") and movimento in ("TODOS", "SEM_MOVIMENTO") and tipo_item in ("TODOS", "PRODUTO"):
        filtros_cad = ["P.CODEMP = ?"]
        params_cad: list = [EMPRESA_PADRAO]
        if codigo_item:
            filtros_cad.append("P.CODPRO LIKE ?")
            params_cad.append(f"%{codigo_item.strip()}%")
        if codigo_produto:
            filtros_cad.append("P.CODPRO LIKE ?")
            params_cad.append(f"%{codigo_produto.strip()}%")
        if descricao:
            filtros_cad.append("P.DESPRO LIKE ?")
            params_cad.append(f"%{descricao.strip()}%")
        if familia:
            filtros_cad.append("P.CODFAM = ?")
            params_cad.append(familia.strip())
        if origem:
            filtros_cad.append("P.CODORI = ?")
            params_cad.append(origem.strip())
        if categoria_material and categoria_material.strip().upper() not in ("", "TODOS"):
            filtros_cad.append(f"{CATEGORIA_SQL.strip()} = ?")
            params_cad.append(categoria_material.strip().upper())
        sql_cad_prod = f"""
            SELECT
                'CADASTRO' AS origem_auditoria,
                'SEM_MOVIMENTO' AS movimento,
                'CADASTRO_FISCAL' AS documento_tipo,
                'PRODUTO' AS tipo_item,
                P.CODEMP AS empresa,
                CAST(NULL AS INT) AS filial,
                CAST(NULL AS INT) AS fornecedor_codigo,
                CAST(NULL AS VARCHAR(120)) AS fornecedor_nome,
                CAST(NULL AS VARCHAR(10)) AS fornecedor_uf,
                CAST(NULL AS VARCHAR(20)) AS fornecedor_codtri,
                CAST(NULL AS VARCHAR(1)) AS fornecedor_tipfor,
                CAST(NULL AS VARCHAR(1)) AS fornecedor_situacao,
                CAST(NULL AS INT) AS cliente_codigo,
                CAST(NULL AS VARCHAR(120)) AS cliente_nome,
                CAST(NULL AS VARCHAR(10)) AS cliente_uf,
                CAST(NULL AS VARCHAR(1)) AS cliente_situacao,
                CAST(NULL AS VARCHAR(120)) AS cliente_endereco,
                CAST(NULL AS VARCHAR(120)) AS cliente_complemento,
                CAST(NULL AS VARCHAR(20)) AS cliente_cep,
                CAST(NULL AS VARCHAR(120)) AS cliente_cidade,
                CAST(NULL AS VARCHAR(120)) AS cliente_bairro,
                CAST(NULL AS NUMERIC(15,4)) AS cliente_redsai_pis,
                CAST(NULL AS NUMERIC(15,4)) AS cliente_redsai_cofins,
                CAST(NULL AS INT) AS numero_documento,
                CAST(NULL AS VARCHAR(10)) AS serie,
                CAST(NULL AS DATETIME) AS data_emissao,
                ROW_NUMBER() OVER (ORDER BY P.CODPRO) AS seq_item,
                P.CODPRO AS codigo_item,
                CAST(NULL AS VARCHAR(20)) AS derivacao,
                P.DESPRO AS descricao_item,
                P.CODFAM AS familia_codigo,
                FAM.DESFAM AS familia_descricao,
                P.CODORI AS origem_codigo,
                ORI.DESORI AS origem_descricao,
                ORI.CODREG AS ori_codreg,
                ORI.CODMS1 AS ori_codms1,
                ORI.CODMS2 AS ori_codms2,
                ORI.CODMS3 AS ori_codms3,
                ORI.CODMS4 AS ori_codms4,
                ORI.PROIMP AS ori_proimp,
                P.CODCLF AS cod_classificacao,
                CLF.CLAFIS AS ncm,
                CAST(NULL AS VARCHAR(20)) AS transacao,
                CAST(NULL AS VARCHAR(10)) AS item_cst_pis,
                CAST(NULL AS VARCHAR(10)) AS item_cst_cofins,
                CAST(NULL AS NUMERIC(15,4)) AS item_bascre,
                CAST(NULL AS NUMERIC(15,4)) AS item_base_pis,
                CAST(NULL AS NUMERIC(15,4)) AS item_base_cofins,
                CAST(NULL AS NUMERIC(15,4)) AS item_valor_pis,
                CAST(NULL AS NUMERIC(15,4)) AS item_valor_cofins,
                CAST(NULL AS VARCHAR(10)) AS tns_cst_pis,
                CAST(NULL AS VARCHAR(10)) AS tns_cst_cofins,
                CAST(NULL AS NUMERIC(15,4)) AS tns_bascre,
                P.RECPIS AS cad_recpis,
                P.RECCOF AS cad_reccof,
                {FAMILIA_SQL_PROD},
                P.PERIPI AS cad_peripi,
                P.RECIPI AS cad_recipi,
                P.TEMICM AS cad_temicm,
                P.CODTRD AS cad_codtrd,
                P.CODTST AS cad_codtst,
                P.CODSTP AS cad_codstp,
                P.RECICM AS cad_recicm,
                {PRODUTO_FISCAL_SQL},
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_icms_st,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_inss,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_inss,
                CAST(NULL AS VARCHAR(20)) AS tns_inss_ref,
                CAST(NULL AS VARCHAR(10)) AS item_cst_ipi,
                CAST(NULL AS NUMERIC(15,4)) AS item_aliq_ipi,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_ipi,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_ipi,
                CAST(NULL AS NUMERIC(15,4)) AS item_aliq_icms,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_icms,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_icms,
                CAST(NULL AS VARCHAR(20)) AS item_cst_icms,
                CAST(NULL AS VARCHAR(20)) AS item_cod_tst_st,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_iss,
                CAST(NULL AS NUMERIC(15,4)) AS item_aliq_iss,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_iss,
                CAST(NULL AS VARCHAR(120)) AS municipio_iss,
                CAST(NULL AS VARCHAR(1)) AS iss_retido,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_irrf,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_csll,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_pis_ret,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_cofins_ret,
                CAST(NULL AS NUMERIC(15,2)) AS item_base_difal,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_difal,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_fcp,
                CAST(NULL AS NUMERIC(15,2)) AS item_valor_fcp_st,
                CAST(NULL AS VARCHAR(20)) AS cfop,
                CAST(NULL AS VARCHAR(120)) AS natureza_operacao,
                CAST(NULL AS VARCHAR(20)) AS cest,
                {CATEGORIA_SQL.strip()} AS categoria_material
            FROM E075PRO P
            LEFT JOIN E012FAM FAM ON FAM.CODEMP = P.CODEMP AND FAM.CODFAM = P.CODFAM
            LEFT JOIN E083ORI ORI ON ORI.CODEMP = P.CODEMP AND ORI.CODORI = P.CODORI
            LEFT JOIN E022CLF CLF ON CLF.CODCLF = P.CODCLF
            WHERE {" AND ".join(filtros_cad)}
        """
        sql_parts.append((sql_cad_prod, params_cad))

    if not sql_parts:
        return {"resumo": {}, "itens": [], "pagina": pagina, "tamanho_pagina": tamanho_pagina, "total_registros": 0, "total_paginas": 1}

    union_sql = " UNION ALL ".join([part[0] for part in sql_parts])
    params = []
    for _, p in sql_parts:
        params.extend(p)

    # Filtro de apenas_divergencia no SQL (evita paginacao com paginas vazias)
    # Nota: divergencia e calculada no Python, entao usamos camada intermediaria
    # Mantemos o cat_where para caso de filtro residual (categoria ja entrou no WHERE de cada sub-query)
    outer_wheres = []
    outer_params: list = []

    sql_total = f"""
        SELECT COUNT(1) AS total
        FROM (
            {union_sql}
        ) X
    """

    final_sql = f"""
        SELECT *
        FROM (
            SELECT
                X.*,
                ROW_NUMBER() OVER (ORDER BY X.numero_documento DESC, X.seq_item DESC) AS _rn
            FROM (
                {union_sql}
            ) X
        ) Z
        WHERE Z._rn > ? AND Z._rn <= ?
        ORDER BY Z._rn
    """

    inicio = offset
    fim = offset + tamanho_pagina

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(sql_total, params)
    total_registros = int(cursor.fetchone()[0])
    cursor.execute(final_sql, params + [inicio, fim])
    rows = cursor.fetchall()
    registros = [row_to_dict(cursor, row) for row in rows]
    conn.close()

    itens = []
    for r in registros:
        divergencias_reais = []
        avisos_cadastrais = []
        pendencias_mapeamento = []
        impostos = {}

        familia_parametrizacao = pick_prefixed_fields(
            r, "fam_", {"familia_codigo", "familia_descricao"}
        )
        origem_parametrizacao = pick_prefixed_fields(
            r, "ori_", {"origem_codigo", "origem_descricao"}
        )
        if is_str_diff(r.get("tns_cst_pis"), r.get("tns_cst_cofins")):
            divergencias_reais.append("Transação com CST PIS diferente de CST COFINS")
        if is_str_diff(r.get("cad_recpis"), r.get("cad_reccof")):
            avisos_cadastrais.append("Cadastro com Recupera PIS diferente de Recupera COFINS")
        if is_str_diff(r.get("item_cst_pis"), r.get("item_cst_cofins")):
            divergencias_reais.append("Item com CST PIS diferente de CST COFINS")
        if is_num_diff(r.get("item_base_pis"), r.get("item_base_cofins")):
            divergencias_reais.append("Item com base de PIS diferente da base de COFINS")
        if is_num_diff(r.get("tns_bascre"), r.get("item_bascre")):
            divergencias_reais.append("Base de crédito do item diferente da base de crédito da transação")

        impostos["pis_cofins"] = {
            "item_cst_pis": r.get("item_cst_pis"), "item_cst_cofins": r.get("item_cst_cofins"),
            "item_base_pis": r.get("item_base_pis"), "item_base_cofins": r.get("item_base_cofins"),
            "item_valor_pis": r.get("item_valor_pis"), "item_valor_cofins": r.get("item_valor_cofins"),
            "tns_cst_pis": r.get("tns_cst_pis"), "tns_cst_cofins": r.get("tns_cst_cofins"),
            "tns_bascre": r.get("tns_bascre"), "item_bascre": r.get("item_bascre"),
            "cad_recpis": r.get("cad_recpis"), "cad_reccof": r.get("cad_reccof"),
            "cad_cstpis_produto": r.get("cad_cstpis_produto"),
            "cad_cstcof_produto": r.get("cad_cstcof_produto"),
            "fam_cst_pis": r.get("fam_cst_pis"), "fam_cst_cofins": r.get("fam_cst_cofins"),
        }

        ipi_motivos = []
        ipi_divergencias = []
        ipi_avisos = []
        if r.get("cad_peripi") is None and r.get("fam_cst_ipi") and r.get("tipo_item") != "SERVICO":
            ipi_avisos.append("Família possui CST IPI, mas cadastro do item não trouxe driver IPI")
        if r.get("cad_recipi") is None:
            ipi_avisos.append("Cadastro sem flag de recuperação de IPI")
        if r.get("fam_proimp") is None and r.get("tipo_item") == "PRODUTO":
            ipi_avisos.append("Família sem tipo de produto para impostos (ProImp)")
        ipi_motivos = ipi_divergencias + ipi_avisos
        impostos["ipi"] = {
            "item_cst_ipi": r.get("item_cst_ipi"), "item_aliq_ipi": r.get("item_aliq_ipi"),
            "item_base_ipi": r.get("item_base_ipi"), "item_valor_ipi": r.get("item_valor_ipi"),
            "cad_peripi": r.get("cad_peripi"), "cad_recipi": r.get("cad_recipi"),
            "fam_cst_ipi": r.get("fam_cst_ipi"), "fam_proimp": r.get("fam_proimp"),
            "motivos": ipi_motivos,
            "divergencias_reais": ipi_divergencias,
            "avisos_cadastrais": ipi_avisos,
        }
        divergencias_reais.extend(ipi_divergencias)
        avisos_cadastrais.extend(ipi_avisos)

        icms_motivos = []
        icms_divergencias = []
        icms_avisos = []
        if r.get("cad_temicm") is None:
            icms_avisos.append("Cadastro sem indicador de incidência de ICMS")
        if r.get("tipo_item") == "PRODUTO" and r.get("cad_codtrd") is None:
            icms_avisos.append("Cadastro sem código de tributação/diferencial (CodTrd)")
        if r.get("tipo_item") == "PRODUTO" and r.get("cad_recicm") is None:
            icms_avisos.append("Cadastro sem flag de recuperação de ICMS")
        if r.get("movimento") in ("ENTRADA", "SAIDA") and r.get("tipo_item") == "PRODUTO":
            if r.get("item_base_icms") is None:
                icms_divergencias.append("Item sem base de ICMS mapeada")
            if r.get("item_aliq_icms") is None:
                icms_divergencias.append("Item sem alíquota de ICMS mapeada")
            if not _clean_str(r.get("item_cst_icms")):
                icms_divergencias.append("Item sem CST/estratégia ICMS mapeada")
        icms_motivos = icms_divergencias + icms_avisos
        impostos["icms"] = {
            "item_aliq_icms": r.get("item_aliq_icms"),
            "item_base_icms": r.get("item_base_icms"),
            "item_valor_icms": r.get("item_valor_icms"),
            "item_cst_icms": r.get("item_cst_icms"),
            "cad_temicm": r.get("cad_temicm"),
            "cad_codtrd": r.get("cad_codtrd"),
            "cad_recicm": r.get("cad_recicm"),
            "motivos": icms_motivos,
            "divergencias_reais": icms_divergencias,
            "avisos_cadastrais": icms_avisos,
        }
        divergencias_reais.extend(icms_divergencias)
        avisos_cadastrais.extend(icms_avisos)

        st_motivos = []
        if r.get("tipo_item") == "PRODUTO":
            if not r.get("cad_codtst"):
                st_motivos.append("Cadastro sem Código TST para análise de ICMS ST")
            if not r.get("cad_codstp"):
                st_motivos.append("Cadastro sem Código STP para análise de ICMS ST")
        impostos["icms_st"] = {
            "item_valor_icms_st": r.get("item_valor_icms_st"),
            "cad_codtst": r.get("cad_codtst"), "cad_codstp": r.get("cad_codstp"),
            "motivos": st_motivos,
            "avisos_cadastrais": list(st_motivos),
        }
        avisos_cadastrais.extend(st_motivos)

        origem_motivos = []
        if r.get("tipo_item") == "PRODUTO":
            if not _clean_str(r.get("ori_codreg")) and not _clean_str(r.get("ori_codms1")) and not _clean_str(r.get("ori_codms2")) and not _clean_str(r.get("ori_codms3")) and not _clean_str(r.get("ori_codms4")) and r.get("ori_proimp") is None:
                origem_motivos.append("Origem sem cadastro fiscal visível (CodReg/CodMS1/CodMS2/CodMS3/CodMS4/ProImp)")
        impostos["origem"] = {
            **origem_parametrizacao,
            "motivos": origem_motivos,
            "avisos_cadastrais": list(origem_motivos),
        }
        avisos_cadastrais.extend(origem_motivos)

        inss_motivos = []
        if r.get("tipo_item") == "SERVICO":
            tns_ref = _clean_str(r.get("tns_inss_ref"))
            if tns_ref in {"5949", "5933", "6933", "6949", "1933", "2933", "6933A"}:
                inss_motivos.append("Transação de serviço está entre as regras que pedem alteração de base de INSS")
        impostos["inss"] = {
            "item_base_inss": r.get("item_base_inss"), "item_valor_inss": r.get("item_valor_inss"),
            "tns_inss_ref": r.get("tns_inss_ref"), "motivos": inss_motivos,
            "avisos_cadastrais": list(inss_motivos),
        }
        avisos_cadastrais.extend(inss_motivos)

        cliente_motivos = []
        cliente_divergencias = []
        cliente_avisos = []
        if r.get("movimento") == "SAIDA":
            if not r.get("cliente_codigo"):
                cliente_divergencias.append("NF de saída sem cliente associado")
            if not _clean_str(r.get("cliente_uf")):
                cliente_divergencias.append("Cliente sem UF cadastrada para regras fiscais de saída")
            if _clean_str(r.get("cliente_situacao")) not in ("", "A"):
                cliente_divergencias.append("Cliente inativo")
            if not _clean_str(r.get("cliente_endereco")):
                cliente_avisos.append("Cliente sem endereço cadastrado")
            if not _clean_str(r.get("cliente_cidade")):
                cliente_avisos.append("Cliente sem cidade cadastrada")
            origem_saida = _clean_str(r.get("origem_codigo")).upper()
            cod_clf = _clean_str(r.get("cod_classificacao")).upper()
            if origem_saida in {"250", "280", "PPG"} and cod_clf == "101":
                if r.get("cad_codtrd") in (None, ""):
                    cliente_avisos.append("Produto sem CodTrd para regra fiscal de saída baseada no cliente")
                if r.get("cliente_redsai_pis") is None and r.get("cliente_redsai_cofins") is None:
                    cliente_avisos.append("Cliente/UF sem regra RedSai encontrada em E019RED")
        cliente_motivos = cliente_divergencias + cliente_avisos
        impostos["cliente"] = {
            "cliente_codigo": r.get("cliente_codigo"),
            "cliente_nome": r.get("cliente_nome"),
            "cliente_uf": r.get("cliente_uf"),
            "cliente_situacao": r.get("cliente_situacao"),
            "cliente_endereco": r.get("cliente_endereco"),
            "cliente_complemento": r.get("cliente_complemento"),
            "cliente_cep": r.get("cliente_cep"),
            "cliente_cidade": r.get("cliente_cidade"),
            "cliente_bairro": r.get("cliente_bairro"),
            "cliente_redsai_pis": r.get("cliente_redsai_pis"),
            "cliente_redsai_cofins": r.get("cliente_redsai_cofins"),
            "motivos": cliente_motivos,
            "divergencias_reais": cliente_divergencias,
            "avisos_cadastrais": cliente_avisos,
        }
        divergencias_reais.extend(cliente_divergencias)
        avisos_cadastrais.extend(cliente_avisos)

        fornecedor_motivos = []
        fornecedor_divergencias = []
        fornecedor_avisos = []
        if r.get("movimento") == "ENTRADA":
            if not r.get("fornecedor_codigo"):
                fornecedor_divergencias.append("NF de entrada sem fornecedor associado")
            if not _clean_str(r.get("fornecedor_uf")):
                fornecedor_divergencias.append("Fornecedor sem UF cadastrada")
            if _clean_str(r.get("fornecedor_situacao")) not in ("", "A"):
                fornecedor_divergencias.append("Fornecedor inativo")
            if r.get("tipo_item") == "SERVICO":
                if not _clean_str(r.get("fornecedor_codtri")):
                    fornecedor_avisos.append("Fornecedor sem código de tributação (CodTri) para análise fiscal do serviço")
                tns_srv = _clean_str(r.get("transacao")).upper()
                tipfor = _clean_str(r.get("fornecedor_tipfor")).upper()
                if tns_srv in {"1933F", "2933F"} and tipfor == "J":
                    fornecedor_divergencias.append("Transação de serviço PF usada para fornecedor PJ")
                if tns_srv in {"1933", "2933", "1933A", "2933A"} and tipfor == "F":
                    fornecedor_divergencias.append("Transação de serviço PJ usada para fornecedor PF")
        fornecedor_motivos = fornecedor_divergencias + fornecedor_avisos
        impostos["fornecedor"] = {
            "fornecedor_codigo": r.get("fornecedor_codigo"),
            "fornecedor_nome": r.get("fornecedor_nome"),
            "fornecedor_uf": r.get("fornecedor_uf"),
            "fornecedor_codtri": r.get("fornecedor_codtri"),
            "fornecedor_tipfor": r.get("fornecedor_tipfor"),
            "fornecedor_situacao": r.get("fornecedor_situacao"),
            "motivos": fornecedor_motivos,
            "divergencias_reais": fornecedor_divergencias,
            "avisos_cadastrais": fornecedor_avisos,
        }
        divergencias_reais.extend(fornecedor_divergencias)
        avisos_cadastrais.extend(fornecedor_avisos)

        cadastro_produto_motivos = []
        if r.get("tipo_item") == "PRODUTO":
            # Valores efetivos por camada (string normalizada)
            item_cst_pis = _clean_str(r.get("item_cst_pis"))
            tns_cst_pis = _clean_str(r.get("tns_cst_pis"))
            fam_cst_pis_v = _clean_str(r.get("fam_cst_pis"))
            cad_cst_pis = _clean_str(r.get("cad_cstpis_produto"))

            item_cst_cofins = _clean_str(r.get("item_cst_cofins"))
            tns_cst_cofins = _clean_str(r.get("tns_cst_cofins"))
            fam_cst_cofins_v = _clean_str(r.get("fam_cst_cofins"))
            cad_cst_cofins = _clean_str(r.get("cad_cstcof_produto"))

            item_cst_ipi = _clean_str(r.get("item_cst_ipi"))
            fam_cst_ipi_v = _clean_str(r.get("fam_cst_ipi"))
            cad_cst_ipi = _clean_str(r.get("cad_cstipi_produto"))

            # ----- CST PIS (hierarquia Item NF > Transação > Família > Cadastro) -----
            if not cad_cst_pis:
                if item_cst_pis or tns_cst_pis:
                    # Parametrização efetiva veio do item/transação. Não apontar nada.
                    pass
                elif fam_cst_pis_v:
                    cadastro_produto_motivos.append(
                        f"Produto sem CST PIS no cadastro, porém família possui CST PIS {fam_cst_pis_v}"
                    )
                else:
                    divergencias_reais.append(
                        "Produto sem CST PIS no cadastro e sem CST PIS efetivo no item/transação/família"
                    )

            # ----- CST COFINS -----
            if not cad_cst_cofins:
                if item_cst_cofins or tns_cst_cofins:
                    pass
                elif fam_cst_cofins_v:
                    cadastro_produto_motivos.append(
                        f"Produto sem CST COFINS no cadastro, porém família possui CST COFINS {fam_cst_cofins_v}"
                    )
                else:
                    divergencias_reais.append(
                        "Produto sem CST COFINS no cadastro e sem CST COFINS efetivo no item/transação/família"
                    )

            # ----- CST IPI (mesma lógica, sem camada de transação na maioria dos casos) -----
            if not cad_cst_ipi:
                if item_cst_ipi:
                    pass
                elif fam_cst_ipi_v:
                    cadastro_produto_motivos.append(
                        f"Produto sem CST IPI no cadastro, porém família possui CST IPI {fam_cst_ipi_v}"
                    )
                # Sem CST IPI em camada nenhuma para produto: tratamos como aviso (depende do CFOP/NCM)
                else:
                    cadastro_produto_motivos.append(
                        "Produto sem CST IPI no cadastro (verificar se a operação exige IPI)"
                    )

            # ----- Tipos de tributação PIS/COFINS (TriPIS/TriCOF) -----
            # Só vira aviso se também não houver CST efetiva pela camada de operação.
            if not _clean_str(r.get("cad_tripis")) and not (item_cst_pis or tns_cst_pis):
                cadastro_produto_motivos.append("Cadastro do produto sem tipo de tributação de PIS")
            if not _clean_str(r.get("cad_tricof")) and not (item_cst_cofins or tns_cst_cofins):
                cadastro_produto_motivos.append("Cadastro do produto sem tipo de tributação de COFINS")

            # ----- Natureza de receita PIS/COFINS -----
            # Só faz sentido em saídas. Em entradas, não apontar.
            if r.get("movimento") == "SAIDA":
                if not _clean_str(r.get("cad_natpis")):
                    cadastro_produto_motivos.append("Cadastro do produto sem natureza de receita PIS")
                if not _clean_str(r.get("cad_natcof")):
                    cadastro_produto_motivos.append("Cadastro do produto sem natureza de receita COFINS")

            # ----- Demais campos cadastrais: continuam como avisos cadastrais (saneamento) -----
            if not _clean_str(r.get("cad_codstr")):
                cadastro_produto_motivos.append("Cadastro do produto sem Código de Situação/estratégia ICMS (CodStr)")
            if not _clean_str(r.get("cad_codtic")):
                cadastro_produto_motivos.append("Cadastro do produto sem CodTic")
            if not _clean_str(r.get("cad_codstc")):
                cadastro_produto_motivos.append("Cadastro do produto sem CodSTC")
            if r.get("cad_basrec") is None:
                cadastro_produto_motivos.append("Cadastro do produto sem base de recuperação (BasRec)")
            if r.get("cad_bascre_produto") is None:
                cadastro_produto_motivos.append("Cadastro do produto sem base de crédito (BasCre)")
            if not _clean_str(r.get("cad_regtri")):
                cadastro_produto_motivos.append("Cadastro do produto sem regime tributário")
            if not _clean_str(r.get("cad_codenq")):
                cadastro_produto_motivos.append("Cadastro do produto sem código de enquadramento")
            if not _clean_str(r.get("cad_codces")):
                cadastro_produto_motivos.append("Cadastro do produto sem CEST/código correspondente")
        avisos_cadastrais.extend(cadastro_produto_motivos)

        iss_motivos = []
        if r.get("tipo_item") == "SERVICO":
            if r.get("item_base_iss") is None and r.get("item_valor_iss") is None:
                iss_motivos.append("Mapear campos de ISS/ISSQN na base para auditoria completa de serviços")
        impostos["iss"] = {
            "item_base_iss": r.get("item_base_iss"), "item_aliq_iss": r.get("item_aliq_iss"),
            "item_valor_iss": r.get("item_valor_iss"), "municipio_iss": r.get("municipio_iss"),
            "iss_retido": r.get("iss_retido"), "motivos": iss_motivos,
        }
        pendencias_mapeamento.extend(iss_motivos)

        ret_motivos = []
        if r.get("tipo_item") == "SERVICO":
            if (r.get("item_valor_irrf") is None and r.get("item_valor_csll") is None and
                    r.get("item_valor_pis_ret") is None and r.get("item_valor_cofins_ret") is None):
                ret_motivos.append("Mapear IRRF/CSLL/PIS/COFINS retidos na base para auditoria completa")
        impostos["retencoes"] = {
            "item_valor_inss": r.get("item_valor_inss"), "item_valor_irrf": r.get("item_valor_irrf"),
            "item_valor_csll": r.get("item_valor_csll"), "item_valor_pis_ret": r.get("item_valor_pis_ret"),
            "item_valor_cofins_ret": r.get("item_valor_cofins_ret"), "motivos": ret_motivos,
        }
        pendencias_mapeamento.extend(ret_motivos)

        difal_motivos = []
        if r.get("movimento") == "SAIDA" and r.get("cliente_uf"):
            if (r.get("item_base_difal") is None and r.get("item_valor_difal") is None and
                    r.get("item_valor_fcp") is None and r.get("item_valor_fcp_st") is None):
                difal_motivos.append("Mapear DIFAL/FCP/FCP-ST na base para auditoria interestadual completa")
        impostos["difal_fcp"] = {
            "item_base_difal": r.get("item_base_difal"), "item_valor_difal": r.get("item_valor_difal"),
            "item_valor_fcp": r.get("item_valor_fcp"), "item_valor_fcp_st": r.get("item_valor_fcp_st"),
            "motivos": difal_motivos,
        }
        pendencias_mapeamento.extend(difal_motivos)

        # Verificacoes especificas para auditoria cadastral (sem movimento) - todas viram avisos cadastrais
        if r.get("origem_auditoria") == "CADASTRO":
            if not _clean_str(r.get("ncm")):
                avisos_cadastrais.append("[Cadastral] Produto sem NCM")
            if not _clean_str(r.get("familia_codigo")):
                avisos_cadastrais.append("[Cadastral] Produto sem família fiscal")
            if not _clean_str(r.get("origem_codigo")):
                avisos_cadastrais.append("[Cadastral] Produto sem origem")
            if not _clean_str(r.get("cad_codtrd")):
                avisos_cadastrais.append("[Cadastral] Produto sem CodTrd (tratamento diferencial ICMS)")
            if r.get("cad_temicm") is None:
                avisos_cadastrais.append("[Cadastral] Produto sem indicador de incidência de ICMS")
            if r.get("cad_recpis") is None:
                avisos_cadastrais.append("[Cadastral] Produto sem flag de recuperação de PIS")
            if r.get("cad_reccof") is None:
                avisos_cadastrais.append("[Cadastral] Produto sem flag de recuperação de COFINS")
            if not _clean_str(r.get("fam_cst_pis")):
                avisos_cadastrais.append("[Cadastral] Família sem CST PIS")
            if not _clean_str(r.get("fam_cst_cofins")):
                avisos_cadastrais.append("[Cadastral] Família sem CST COFINS")

        # Divergencias familia x cadastro do produto
        familia_vs_cadastro_motivos = []
        if r.get("tipo_item") == "PRODUTO":
            if is_str_diff(r.get("fam_codori"), r.get("origem_codigo")):
                familia_vs_cadastro_motivos.append("Família com origem diferente da origem do produto")
            if is_str_diff(r.get("fam_codclf"), r.get("cod_classificacao")):
                familia_vs_cadastro_motivos.append("Família com classificação fiscal diferente do produto")
            if is_str_diff(r.get("fam_recipi"), r.get("cad_recipi")):
                familia_vs_cadastro_motivos.append("Família com Recupera IPI diferente do cadastro do produto")
            if is_str_diff(r.get("fam_reccof"), r.get("cad_reccof")):
                familia_vs_cadastro_motivos.append("Família com Recupera COFINS diferente do cadastro do produto")
            if is_str_diff(r.get("fam_temicm"), r.get("cad_temicm")):
                familia_vs_cadastro_motivos.append("Família com incidência ICMS diferente do cadastro do produto")
            if is_str_diff(r.get("fam_codtrd"), r.get("cad_codtrd")):
                familia_vs_cadastro_motivos.append("Família com CodTrd diferente do cadastro do produto")
            if is_str_diff(r.get("fam_codtst"), r.get("cad_codtst")):
                familia_vs_cadastro_motivos.append("Família com CodTST diferente do cadastro do produto")
            if is_str_diff(r.get("fam_codstp"), r.get("cad_codstp")):
                familia_vs_cadastro_motivos.append("Família com CodSTP diferente do cadastro do produto")
            if is_str_diff(r.get("fam_recicm"), r.get("cad_recicm")):
                familia_vs_cadastro_motivos.append("Família com Recupera ICMS diferente do cadastro do produto")
            if is_str_diff(r.get("fam_recpis"), r.get("cad_recpis")):
                familia_vs_cadastro_motivos.append("Família com Recupera PIS diferente do cadastro do produto")
        avisos_cadastrais.extend(familia_vs_cadastro_motivos)

        # Divergencias familia x item gravado na NF (com hierarquia Item NF > Transacao > Familia > Cadastro)
        familia_vs_item_divergencias = []
        familia_vs_item_avisos = []
        if r.get("tipo_item") == "PRODUTO" and r.get("movimento") in ("ENTRADA", "SAIDA"):
            fam_pis, item_pis, tns_pis = r.get("fam_cst_pis"), r.get("item_cst_pis"), r.get("tns_cst_pis")
            if fam_pis is not None and item_pis is not None and is_str_diff(fam_pis, item_pis):
                if tns_pis is not None and not is_str_diff(tns_pis, item_pis):
                    familia_vs_item_avisos.append(
                        f"CST PIS da família ({fam_pis}) difere do item ({item_pis}), mas o item está coerente com a transação ({tns_pis})"
                    )
                else:
                    familia_vs_item_divergencias.append(f"CST PIS da família ({fam_pis}) difere do item NF ({item_pis})")

            fam_cof, item_cof, tns_cof = r.get("fam_cst_cofins"), r.get("item_cst_cofins"), r.get("tns_cst_cofins")
            if fam_cof is not None and item_cof is not None and is_str_diff(fam_cof, item_cof):
                if tns_cof is not None and not is_str_diff(tns_cof, item_cof):
                    familia_vs_item_avisos.append(
                        f"CST COFINS da família ({fam_cof}) difere do item ({item_cof}), mas o item está coerente com a transação ({tns_cof})"
                    )
                else:
                    familia_vs_item_divergencias.append(f"CST COFINS da família ({fam_cof}) difere do item NF ({item_cof})")

            fam_ipi, item_ipi = r.get("fam_cst_ipi"), r.get("item_cst_ipi")
            if fam_ipi is not None and item_ipi is not None and is_str_diff(fam_ipi, item_ipi):
                familia_vs_item_avisos.append(f"CST IPI da família ({fam_ipi}) difere do item NF ({item_ipi})")

            fam_peripi = r.get("fam_peripi")
            if fam_peripi is not None and r.get("item_aliq_ipi") is not None and is_num_diff(fam_peripi, r.get("item_aliq_ipi")):
                familia_vs_item_avisos.append(f"Alíquota IPI da família ({fam_peripi}%) difere da aplicada no item ({r.get('item_aliq_ipi')}%)")
            fam_pericm = r.get("fam_pericm")
            if fam_pericm is not None and r.get("item_aliq_icms") is not None and is_num_diff(fam_pericm, r.get("item_aliq_icms")):
                familia_vs_item_avisos.append(f"Alíquota ICMS da família ({fam_pericm}%) difere da aplicada no item ({r.get('item_aliq_icms')}%)")
            if r.get("fam_codstr") is not None and r.get("item_cst_icms") is not None and is_str_diff(r.get("fam_codstr"), r.get("item_cst_icms")):
                familia_vs_item_avisos.append(f"Estratégia ICMS da família ({r.get('fam_codstr')}) difere da aplicada no item ({r.get('item_cst_icms')})")
        divergencias_reais.extend(familia_vs_item_divergencias)
        avisos_cadastrais.extend(familia_vs_item_avisos)
        familia_vs_item_motivos = familia_vs_item_divergencias + familia_vs_item_avisos

        operacao_motivos = []
        operacao_divergencias = []
        operacao_pendencias = []

        transacao_atual = _clean_str(r.get("transacao"))
        cfop_atual = _clean_str(r.get("cfop"))
        natureza_operacao_atual = _clean_str(r.get("natureza_operacao"))

        if r.get("movimento") not in (None, "SEM_MOVIMENTO") and not transacao_atual:
            operacao_divergencias.append("Documento sem transação fiscal")

        if transacao_atual and r.get("tns_cst_pis") is None and r.get("tns_cst_cofins") is None:
            operacao_divergencias.append("Transação sem CST PIS/COFINS mapeado")

        if transacao_atual and r.get("item_bascre") is not None and r.get("tns_bascre") is None:
            operacao_pendencias.append("Transação sem base de crédito mapeada")

        if r.get("movimento") == "SAIDA" and not cfop_atual:
            operacao_divergencias.append("CFOP da transação não mapeado")

        if r.get("movimento") == "SAIDA" and not natureza_operacao_atual:
            operacao_pendencias.append("Natureza da operação não mapeada")

        if r.get("tipo_item") == "SERVICO" and transacao_atual and not _clean_str(r.get("tns_inss_ref")):
            operacao_pendencias.append("Transação de serviço sem referência de INSS")

        operacao_motivos = operacao_divergencias + operacao_pendencias

        transacao_parametros = {
            "transacao": transacao_atual,
            "movimento": r.get("movimento"),
            "documento_tipo": r.get("documento_tipo"),
            "tns_cst_pis": r.get("tns_cst_pis"),
            "tns_cst_cofins": r.get("tns_cst_cofins"),
            "tns_bascre": r.get("tns_bascre"),
            "cfop": cfop_atual,
            "natureza_operacao": natureza_operacao_atual,
            "tns_inss_ref": _clean_str(r.get("tns_inss_ref")),
        }

        impostos["operacao_fiscal"] = {
            **transacao_parametros,
            "motivos": operacao_motivos,
            "divergencias_reais": operacao_divergencias,
            "pendencias_mapeamento": operacao_pendencias,
        }
        divergencias_reais.extend(operacao_divergencias)
        pendencias_mapeamento.extend(operacao_pendencias)

        class_motivos = []
        class_divergencias = []
        class_pendencias = []
        if r.get("tipo_item") == "PRODUTO":
            if not (r.get("ncm") or "").strip():
                class_divergencias.append("Produto sem NCM")
            if r.get("cest") is None:
                class_pendencias.append("Mapear CEST na base para auditoria fiscal mais completa")
        class_motivos = class_divergencias + class_pendencias
        impostos["classificacao_fiscal"] = {
            "ncm": r.get("ncm"), "classificacao": r.get("cod_classificacao"),
            "cest": r.get("cest"), "familia_codigo": r.get("familia_codigo"),
            "origem_codigo": r.get("origem_codigo"), "motivos": class_motivos,
            "divergencias_reais": class_divergencias,
            "pendencias_mapeamento": class_pendencias,
        }
        divergencias_reais.extend(class_divergencias)
        pendencias_mapeamento.extend(class_pendencias)

        if _clean_str(r.get("transacao")) and (
            r.get("tns_cst_pis") is not None
            or r.get("tns_cst_cofins") is not None
            or r.get("tns_bascre") is not None
            or cfop_atual
            or natureza_operacao_atual
        ):
            fonte = "TRANSACAO"
        elif r.get("cad_recpis") is not None or r.get("cad_reccof") is not None:
            fonte = "CADASTRO"
        elif r.get("fam_cst_pis") is not None or r.get("fam_cst_cofins") is not None:
            fonte = "FAMILIA"
        else:
            fonte = "ITEM"

        def _fonte_imposto(item_v, tns_v, fam_v, cad_v):
            # Hierarquia: ITEM_NF > TRANSACAO > FAMILIA > CADASTRO
            if _clean_str(item_v):
                return "ITEM_NF"
            if _clean_str(tns_v):
                return "TRANSACAO"
            if _clean_str(fam_v):
                return "FAMILIA"
            if _clean_str(cad_v):
                return "CADASTRO"
            return None

        fonte_efetiva = {
            "pis": _fonte_imposto(r.get("item_cst_pis"), r.get("tns_cst_pis"), r.get("fam_cst_pis"), r.get("cad_cstpis_produto")),
            "cofins": _fonte_imposto(r.get("item_cst_cofins"), r.get("tns_cst_cofins"), r.get("fam_cst_cofins"), r.get("cad_cstcof_produto")),
            "ipi": _fonte_imposto(r.get("item_cst_ipi"), None, r.get("fam_cst_ipi"), r.get("cad_cstipi_produto")),
            "icms": _fonte_imposto(r.get("item_cst_icms"), None, r.get("fam_codstr"), r.get("cad_codstr")),
        }

        divergencias_unicas = list(dict.fromkeys([m for m in divergencias_reais if m]))
        avisos_unicos = list(dict.fromkeys([m for m in avisos_cadastrais if m]))
        pendencias_unicas = list(dict.fromkeys([m for m in pendencias_mapeamento if m]))

        # IMPORTANTE: "motivos" (campo legado) agora contém SOMENTE divergências reais.
        # Avisos cadastrais ficam em "avisos_cadastrais"; pendências em "pendencias_mapeamento".
        motivos_unicos = list(divergencias_unicas)
        if divergencias_unicas:
            status = "DIVERGENTE"
        elif avisos_unicos:
            status = "OK_COM_AVISO"
        elif pendencias_unicas:
            status = "PENDENTE_MAPEAMENTO"
        else:
            status = "OK"

        impostos["familia_parametrizacao"] = familia_parametrizacao

        cadastro_produto = {
            "cod_classificacao": r.get("cod_classificacao"), "ncm": r.get("ncm"),
            "cad_recpis": r.get("cad_recpis"), "cad_reccof": r.get("cad_reccof"),
            "cad_peripi": r.get("cad_peripi"), "cad_recipi": r.get("cad_recipi"),
            "cad_temicm": r.get("cad_temicm"), "cad_codtrd": r.get("cad_codtrd"),
            "cad_codtst": r.get("cad_codtst"), "cad_codstp": r.get("cad_codstp"),
            "cad_recicm": r.get("cad_recicm"),
            "cad_codstr": r.get("cad_codstr"), "cad_codtic": r.get("cad_codtic"),
            "cad_codstc": r.get("cad_codstc"), "cad_basrec": r.get("cad_basrec"),
            "cad_bascre_produto": r.get("cad_bascre_produto"),
            "cad_tripis": r.get("cad_tripis"), "cad_tricof": r.get("cad_tricof"),
            "cad_cstipi_produto": r.get("cad_cstipi_produto"),
            "cad_cstpis_produto": r.get("cad_cstpis_produto"),
            "cad_cstcof_produto": r.get("cad_cstcof_produto"),
            "cad_tprpis": r.get("cad_tprpis"), "cad_tprcof": r.get("cad_tprcof"),
            "cad_tpripi": r.get("cad_tpripi"), "cad_regtri": r.get("cad_regtri"),
            "cad_cstipc": r.get("cad_cstipc"), "cad_cstpic": r.get("cad_cstpic"),
            "cad_cstcoc": r.get("cad_cstcoc"), "cad_orimer": r.get("cad_orimer"),
            "cad_natpis": r.get("cad_natpis"), "cad_natcof": r.get("cad_natcof"),
            "cad_tprpii": r.get("cad_tprpii"), "cad_tprcoi": r.get("cad_tprcoi"),
            "cad_perifp": r.get("cad_perifp"), "cad_pdifcp": r.get("cad_pdifcp"),
            "cad_codenq": r.get("cad_codenq"), "cad_codces": r.get("cad_codces"),
            "cad_coddfs": r.get("cad_coddfs"), "cad_origti": r.get("cad_origti"),
            "cad_catpro": r.get("cad_catpro"), "cad_itefis": r.get("cad_itefis"),
            "cad_desfis": r.get("cad_desfis"), "cad_impscf": r.get("cad_impscf"),
            "cad_perdif": r.get("cad_perdif"), "cad_emirec": r.get("cad_emirec"),
            "cad_idepar": r.get("cad_idepar"), "cad_tipcic": r.get("cad_tipcic"),
            "cad_ficcat": r.get("cad_ficcat"), "cad_usu_mcgrcp": r.get("cad_usu_mcgrcp"),
        }
        impostos["cadastro_produto"] = cadastro_produto

        comparativo_camadas = {
            "transacao": {
                "movimento": transacao_parametros["movimento"],
                "documento_tipo": transacao_parametros["documento_tipo"],
                "transacao": transacao_parametros["transacao"],
                "tns_cst_pis": transacao_parametros["tns_cst_pis"],
                "tns_cst_cofins": transacao_parametros["tns_cst_cofins"],
                "tns_bascre": transacao_parametros["tns_bascre"],
                "cfop": transacao_parametros["cfop"],
                "natureza_operacao": transacao_parametros["natureza_operacao"],
                "tns_inss_ref": transacao_parametros["tns_inss_ref"],
            },
            "cadastro": build_enriched_block(r, [
                "cod_classificacao", "ncm",
                "cad_recpis", "cad_reccof", "cad_peripi", "cad_recipi",
                "cad_temicm", "cad_codtrd", "cad_codtst", "cad_codstp", "cad_recicm",
                "cad_codstr", "cad_codtic", "cad_codstc",
                "cad_basrec", "cad_bascre_produto",
                "cad_tripis", "cad_tricof",
                "cad_cstipi_produto", "cad_cstpis_produto", "cad_cstcof_produto",
                "cad_tprpis", "cad_tprcof", "cad_tpripi", "cad_regtri",
                "cad_natpis", "cad_natcof",
                "cad_codenq", "cad_codces", "cad_perdif", "cad_pdifcp",
                "cad_usu_mcgrcp"
            ]),
            "familia": build_enriched_block(r, [
                "familia_codigo", "familia_descricao",
                "fam_cst_pis", "fam_cst_cofins", "fam_cst_ipi",
                "fam_recpis", "fam_reccof", "fam_perpis", "fam_percof", "fam_peripi",
                "fam_pericm", "fam_codstr",
                "fam_codtrd", "fam_codtst", "fam_codstp",
                "fam_temicm", "fam_recicm", "fam_recipi", "fam_proimp",
                "fam_tippro", "fam_codori"
            ]),
            "origem": build_enriched_block(r, [
                "origem_codigo", "origem_descricao",
                "ori_codreg", "ori_codms1", "ori_codms2", "ori_codms3", "ori_codms4", "ori_proimp"
            ]),
            "produto": {"codigo": r.get("codigo_item"), "descricao": r.get("descricao_item"), "derivacao": r.get("derivacao"), "ncm": r.get("ncm"), "classificacao": r.get("cod_classificacao"), "cest": r.get("cest")},
            "cliente": build_enriched_block(r, [
                "cliente_codigo", "cliente_nome", "cliente_uf", "cliente_situacao",
                "cliente_endereco", "cliente_cidade", "cliente_cep", "cliente_bairro",
                "cliente_redsai_pis", "cliente_redsai_cofins"
            ]),
            "fornecedor": build_enriched_block(r, [
                "fornecedor_codigo", "fornecedor_nome", "fornecedor_uf",
                "fornecedor_codtri", "fornecedor_tipfor", "fornecedor_situacao"
            ]),
            "item_gravado": {"cst_pis": r.get("item_cst_pis"), "cst_cofins": r.get("item_cst_cofins"), "base_pis": r.get("item_base_pis"), "base_cofins": r.get("item_base_cofins"), "valor_pis": r.get("item_valor_pis"), "valor_cofins": r.get("item_valor_cofins"), "cst_ipi": r.get("item_cst_ipi"), "aliq_ipi": r.get("item_aliq_ipi"), "base_ipi": r.get("item_base_ipi"), "valor_ipi": r.get("item_valor_ipi"), "aliq_icms": r.get("item_aliq_icms"), "base_icms": r.get("item_base_icms"), "valor_icms": r.get("item_valor_icms"), "valor_icms_st": r.get("item_valor_icms_st"), "base_inss": r.get("item_base_inss"), "valor_inss": r.get("item_valor_inss")}
        }

        item = {
            **r,
            "origem_auditoria": r.get("origem_auditoria", "MOVIMENTO"),
            "situacao_movimento": "SEM_MOVIMENTO" if r.get("origem_auditoria") == "CADASTRO" else "COM_MOVIMENTO",
            "permite_auditoria_documental": r.get("origem_auditoria") != "CADASTRO",
            "permite_auditoria_cadastral": True,
            "parceiro_codigo": r.get("cliente_codigo") or r.get("fornecedor_codigo"),
            "parceiro_nome": r.get("cliente_nome") or r.get("fornecedor_nome"),
            "status_auditoria": status,
            "fonte_prioritaria": fonte,
            "fonte_efetiva": fonte_efetiva,
            "motivos": motivos_unicos,
            "qtd_motivos": len(motivos_unicos),
            "divergencias_reais": divergencias_unicas,
            "qtd_divergencias_reais": len(divergencias_unicas),
            "avisos_cadastrais": avisos_unicos,
            "qtd_avisos_cadastrais": len(avisos_unicos),
            "pendencias_mapeamento": pendencias_unicas,
            "impostos": impostos,
            "comparativo_camadas": comparativo_camadas,
            "trilha_decisao": [
                {"ordem": 1, "camada": "Transação", **transacao_parametros},
                {"ordem": 2, "camada": "Cadastro", **cadastro_produto},
                {"ordem": 3, "camada": "Família", **familia_parametrizacao},
                {"ordem": 4, "camada": "Origem", **origem_parametrizacao},
                {"ordem": 5, "camada": "Produto", "codigo": r.get("codigo_item"), "descricao": r.get("descricao_item"), "derivacao": r.get("derivacao"), "ncm": r.get("ncm"), "cest": r.get("cest")},
                {"ordem": 6, "camada": "Cliente",
                 "codigo": r.get("cliente_codigo"), "nome": r.get("cliente_nome"),
                 "uf": r.get("cliente_uf"), "situacao": r.get("cliente_situacao"),
                 "endereco": r.get("cliente_endereco"), "complemento": r.get("cliente_complemento"),
                 "cep": r.get("cliente_cep"), "cidade": r.get("cliente_cidade"),
                 "bairro": r.get("cliente_bairro"),
                 "redsai_pis": r.get("cliente_redsai_pis"), "redsai_cofins": r.get("cliente_redsai_cofins")},
                {"ordem": 7, "camada": "Fornecedor",
                 "codigo": r.get("fornecedor_codigo"), "nome": r.get("fornecedor_nome"),
                 "uf": r.get("fornecedor_uf"), "codtri": r.get("fornecedor_codtri"),
                 "tipfor": r.get("fornecedor_tipfor"), "situacao": r.get("fornecedor_situacao")},
                {"ordem": 8, "camada": "Item gravado", "cst_pis": r.get("item_cst_pis"), "cst_cofins": r.get("item_cst_cofins"), "base_pis": r.get("item_base_pis"), "base_cofins": r.get("item_base_cofins"), "valor_pis": r.get("item_valor_pis"), "valor_cofins": r.get("item_valor_cofins"), "cst_ipi": r.get("item_cst_ipi"), "aliq_ipi": r.get("item_aliq_ipi"), "aliq_icms": r.get("item_aliq_icms")},
            ]
        }
        item["risco"] = calcular_score_risco(item)
        itens.append(item)

    if apenas_divergencia:
        itens = [x for x in itens if x["status_auditoria"] == "DIVERGENTE"]

    total_pagina = len(itens)
    resumo = {
        "total_itens": total_registros,
        "entradas": len([x for x in itens if x.get("movimento") == "ENTRADA"]),
        "saidas": len([x for x in itens if x.get("movimento") == "SAIDA"]),
        "divergentes": len([x for x in itens if x["status_auditoria"] == "DIVERGENTE"]),
        "ok_com_aviso": len([x for x in itens if x["status_auditoria"] == "OK_COM_AVISO"]),
        "pendente_mapeamento": len([x for x in itens if x["status_auditoria"] == "PENDENTE_MAPEAMENTO"]),
        "ok": len([x for x in itens if x["status_auditoria"] == "OK"]),
        "fontes_transacao": len([x for x in itens if x["fonte_prioritaria"] == "TRANSACAO"]),
        "fontes_cadastro": len([x for x in itens if x["fonte_prioritaria"] == "CADASTRO"]),
        "itens_pagina": total_pagina,
    }

    total_paginas = max(1, (total_registros + tamanho_pagina - 1) // tamanho_pagina) if total_registros else 1

    return {
        "resumo": resumo,
        "itens": itens,
        "pagina": pagina,
        "tamanho_pagina": tamanho_pagina,
        "total_registros": total_registros,
        "total_paginas": total_paginas
    }




# =========================================================
# FASE 2 - CONSULTA TRIBUTÁRIA IA
# =========================================================

class ConsultaTributariaIARequest(BaseModel):
    codigo_produto: Optional[str] = None
    descricao_operacao: str
    uf_origem: str
    uf_destino: str
    tipo_cliente: Optional[str] = None
    finalidade: Optional[str] = None
    transacao: Optional[str] = None
    familia: Optional[str] = None
    origem: Optional[str] = None
    limite_erp: int = 20


def _clean_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _upper(v: Any) -> str:
    return _clean_str(v).upper()


def interpretar_operacao_fiscal(
    descricao_operacao: str,
    uf_origem: str,
    uf_destino: str,
    tipo_cliente: Optional[str],
    finalidade: Optional[str]
) -> Dict[str, Any]:
    texto = _clean_str(descricao_operacao).lower()
    texto_sem_acentos = (
        texto.replace("\u00e1", "a").replace("\u00e0", "a").replace("\u00e3", "a")
             .replace("\u00e2", "a").replace("\u00e9", "e").replace("\u00ea", "e")
             .replace("\u00ed", "i").replace("\u00f3", "o").replace("\u00f4", "o")
             .replace("\u00f5", "o").replace("\u00fa", "u").replace("\u00e7", "c")
    )

    tipo_operacao = "OUTROS"

    if "remessa" in texto_sem_acentos and "conserto" in texto_sem_acentos:
        tipo_operacao = "REMESSA_CONSERTO"
    elif "retorno" in texto_sem_acentos and "conserto" in texto_sem_acentos:
        tipo_operacao = "RETORNO_CONSERTO"
    elif "conserto" in texto_sem_acentos:
        tipo_operacao = "CONSERTO"
    elif "devolucao" in texto_sem_acentos or "devolucao" in texto:
        tipo_operacao = "DEVOLUCAO"
    elif "transferencia" in texto_sem_acentos or "transferencia" in texto:
        tipo_operacao = "TRANSFERENCIA"
    elif "demonstracao" in texto_sem_acentos or "demonstracao" in texto:
        tipo_operacao = "DEMONSTRACAO"
    elif "industrializacao" in texto_sem_acentos or "industrializacao" in texto:
        tipo_operacao = "INDUSTRIALIZACAO"
    elif "retorno" in texto_sem_acentos:
        tipo_operacao = "RETORNO"
    elif "remessa" in texto_sem_acentos:
        tipo_operacao = "REMESSA"
    elif "venda" in texto_sem_acentos:
        tipo_operacao = "VENDA"
    elif "compra" in texto_sem_acentos or "aquisicao" in texto_sem_acentos:
        tipo_operacao = "COMPRA"

    natureza = "INTERNA" if _upper(uf_origem) == _upper(uf_destino) else "INTERESTADUAL"

    movimento = "SAIDA"
    if tipo_operacao in {"COMPRA"}:
        movimento = "ENTRADA"
    elif any(x in texto_sem_acentos for x in ["entrada", "compra", "aquisicao"]):
        movimento = "ENTRADA"

    confianca = "MEDIA"
    if tipo_operacao != "OUTROS" and _clean_str(uf_origem) and _clean_str(uf_destino):
        confianca = "ALTA"

    return {
        "descricao_original": descricao_operacao,
        "descricao_normalizada": texto_sem_acentos,
        "tipo_operacao": tipo_operacao,
        "natureza": natureza,
        "movimento": movimento,
        "uf_origem": _upper(uf_origem),
        "uf_destino": _upper(uf_destino),
        "tipo_cliente": _upper(tipo_cliente),
        "finalidade": _upper(finalidade),
        "confianca_interpretacao": confianca
    }


def buscar_contexto_erp_para_consulta_tributaria(payload: ConsultaTributariaIARequest, movimento: str) -> Dict[str, Any]:
    limite = max(1, min(payload.limite_erp or 20, 100))
    return _auditoria_tributaria_inner(
        tipo_item="TODOS",
        movimento=movimento,
        numero_documento=None,
        serie=None,
        parceiro=None,
        codigo_item=None,
        codigo_produto=payload.codigo_produto,
        descricao=None,
        familia=payload.familia,
        origem=payload.origem,
        transacao=payload.transacao,
        data_emissao_ini=None,
        data_emissao_fim=None,
        categoria_material=None,
        base_auditoria="MOVIMENTOS",
        apenas_divergencia=False,
        pagina=1,
        tamanho_pagina=limite
    )


def extrair_item_referencia_erp(contexto_erp: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    itens = contexto_erp.get("itens", []) if isinstance(contexto_erp, dict) else []
    if not itens:
        return None

    codigo_buscado = _clean_str(contexto_erp.get("consulta_codigo_produto"))
    if codigo_buscado:
        for item in itens:
            if _clean_str(item.get("codigo_item")) == codigo_buscado:
                return item

    for item in itens:
        if _upper(item.get("movimento")) == "SAIDA":
            return item

    return itens[0]


def buscar_regra_fiscal_sugerida(
    interpretacao: Dict[str, Any],
    payload: ConsultaTributariaIARequest,
    item_erp: Optional[Dict[str, Any]]
) -> Optional[Dict[str, Any]]:
    ncm = ""
    familia = _clean_str(payload.familia)
    origem_prod = _clean_str(payload.origem)
    transacao = _clean_str(payload.transacao)

    if item_erp:
        ncm = _clean_str(item_erp.get("ncm"))
        if not familia:
            familia = _clean_str(item_erp.get("familia_codigo"))
        if not origem_prod:
            origem_prod = _clean_str(item_erp.get("origem_codigo"))
        if not transacao:
            transacao = _clean_str(item_erp.get("transacao"))

    conn = get_connection()
    cursor = conn.cursor()

    sql = """
        SELECT TOP 1
            ID,
            UF_ORIGEM,
            UF_DESTINO,
            TIPO_OPERACAO,
            NATUREZA,
            TIPO_CLIENTE,
            FINALIDADE,
            NCM,
            FAMILIA,
            ORIGEM_PRODUTO,
            TRANSACAO,
            CST_ICMS,
            CFOP,
            BENEFICIO_FISCAL,
            BASE_LEGAL,
            OBSERVACOES,
            PRIORIDADE
        FROM USU_TBTRIB_REGRA
        WHERE ATIVO = 1
          AND UF_ORIGEM = ?
          AND UF_DESTINO = ?
          AND TIPO_OPERACAO = ?
          AND NATUREZA = ?
          AND (TIPO_CLIENTE = ? OR TIPO_CLIENTE IS NULL OR TIPO_CLIENTE = '')
          AND (FINALIDADE = ? OR FINALIDADE IS NULL OR FINALIDADE = '')
          AND (NCM = ? OR NCM IS NULL OR NCM = '')
          AND (FAMILIA = ? OR FAMILIA IS NULL OR FAMILIA = '')
          AND (ORIGEM_PRODUTO = ? OR ORIGEM_PRODUTO IS NULL OR ORIGEM_PRODUTO = '')
          AND (TRANSACAO = ? OR TRANSACAO IS NULL OR TRANSACAO = '')
        ORDER BY PRIORIDADE ASC, ID ASC
    """

    params = [
        _upper(interpretacao.get("uf_origem")),
        _upper(interpretacao.get("uf_destino")),
        _upper(interpretacao.get("tipo_operacao")),
        _upper(interpretacao.get("natureza")),
        _upper(interpretacao.get("tipo_cliente")),
        _upper(interpretacao.get("finalidade")),
        ncm,
        familia,
        origem_prod,
        transacao,
    ]

    try:
        cursor.execute(sql, params)
        row = cursor.fetchone()
    except Exception:
        # Tabela USU_TBTRIB_REGRA ainda não existe â€” retorna None sem quebrar
        conn.close()
        return None
    conn.close()

    if not row:
        return None

    return {
        "id_regra": row[0],
        "uf_origem": row[1],
        "uf_destino": row[2],
        "tipo_operacao": row[3],
        "natureza": row[4],
        "tipo_cliente": row[5],
        "finalidade": row[6],
        "ncm": row[7],
        "familia": row[8],
        "origem_produto": row[9],
        "transacao": row[10],
        "cst_icms": row[11],
        "cfop": row[12],
        "beneficio_fiscal": row[13],
        "base_legal": row[14],
        "observacoes": row[15],
        "prioridade": row[16],
        "fonte_regra": "USU_TBTRIB_REGRA"
    }


def montar_resumo_erp(item_erp: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not item_erp:
        return {}

    impostos = item_erp.get("impostos", {}) or {}
    opf = impostos.get("operacao_fiscal", {}) or {}
    icms = impostos.get("icms", {}) or {}
    classificacao = impostos.get("classificacao_fiscal", {}) or {}

    return {
        "movimento": item_erp.get("movimento"),
        "documento_tipo": item_erp.get("documento_tipo"),
        "numero_documento": item_erp.get("numero_documento"),
        "serie": item_erp.get("serie"),
        "codigo_item": item_erp.get("codigo_item"),
        "descricao_item": item_erp.get("descricao_item"),
        "familia_codigo": item_erp.get("familia_codigo"),
        "familia_descricao": item_erp.get("familia_descricao"),
        "origem_codigo": item_erp.get("origem_codigo"),
        "origem_descricao": item_erp.get("origem_descricao"),
        "transacao": item_erp.get("transacao") or opf.get("transacao"),
        "cfop": opf.get("cfop"),
        "natureza_operacao": opf.get("natureza_operacao"),
        "tns_cst_pis": item_erp.get("tns_cst_pis") or opf.get("tns_cst_pis"),
        "tns_cst_cofins": item_erp.get("tns_cst_cofins") or opf.get("tns_cst_cofins"),
        "tns_bascre": item_erp.get("tns_bascre") or opf.get("tns_bascre"),
        "tns_inss_ref": item_erp.get("tns_inss_ref") or opf.get("tns_inss_ref"),
        "ncm": item_erp.get("ncm") or classificacao.get("ncm"),
        "cest": item_erp.get("cest") or classificacao.get("cest"),
        "cad_codtrd": item_erp.get("cad_codtrd") or icms.get("cad_codtrd"),
        "item_cst_icms": icms.get("item_cst_icms"),
        "item_aliq_icms": icms.get("item_aliq_icms"),
        "item_base_icms": icms.get("item_base_icms"),
        "item_valor_icms": icms.get("item_valor_icms"),
        "status_auditoria": item_erp.get("status_auditoria"),
        "fonte_prioritaria": item_erp.get("fonte_prioritaria"),
        "fonte_efetiva": item_erp.get("fonte_efetiva", {}),
        "motivos": item_erp.get("motivos", []),
        "divergencias_reais": item_erp.get("divergencias_reais", []),
        "avisos_cadastrais": item_erp.get("avisos_cadastrais", []),
        "pendencias_mapeamento": item_erp.get("pendencias_mapeamento", []),
    }


def comparar_sugestao_com_erp(
    regra: Optional[Dict[str, Any]],
    item_erp: Optional[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    resumo_erp = montar_resumo_erp(item_erp)

    erp_cst = _clean_str(resumo_erp.get("item_cst_icms"))
    erp_cfop = _clean_str(resumo_erp.get("cfop"))
    erp_benef = _clean_str(resumo_erp.get("cad_codtrd"))

    sug_cst = _clean_str((regra or {}).get("cst_icms"))
    sug_cfop = _clean_str((regra or {}).get("cfop"))
    sug_benef = _clean_str((regra or {}).get("beneficio_fiscal"))
    sug_base = _clean_str((regra or {}).get("base_legal"))

    return [
        {
            "campo": "CST_ICMS",
            "erp": erp_cst or None,
            "sugestao": sug_cst or None,
            "divergente": bool(erp_cst and sug_cst and erp_cst != sug_cst)
        },
        {
            "campo": "CFOP",
            "erp": erp_cfop or None,
            "sugestao": sug_cfop or None,
            "divergente": bool(erp_cfop and sug_cfop and erp_cfop != sug_cfop)
        },
        {
            "campo": "BENEFICIO_FISCAL",
            "erp": erp_benef or None,
            "sugestao": sug_benef or None,
            "divergente": bool(erp_benef and sug_benef and erp_benef != sug_benef)
        },
        {
            "campo": "BASE_LEGAL",
            "erp": None,
            "sugestao": sug_base or None,
            "divergente": False
        },
    ]


def classificar_status_consulta(
    regra: Optional[Dict[str, Any]],
    comparativo: List[Dict[str, Any]],
    item_erp: Optional[Dict[str, Any]]
) -> str:
    if not regra:
        return "SEM_REGRA"
    if not item_erp:
        return "REGRA_SEM_ERP"
    if any(x.get("divergente") for x in comparativo):
        return "DIVERGENTE"
    return "ADERENTE"


@app.post("/api/consulta-tributaria-ia")
def consulta_tributaria_ia(
    payload: ConsultaTributariaIARequest,
    usuario=Depends(validar_token)
):
    try:
        descricao = _clean_str(payload.descricao_operacao)
        if len(descricao) < 5:
            raise HTTPException(status_code=400, detail="Informe uma descrição de operação mais completa")

        if not _clean_str(payload.uf_origem):
            raise HTTPException(status_code=400, detail="UF de origem é obrigatória")

        if not _clean_str(payload.uf_destino):
            raise HTTPException(status_code=400, detail="UF de destino é obrigatória")

        interpretacao = interpretar_operacao_fiscal(
            descricao_operacao=payload.descricao_operacao,
            uf_origem=payload.uf_origem,
            uf_destino=payload.uf_destino,
            tipo_cliente=payload.tipo_cliente,
            finalidade=payload.finalidade
        )

        contexto_erp = buscar_contexto_erp_para_consulta_tributaria(
            payload,
            movimento=interpretacao["movimento"]
        )

        if isinstance(contexto_erp, dict):
            contexto_erp["consulta_codigo_produto"] = payload.codigo_produto

        item_erp = extrair_item_referencia_erp(contexto_erp)
        regra = buscar_regra_fiscal_sugerida(interpretacao, payload, item_erp)
        comparativo = comparar_sugestao_com_erp(regra, item_erp)
        status_consulta = classificar_status_consulta(regra, comparativo, item_erp)

        alertas = []
        if not regra:
            alertas.append("Nenhuma regra fiscal encontrada na tabela USU_TBTRIB_REGRA para o cenário informado")
        if not item_erp:
            alertas.append("Nenhum item de referência foi encontrado no ERP com os filtros informados")
        if item_erp:
            pendencias = item_erp.get("pendencias_mapeamento", []) or []
            alertas.extend(pendencias)

        return {
            "status_consulta": status_consulta,
            "consulta": payload.dict(),
            "interpretacao": interpretacao,
            "sugestao_tributaria": regra or {
                "cst_icms": None,
                "cfop": None,
                "beneficio_fiscal": None,
                "base_legal": None,
                "observacoes": "Sem regra encontrada"
            },
            "erp_encontrado": {
                "resumo": (contexto_erp or {}).get("resumo", {}),
                "total_itens": len((contexto_erp or {}).get("itens", [])),
                "item_referencia": montar_resumo_erp(item_erp)
            },
            "comparativo_erp_vs_sugestao": comparativo,
            "alertas": list(dict.fromkeys([a for a in alertas if a]))
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro na consulta tributária IA: {str(e)}")


# =========================================================
# FASE 3 - ENDPOINTS IA AVANÁ‡ADA
# =========================================================

class ParecerIARequest(BaseModel):
    resumo_erp: Dict[str, Any]
    motivos: List[str]
    contexto_extra: Optional[str] = None

@app.post("/api/auditoria/parecer-ia")
def parecer_ia(payload: ParecerIARequest, usuario=Depends(validar_token)):
    """Gera parecer fiscal em linguagem natural para um item auditado."""
    try:
        motivos_txt = "\n".join(f"- {m}" for m in (payload.motivos or []))
        resumo_txt = json.dumps(payload.resumo_erp, ensure_ascii=False, indent=2)
        contexto = f"\nContexto adicional: {payload.contexto_extra}" if payload.contexto_extra else ""

        prompt = f"""Analise os dados fiscais do item abaixo e as divergências encontradas.
Retorne um JSON com exatamente estas chaves:
{{
  "causa_provavel": "explicação clara da causa raiz em 2-3 frases",
  "risco_fiscal": "BAIXO|MEDIO|ALTO|CRITICO",
  "acao_recomendada": "o que a equipe fiscal deve fazer agora",
  "base_legal_sugerida": "lei/instrução normativa relevante ou null",
  "campos_para_corrigir": ["lista de campos ERP que provavelmente precisam ajuste"],
  "confianca": "ALTA|MEDIA|BAIXA"
}}

Dados do item:
{resumo_txt}

Divergências identificadas:
{motivos_txt or "Nenhuma divergência"}
{contexto}"""

        resultado = chamar_gemini(prompt)
        if not resultado:
            return {
                "disponivel": False,
                "motivo": "Gemini indisponível â€” verifique a chave de API",
                "parecer": None
            }
        return {"disponivel": True, "parecer": resultado}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro no parecer IA: {str(e)}")


class PadraoIARequest(BaseModel):
    itens: List[Dict[str, Any]]
    filtros_aplicados: Optional[Dict[str, Any]] = None

@app.post("/api/auditoria/padrao-ia")
def padrao_ia(payload: PadraoIARequest, usuario=Depends(validar_token)):
    """Detecta padrões sistêmicos de divergência fiscal em um conjunto de itens."""
    try:
        itens = payload.itens or []
        if not itens:
            raise HTTPException(status_code=400, detail="Informe ao menos 1 item para análise de padrões")

        # Agrega motivos e riscos
        from collections import Counter
        todos_motivos = []
        riscos: Dict[str, int] = {"CRITICO": 0, "ALTO": 0, "MEDIO": 0, "BAIXO": 0}
        familias: Dict[str, int] = {}
        transacoes: Dict[str, int] = {}

        for it in itens:
            todos_motivos.extend(it.get("motivos", []) or [])
            nivel = (it.get("risco") or {}).get("nivel_risco", "BAIXO")
            riscos[nivel] = riscos.get(nivel, 0) + 1
            fam = _clean_str(it.get("familia_codigo"))
            if fam:
                familias[fam] = familias.get(fam, 0) + 1
            tns = _clean_str(it.get("transacao"))
            if tns:
                transacoes[tns] = transacoes.get(tns, 0) + 1

        contagem = Counter(todos_motivos)
        top_motivos = [{"motivo": m, "ocorrencias": c} for m, c in contagem.most_common(15)]
        top_familias = sorted(familias.items(), key=lambda x: -x[1])[:5]
        top_transacoes = sorted(transacoes.items(), key=lambda x: -x[1])[:5]

        total = len(itens)
        divergentes = sum(1 for it in itens if it.get("status_auditoria") == "DIVERGENTE")

        prompt = f"""Analise os padrões de divergência fiscal abaixo e identifique problemas sistêmicos.
Total de itens: {total} | Divergentes: {divergentes} ({round(divergentes/max(1,total)*100)}%)

Distribuição de risco: {json.dumps(riscos, ensure_ascii=False)}

Top 15 motivos de divergência (motivo: ocorrências):
{json.dumps(top_motivos, ensure_ascii=False, indent=2)}

Famílias com mais ocorrências: {json.dumps(top_familias, ensure_ascii=False)}
Transações com mais ocorrências: {json.dumps(top_transacoes, ensure_ascii=False)}

Retorne um JSON com:
{{
  "resumo_executivo": "2-4 frases descrevendo o estado geral da auditoria",
  "padroes_detectados": [
    {{
      "descricao": "descrição clara do padrão",
      "tipo": "PARAMETRIZACAO_FAMILIA|CADASTRO_PRODUTO|TRANSACAO|REGRA_FISCAL|OUTRO",
      "impacto": "SISTEMICO|RECORRENTE|PONTUAL",
      "itens_afetados": numero,
      "acao_corretiva": "o que fazer para resolver",
      "urgencia": "IMEDIATA|CURTO_PRAZO|PLANEJAMENTO"
    }}
  ],
  "recomendacao_prioritaria": "a ação mais importante a tomar agora"
}}"""

        resultado = chamar_gemini(prompt)
        if not resultado:
            return {
                "disponivel": False,
                "agregado": {
                    "total_itens": total,
                    "divergentes": divergentes,
                    "distribuicao_risco": riscos,
                    "top_motivos": top_motivos
                }
            }

        return {
            "disponivel": True,
            "agregado": {
                "total_itens": total,
                "divergentes": divergentes,
                "distribuicao_risco": riscos,
                "top_motivos": top_motivos,
                "top_familias": [{"familia": k, "ocorrencias": v} for k, v in top_familias],
                "top_transacoes": [{"transacao": k, "ocorrencias": v} for k, v in top_transacoes],
            },
            "analise_ia": resultado
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro na análise de padrões IA: {str(e)}")


class CorrecaoIARequest(BaseModel):
    codigo_produto: str
    descricao_produto: Optional[str] = None
    ncm: Optional[str] = None
    familia_codigo: Optional[str] = None
    origem_codigo: Optional[str] = None
    uf_operacao: Optional[str] = None
    motivos_divergencia: Optional[List[str]] = None
    campos_atual: Optional[Dict[str, Any]] = None

@app.post("/api/produto/sugestao-correcao-ia")
def sugestao_correcao_ia(payload: CorrecaoIARequest, usuario=Depends(validar_token)):
    """Sugere correções nos campos fiscais do cadastro do produto (E075PRO)."""
    try:
        campos_txt = json.dumps(payload.campos_atual or {}, ensure_ascii=False, indent=2)
        motivos_txt = "\n".join(f"- {m}" for m in (payload.motivos_divergencia or []))

        prompt = f"""Analise o cadastro fiscal do produto abaixo e sugira correções nos campos do ERP Senior.

Produto: {payload.codigo_produto} â€” {payload.descricao_produto or "sem descrição"}
NCM: {payload.ncm or "não informado"}
Família: {payload.familia_codigo or "não informada"}
Origem: {payload.origem_codigo or "não informada"}
UF principal de operação: {payload.uf_operacao or "não informada"}

Campos fiscais atuais no ERP (E075PRO):
{campos_txt}

Divergências identificadas na auditoria:
{motivos_txt or "Sem divergências informadas"}

Retorne um JSON com:
{{
  "sugestoes": [
    {{
      "campo_erp": "nome do campo no ERP (ex: CSTPIS)",
      "alias_sistema": "alias no sistema (ex: cad_cstpis_produto)",
      "valor_atual": "valor atual ou null",
      "valor_sugerido": "valor recomendado",
      "justificativa": "fundamentação técnica/legal",
      "base_legal": "lei/IN relevante ou null",
      "confianca": "ALTA|MEDIA|BAIXA"
    }}
  ],
  "observacao_geral": "contexto adicional sobre o produto e sua tributação"
}}"""

        resultado = chamar_gemini(prompt)
        if not resultado:
            return {"disponivel": False, "sugestoes": None}

        return {"disponivel": True, "produto": payload.codigo_produto, "resultado": resultado}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro na sugestão de correção IA: {str(e)}")


class ConsultaNaturalRequest(BaseModel):
    pergunta: str
    contexto: Optional[str] = None

@app.post("/api/auditoria/consulta-natural")
def consulta_natural(payload: ConsultaNaturalRequest, usuario=Depends(validar_token)):
    """Traduz pergunta em linguagem natural para filtros da auditoria e executa a busca."""
    try:
        pergunta = _clean_str(payload.pergunta)
        if len(pergunta) < 8:
            raise HTTPException(status_code=400, detail="Pergunta muito curta")

        schema_filtros = {
            "tipo_item": "PRODUTO|SERVICO|TODOS",
            "movimento": "ENTRADA|SAIDA|TODOS",
            "numero_documento": "número da NF (string) ou null",
            "parceiro": "código do cliente ou fornecedor ou null",
            "codigo_item": "código do produto/serviço ou null",
            "descricao": "parte da descrição do produto ou null",
            "familia": "código da família (ex: MAQU) ou null",
            "origem": "código de origem ou null",
            "transacao": "código da transação ou null",
            "data_emissao_ini": "YYYY-MM-DD ou null",
            "data_emissao_fim": "YYYY-MM-DD ou null",
            "base_auditoria": "MOVIMENTOS|CADASTRO|AMBOS",
            "apenas_divergencia": "true|false"
        }

        prompt = f"""O usuário quer consultar a auditoria tributária fiscal de um ERP.
Traduza a pergunta abaixo para os filtros disponíveis.

Pergunta: "{pergunta}"
{f'Contexto: {payload.contexto}' if payload.contexto else ''}

Filtros disponíveis:
{json.dumps(schema_filtros, ensure_ascii=False, indent=2)}

Retorne SOMENTE um JSON com os filtros aplicáveis (omita campos null):
{{
  "filtros": {{ ... }},
  "interpretacao": "o que você entendeu da pergunta em 1 frase",
  "confianca": "ALTA|MEDIA|BAIXA"
}}"""

        resultado = chamar_gemini(prompt)
        if not resultado:
            return {
                "disponivel": False,
                "interpretacao": "IA indisponível",
                "resultados": None
            }

        filtros = resultado.get("filtros", {})
        # Executa a auditoria com os filtros interpretados
        try:
            resultados = _auditoria_tributaria_inner(
                tipo_item=filtros.get("tipo_item", "TODOS"),
                movimento=filtros.get("movimento", "TODOS"),
                numero_documento=filtros.get("numero_documento"),
                serie=None,
                parceiro=filtros.get("parceiro"),
                codigo_item=filtros.get("codigo_item"),
                codigo_produto=None,
                descricao=filtros.get("descricao"),
                familia=filtros.get("familia"),
                origem=filtros.get("origem"),
                transacao=filtros.get("transacao"),
                data_emissao_ini=filtros.get("data_emissao_ini"),
                data_emissao_fim=filtros.get("data_emissao_fim"),
                categoria_material=None,
                base_auditoria=filtros.get("base_auditoria", "MOVIMENTOS"),
                apenas_divergencia=str(filtros.get("apenas_divergencia", "false")).lower() == "true",
                pagina=1,
                tamanho_pagina=50
            )
        except Exception:
            resultados = None

        return {
            "disponivel": True,
            "pergunta_original": pergunta,
            "interpretacao": resultado.get("interpretacao"),
            "confianca": resultado.get("confianca"),
            "filtros_aplicados": filtros,
            "resultados": resultados
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro na consulta natural IA: {str(e)}")


# =========================================================
# UI
# =========================================================

@app.get("/", response_class=HTMLResponse)
def home():
    return """<!DOCTYPE html>
<html lang="pt-br">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1.0" />
<title>Auditoria Tributária ERP Senior</title>
<style>
    :root{--bg:#081226;--card:#0d1b34;--line:#233a64;--text:#eaf1ff;--muted:#9bb0d3;--primary:#3b82f6;--danger:#ef4444;--ok:#22c55e;}
    *{box-sizing:border-box}
    body{margin:0;font-family:Arial,Helvetica,sans-serif;background:linear-gradient(180deg,#07101f 0%,#081226 100%);color:var(--text);}
    .wrap{max-width:1750px;margin:0 auto;padding:24px}
    .title{font-size:38px;font-weight:700;margin-bottom:6px}
    .sub{color:var(--muted);margin-bottom:20px}
    .card{background:rgba(13,27,52,.96);border:1px solid var(--line);border-radius:18px;padding:18px;box-shadow:0 12px 30px rgba(0,0,0,.18);}
    .grid{display:grid;gap:14px}
    .login-box{max-width:460px;margin:60px auto}
    label{display:block;color:var(--muted);font-size:13px;margin-bottom:6px}
    input,select{width:100%;background:#07101f;color:var(--text);border:1px solid var(--line);border-radius:12px;padding:12px 14px;outline:none;}
    input::placeholder{color:#6f86ad}
    .btn{background:var(--primary);color:white;border:none;border-radius:12px;padding:12px 16px;cursor:pointer;font-weight:700;}
    .btn.secondary{background:#243755}
    .btn.ghost{background:#13213d}
    .row{display:grid;grid-template-columns:repeat(6,1fr);gap:14px;}
    .actions{display:flex;gap:10px;flex-wrap:wrap;margin-top:8px;}
    .cards{display:grid;grid-template-columns:repeat(7,1fr);gap:14px;margin-top:18px;}
    .metric{background:rgba(9,19,40,.95);border:1px solid var(--line);border-radius:18px;padding:16px;}
    .metric .k{color:var(--muted);font-size:13px;margin-bottom:8px;}
    .metric .v{font-size:34px;font-weight:800;}
    .status{display:inline-block;padding:6px 10px;border-radius:999px;font-size:12px;font-weight:700;}
    .status.ok{background:rgba(34,197,94,.15);color:#7ef3a0;border:1px solid rgba(34,197,94,.35)}
    .status.bad{background:rgba(239,68,68,.15);color:#ff9ea2;border:1px solid rgba(239,68,68,.35)}
    .status.ent{background:rgba(59,130,246,.15);color:#93c5fd;border:1px solid rgba(59,130,246,.35)}
    .status.sai{background:rgba(234,179,8,.15);color:#fde047;border:1px solid rgba(234,179,8,.35)}
    table{width:100%;border-collapse:collapse;margin-top:16px;font-size:13px;}
    th,td{padding:12px 10px;border-bottom:1px solid rgba(35,58,100,.7);text-align:left;vertical-align:top;}
    th{color:#afc2e4;font-size:12px;position:sticky;top:0;background:#0d1b34;z-index:1;}
    tbody tr{cursor:pointer;transition:.15s ease}
    tbody tr:hover{background:rgba(255,255,255,.03)}
    .table-wrap{max-height:470px;overflow:auto;border:1px solid var(--line);border-radius:16px;}
    .muted{color:var(--muted)}
    .panel{margin-top:18px;display:grid;grid-template-columns:1.15fr .85fr;gap:16px;}
    .chips{display:flex;flex-wrap:wrap;gap:8px}
    .chip{padding:8px 10px;border-radius:999px;border:1px solid var(--line);background:#0b1830;color:#cddbf4;font-size:12px;}
    .hidden{display:none}
    .topbar{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:18px;}
    .userbox{display:flex;gap:10px;align-items:center}
    .small{font-size:12px}
    .tabs{display:flex;gap:8px;flex-wrap:wrap;margin-top:16px}
    .tab{padding:10px 12px;border-radius:12px;border:1px solid var(--line);background:#0b1830;color:#dbe7ff;cursor:pointer;font-size:13px;font-weight:700;}
    .tab.active{background:#17315f;border-color:#3b82f6}
    .layer-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:12px;margin-top:14px;}
    .layer-card{border:1px solid var(--line);border-radius:14px;padding:12px;background:#0b1830;max-height:340px;overflow:auto;}
    .layer-title{font-weight:700;margin-bottom:8px;color:#dbe7ff;}
    .kv{font-size:12px;line-height:1.6;color:#dbe7ff}
    .kv b{color:#fff}
    .field-ref{color:#5a7aa8;font-size:10px;font-weight:400;margin-left:2px}
    .field-group{display:inline-block;padding:1px 6px;border-radius:6px;font-size:10px;font-weight:700;margin-right:4px;background:#17315f;color:#7eb8ff}
    .field-group-ICMS{background:#1a3a1a;color:#5eda8e}
    .field-group-PIS{background:#2a1a3a;color:#b07ef8}
    @media(max-width:1400px){.cards{grid-template-columns:repeat(4,1fr)}}
    @media(max-width:1200px){.row{grid-template-columns:repeat(3,1fr)}.cards{grid-template-columns:repeat(3,1fr)}.panel{grid-template-columns:1fr}.layer-grid{grid-template-columns:1fr}}
    @media(max-width:700px){.row,.cards{grid-template-columns:1fr}.title{font-size:28px}}
</style>
</head>
<body>
<div class="wrap">
<div id="loginView" class="login-box card">
    <div class="title" style="font-size:30px">Auditoria Tributária</div>
    <div class="sub">Login para consultar divergências fiscais no ERP Senior.</div>
    <div class="grid">
        <div><label>Usuário</label><input id="loginUsuario" placeholder="Ex.: RENATO" value="RENATO" /></div>
        <div><label>Senha</label><input id="loginSenha" type="password" placeholder="Senha" value="123" /></div>
        <div class="actions"><button class="btn" onclick="entrar()">Entrar</button></div>
        <div id="loginMsg" class="muted"></div>
    </div>
</div>

<div id="appView" class="hidden">
    <div class="topbar">
        <div>
            <div class="title">Auditoria Tributária ERP Senior</div>
            <div class="sub">Quadro comparativo: transação, cadastro fiscal, família, origem, produto, cliente, fornecedor e tributação gravada na NF.</div>
        </div>
        <div class="userbox"><span id="userLabel" class="chip"></span><button class="btn secondary" onclick="window.location.href='/controle-fiscal-produtos'">Controle Fiscal Produtos</button><button class="btn secondary" onclick="window.location.href='/inteligencia-tributaria'">Inteligência Tributária</button><button class="btn ghost" onclick="sair()">Sair</button></div>
    </div>

    <div class="card">
        <div class="row">
            <div><label>Movimento</label><select id="f_movimento"><option value="TODOS">Todos</option><option value="ENTRADA">Entrada</option><option value="SAIDA">Saída</option></select></div>
            <div><label>Tipo item</label><select id="f_tipo_item"><option value="TODOS">Todos</option><option value="PRODUTO">Produto</option><option value="SERVICO">Serviço</option></select></div>
            <div><label>Número documento</label><input id="f_numero_documento" placeholder="Ex.: 81502676" /></div>
            <div><label>Série</label><input id="f_serie" placeholder="Ex.: NFE" /></div>
            <div><label>Fornecedor / Cliente</label><input id="f_parceiro" placeholder="Código parceiro" /></div>
            <div><label>Código item</label><input id="f_codigo_item" placeholder="Produto ou serviço" /></div>
        </div>
        <div class="row" style="margin-top:14px">
            <div><label>Código produto</label><input id="f_codigo_produto" placeholder="Ex.: 104004" /></div>
            <div><label>Descrição</label><input id="f_descricao" placeholder="Somente produto" /></div>
            <div><label>Família</label><input id="f_familia" list="listaFamilias" placeholder="Ex.: BR-CHA" /><datalist id="listaFamilias"></datalist></div>
            <div><label>Origem</label><input id="f_origem" list="listaOrigens" placeholder="Ex.: 250" /><datalist id="listaOrigens"></datalist></div>
            <div><label>Transação</label><input id="f_transacao" placeholder="Ex.: 1101E" /></div>
            <div><label>Apenas divergência</label><select id="f_apenas_divergencia"><option value="false">Não</option><option value="true">Sim</option></select></div>
        </div>
        <div class="row" style="margin-top:14px">
            <div><label>Emissão de</label><input id="f_data_emissao_ini" type="date" /></div>
            <div><label>Emissão até</label><input id="f_data_emissao_fim" type="date" /></div>
            <div>
                <label>Base da auditoria</label>
                <select id="f_base_auditoria">
                    <option value="MOVIMENTOS">Movimentações (NFs)</option>
                    <option value="CADASTRO">Cadastro fiscal (sem NF)</option>
                    <option value="AMBOS">Ambos</option>
                </select>
            </div>
            <div>
                <label>Categoria do material</label>
                <select id="f_categoria_material">
                    <option value="">Todos</option>
                    <option value="MATERIA_PRIMA">Matéria-prima fabricação</option>
                    <option value="PRODUTO_PRODUZIDO">Produto produzido</option>
                    <option value="CONSUMO">Material de consumo</option>
                    <option value="IMOBILIZADO">Imobilizado</option>
                    <option value="EPI">EPI</option>
                    <option value="REVENDA">Revenda</option>
                    <option value="SERVICO">Serviço</option>
                    <option value="OUTROS">Outros</option>
                </select>
            </div>
            <div><label>Itens por página</label><select id="f_tamanho_pagina"><option value="50">50</option><option value="100" selected>100</option><option value="200">200</option></select></div>
        </div>
        <div class="actions">
            <button class="btn" onclick="consultar(1)">Consultar</button>
            <button class="btn secondary" onclick="limparFiltros()">Limpar</button>
        </div>
    </div>

    <div class="cards">
        <div class="metric"><div class="k">Itens analisados</div><div class="v" id="m_total">0</div></div>
        <div class="metric"><div class="k">Entradas</div><div class="v" id="m_ent">0</div></div>
        <div class="metric"><div class="k">Saídas</div><div class="v" id="m_sai">0</div></div>
        <div class="metric"><div class="k">Divergentes</div><div class="v" id="m_div">0</div></div>
        <div class="metric"><div class="k">Sem divergência</div><div class="v" id="m_ok">0</div></div>
        <div class="metric"><div class="k">Fonte: transação</div><div class="v" id="m_tns">0</div></div>
        <div class="metric"><div class="k">Fonte: cadastro fiscal</div><div class="v" id="m_cad">0</div></div>
    </div>

    <div class="panel">
        <div class="card">
            <div style="font-size:22px;font-weight:700">Resultado da auditoria</div>
            <div class="muted small" style="margin-top:4px">Clique em uma linha para ver o quadro comparativo, os impostos e a trilha da decisão.</div>
            <div id="paginacaoResultados" style="margin-top:10px"></div>
            <div class="table-wrap">
                <table>
                    <thead><tr><th>Status</th><th>Movto</th><th>Tipo</th><th>Categoria</th><th>Doc</th><th>Seq</th><th>Item</th><th>Descrição</th><th>Transação</th><th>Fonte</th><th>Motivos</th></tr></thead>
                    <tbody id="tbodyResultados"><tr><td colspan="11" class="muted">Nenhum dado carregado.</td></tr></tbody>
                </table>
            </div>
        </div>

        <div class="card">
            <div style="font-size:22px;font-weight:700">Detalhe do item</div>
            <div id="detalheVazio" class="muted" style="margin-top:12px">Selecione um item para visualizar as camadas da decisão tributária.</div>
            <div id="detalheItem" class="hidden">
                <div id="detalheHeader" style="margin-top:14px"></div>
                <div style="margin-top:14px;font-weight:700">Motivos</div>
                <div id="detalheMotivos" class="chips" style="margin-top:10px"></div>
                <div class="tabs">
                    <button class="tab" onclick="trocarTab('tabQuadro',this)">Quadro Comparativo</button>
                    <button class="tab" onclick="trocarTab('tabTrilha',this)">Trilha</button>
                    <button class="tab" onclick="trocarTab('tabFamilia',this)">Família Fiscal</button>
                    <button class="tab" onclick="trocarTab('tabPisCofins',this)">PIS/COFINS</button>
                    <button class="tab" onclick="trocarTab('tabIpi',this)">IPI</button>
                    <button class="tab" onclick="trocarTab('tabIcms',this)">ICMS</button>
                    <button class="tab" onclick="trocarTab('tabSt',this)">ICMS ST</button>
                    <button class="tab" onclick="trocarTab('tabOrigem',this)">Origem</button>
                    <button class="tab" onclick="trocarTab('tabCliente',this)">Cliente</button>
                    <button class="tab" onclick="trocarTab('tabInss',this)">INSS</button>
                    <button class="tab" onclick="trocarTab('tabIss',this)">ISS</button>
                    <button class="tab" onclick="trocarTab('tabRet',this)">Retenções</button>
                    <button class="tab" onclick="trocarTab('tabDifal',this)">DIFAL / FCP</button>
                    <button class="tab" onclick="trocarTab('tabOperacao',this)">Operação Fiscal</button>
                    <button class="tab" onclick="trocarTab('tabClassFiscal',this)">Classificação Fiscal</button>
                </div>
                <div id="tabQuadro" class="tab-pane" style="margin-top:14px"></div>
                <div id="tabTrilha" class="tab-pane hidden" style="margin-top:14px"></div>
                <div id="tabFamilia" class="tab-pane hidden" style="margin-top:14px"></div>
                <div id="tabPisCofins" class="tab-pane hidden" style="margin-top:14px"></div>
                <div id="tabIpi" class="tab-pane hidden" style="margin-top:14px"></div>
                <div id="tabIcms" class="tab-pane hidden" style="margin-top:14px"></div>
                <div id="tabSt" class="tab-pane hidden" style="margin-top:14px"></div>
                <div id="tabOrigem" class="tab-pane hidden" style="margin-top:14px"></div>
                <div id="tabCliente" class="tab-pane hidden" style="margin-top:14px"></div>
                <div id="tabInss" class="tab-pane hidden" style="margin-top:14px"></div>
                <div id="tabIss" class="tab-pane hidden" style="margin-top:14px"></div>
                <div id="tabRet" class="tab-pane hidden" style="margin-top:14px"></div>
                <div id="tabDifal" class="tab-pane hidden" style="margin-top:14px"></div>
                <div id="tabOperacao" class="tab-pane hidden" style="margin-top:14px"></div>
                <div id="tabClassFiscal" class="tab-pane hidden" style="margin-top:14px"></div>
            </div>
        </div>
    </div>
</div>
</div>

<script>
let token=localStorage.getItem("token")||"";
let usuarioLogado=localStorage.getItem("usuario")||"";
let resultados=[];

function showLogin(msg=""){document.getElementById("loginView").classList.remove("hidden");document.getElementById("appView").classList.add("hidden");document.getElementById("loginMsg").innerText=msg;}
function showApp(){document.getElementById("loginView").classList.add("hidden");document.getElementById("appView").classList.remove("hidden");document.getElementById("userLabel").innerText=usuarioLogado||"Usuário";}

async function apiFetch(url,options={}){
    const headers=options.headers||{};
    if(token)headers["Authorization"]="Bearer "+token;
    const response=await fetch(url,{...options,headers});
    const raw=await response.text();
    let data;
    try{data=raw?JSON.parse(raw):{}}catch{data={detail:raw||`HTTP ${response.status}`}}
    if(response.status===401){sair();throw new Error(data.detail||"Sessão expirada");}
    if(!response.ok)throw new Error(data.detail||`Erro ${response.status}`);
    return data;
}

async function entrar(){
    const usuario=document.getElementById("loginUsuario").value.trim();
    const senha=document.getElementById("loginSenha").value.trim();
    const msg=document.getElementById("loginMsg");
    msg.innerText="Entrando...";
    try{
        const resp=await fetch("/login",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({usuario,senha})});
        const raw=await resp.text();
        let data;try{data=raw?JSON.parse(raw):{}}catch{data={detail:raw||`HTTP ${resp.status}`}}
        if(!resp.ok)throw new Error(data.detail||"Erro login");
        token=data.access_token;usuarioLogado=data.usuario;
        localStorage.setItem("token",token);localStorage.setItem("usuario",usuarioLogado);
        showApp();await carregarCombos();
    }catch(e){msg.innerText=e.message;}
}

function sair(){
    token="";usuarioLogado="";
    localStorage.removeItem("token");localStorage.removeItem("usuario");
    resultados=[];renderTabela([]);
    renderResumo({total_itens:0,entradas:0,saidas:0,divergentes:0,ok:0,fontes_transacao:0,fontes_cadastro:0});
    document.getElementById("detalheItem").classList.add("hidden");
    document.getElementById("detalheVazio").classList.remove("hidden");
    showLogin("Faça login novamente.");
}

async function carregarCombos(){
    try{
        const [familias,origens]=await Promise.all([apiFetch("/api/familias?limite=200"),apiFetch("/api/origens?limite=200")]);
        document.getElementById("listaFamilias").innerHTML=familias.map(f=>`<option value="${f.codigo}">${f.label}</option>`).join("");
        document.getElementById("listaOrigens").innerHTML=origens.map(o=>`<option value="${o.codigo}">${o.label}</option>`).join("");
    }catch(e){console.error(e);}
}

function limparFiltros(){
    ["f_movimento","f_tipo_item","f_numero_documento","f_serie","f_parceiro","f_codigo_item","f_codigo_produto","f_descricao","f_familia","f_origem","f_transacao","f_data_emissao_ini","f_data_emissao_fim","f_categoria_material","f_apenas_divergencia"].forEach(id=>{
        const el=document.getElementById(id);
        if(!el) return;
        if(id==="f_movimento"||id==="f_tipo_item")el.value="TODOS";
        else if(id==="f_apenas_divergencia")el.value="false";
        else el.value="";
    });
    const tp=document.getElementById("f_tamanho_pagina");if(tp)tp.value="100";
    const ba=document.getElementById("f_base_auditoria");if(ba)ba.value="MOVIMENTOS";
    paginaAtual=1;
    resultados=[];renderTabela([]);
    renderResumo({total_itens:0,entradas:0,saidas:0,divergentes:0,ok:0,fontes_transacao:0,fontes_cadastro:0});
    const box=document.getElementById("paginacaoResultados");if(box)box.innerHTML="";
    document.getElementById("detalheItem").classList.add("hidden");
    document.getElementById("detalheVazio").classList.remove("hidden");
}

let paginaAtual=1;

async function consultar(pagina=1){
    paginaAtual=pagina;
    const params=new URLSearchParams();
    const campos={movimento:"f_movimento",tipo_item:"f_tipo_item",numero_documento:"f_numero_documento",serie:"f_serie",parceiro:"f_parceiro",codigo_item:"f_codigo_item",codigo_produto:"f_codigo_produto",descricao:"f_descricao",familia:"f_familia",origem:"f_origem",transacao:"f_transacao",data_emissao_ini:"f_data_emissao_ini",data_emissao_fim:"f_data_emissao_fim",categoria_material:"f_categoria_material",base_auditoria:"f_base_auditoria",apenas_divergencia:"f_apenas_divergencia",tamanho_pagina:"f_tamanho_pagina"};
    Object.entries(campos).forEach(([k,id])=>{const el=document.getElementById(id);if(!el)return;const v=el.value;if(v!=="")params.append(k,v);});
    params.append("pagina",String(paginaAtual));
    renderTabela([]);
    document.getElementById("tbodyResultados").innerHTML=`<tr><td colspan="10" class="muted">Consultando página ${paginaAtual}...</td></tr>`;
    try{
        const data=await apiFetch("/api/auditoria-tributaria?"+params.toString());
        resultados=data.itens||[];renderResumo(data.resumo||{});renderTabela(resultados);renderPaginacao(data);
        document.getElementById("detalheItem").classList.add("hidden");
        document.getElementById("detalheVazio").classList.remove("hidden");
    }catch(e){document.getElementById("tbodyResultados").innerHTML=`<tr><td colspan="10" class="muted">${e.message}</td></tr>`;}
}

function renderPaginacao(data){
    const box=document.getElementById("paginacaoResultados");if(!box)return;
    const pg=data.pagina||1,total=data.total_paginas||1,regs=data.total_registros||0;
    box.innerHTML=`
        <div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:10px">
            <span class="muted small">Total: <strong>${regs}</strong> registro(s) &nbsp;|&nbsp; Página <strong>${pg}</strong> de <strong>${total}</strong></span>
            <div style="display:flex;gap:8px;flex-wrap:wrap">
                <button class="btn ghost" style="padding:8px 14px" ${pg<=1?"disabled":""} onclick="consultar(${pg-1})">&#8592; Anterior</button>
                <button class="btn ghost" style="padding:8px 14px" ${pg>=total?"disabled":""} onclick="consultar(${pg+1})">Próxima &#8594;</button>
            </div>
        </div>`;
}

function rotuloFonte(f){return{TRANSACAO:"TRANSAÁ‡ÁƒO",CADASTRO:"CADASTRO FISCAL",FAMILIA:"FAMÁLIA",ITEM:"TRIBUTAÁ‡ÁƒO GRAVADA NA NF"}[f]||f||"";}
function rotuloCamada(c){return{"Transação":"Transação","Cadastro":"Cadastro fiscal","Família":"Família","Origem":"Origem","Produto":"Produto","Cliente":"Cliente","Fornecedor":"Fornecedor","Item gravado":"Tributação gravada na NF"}[c]||c||"";}

function renderResumo(r){
    document.getElementById("m_total").innerText=r.total_itens||0;
    document.getElementById("m_ent").innerText=r.entradas||0;
    document.getElementById("m_sai").innerText=r.saidas||0;
    document.getElementById("m_div").innerText=r.divergentes||0;
    document.getElementById("m_ok").innerText=r.ok||0;
    document.getElementById("m_tns").innerText=r.fontes_transacao||0;
    document.getElementById("m_cad").innerText=r.fontes_cadastro||0;
}

function badgeStatus(s){
    if(s==="OK") return `<span class="status ok">OK</span>`;
    if(s==="OK_COM_AVISO") return `<span class="status" style="background:#78350f;border:1px solid #f59e0b;color:#fde68a">OK COM AVISO</span>`;
    if(s==="PENDENTE_MAPEAMENTO") return `<span class="status" style="background:#4c1d95;border:1px solid #7c3aed;color:#d8b4fe">PENDENTE MAPEAMENTO</span>`;
    if(s==="DIVERGENTE") return `<span class="status bad">DIVERGENTE</span>`;
    return `<span class="status">${s||"-"}</span>`;
}
function badgeMovimento(m){return m==="ENTRADA"?`<span class="status ent">ENTRADA</span>`:`<span class="status sai">SAÁDA</span>`;}

function renderTabela(lista){
    const tbody=document.getElementById("tbodyResultados");
    if(!lista.length){tbody.innerHTML=`<tr><td colspan="11" class="muted">Nenhum registro encontrado.</td></tr>`;return;}
    tbody.innerHTML=lista.map((item,idx)=>`
        <tr onclick="abrirDetalhe(${idx})">
            <td>${badgeStatus(item.status_auditoria)}</td>
            <td>${badgeMovimento(item.movimento)}</td>
            <td>${item.tipo_item||""}</td>
            <td><span class="muted small">${item.categoria_material||""}</span></td>
            <td>${item.numero_documento||""}<br><span class="muted small">${item.serie||""} | ${item.documento_tipo||""} | ${item.data_emissao ? String(item.data_emissao).slice(0,10) : ""}</span></td>
            <td>${item.seq_item||""}</td>
            <td>${item.codigo_item||""}<br><span class="muted small">${item.derivacao||""}</span></td>
            <td>${item.descricao_item||""}</td>
            <td>${item.transacao||""}</td>
            <td>${rotuloFonte(item.fonte_prioritaria)}</td>
            <td>${(item.motivos||[]).slice(0,2).map(m=>`<div class="small">${m}</div>`).join("")||'<span class="muted">Sem divergência</span>'}</td>
        </tr>`).join("");
}

function tableRow(label,a,b,c=""){return`<tr><td>${label}</td><td>${a??""}</td><td>${b??""}</td><td>${c??""}</td></tr>`;}
function layerCard(title,obj){
    const entries=Object.entries(obj||{}).filter(([,v])=>{
        if(v===null||v===undefined||v==="")return false;
        if(typeof v==="object"&&"value" in v)return v.value!==null&&v.value!==undefined&&v.value!=="";
        return true;
    });
    const html=entries.length
        ?entries.map(([k,v])=>{
            if(v&&typeof v==="object"&&"value" in v){
                const titulo=v.label||k;
                const origem=v.table&&v.field?` <span class="field-ref">(${v.table}.${v.field})</span>`:"";
                const grupo=v.group?`<span class="field-group">${v.group}</span> `:"";
                return`<div>${grupo}<b>${titulo}</b>${origem}: ${v.value??""}</div>`;
            }
            return`<div><b>${k}</b>: ${v}</div>`;
        }).join("")
        :`<div class="muted">Sem dados</div>`;
    return`<div class="layer-card"><div class="layer-title">${title}</div><div class="kv">${html}</div></div>`;
}
function trocarTab(id,btn){
    document.querySelectorAll(".tab-pane").forEach(el=>el.classList.add("hidden"));
    document.querySelectorAll(".tab").forEach(el=>el.classList.remove("active"));
    document.getElementById(id).classList.remove("hidden");btn.classList.add("active");
}

function abrirDetalhe(index){
    const item=resultados[index];if(!item)return;
    document.getElementById("detalheVazio").classList.add("hidden");
    document.getElementById("detalheItem").classList.remove("hidden");
    const parceiroLinha=item.movimento==="SAIDA"
      ?`Cliente: ${item.cliente_codigo||"-"} ${item.cliente_nome||""} | Situação: ${item.cliente_situacao||"-"} | ${item.cliente_cidade||""} ${item.cliente_uf||""}`
      :`Fornecedor: ${item.fornecedor_codigo||"-"} ${item.fornecedor_nome||""} | Situação: ${item.fornecedor_situacao||"-"} | Tipo: ${item.fornecedor_tipfor||"-"}`;
    document.getElementById("detalheHeader").innerHTML=`
        <div class="chips">
            <span class="chip">${item.movimento||""}</span><span class="chip">${item.documento_tipo||""}</span>
            <span class="chip">${item.tipo_item||""}</span><span class="chip">${item.codigo_item||""}</span>
            <span class="chip">Doc ${item.numero_documento||""}</span><span class="chip">Seq ${item.seq_item||""}</span>
            <span class="chip">${item.categoria_material||""}</span>
            <span class="chip">Fonte ${rotuloFonte(item.fonte_prioritaria)}</span>
        </div>
        <div style="margin-top:10px;font-weight:700">${item.descricao_item||""}</div>
        <div class="muted small" style="margin-top:6px">${parceiroLinha} | Família: ${item.familia_codigo||"-"} ${item.familia_descricao||""} | Origem: ${item.origem_codigo||"-"} ${item.origem_descricao||""} | NCM: ${item.ncm||"-"} | Transação: ${item.transacao||"-"}</div>`;

    const divs=(item.divergencias_reais||[]);
    const avs=(item.avisos_cadastrais||[]);
    const pendsArr=(item.pendencias_mapeamento||[]);
    const chipsD=divs.map(m=>`<span class="chip" style="border-color:#dc2626;color:#fecaca" title="Divergência fiscal real">${m}</span>`);
    const chipsA=avs.map(m=>`<span class="chip" style="border-color:#f59e0b;color:#fde68a" title="Aviso cadastral (saneamento)">${m}</span>`);
    const chipsP=pendsArr.map(m=>`<span class="chip" style="border-color:#7c3aed;color:#d8b4fe" title="Pendência de mapeamento">${m}</span>`);
    const lbl=`<div class="muted small" style="margin-bottom:6px">Divergências reais: <b style="color:#fecaca">${divs.length}</b> | Avisos cadastrais: <b style="color:#fde68a">${avs.length}</b> | Pendências: <b style="color:#d8b4fe">${pendsArr.length}</b></div>`;
    document.getElementById("detalheMotivos").innerHTML=(chipsD.length||chipsA.length||chipsP.length)?lbl+[...chipsD,...chipsA,...chipsP].join(""):`<span class="chip">Sem divergência</span>`;

    const cam=item.comparativo_camadas||{};
    document.getElementById("tabQuadro").innerHTML=`<div class="layer-grid">${["Transação","Cadastro","Família","Origem","Produto","Cliente","Fornecedor","Item gravado"].map(c=>layerCard(rotuloCamada(c),cam[c.toLowerCase().replace(/ /g,"_").replace("ção","cao").replace("ília","ilia").replace("Item gravado","item_gravado")]||cam[{Transação:"transacao",Cadastro:"cadastro",Família:"familia",Origem:"origem",Produto:"produto",Cliente:"cliente",Fornecedor:"fornecedor","Item gravado":"item_gravado"}[c]])).join("")}</div>`;

    document.getElementById("tabTrilha").innerHTML=(item.trilha_decisao||[]).map(t=>`<div style="padding:10px 12px;border:1px solid var(--line);border-radius:12px;background:#0b1830;margin-bottom:8px"><div style="font-weight:700">${t.ordem}. ${rotuloCamada(t.camada)}</div><div class="small muted" style="margin-top:6px;line-height:1.6">${Object.entries(t).filter(([k])=>!["ordem","camada"].includes(k)).map(([k,v])=>`<div>${k}: <strong>${v??"-"}</strong></div>`).join("")}</div></div>`).join("");

    const pis=item.impostos?.pis_cofins||{};
    const fontePis = (item.fonte_efetiva||{}).pis || "-";
    const fonteCof = (item.fonte_efetiva||{}).cofins || "-";
    document.getElementById("tabPisCofins").innerHTML=`<div class="muted small" style="margin-bottom:6px">Fonte efetiva — PIS: <strong>${fontePis}</strong> | COFINS: <strong>${fonteCof}</strong></div><div class="table-wrap" style="max-height:none"><table><thead><tr><th>Campo</th><th>Transação/Família</th><th>Cadastro</th><th>Item NF</th></tr></thead><tbody>${tableRow("CST PIS",pis.tns_cst_pis??pis.fam_cst_pis??"",pis.cad_cstpis_produto??"",pis.item_cst_pis??"")}${tableRow("CST COFINS",pis.tns_cst_cofins??pis.fam_cst_cofins??"",pis.cad_cstcof_produto??"",pis.item_cst_cofins??"")}${tableRow("Recupera PIS","",pis.cad_recpis??"","")}${tableRow("Recupera COFINS","",pis.cad_reccof??"","")}${tableRow("Base crédito",pis.tns_bascre??"","",pis.item_bascre??"")}${tableRow("Base PIS","","",pis.item_base_pis??"")}${tableRow("Base COFINS","","",pis.item_base_cofins??"")}${tableRow("Valor PIS","","",pis.item_valor_pis??"")}${tableRow("Valor COFINS","","",pis.item_valor_cofins??"")}</tbody></table></div>`;

    const ipi=item.impostos?.ipi||{};
    document.getElementById("tabIpi").innerHTML=`<div class="table-wrap" style="max-height:none"><table><thead><tr><th>Campo</th><th>Família</th><th>Cadastro</th><th>Item</th></tr></thead><tbody>${tableRow("CST IPI",ipi.fam_cst_ipi??"","",ipi.item_cst_ipi??"")}${tableRow("ProImp",ipi.fam_proimp??"","","")}${tableRow("Perc. IPI cad","",ipi.cad_peripi??"","")}${tableRow("Recupera IPI","",ipi.cad_recipi??"","")}${tableRow("Alíquota IPI","","",ipi.item_aliq_ipi??"")}${tableRow("Base IPI","","",ipi.item_base_ipi??"")}${tableRow("Valor IPI","","",ipi.item_valor_ipi??"")}</tbody></table></div><div class="chips" style="margin-top:10px">${(ipi.motivos||[]).map(m=>`<span class="chip">${m}</span>`).join("")||'<span class="chip">Sem observações</span>'}</div>`;

    const icms=item.impostos?.icms||{};
    document.getElementById("tabIcms").innerHTML=`<div class="table-wrap" style="max-height:none"><table><thead><tr><th>Campo</th><th>Cadastro</th><th>Item</th><th></th></tr></thead><tbody>${tableRow("Incide ICMS",icms.cad_temicm??"","","")}${tableRow("CodTrd",icms.cad_codtrd??"","","")}${tableRow("Recupera ICMS",icms.cad_recicm??"","","")}${tableRow("Alíquota ICMS","",icms.item_aliq_icms??"","")}${tableRow("Base ICMS","",icms.item_base_icms??"","")}${tableRow("Valor ICMS","",icms.item_valor_icms??"","")}${tableRow("CST ICMS","",icms.item_cst_icms??"","")}</tbody></table></div><div class="chips" style="margin-top:10px">${(icms.motivos||[]).map(m=>`<span class="chip">${m}</span>`).join("")||'<span class="chip">Sem observações</span>'}</div>`;

    const st=item.impostos?.icms_st||{};
    document.getElementById("tabSt").innerHTML=`<div class="table-wrap" style="max-height:none"><table><thead><tr><th>Campo</th><th>Cadastro</th><th>Item</th></tr></thead><tbody>${tableRow("CodTST",st.cad_codtst??"",st.item_cod_tst_st??"","")}${tableRow("CodSTP",st.cad_codstp??"","","")}${tableRow("Valor ICMS ST","",st.item_valor_icms_st??"","")}</tbody></table></div><div class="chips" style="margin-top:10px">${(st.motivos||[]).map(m=>`<span class="chip">${m}</span>`).join("")||'<span class="chip">Sem observações</span>'}</div>`;

    const ori=item.impostos?.origem||{};
    document.getElementById("tabOrigem").innerHTML=`<div class="table-wrap" style="max-height:none"><table><thead><tr><th>Campo</th><th>Valor</th></tr></thead><tbody>${tableRow("Código",ori.origem_codigo??"","","")}${tableRow("Descrição",ori.origem_descricao??"","","")}${tableRow("CodReg",ori.ori_codreg??"","","")}${tableRow("CodMS1",ori.ori_codms1??"","","")}${tableRow("CodMS2",ori.ori_codms2??"","","")}${tableRow("CodMS3",ori.ori_codms3??"","","")}${tableRow("CodMS4",ori.ori_codms4??"","","")}${tableRow("ProImp",ori.ori_proimp??"","","")}</tbody></table></div><div class="chips" style="margin-top:10px">${(ori.motivos||[]).map(m=>`<span class="chip">${m}</span>`).join("")||'<span class="chip">Sem observações</span>'}</div>`;

    const cli=item.impostos?.cliente||{};
    document.getElementById("tabCliente").innerHTML=`<div class="table-wrap" style="max-height:none"><table><thead><tr><th>Campo</th><th>Valor</th></tr></thead><tbody>${tableRow("Código",cli.cliente_codigo??"","","")}${tableRow("Nome",cli.cliente_nome??"","","")}${tableRow("UF",cli.cliente_uf??"","","")}${tableRow("RedSai PIS (41)",cli.cliente_redsai_pis??"","","")}${tableRow("RedSai COFINS (42)",cli.cliente_redsai_cofins??"","","")}</tbody></table></div><div class="chips" style="margin-top:10px">${(cli.motivos||[]).map(m=>`<span class="chip">${m}</span>`).join("")||'<span class="chip">Sem observações</span>'}</div>`;

    const inss=item.impostos?.inss||{};
    document.getElementById("tabInss").innerHTML=`<div class="table-wrap" style="max-height:none"><table><thead><tr><th>Campo</th><th>Transação</th><th>Item</th></tr></thead><tbody>${tableRow("Transação INSS",inss.tns_inss_ref??"","","")}${tableRow("Base INSS","",inss.item_base_inss??"","")}${tableRow("Valor INSS","",inss.item_valor_inss??"","")}</tbody></table></div><div class="chips" style="margin-top:10px">${(inss.motivos||[]).map(m=>`<span class="chip">${m}</span>`).join("")||'<span class="chip">Sem observações</span>'}</div>`;

    const iss=item.impostos?.iss||{};
    document.getElementById("tabIss").innerHTML=`<div class="table-wrap" style="max-height:none"><table><thead><tr><th>Campo</th><th>Valor</th></tr></thead><tbody>${tableRow("Base ISS",iss.item_base_iss??"","","")}${tableRow("Alíquota ISS",iss.item_aliq_iss??"","","")}${tableRow("Valor ISS",iss.item_valor_iss??"","","")}${tableRow("Município ISS",iss.municipio_iss??"","","")}${tableRow("ISS Retido",iss.iss_retido??"","","")}</tbody></table></div><div class="chips" style="margin-top:10px">${(iss.motivos||[]).map(m=>`<span class="chip">${m}</span>`).join("")||'<span class="chip">Sem observações</span>'}</div>`;

    const ret=item.impostos?.retencoes||{};
    document.getElementById("tabRet").innerHTML=`<div class="table-wrap" style="max-height:none"><table><thead><tr><th>Campo</th><th>Valor</th></tr></thead><tbody>${tableRow("INSS",ret.item_valor_inss??"","","")}${tableRow("IRRF",ret.item_valor_irrf??"","","")}${tableRow("CSLL",ret.item_valor_csll??"","","")}${tableRow("PIS Retido",ret.item_valor_pis_ret??"","","")}${tableRow("COFINS Retido",ret.item_valor_cofins_ret??"","","")}</tbody></table></div><div class="chips" style="margin-top:10px">${(ret.motivos||[]).map(m=>`<span class="chip">${m}</span>`).join("")||'<span class="chip">Sem observações</span>'}</div>`;

    const difal=item.impostos?.difal_fcp||{};
    document.getElementById("tabDifal").innerHTML=`<div class="table-wrap" style="max-height:none"><table><thead><tr><th>Campo</th><th>Valor</th></tr></thead><tbody>${tableRow("Base DIFAL",difal.item_base_difal??"","","")}${tableRow("Valor DIFAL",difal.item_valor_difal??"","","")}${tableRow("Valor FCP",difal.item_valor_fcp??"","","")}${tableRow("Valor FCP-ST",difal.item_valor_fcp_st??"","","")}</tbody></table></div><div class="chips" style="margin-top:10px">${(difal.motivos||[]).map(m=>`<span class="chip">${m}</span>`).join("")||'<span class="chip">Sem observações</span>'}</div>`;

    const opf=item.impostos?.operacao_fiscal||{};
    document.getElementById("tabOperacao").innerHTML=`<div class="table-wrap" style="max-height:none"><table><thead><tr><th>Campo</th><th>Valor</th></tr></thead><tbody>${tableRow("Movimento",opf.movimento??"","","")}${tableRow("Tipo Documento",opf.documento_tipo??"","","")}${tableRow("Transação",opf.transacao??"","","")}${tableRow("CFOP",opf.cfop??"","","")}${tableRow("Natureza Operação",opf.natureza_operacao??"","","")}</tbody></table></div><div class="chips" style="margin-top:10px">${(opf.motivos||[]).map(m=>`<span class="chip">${m}</span>`).join("")||'<span class="chip">Sem observações</span>'}</div>`;

    const clf=item.impostos?.classificacao_fiscal||{};
    document.getElementById("tabClassFiscal").innerHTML=`<div class="table-wrap" style="max-height:none"><table><thead><tr><th>Campo</th><th>Valor</th></tr></thead><tbody>${tableRow("NCM",clf.ncm??"","","")}${tableRow("Classificação",clf.classificacao??"","","")}${tableRow("CEST",clf.cest??"","","")}${tableRow("Família",clf.familia_codigo??"","","")}${tableRow("Origem",clf.origem_codigo??"","","")}</tbody></table></div><div class="chips" style="margin-top:10px">${(clf.motivos||[]).map(m=>`<span class="chip">${m}</span>`).join("")||'<span class="chip">Sem observações</span>'}</div>`;

    // ---- Tab Família Fiscal: tabela comparativa fam_ Á— cad Á— item ----
    const fam=item.impostos?.familia_parametrizacao||{};
    const famRows=[
        tableRow("CST PIS",         fam.fam_cst_pis??"",       item.cad_recpis??"",  item.item_cst_pis??""),
        tableRow("CST COFINS",      fam.fam_cst_cofins??"",    item.cad_reccof??"",  item.item_cst_cofins??""),
        tableRow("CST IPI",         fam.fam_cst_ipi??"",       item.cad_peripi??"", item.item_cst_ipi??""),
        tableRow("% PIS",           fam.fam_perpis??"",        "",                  ""),
        tableRow("% COFINS",        fam.fam_percof??"",        "",                  ""),
        tableRow("% IPI",           fam.fam_peripi??"",        item.cad_peripi??"", item.item_aliq_ipi??""),
        tableRow("ProImp",          fam.fam_proimp??"",        "",                  ""),
        tableRow("TipPro (fam)",    fam.fam_tippro??"",        "",                  ""),
        tableRow("Origem (fam)",    fam.fam_codori??"",        item.origem_codigo??"", ""),
        tableRow("Rec PIS",         fam.fam_recpis??"",        item.cad_recpis??"",  ""),
        tableRow("Rec COFINS",      fam.fam_reccof??"",        item.cad_reccof??"",  ""),
        tableRow("Rec IPI",         fam.fam_recipi??"",        item.cad_recipi??"",  ""),
        tableRow("Rec ICMS",        fam.fam_recicm??"",        item.cad_recicm??"",  ""),
        tableRow("Tem ICMS",        fam.fam_temicm??"",        item.cad_temicm??"",  ""),
        tableRow("CodTrd (fam)",    fam.fam_codtrd??"",        item.cad_codtrd??"",  ""),
        tableRow("CodTST (fam)",    fam.fam_codtst??"",        item.cad_codtst??"",  ""),
        tableRow("CodSTP (fam)",    fam.fam_codstp??"",        item.cad_codstp??"",  ""),
        tableRow("Clf.Fiscal (fam)",fam.fam_codclf??"",        item.cod_classificacao??"", ""),
        tableRow("VarPro",          fam.fam_varpro??"",        "",                  ""),
        tableRow("FinCrP",          fam.fam_fincrp??"",        "",                  ""),
        tableRow("FinCdP",          fam.fam_fincdp??"",        "",                  ""),
        tableRow("% ICMS fam",      fam.fam_pericm??"",        item.cad_temicm??"",  item.item_aliq_icms??""),
        tableRow("% IPI fam",       fam.fam_peripi??"",        item.cad_peripi??"",  item.item_aliq_ipi??""),
        tableRow("% IRF fam",       fam.fam_perirf??"",        "",                  ""),
        tableRow("% IMP fam",       fam.fam_perpim??"",        "",                  ""),
        tableRow("RegTri",          fam.fam_regtri??"",        "",                  ""),
        tableRow("Unid. Medida",    fam.fam_unimed??"",        "",                  ""),
        tableRow("DescrFam",        fam.fam_desfam??item.familia_descricao??"", "", ""),
    ].join("");
    document.getElementById("tabFamilia").innerHTML=
        `<div class="table-wrap" style="max-height:420px"><table><thead><tr><th>Parâmetro</th><th>Família (E012FAM)</th><th>Cadastro Produto</th><th>Item NF</th></tr></thead><tbody>${famRows}</tbody></table></div>`;

    document.querySelectorAll(".tab").forEach(el=>el.classList.remove("active"));
    document.querySelectorAll(".tab-pane").forEach(el=>el.classList.add("hidden"));
    document.querySelectorAll(".tab")[0].classList.add("active");
    document.getElementById("tabQuadro").classList.remove("hidden");
}

(async function init(){
    if(token){showApp();await carregarCombos();}else{showLogin();}
})();
</script>
</body>
</html>"""


# =========================================================
# CONTROLE FISCAL DE PRODUTOS - PÁGINA
# =========================================================

@app.get("/controle-fiscal-produtos", response_class=HTMLResponse)
def controle_fiscal_produtos_page():
    return """<!DOCTYPE html>
<html lang="pt-br">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1.0" />
<title>Controle Fiscal de Produtos</title>
<style>
:root{--bg:#081226;--card:#0d1b34;--line:#233a64;--text:#eaf1ff;--muted:#9bb0d3;--primary:#3b82f6;--ok:#22c55e;--warn:#eab308;--danger:#ef4444;}
*{box-sizing:border-box}
body{margin:0;font-family:Arial,Helvetica,sans-serif;background:linear-gradient(180deg,#07101f 0%,#081226 100%);color:var(--text);}
.wrap{max-width:1700px;margin:0 auto;padding:24px}
.title{font-size:34px;font-weight:700;margin-bottom:6px}
.sub{color:var(--muted);margin-bottom:20px}
.card{background:rgba(13,27,52,.96);border:1px solid var(--line);border-radius:18px;padding:18px;box-shadow:0 12px 30px rgba(0,0,0,.18);}
.topbar{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:18px;flex-wrap:wrap}
.userbox{display:flex;gap:10px;align-items:center;flex-wrap:wrap}
.chip{padding:8px 10px;border-radius:999px;border:1px solid var(--line);background:#0b1830;color:#cddbf4;font-size:12px;}
.row{display:grid;grid-template-columns:repeat(5,1fr);gap:14px}
.grid{display:grid;gap:14px}
label{display:block;color:var(--muted);font-size:13px;margin-bottom:6px}
input,select{width:100%;background:#07101f;color:var(--text);border:1px solid var(--line);border-radius:12px;padding:12px 14px;outline:none;}
.btn{background:var(--primary);color:white;border:none;border-radius:12px;padding:12px 16px;cursor:pointer;font-weight:700;}
.btn.secondary{background:#243755}
.btn.ghost{background:#13213d}
.actions{display:flex;gap:10px;flex-wrap:wrap;margin-top:8px}
.cards{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-top:18px;}
.metric{background:rgba(9,19,40,.95);border:1px solid var(--line);border-radius:18px;padding:16px;}
.metric .k{color:var(--muted);font-size:13px;margin-bottom:8px;}
.metric .v{font-size:32px;font-weight:800;}
.panel{margin-top:18px;display:grid;grid-template-columns:1.15fr .85fr;gap:16px;}
table{width:100%;border-collapse:collapse;margin-top:16px;font-size:13px;}
th,td{padding:12px 10px;border-bottom:1px solid rgba(35,58,100,.7);text-align:left;vertical-align:top;}
th{color:#afc2e4;font-size:12px;position:sticky;top:0;background:#0d1b34;z-index:1;}
tbody tr{cursor:pointer;transition:.15s ease}
tbody tr:hover{background:rgba(255,255,255,.03)}
.table-wrap{max-height:520px;overflow:auto;border:1px solid var(--line);border-radius:16px;}
.muted{color:var(--muted)}
.status{display:inline-block;padding:6px 10px;border-radius:999px;font-size:12px;font-weight:700;}
.status.bad{background:rgba(239,68,68,.15);color:#ff9ea2;border:1px solid rgba(239,68,68,.35)}
.status.warn{background:rgba(234,179,8,.15);color:#fde047;border:1px solid rgba(234,179,8,.35)}
.hidden{display:none}
.small{font-size:12px}
.kv{font-size:12px;line-height:1.7}
@media(max-width:1200px){.row,.cards{grid-template-columns:repeat(2,1fr)}.panel{grid-template-columns:1fr}}
@media(max-width:700px){.row,.cards{grid-template-columns:1fr}.title{font-size:28px}}
.btn.ok{background:var(--ok);color:#071020}
.btn.danger{background:var(--danger)}
.detail-actions{display:flex;gap:10px;flex-wrap:wrap;margin-top:14px;padding-top:14px;border-top:1px solid var(--line)}
.modal-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.65);z-index:100;align-items:center;justify-content:center}
.modal-overlay.show{display:flex}
.modal-box{background:var(--card);border:1px solid var(--line);border-radius:18px;padding:24px;width:95%;max-width:720px;max-height:85vh;overflow:auto;box-shadow:0 20px 60px rgba(0,0,0,.5)}
.modal-box h2{margin:0 0 18px;font-size:22px}
.modal-grid{display:grid;grid-template-columns:1fr 1fr;gap:10px}
.modal-grid label{font-size:12px;color:var(--muted)}
.modal-grid input{padding:10px 12px;font-size:13px}
.modal-footer{display:flex;gap:10px;margin-top:18px;justify-content:flex-end;flex-wrap:wrap}
.msg-ok{color:var(--ok);font-weight:700;font-size:13px;margin-top:10px}
.msg-err{color:var(--danger);font-weight:700;font-size:13px;margin-top:10px}
</style>
</head>
<body>
<div class="wrap">
    <div class="topbar">
        <div>
            <div class="title">Controle Fiscal de Produtos</div>
            <div class="sub">Fila para conferencia dos parametros fiscais em E075PRO e E075DER antes da ativacao.</div>
        </div>
        <div class="userbox">
            <span id="userLabel" class="chip"></span>
            <button class="btn ghost" onclick="window.location.href='/'">Auditoria Tributaria</button>
            <button class="btn secondary" onclick="window.location.href='/inteligencia-tributaria'">Inteligencia Tributaria</button>
            <button class="btn ghost" onclick="sair()">Sair</button>
        </div>
    </div>

    <div class="card">
        <div class="row">
            <div><label>Visao</label>
                <select id="f_visao">
                    <option value="AMBOS">Produto e/ou derivacao inativos</option>
                    <option value="PRODUTO">Somente produto inativo</option>
                    <option value="DERIVACAO">Somente derivacao inativa</option>
                </select>
            </div>
            <div><label>Produto</label><input id="f_codigo_produto" placeholder="Codigo do produto" /></div>
            <div><label>Descricao</label><input id="f_descricao" placeholder="Descricao do produto" /></div>
            <div><label>Derivacao</label><input id="f_derivacao" placeholder="Codigo derivacao" /></div>
            <div><label>Tipo produto</label><input id="f_tipo_produto" placeholder="Ex.: P" /></div>
        </div>
        <div class="row" style="margin-top:14px">
            <div><label>Origem</label><input id="f_origem" list="listaOrigens" placeholder="Ex.: 250" /><datalist id="listaOrigens"></datalist></div>
            <div><label>Família</label><input id="f_familia" list="listaFamilias" placeholder="Ex.: BR-CHA" /><datalist id="listaFamilias"></datalist></div>
            <div><label>Usuario geracao</label><input id="f_usuario_geracao" type="number" placeholder="Codigo usuario" /></div>
            <div><label>Data geracao de</label><input id="f_data_ini" type="date" /></div>
            <div><label>Data geracao ate</label><input id="f_data_fim" type="date" /></div>
        </div>
        <div class="row" style="margin-top:14px">
            <div>
                <label>Mês de geração</label>
                <input id="f_mes_geracao" type="month" />
            </div>
            <div style="display:flex;align-items:end">
                <label style="display:flex;gap:8px;align-items:center;margin:0">
                    <input id="f_fixar_mes" type="checkbox" style="width:auto" />
                    <span>Fixar mês na grade</span>
                </label>
            </div>
        </div>
        <div class="actions">
            <button class="btn" onclick="consultarControleFiscal(1)">Consultar</button>
            <button class="btn secondary" onclick="limparFiltrosControleFiscal()">Limpar</button>
        </div>
    </div>

    <div class="cards">
        <div class="metric"><div class="k">Total</div><div class="v" id="m_total">0</div></div>
        <div class="metric"><div class="k">Produto inativo</div><div class="v" id="m_produto">0</div></div>
        <div class="metric"><div class="k">Derivacao inativa</div><div class="v" id="m_derivacao">0</div></div>
        <div class="metric"><div class="k">Ambos inativos</div><div class="v" id="m_ambos">0</div></div>
    </div>

    <div class="panel">
        <div class="card">
            <div style="font-size:22px;font-weight:700">Pendencias fiscais</div>
            <div id="paginacao" class="muted small" style="margin-top:8px"></div>
            <div id="bannerMesFixadoControleFiscal" class="muted small" style="margin-top:8px"></div>
            <div class="table-wrap">
                <table>
                    <thead>
                        <tr>
                            <th>Status</th>
                            <th>Produto</th>
                            <th>Descricao</th>
                            <th>Der.</th>
                            <th>NCM</th>
                            <th>Origem</th>
                            <th>Família</th>
                            <th>Usuario geracao</th>
                            <th>Data geracao</th>
                        </tr>
                    </thead>
                    <tbody id="tbodyResultados">
                        <tr><td colspan="9" class="muted">Nenhum dado carregado.</td></tr>
                    </tbody>
                </table>
            </div>
        </div>

        <div class="card">
            <div style="font-size:22px;font-weight:700">Detalhe fiscal</div>
            <div id="detalheVazio" class="muted" style="margin-top:12px">Selecione uma linha para ver os campos fiscais.</div>
            <div id="detalheItem" class="hidden">
                <div id="detalheHeader" style="margin-top:10px"></div>
                <div style="margin-top:14px;font-weight:700">E075PRO</div>
                <div id="blocoPro" class="kv" style="margin-top:8px"></div>
                <div style="margin-top:14px;font-weight:700">E075DER</div>
                <div id="blocoDer" class="kv" style="margin-top:8px"></div>
                <div class="actions" style="margin-top:14px">
                    <button class="btn" onclick="abrirModalEdicaoControleFiscal()">Editar</button>
                    <button class="btn secondary" onclick="ativarItemControleFiscal()">Ativar</button>
                </div>
            </div>
        </div>
    </div>
</div>

<div id="modalEdicaoControleFiscal" class="hidden" style="position:fixed;inset:0;background:rgba(0,0,0,.65);z-index:9999;padding:24px;overflow:auto">
    <div style="max-width:1100px;margin:0 auto;background:#0d1b34;border:1px solid #233a64;border-radius:18px;padding:18px">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap">
            <div>
                <div style="font-size:24px;font-weight:700">Edicao Fiscal</div>
                <div class="muted" id="modalTituloControleFiscal">Produto</div>
            </div>
            <button class="btn ghost" onclick="fecharModalEdicaoControleFiscal()">Fechar</button>
        </div>
        <div style="margin-top:18px;font-weight:700">E075PRO</div>
        <div class="row" style="margin-top:10px">
            <div><label>CODCLF</label><input id="edit_CODCLF"></div>
            <div><label>CODTRD</label><input id="edit_CODTRD"></div>
            <div><label>CODTST</label><input id="edit_CODTST"></div>
            <div><label>CODSTP</label><input id="edit_CODSTP"></div>
        </div>
        <div class="row">
            <div><label>RECPIS</label><input id="edit_RECPIS"></div>
            <div><label>RECCOF</label><input id="edit_RECCOF"></div>
            <div><label>CSTPIS</label><input id="edit_CSTPIS"></div>
            <div><label>CSTCOF</label><input id="edit_CSTCOF"></div>
        </div>
        <div class="row">
            <div><label>BASCRE</label><input id="edit_BASCRE"></div>
            <div><label>CODSTR</label><input id="edit_CODSTR"></div>
            <div><label>CODTIC</label><input id="edit_CODTIC"></div>
            <div><label>CODSTC</label><input id="edit_CODSTC"></div>
        </div>
        <div class="row">
            <div><label>PERIPI</label><input id="edit_PERIPI"></div>
            <div><label>RECIPI</label><input id="edit_RECIPI"></div>
            <div><label>TEMICM</label><input id="edit_TEMICM"></div>
            <div><label>RECICM</label><input id="edit_RECICM"></div>
        </div>
        <div class="row">
            <div><label>TRIPIS</label><input id="edit_TRIPIS"></div>
            <div><label>TRICOF</label><input id="edit_TRICOF"></div>
            <div><label>ORIMER</label><input id="edit_ORIMER"></div>
            <div><label>CODANP</label><input id="edit_CODANP"></div>
        </div>
        <div class="row">
            <div><label>NATPIS</label><input id="edit_NATPIS"></div>
            <div><label>NATCOF</label><input id="edit_NATCOF"></div>
            <div><label>PROIMP</label><input id="edit_PROIMP"></div>
            <div><label>REGTRI</label><input id="edit_REGTRI"></div>
        </div>
        <div style="margin-top:18px;font-weight:700">E075DER</div>
        <div class="row" style="margin-top:10px">
            <div><label>ITEFIS</label><input id="edit_ITEFIS"></div>
            <div><label>DESFIS</label><input id="edit_DESFIS"></div>
            <div><label>CODFIF</label><input id="edit_CODFIF"></div>
            <div><label>CODFIE</label><input id="edit_CODFIE"></div>
        </div>
        <div class="row">
            <div><label>CODFIM</label><input id="edit_CODFIM"></div>
            <div><label>BSTUFC</label><input id="edit_BSTUFC"></div>
            <div><label>ASTFCP</label><input id="edit_ASTFCP"></div>
            <div><label>VSTUFC</label><input id="edit_VSTUFC"></div>
        </div>
        <div class="row">
            <div><label>CODCES</label><input id="edit_CODCES"></div>
        </div>
        <div class="actions" style="margin-top:18px">
            <button class="btn" onclick="salvarItemControleFiscal(false)">Salvar</button>
            <button class="btn secondary" onclick="salvarItemControleFiscal(true)">Salvar e ativar</button>
            <button class="btn secondary" onclick="sugerirComIAControleFiscal()">Sugerir com IA</button>
        </div>
        <div id="msgModalControleFiscal" class="muted" style="margin-top:12px"></div>
        <div id="boxSugestaoIAControleFiscal" class="hidden" style="margin-top:18px;border:1px solid #233a64;border-radius:14px;padding:14px;background:#0b1830">
            <div style="font-size:18px;font-weight:700;margin-bottom:8px">Sugestão da IA</div>
            <div id="resumoIAControleFiscal" class="muted" style="margin-bottom:10px"></div>
            <div id="alertasIAControleFiscal" class="muted" style="margin-bottom:10px"></div>
            <div id="listaSugestoesIAControleFiscal"></div>
            <div class="actions" style="margin-top:14px">
                <button class="btn" onclick="aplicarSugestoesIAControleFiscal()">Aplicar sugestões</button>
            </div>
        </div>
    </div>
</div>

<script>
let token = localStorage.getItem("token") || "";
let usuarioLogado = localStorage.getItem("usuario") || "";
let resultadosControleFiscal = [];
let paginaAtualControleFiscal = 1;

function showUserControleFiscal(){
    var el = document.getElementById("userLabel");
    if(el) el.innerText = usuarioLogado || "Usuario";
}

async function apiFetchControleFiscal(url, options){
    options = options || {};
    var headers = options.headers || {};
    if(token) headers["Authorization"] = "Bearer " + token;
    var response = await fetch(url, Object.assign({}, options, {headers: headers}));
    var raw = await response.text();
    var data;
    try { data = raw ? JSON.parse(raw) : {}; } catch(ex) { data = {detail: raw || ("HTTP " + response.status)}; }
    if(response.status === 401){
        localStorage.removeItem("token");
        localStorage.removeItem("usuario");
        window.location.href = "/";
        throw new Error(data.detail || "Sessao expirada");
    }
    if(!response.ok) throw new Error(data.detail || ("Erro " + response.status));
    return data;
}

async function carregarCombosControleFiscal(){
    try{
        var familias = await apiFetchControleFiscal("/api/familias?limite=200");
        var origens = await apiFetchControleFiscal("/api/origens?limite=200");
        var lf = document.getElementById("listaFamilias");
        var lo = document.getElementById("listaOrigens");
        if(lf) lf.innerHTML = familias.map(function(f){ return '<option value="' + f.codigo + '">' + f.label + '</option>'; }).join("");
        if(lo) lo.innerHTML = origens.map(function(o){ return '<option value="' + o.codigo + '">' + o.label + '</option>'; }).join("");
    }catch(e){
        console.error("Erro ao carregar combos:", e);
    }
}

function limparFiltrosControleFiscal(){
    [
        "f_codigo_produto",
        "f_descricao",
        "f_derivacao",
        "f_tipo_produto",
        "f_origem",
        "f_familia",
        "f_usuario_geracao",
        "f_data_ini",
        "f_data_fim"
    ].forEach(id => {
        const el = document.getElementById(id);
        if (el) el.value = "";
    });

    const visao = document.getElementById("f_visao");
    if (visao) visao.value = "AMBOS";

    limparMesControleFiscal();

    resultadosControleFiscal = [];
    renderTabelaControleFiscal([]);
    renderResumoControleFiscal({});

    const pag = document.getElementById("paginacao");
    if (pag) pag.innerHTML = "";

    const detalheItem = document.getElementById("detalheItem");
    const detalheVazio = document.getElementById("detalheVazio");

    if (detalheItem) detalheItem.classList.add("hidden");
    if (detalheVazio) detalheVazio.classList.remove("hidden");
}

async function consultarControleFiscal(pagina = 1){
    paginaAtualControleFiscal = pagina;

    const params = new URLSearchParams();

    const visao = document.getElementById("f_visao")?.value || "AMBOS";
    const codigoProduto = document.getElementById("f_codigo_produto")?.value || "";
    const descricao = document.getElementById("f_descricao")?.value || "";
    const derivacao = document.getElementById("f_derivacao")?.value || "";
    const tipoProduto = document.getElementById("f_tipo_produto")?.value || "";
    const origem = document.getElementById("f_origem")?.value || "";
    const familia = document.getElementById("f_familia")?.value || "";
    const usuarioGeracao = document.getElementById("f_usuario_geracao")?.value || "";
    const mes = document.getElementById("f_mes_geracao")?.value || "";
    const fixarMes = document.getElementById("f_fixar_mes")?.checked || false;

    let dataIni = document.getElementById("f_data_ini")?.value || "";
    let dataFim = document.getElementById("f_data_fim")?.value || "";

    if (mes) {
        const periodo = montarPeriodoDoMesControleFiscal(mes);
        dataIni = periodo.dataIni;
        dataFim = periodo.dataFim;

        const elIni = document.getElementById("f_data_ini");
        const elFim = document.getElementById("f_data_fim");
        if (elIni) elIni.value = dataIni;
        if (elFim) elFim.value = dataFim;
    }

    if (visao) params.append("visao", visao);
    if (codigoProduto.trim() !== "") params.append("codigo_produto", codigoProduto.trim());
    if (descricao.trim() !== "") params.append("descricao", descricao.trim());
    if (derivacao.trim() !== "") params.append("derivacao", derivacao.trim());
    if (tipoProduto.trim() !== "") params.append("tipo_produto", tipoProduto.trim());
    if (origem.trim() !== "") params.append("origem", origem.trim());
    if (familia.trim() !== "") params.append("familia", familia.trim());
    if (usuarioGeracao.trim() !== "") params.append("usuario_geracao", usuarioGeracao.trim());
    if (dataIni.trim() !== "") params.append("data_ini", dataIni);
    if (dataFim.trim() !== "") params.append("data_fim", dataFim);

    params.append("pagina", String(paginaAtualControleFiscal));
    params.append("tamanho_pagina", "100");

    const tbody = document.getElementById("tbodyResultados");
    if (tbody) {
        tbody.innerHTML = `<tr><td colspan="9" class="muted">Consultando...</td></tr>`;
    }

    try {
        const data = await apiFetchControleFiscal("/api/controle-fiscal-produtos?" + params.toString());

        resultadosControleFiscal = data.itens || [];
        renderResumoControleFiscal(data.resumo || {});
        renderTabelaControleFiscal(resultadosControleFiscal);

        const pag = document.getElementById("paginacao");
        if (pag) {
            pag.innerHTML = `Página <strong>${data.pagina || 1}</strong> de <strong>${data.total_paginas || 1}</strong> | Total: <strong>${data.total_registros || 0}</strong>`;
        }

        if (fixarMes && mes) {
            salvarPreferenciaMesControleFiscal();
        } else if (!fixarMes) {
            localStorage.removeItem("mesControleFiscal");
            localStorage.removeItem("fixarMesControleFiscal");
            atualizarBannerMesControleFiscal();
        }

    } catch (e) {
        if (tbody) {
            tbody.innerHTML = `<tr><td colspan="9" class="muted">${e.message}</td></tr>`;
        }
    }
}

function renderResumoControleFiscal(r){
    var a = document.getElementById("m_total");
    var b = document.getElementById("m_produto");
    var c = document.getElementById("m_derivacao");
    var d = document.getElementById("m_ambos");
    if(a) a.innerText = r.total_registros || 0;
    if(b) b.innerText = r.produto_inativo || 0;
    if(c) c.innerText = r.derivacao_inativa || 0;
    if(d) d.innerText = r.produto_e_derivacao_inativos || 0;
}

function badgeStatusControleFiscal(v){
    if(v === "PRODUTO_E_DERIVACAO_INATIVOS") return '<span class="status bad">Produto + derivacao</span>';
    if(v === "PRODUTO_INATIVO") return '<span class="status warn">Produto</span>';
    if(v === "DERIVACAO_INATIVA") return '<span class="status warn">Derivacao</span>';
    return '<span class="status">Ativo</span>';
}

function renderTabelaControleFiscal(lista){
    var tbody = document.getElementById("tbodyResultados");
    if(!tbody) return;
    if(!lista.length){
        tbody.innerHTML = '<tr><td colspan="9" class="muted">Nenhum registro encontrado.</td></tr>';
        return;
    }
    tbody.innerHTML = lista.map(function(item, idx){
        return '<tr onclick="abrirDetalheControleFiscal(' + idx + ')">' +
            '<td>' + badgeStatusControleFiscal(item.STATUS_CONTROLE) + '</td>' +
            '<td>' + (item.CODPRO || '') + '</td>' +
            '<td>' + (item.DESPRO || '') + '</td>' +
            '<td>' + (item.CODDER || '') + '</td>' +
            '<td>' + (item.NCM || '') + '</td>' +
            '<td>' + (item.CODORI || '') + '</td>' +
            '<td>' + (item.CODFAM || '') + '</td>' +
            '<td>' + (item.NOME_USUGER_DER || item.NOME_USUGER_PRO || '') + '</td>' +
            '<td>' + ((item.DATGER_DER || item.DATGER_PRO || '').toString().slice(0,10)) + '</td>' +
            '</tr>';
    }).join('');
}

function linhaControleFiscal(label, valor){
    return '<div><b>' + label + '</b>: ' + (valor != null ? valor : '') + '</div>';
}

var itemSelecionadoControleFiscal = null;

function abrirModalEdicaoControleFiscal(){
    if(!itemSelecionadoControleFiscal){
        alert("Selecione um item primeiro.");
        return;
    }
    var it = itemSelecionadoControleFiscal;
    document.getElementById("modalEdicaoControleFiscal").classList.remove("hidden");
    document.getElementById("modalTituloControleFiscal").innerText =
        (it.CODPRO || "") + " " + (it.DESPRO || "") + " | Derivacao: " + (it.CODDER || "-");
    var map = [
        "CODCLF","CODTRD","CODTST","CODSTP","RECPIS","RECCOF","CSTPIS","CSTCOF",
        "BASCRE","CODSTR","CODTIC","CODSTC","PERIPI","RECIPI","TEMICM","RECICM",
        "TRIPIS","TRICOF","ORIMER","CODANP","NATPIS","NATCOF","PROIMP","REGTRI",
        "ITEFIS","DESFIS","CODFIF","CODFIE","CODFIM","BSTUFC","ASTFCP","VSTUFC","CODCES"
    ];
    map.forEach(function(campo){
        var el = document.getElementById("edit_" + campo);
        if(el) el.value = (it[campo] != null ? it[campo] : "");
    });
    document.getElementById("msgModalControleFiscal").innerText = "";
}

function fecharModalEdicaoControleFiscal(){
    document.getElementById("modalEdicaoControleFiscal").classList.add("hidden");
}

function abrirDetalheControleFiscal(idx){
    var item = resultadosControleFiscal[idx];
    if(!item) return;
    itemSelecionadoControleFiscal = item;
    var dv = document.getElementById("detalheVazio");
    var di = document.getElementById("detalheItem");
    if(dv) dv.classList.add("hidden");
    if(di) di.classList.remove("hidden");
    var header = document.getElementById("detalheHeader");
    if(header) header.innerHTML =
        '<div class="chip">' + (item.CODPRO || "") + '</div>' +
        '<div style="margin-top:8px;font-weight:700">' + (item.DESPRO || "") + '</div>' +
        '<div class="muted small" style="margin-top:6px">Derivacao: ' + (item.CODDER || "-") + ' | NCM: ' + (item.NCM || "-") + ' | Status: ' + (item.STATUS_CONTROLE || "-") + '</div>';
    var blocoPro = document.getElementById("blocoPro");
    if(blocoPro) blocoPro.innerHTML = [
        linhaControleFiscal("Situacao produto", item.SITPRO),
        linhaControleFiscal("Tipo produto", item.TIPPRO),
        linhaControleFiscal("Origem", item.CODORI),
        linhaControleFiscal("Familia", item.CODFAM),
        linhaControleFiscal("Usuario geracao", item.NOME_USUGER_PRO || item.USUGER_PRO),
        linhaControleFiscal("Data geracao", item.DATGER_PRO),
        linhaControleFiscal("CODCLF", item.CODCLF),
        linhaControleFiscal("CODSTR", item.CODSTR),
        linhaControleFiscal("CODTIC", item.CODTIC),
        linhaControleFiscal("CODTRD", item.CODTRD),
        linhaControleFiscal("CODTST", item.CODTST),
        linhaControleFiscal("CODSTP", item.CODSTP),
        linhaControleFiscal("CODSTC", item.CODSTC),
        linhaControleFiscal("PERIPI", item.PERIPI),
        linhaControleFiscal("RECIPI", item.RECIPI),
        linhaControleFiscal("TEMICM", item.TEMICM),
        linhaControleFiscal("RECICM", item.RECICM),
        linhaControleFiscal("RECPIS", item.RECPIS),
        linhaControleFiscal("TRIPIS", item.TRIPIS),
        linhaControleFiscal("TRICOF", item.TRICOF),
        linhaControleFiscal("RECCOF", item.RECCOF),
        linhaControleFiscal("BASCRE", item.BASCRE),
        linhaControleFiscal("BASREC", item.BASREC),
        linhaControleFiscal("CSTIPI", item.CSTIPI),
        linhaControleFiscal("CSTPIS", item.CSTPIS),
        linhaControleFiscal("CSTCOF", item.CSTCOF),
        linhaControleFiscal("TPRPIS", item.TPRPIS),
        linhaControleFiscal("TPRCOF", item.TPRCOF),
        linhaControleFiscal("TPRIPI", item.TPRIPI),
        linhaControleFiscal("REGTRI", item.REGTRI),
        linhaControleFiscal("CSTIPC", item.CSTIPC),
        linhaControleFiscal("CSTPIC", item.CSTPIC),
        linhaControleFiscal("CSTCOC", item.CSTCOC),
        linhaControleFiscal("ORIMER", item.ORIMER),
        linhaControleFiscal("NATPIS", item.NATPIS),
        linhaControleFiscal("NATCOF", item.NATCOF),
        linhaControleFiscal("CODANP", item.CODANP),
        linhaControleFiscal("PROIMP", item.PROIMP)
    ].join("");
    var blocoDer = document.getElementById("blocoDer");
    if(blocoDer) blocoDer.innerHTML = [
        linhaControleFiscal("Situacao derivacao", item.SITDER),
        linhaControleFiscal("Descricao derivacao", item.DESDER),
        linhaControleFiscal("Usuario geracao", item.NOME_USUGER_DER || item.USUGER_DER),
        linhaControleFiscal("Data geracao", item.DATGER_DER),
        linhaControleFiscal("ITEFIS", item.ITEFIS),
        linhaControleFiscal("DESFIS", item.DESFIS),
        linhaControleFiscal("CODFIF", item.CODFIF),
        linhaControleFiscal("CODFIE", item.CODFIE),
        linhaControleFiscal("CODFIM", item.CODFIM),
        linhaControleFiscal("BSTUFC", item.BSTUFC),
        linhaControleFiscal("ASTFCP", item.ASTFCP),
        linhaControleFiscal("VSTUFC", item.VSTUFC),
        linhaControleFiscal("CODCES", item.CODCES)
    ].join("");
}

async function salvarItemControleFiscal(ativarAposSalvar){
    if(!itemSelecionadoControleFiscal){
        alert("Selecione um item primeiro.");
        return;
    }
    var msg = document.getElementById("msgModalControleFiscal");
    msg.innerText = "Salvando...";
    var camposPro = {
        CODCLF: document.getElementById("edit_CODCLF").value,
        CODTRD: document.getElementById("edit_CODTRD").value,
        CODTST: document.getElementById("edit_CODTST").value,
        CODSTP: document.getElementById("edit_CODSTP").value,
        RECPIS: document.getElementById("edit_RECPIS").value,
        RECCOF: document.getElementById("edit_RECCOF").value,
        CSTPIS: document.getElementById("edit_CSTPIS").value,
        CSTCOF: document.getElementById("edit_CSTCOF").value,
        BASCRE: document.getElementById("edit_BASCRE").value,
        CODSTR: document.getElementById("edit_CODSTR").value,
        CODTIC: document.getElementById("edit_CODTIC").value,
        CODSTC: document.getElementById("edit_CODSTC").value,
        PERIPI: document.getElementById("edit_PERIPI").value,
        RECIPI: document.getElementById("edit_RECIPI").value,
        TEMICM: document.getElementById("edit_TEMICM").value,
        RECICM: document.getElementById("edit_RECICM").value,
        TRIPIS: document.getElementById("edit_TRIPIS").value,
        TRICOF: document.getElementById("edit_TRICOF").value,
        ORIMER: document.getElementById("edit_ORIMER").value,
        CODANP: document.getElementById("edit_CODANP").value,
        NATPIS: document.getElementById("edit_NATPIS").value,
        NATCOF: document.getElementById("edit_NATCOF").value,
        PROIMP: document.getElementById("edit_PROIMP").value,
        REGTRI: document.getElementById("edit_REGTRI").value
    };
    var camposDer = {
        ITEFIS: document.getElementById("edit_ITEFIS").value,
        DESFIS: document.getElementById("edit_DESFIS").value,
        CODFIF: document.getElementById("edit_CODFIF").value,
        CODFIE: document.getElementById("edit_CODFIE").value,
        CODFIM: document.getElementById("edit_CODFIM").value,
        BSTUFC: document.getElementById("edit_BSTUFC").value,
        ASTFCP: document.getElementById("edit_ASTFCP").value,
        VSTUFC: document.getElementById("edit_VSTUFC").value,
        CODCES: document.getElementById("edit_CODCES").value
    };
    try{
        var data = await apiFetchControleFiscal("/api/controle-fiscal-produtos/salvar", {
            method: "POST",
            headers: {"Content-Type": "application/json"},
            body: JSON.stringify({
                codemp: itemSelecionadoControleFiscal.CODEMP,
                codpro: itemSelecionadoControleFiscal.CODPRO,
                codder: itemSelecionadoControleFiscal.CODDER || null,
                campos_pro: camposPro,
                campos_der: camposDer,
                ativar_apos_salvar: ativarAposSalvar
            })
        });
        msg.innerText = data.mensagem || "Salvo com sucesso.";
        fecharModalEdicaoControleFiscal();
        await consultarControleFiscal(paginaAtualControleFiscal);
    }catch(e){
        msg.innerText = e.message;
    }
}

let sugestoesIAControleFiscal = [];

async function sugerirComIAControleFiscal(){
    if(!itemSelecionadoControleFiscal){
        alert("Selecione um item primeiro.");
        return;
    }

    const msg = document.getElementById("msgModalControleFiscal");
    const box = document.getElementById("boxSugestaoIAControleFiscal");
    const resumo = document.getElementById("resumoIAControleFiscal");
    const alertas = document.getElementById("alertasIAControleFiscal");
    const lista = document.getElementById("listaSugestoesIAControleFiscal");

    msg.innerText = "Consultando IA...";
    box.classList.add("hidden");
    resumo.innerHTML = "";
    alertas.innerHTML = "";
    lista.innerHTML = "";
    sugestoesIAControleFiscal = [];

    try{
        const payload = {
            codemp: itemSelecionadoControleFiscal.CODEMP,
            codpro: itemSelecionadoControleFiscal.CODPRO,
            codder: itemSelecionadoControleFiscal.CODDER || null,
            campos_alvo: ["CODTRD", "CSTPIS", "CSTCOF", "BASCRE"],
            uf_origem: "SC",
            uf_destino: "SC"
        };

        const data = await apiFetchControleFiscal("/api/controle-fiscal-produtos/sugerir-ia", {
            method: "POST",
            headers: {"Content-Type": "application/json"},
            body: JSON.stringify(payload)
        });

        msg.innerText = "";
        box.classList.remove("hidden");

        const resultado = data.resultado_ia || {};
        sugestoesIAControleFiscal = resultado.campos_sugeridos || [];

        resumo.innerHTML = `<b>Resumo:</b> ${resultado.resumo || "Sem resumo retornado."}`;

        const listaAlertas = resultado.alertas || [];
        alertas.innerHTML = listaAlertas.length
            ? `<div><b>Alertas:</b></div>` + listaAlertas.map(a => `<div class="muted">- ${a}</div>`).join("")
            : "";

        if(!sugestoesIAControleFiscal.length){
            lista.innerHTML = `<div class="muted">A IA não sugeriu preenchimento automático para os campos solicitados.</div>`;
            return;
        }

        lista.innerHTML = sugestoesIAControleFiscal.map((s, idx) => `
            <div style="border:1px solid #233a64;border-radius:12px;padding:12px;margin-top:10px">
                <div><b>${s.campo_erp || "-"}</b></div>
                <div class="muted small" style="margin-top:4px">Atual: ${s.valor_atual ?? "-"}</div>
                <div class="muted small">Sugerido: <b>${s.valor_sugerido ?? "-"}</b></div>
                <div class="muted small" style="margin-top:6px">${s.justificativa || ""}</div>
                <div class="muted small">Base legal/regra: ${s.base_legal_ou_regra || "-"}</div>
                <div class="muted small">Confiança: ${s.confianca || "-"}</div>
            </div>
        `).join("");

    } catch(e){
        msg.innerText = e.message;
    }
}

function aplicarSugestoesIAControleFiscal(){
    if(!sugestoesIAControleFiscal.length){
        alert("Não há sugestões para aplicar.");
        return;
    }

    const camposPermitidos = new Set([
        "CODCLF","CODTRD","CODTST","CODSTP","RECPIS","RECCOF","CSTPIS","CSTCOF",
        "BASCRE","CODSTR","CODTIC","CODSTC","PERIPI","RECIPI","TEMICM","RECICM",
        "TRIPIS","TRICOF","ORIMER","CODANP","NATPIS","NATCOF","PROIMP","REGTRI",
        "ITEFIS","DESFIS","CODFIF","CODFIE","CODFIM","BSTUFC","ASTFCP","VSTUFC","CODCES"
    ]);

    sugestoesIAControleFiscal.forEach(s => {
        const campo = (s.campo_erp || "").toUpperCase().trim();
        const valor = s.valor_sugerido;

        if(!campo || !camposPermitidos.has(campo)) return;
        if(valor === null || valor === undefined || valor === "") return;

        const el = document.getElementById("edit_" + campo);
        if(el){
            el.value = valor;
        }
    });

    document.getElementById("msgModalControleFiscal").innerText = "Sugestões aplicadas no formulário. Revise antes de salvar.";
}

async function ativarItemControleFiscal(){
    if(!itemSelecionadoControleFiscal){
        alert("Selecione um item primeiro.");
        return;
    }
    if(!confirm("Deseja ativar este produto/derivacao?")) return;
    try{
        var data = await apiFetchControleFiscal("/api/controle-fiscal-produtos/ativar", {
            method: "POST",
            headers: {"Content-Type": "application/json"},
            body: JSON.stringify({
                codemp: itemSelecionadoControleFiscal.CODEMP,
                codpro: itemSelecionadoControleFiscal.CODPRO,
                codder: itemSelecionadoControleFiscal.CODDER || null
            })
        });
        alert(data.mensagem || "Ativado com sucesso.");
        await consultarControleFiscal(paginaAtualControleFiscal);
    }catch(e){
        alert(e.message);
    }
}

function montarPeriodoDoMesControleFiscal(yyyyMm) {
    if (!yyyyMm) return { dataIni: "", dataFim: "" };

    const [ano, mes] = yyyyMm.split("-").map(Number);
    if (!ano || !mes) return { dataIni: "", dataFim: "" };

    const inicio = new Date(ano, mes - 1, 1);
    const fim = new Date(ano, mes, 0);

    const fmt = (d) => {
        const y = d.getFullYear();
        const m = String(d.getMonth() + 1).padStart(2, "0");
        const day = String(d.getDate()).padStart(2, "0");
        return `${y}-${m}-${day}`;
    };

    return {
        dataIni: fmt(inicio),
        dataFim: fmt(fim)
    };
}

function atualizarBannerMesControleFiscal() {
    const mes = document.getElementById("f_mes_geracao")?.value || "";
    const fixar = document.getElementById("f_fixar_mes")?.checked || false;
    const banner = document.getElementById("bannerMesFixadoControleFiscal");

    if (!banner) return;

    if (mes && fixar) {
        banner.innerHTML = `<strong>Mês fixado:</strong> ${mes}`;
    } else if (mes) {
        banner.innerHTML = `<strong>Mês selecionado:</strong> ${mes}`;
    } else {
        banner.innerHTML = "";
    }
}

function salvarPreferenciaMesControleFiscal() {
    const mes = document.getElementById("f_mes_geracao")?.value || "";
    const fixar = document.getElementById("f_fixar_mes")?.checked || false;

    if (fixar && mes) {
        localStorage.setItem("mesControleFiscal", mes);
        localStorage.setItem("fixarMesControleFiscal", "true");
    } else {
        localStorage.removeItem("mesControleFiscal");
        localStorage.removeItem("fixarMesControleFiscal");
    }

    atualizarBannerMesControleFiscal();
}

function carregarPreferenciaMesControleFiscal() {
    const mes = localStorage.getItem("mesControleFiscal") || "";
    const fixar = localStorage.getItem("fixarMesControleFiscal") === "true";

    const inputMes = document.getElementById("f_mes_geracao");
    const inputFixar = document.getElementById("f_fixar_mes");

    if (inputMes) inputMes.value = mes;
    if (inputFixar) inputFixar.checked = fixar;

    if (fixar && mes) {
        const periodo = montarPeriodoDoMesControleFiscal(mes);
        const dataIni = document.getElementById("f_data_ini");
        const dataFim = document.getElementById("f_data_fim");

        if (dataIni) dataIni.value = periodo.dataIni;
        if (dataFim) dataFim.value = periodo.dataFim;
    }

    atualizarBannerMesControleFiscal();
}

function aplicarMesControleFiscalNosFiltros() {
    const mes = document.getElementById("f_mes_geracao")?.value || "";
    if (!mes) return;

    const periodo = montarPeriodoDoMesControleFiscal(mes);

    const dataIni = document.getElementById("f_data_ini");
    const dataFim = document.getElementById("f_data_fim");

    if (dataIni) dataIni.value = periodo.dataIni;
    if (dataFim) dataFim.value = periodo.dataFim;
}

function limparMesControleFiscal() {
    const inputMes = document.getElementById("f_mes_geracao");
    const inputFixar = document.getElementById("f_fixar_mes");
    const dataIni = document.getElementById("f_data_ini");
    const dataFim = document.getElementById("f_data_fim");

    if (inputMes) inputMes.value = "";
    if (inputFixar) inputFixar.checked = false;
    if (dataIni) dataIni.value = "";
    if (dataFim) dataFim.value = "";

    localStorage.removeItem("mesControleFiscal");
    localStorage.removeItem("fixarMesControleFiscal");

    atualizarBannerMesControleFiscal();
}

function bindEventosMesControleFiscal() {
    const inputMes = document.getElementById("f_mes_geracao");
    const inputFixar = document.getElementById("f_fixar_mes");

    if (inputMes) {
        inputMes.addEventListener("change", () => {
            aplicarMesControleFiscalNosFiltros();
            salvarPreferenciaMesControleFiscal();
        });
    }

    if (inputFixar) {
        inputFixar.addEventListener("change", () => {
            if (inputFixar.checked) {
                aplicarMesControleFiscalNosFiltros();
            }
            salvarPreferenciaMesControleFiscal();
        });
    }
}

(async function initControleFiscal(){
    if (!token) {
        window.location.href = "/";
        return;
    }

    showUserControleFiscal();
    await carregarCombosControleFiscal();
    carregarPreferenciaMesControleFiscal();
    bindEventosMesControleFiscal();
})();
</script>
</body>
</html>"""



@app.get("/inteligencia-tributaria", response_class=HTMLResponse)
def inteligencia_tributaria_page():
    return """<!DOCTYPE html>
<html lang="pt-br">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1.0" />
<title>Inteligência Tributária ERP Senior</title>
<style>
    :root{--bg:#081226;--card:#0d1b34;--line:#233a64;--text:#eaf1ff;--muted:#9bb0d3;--primary:#3b82f6;--ok:#22c55e;--warn:#eab308;--danger:#ef4444;}
    *{box-sizing:border-box}
    body{margin:0;font-family:Arial,Helvetica,sans-serif;background:linear-gradient(180deg,#07101f 0%,#081226 100%);color:var(--text);}
    .wrap{max-width:1700px;margin:0 auto;padding:24px}
    .title{font-size:34px;font-weight:700;margin-bottom:6px}
    .sub{color:var(--muted);margin-bottom:20px}
    .card{background:rgba(13,27,52,.96);border:1px solid var(--line);border-radius:18px;padding:18px;box-shadow:0 12px 30px rgba(0,0,0,.18);}
    .topbar{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:18px;}
    .userbox{display:flex;gap:10px;align-items:center;flex-wrap:wrap}
    .chip{padding:8px 10px;border-radius:999px;border:1px solid var(--line);background:#0b1830;color:#cddbf4;font-size:12px;}
    .grid{display:grid;gap:14px}
    .row{display:grid;grid-template-columns:repeat(4,1fr);gap:14px}
    label{display:block;color:var(--muted);font-size:13px;margin-bottom:6px}
    input,select,textarea{width:100%;background:#07101f;color:var(--text);border:1px solid var(--line);border-radius:12px;padding:12px 14px;outline:none;font-family:inherit;}
    textarea{min-height:110px;resize:vertical}
    .btn{background:var(--primary);color:white;border:none;border-radius:12px;padding:12px 16px;cursor:pointer;font-weight:700;}
    .btn.secondary{background:#243755}
    .btn.ghost{background:#13213d}
    .actions{display:flex;gap:10px;flex-wrap:wrap;margin-top:8px}
    .cards{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-top:18px}
    .metric{background:rgba(9,19,40,.95);border:1px solid var(--line);border-radius:18px;padding:16px;}
    .metric .k{color:var(--muted);font-size:13px;margin-bottom:8px}
    .metric .v{font-size:28px;font-weight:800}
    .badge{display:inline-block;padding:6px 10px;border-radius:999px;font-size:12px;font-weight:700;border:1px solid transparent;}
    .badge.blue{background:rgba(59,130,246,.15);color:#93c5fd;border-color:rgba(59,130,246,.35)}
    .badge.green{background:rgba(34,197,94,.15);color:#86efac;border-color:rgba(34,197,94,.35)}
    .badge.yellow{background:rgba(234,179,8,.15);color:#fde047;border-color:rgba(234,179,8,.35)}
    .badge.red{background:rgba(239,68,68,.15);color:#fca5a5;border-color:rgba(239,68,68,.35)}
    .panel{margin-top:18px;display:grid;grid-template-columns:1fr 1fr;gap:16px}
    .muted{color:var(--muted)}
    .small{font-size:12px}
    .table-wrap{overflow:auto;border:1px solid var(--line);border-radius:16px;margin-top:12px}
    table{width:100%;border-collapse:collapse;font-size:13px}
    th,td{padding:12px 10px;border-bottom:1px solid rgba(35,58,100,.7);text-align:left;vertical-align:top}
    th{color:#afc2e4;font-size:12px;background:#0d1b34}
    .section-title{font-size:20px;font-weight:700;margin-bottom:8px}
    .kv-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:12px;margin-top:12px}
    .kv-card{border:1px solid var(--line);border-radius:14px;padding:12px;background:#0b1830}
    .kv-card div{font-size:13px;line-height:1.6}
    .history-list{display:grid;gap:10px;margin-top:12px}
    .history-item{border:1px solid var(--line);border-radius:12px;padding:12px;background:#0b1830;cursor:pointer;}
    .history-item:hover{background:#102041}
    .hidden{display:none}
    @media(max-width:1200px){.row,.cards,.panel,.kv-grid{grid-template-columns:1fr}}
</style>
</head>
<body>
<div class="wrap">
    <div class="topbar">
        <div>
            <div class="title">Inteligência Tributária</div>
            <div class="sub">Consulta fiscal por operação descrita e comparação com os dados atuais do ERP.</div>
        </div>
        <div class="userbox">
            <span id="userLabel" class="chip"></span>
            <button class="btn ghost" onclick="window.location.href='/'">Auditoria Tributária</button>
            <button class="btn secondary" onclick="window.location.href='/controle-fiscal-produtos'">Controle Fiscal Produtos</button>
            <button class="btn ghost" onclick="sair()">Sair</button>
        </div>
    </div>

    <div class="card">
        <div class="section-title">Consulta da Operação</div>
        <div class="row">
            <div><label>Código do produto</label><input id="codigo_produto" placeholder="Ex.: 104004" /></div>
            <div><label>UF origem</label><input id="uf_origem" placeholder="Ex.: SC" maxlength="2" /></div>
            <div><label>UF destino</label><input id="uf_destino" placeholder="Ex.: PR" maxlength="2" /></div>
            <div><label>Transação ERP</label><input id="transacao" placeholder="Ex.: 5101" /></div>
        </div>
        <div class="row" style="margin-top:14px">
            <div>
                <label>Tipo do cliente</label>
                <select id="tipo_cliente">
                    <option value="">Selecione</option>
                    <option value="PJ_CONTRIBUINTE">Pessoa jurídica contribuinte</option>
                    <option value="PJ_NAO_CONTRIBUINTE">Pessoa jurídica não contribuinte</option>
                    <option value="PRODUTOR_RURAL">Produtor rural</option>
                    <option value="CONSUMIDOR_FINAL">Consumidor final</option>
                </select>
            </div>
            <div>
                <label>Finalidade</label>
                <select id="finalidade">
                    <option value="">Selecione</option>
                    <option value="REVENDA">Revenda</option>
                    <option value="USO_CONSUMO">Uso e consumo</option>
                    <option value="ATIVO_IMOBILIZADO">Ativo imobilizado</option>
                    <option value="INDUSTRIALIZACAO">Industrialização</option>
                    <option value="CONSERTO">Conserto</option>
                    <option value="DEMONSTRACAO">Demonstração</option>
                    <option value="REMESSA">Remessa</option>
                    <option value="RETORNO">Retorno</option>
                </select>
            </div>
            <div><label>Família</label><input id="familia" list="listaFamilias" placeholder="Ex.: BR-CHA" /><datalist id="listaFamilias"></datalist></div>
            <div><label>Origem</label><input id="origem" list="listaOrigens" placeholder="Ex.: 250" /><datalist id="listaOrigens"></datalist></div>
        </div>
        <div style="margin-top:14px">
            <label>Descrição da operação</label>
            <textarea id="descricao_operacao" placeholder="Ex.: Venda de produção de máquina agrícola para pessoa jurídica contribuinte de ICMS interno para revenda"></textarea>
        </div>
        <div class="actions">
            <button class="btn" onclick="analisarOperacao()">Analisar Operação</button>
            <button class="btn secondary" onclick="limparTela()">Limpar</button>
        </div>
    </div>

    <div class="cards">
        <div class="metric"><div class="k">Status da consulta</div><div class="v" id="m_status">-</div></div>
        <div class="metric"><div class="k">Total itens ERP</div><div class="v" id="m_total_erp">0</div></div>
        <div class="metric"><div class="k">CFOP sugerido</div><div class="v" id="m_cfop">-</div></div>
        <div class="metric"><div class="k">CST sugerido</div><div class="v" id="m_cst">-</div></div>
    </div>

    <div class="panel">
        <div class="card"><div class="section-title">Resumo da Operação</div><div id="resumo_operacao" class="muted">Nenhuma análise executada.</div></div>
        <div class="card"><div class="section-title">Sugestão Tributária</div><div id="sugestao_tributaria" class="muted">Aguardando análise.</div></div>
    </div>

    <div class="panel">
        <div class="card"><div class="section-title">ERP encontrado</div><div id="erp_encontrado" class="muted">Nenhum dado carregado.</div></div>
        <div class="card"><div class="section-title">Alertas</div><div id="alertas" class="muted">Sem alertas.</div></div>
    </div>

    <div class="card" style="margin-top:18px">
        <div class="section-title">Comparativo ERP x Sugestão</div>
        <div class="table-wrap">
            <table>
                <thead><tr><th>Campo</th><th>ERP</th><th>Sugestão</th><th>Divergente</th></tr></thead>
                <tbody id="tbodyComparativo"><tr><td colspan="4" class="muted">Nenhum comparativo disponível.</td></tr></tbody>
            </table>
        </div>
    </div>

    <div class="card" style="margin-top:18px">
        <div class="section-title">Histórico de Consultas</div>
        <div id="historyList" class="history-list"><div class="muted">Nenhuma consulta realizada.</div></div>
    </div>
</div>

<script>
let token = localStorage.getItem("token") || "";
let usuarioLogado = localStorage.getItem("usuario") || "";
const HISTORY_KEY = "historico_consulta_tributaria_ia";

function sair(){
    localStorage.removeItem("token");
    localStorage.removeItem("usuario");
    window.location.href = "/";
}

function apiHeaders(extra={}){
    const headers = {...extra};
    if(token) headers["Authorization"] = "Bearer " + token;
    return headers;
}

async function apiFetch(url, options={}){
    const response = await fetch(url, {...options, headers: apiHeaders(options.headers || {})});
    const raw = await response.text();
    let data;
    try { data = raw ? JSON.parse(raw) : {}; } catch { data = { detail: raw || ("HTTP " + response.status) }; }
    if(response.status === 401){ sair(); throw new Error(data.detail || "Sessão expirada"); }
    if(!response.ok) throw new Error(data.detail || ("Erro " + response.status));
    return data;
}

async function carregarCombos(){
    try{
        const [familias, origens] = await Promise.all([
            apiFetch("/api/familias?limite=200"),
            apiFetch("/api/origens?limite=200")
        ]);
        document.getElementById("listaFamilias").innerHTML = familias.map(f => `<option value="${f.codigo}">${f.label}</option>`).join("");
        document.getElementById("listaOrigens").innerHTML = origens.map(o => `<option value="${o.codigo}">${o.label}</option>`).join("");
    }catch(e){ console.error(e); }
}

function limparTela(){
    ["codigo_produto","uf_origem","uf_destino","transacao","familia","origem","descricao_operacao"]
        .forEach(id => document.getElementById(id).value = "");
    document.getElementById("tipo_cliente").value = "";
    document.getElementById("finalidade").value = "";
    document.getElementById("m_status").innerText = "-";
    document.getElementById("m_total_erp").innerText = "0";
    document.getElementById("m_cfop").innerText = "-";
    document.getElementById("m_cst").innerText = "-";
    document.getElementById("resumo_operacao").innerHTML = '<span class="muted">Nenhuma análise executada.</span>';
    document.getElementById("sugestao_tributaria").innerHTML = '<span class="muted">Aguardando análise.</span>';
    document.getElementById("erp_encontrado").innerHTML = '<span class="muted">Nenhum dado carregado.</span>';
    document.getElementById("alertas").innerHTML = '<span class="muted">Sem alertas.</span>';
    document.getElementById("tbodyComparativo").innerHTML = '<tr><td colspan="4" class="muted">Nenhum comparativo disponível.</td></tr>';
}

function salvarHistorico(payload){
    const atual = JSON.parse(localStorage.getItem(HISTORY_KEY) || "[]");
    const novo = [{datahora: new Date().toLocaleString("pt-BR"), ...payload}, ...atual].slice(0, 15);
    localStorage.setItem(HISTORY_KEY, JSON.stringify(novo));
    renderHistorico();
}

function renderHistorico(){
    const lista = JSON.parse(localStorage.getItem(HISTORY_KEY) || "[]");
    const el = document.getElementById("historyList");
    if(!lista.length){ el.innerHTML = '<div class="muted">Nenhuma consulta realizada.</div>'; return; }
    el.innerHTML = lista.map((item, idx) => `
        <div class="history-item" onclick="reaplicarHistorico(${idx})">
            <div style="font-weight:700">${item.descricao_operacao || "-"}</div>
            <div class="muted small" style="margin-top:6px">
                ${item.datahora} | Produto: ${item.codigo_produto || "-"} | ${item.uf_origem || "-"} â†’ ${item.uf_destino || "-"}
            </div>
        </div>`).join("");
}

function reaplicarHistorico(idx){
    const lista = JSON.parse(localStorage.getItem(HISTORY_KEY) || "[]");
    const item = lista[idx];
    if(!item) return;
    document.getElementById("codigo_produto").value = item.codigo_produto || "";
    document.getElementById("descricao_operacao").value = item.descricao_operacao || "";
    document.getElementById("uf_origem").value = item.uf_origem || "";
    document.getElementById("uf_destino").value = item.uf_destino || "";
    document.getElementById("tipo_cliente").value = item.tipo_cliente || "";
    document.getElementById("finalidade").value = item.finalidade || "";
    document.getElementById("transacao").value = item.transacao || "";
    document.getElementById("familia").value = item.familia || "";
    document.getElementById("origem").value = item.origem || "";
}

function badgeStatus(status){
    if(status === "ADERENTE") return '<span class="badge green">ADERENTE</span>';
    if(status === "DIVERGENTE") return '<span class="badge red">DIVERGENTE</span>';
    if(status === "REGRA_SEM_ERP") return '<span class="badge yellow">REGRA SEM ERP</span>';
    if(status === "SEM_REGRA") return '<span class="badge blue">SEM REGRA</span>';
    return '<span class="badge blue">' + (status || "-") + '</span>';
}

async function analisarOperacao(){
    const payload = {
        codigo_produto: document.getElementById("codigo_produto").value.trim(),
        descricao_operacao: document.getElementById("descricao_operacao").value.trim(),
        uf_origem: document.getElementById("uf_origem").value.trim().toUpperCase(),
        uf_destino: document.getElementById("uf_destino").value.trim().toUpperCase(),
        tipo_cliente: document.getElementById("tipo_cliente").value,
        finalidade: document.getElementById("finalidade").value,
        transacao: document.getElementById("transacao").value.trim(),
        familia: document.getElementById("familia").value.trim(),
        origem: document.getElementById("origem").value.trim(),
        limite_erp: 20
    };

    document.getElementById("resumo_operacao").innerHTML = '<span class="muted">Analisando...</span>';
    document.getElementById("sugestao_tributaria").innerHTML = '<span class="muted">Consultando motor tributário...</span>';
    document.getElementById("erp_encontrado").innerHTML = '<span class="muted">Consultando ERP...</span>';
    document.getElementById("alertas").innerHTML = '<span class="muted">Processando...</span>';
    document.getElementById("tbodyComparativo").innerHTML = '<tr><td colspan="4" class="muted">Gerando comparativo...</td></tr>';

    try{
        const data = await apiFetch("/api/consulta-tributaria-ia", {
            method: "POST",
            headers: {"Content-Type": "application/json"},
            body: JSON.stringify(payload)
        });

        salvarHistorico(payload);

        document.getElementById("m_status").innerHTML = badgeStatus(data.status_consulta);
        document.getElementById("m_total_erp").innerText = data.erp_encontrado?.total_itens || 0;
        document.getElementById("m_cfop").innerText = data.sugestao_tributaria?.cfop || "-";
        document.getElementById("m_cst").innerText = data.sugestao_tributaria?.cst_icms || "-";

        const interp = data.interpretacao || {};
        document.getElementById("resumo_operacao").innerHTML = `
            <div class="kv-grid">
                <div class="kv-card"><div><b>Tipo da operação:</b> ${interp.tipo_operacao || "-"}</div></div>
                <div class="kv-card"><div><b>Natureza:</b> ${interp.natureza || "-"}</div></div>
                <div class="kv-card"><div><b>Movimento:</b> ${interp.movimento || "-"}</div></div>
                <div class="kv-card"><div><b>Confiança:</b> ${interp.confianca_interpretacao || "-"}</div></div>
                <div class="kv-card"><div><b>UF origem:</b> ${interp.uf_origem || "-"}</div></div>
                <div class="kv-card"><div><b>UF destino:</b> ${interp.uf_destino || "-"}</div></div>
                <div class="kv-card"><div><b>Tipo cliente:</b> ${interp.tipo_cliente || "-"}</div></div>
                <div class="kv-card"><div><b>Finalidade:</b> ${interp.finalidade || "-"}</div></div>
            </div>`;

        const sug = data.sugestao_tributaria || {};
        document.getElementById("sugestao_tributaria").innerHTML = `
            <div class="kv-grid">
                <div class="kv-card"><div><b>CST ICMS:</b> ${sug.cst_icms || "-"}</div></div>
                <div class="kv-card"><div><b>CFOP:</b> ${sug.cfop || "-"}</div></div>
                <div class="kv-card"><div><b>Benefício fiscal:</b> ${sug.beneficio_fiscal || "-"}</div></div>
                <div class="kv-card"><div><b>Fonte:</b> ${sug.fonte_regra || "-"}</div></div>
                <div class="kv-card" style="grid-column:1 / -1"><div><b>Base legal:</b> ${sug.base_legal || "-"}</div></div>
                <div class="kv-card" style="grid-column:1 / -1"><div><b>Observações:</b> ${sug.observacoes || "-"}</div></div>
            </div>`;

        const erp = data.erp_encontrado?.item_referencia || {};
        document.getElementById("erp_encontrado").innerHTML = `
            <div class="kv-grid">
                <div class="kv-card"><div><b>Produto:</b> ${erp.codigo_item || "-"}</div></div>
                <div class="kv-card"><div><b>Descrição:</b> ${erp.descricao_item || "-"}</div></div>
                <div class="kv-card"><div><b>Família:</b> ${erp.familia_codigo || "-"} ${erp.familia_descricao || ""}</div></div>
                <div class="kv-card"><div><b>Origem:</b> ${erp.origem_codigo || "-"} ${erp.origem_descricao || ""}</div></div>
                <div class="kv-card"><div><b>Transação:</b> ${erp.transacao || "-"}</div></div>
                <div class="kv-card"><div><b>CFOP:</b> ${erp.cfop || "-"}</div></div>
                <div class="kv-card"><div><b>NCM:</b> ${erp.ncm || "-"}</div></div>
                <div class="kv-card"><div><b>CEST:</b> ${erp.cest || "-"}</div></div>
                <div class="kv-card"><div><b>CodTrd:</b> ${erp.cad_codtrd || "-"}</div></div>
                <div class="kv-card"><div><b>CST ICMS:</b> ${erp.item_cst_icms || "-"}</div></div>
                <div class="kv-card"><div><b>Status auditoria:</b> ${erp.status_auditoria || "-"}</div></div>
                <div class="kv-card"><div><b>Fonte prioritária:</b> ${erp.fonte_prioritaria || "-"}</div></div>
            </div>`;

        const alertas = data.alertas || [];
        document.getElementById("alertas").innerHTML = alertas.length
            ? alertas.map(a => `<div class="badge yellow" style="display:block;margin-bottom:8px">${a}</div>`).join("")
            : '<span class="badge green">Sem alertas relevantes</span>';

        const comp = data.comparativo_erp_vs_sugestao || [];
        document.getElementById("tbodyComparativo").innerHTML = comp.length
            ? comp.map(l => `<tr>
                <td>${l.campo || ""}</td>
                <td>${l.erp ?? "-"}</td>
                <td>${l.sugestao ?? "-"}</td>
                <td>${l.divergente ? '<span class="badge red">SIM</span>' : '<span class="badge green">NÁƒO</span>'}</td>
              </tr>`).join("")
            : '<tr><td colspan="4" class="muted">Nenhum comparativo disponível.</td></tr>';

    }catch(e){
        document.getElementById("resumo_operacao").innerHTML = '<span class="badge red">' + e.message + '</span>';
        document.getElementById("sugestao_tributaria").innerHTML = '<span class="muted">Falha na consulta.</span>';
        document.getElementById("erp_encontrado").innerHTML = '<span class="muted">Falha na consulta.</span>';
        document.getElementById("alertas").innerHTML = '<span class="badge red">' + e.message + '</span>';
        document.getElementById("tbodyComparativo").innerHTML = '<tr><td colspan="4" class="muted">Falha ao montar comparativo.</td></tr>';
    }
}

(function init(){
    if(!token){ window.location.href = "/"; return; }
    document.getElementById("userLabel").innerText = usuarioLogado || "Usuário";
    carregarCombos();
    renderHistorico();
})();
</script>
</body>
</html>"""


# =========================================================
# RUN
# =========================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host=API_HOST, port=API_PORT)
